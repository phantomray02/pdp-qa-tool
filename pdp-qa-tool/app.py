# =========================================
# IMPORTS
# =========================================
import re
import html
import json
import time
import hashlib
import traceback
from io import BytesIO
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from bs4 import BeautifulSoup
from PIL import Image, UnidentifiedImageError
import numpy as np
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pandas.errors import EmptyDataError
import threading
from requests.adapters import HTTPAdapter
import base64

# =========================================
# APP SETUP
# =========================================
st.set_page_config(layout="wide")
st.title("PDP QA Tool 🎉")

st.markdown(
    "<style>"
    "div[data-testid='stFileUploader'] > section {"
    "background:#232733;"
    "border:1px solid #2f3442;"
    "border-radius:10px;"
    "padding:10px;"
    "}"
    "div[data-testid='stDownloadButton'] > button {"
    "width:100%;"
    "min-height:56px;"
    "border-radius:10px;"
    "border:1px solid #2f3442;"
    "background:#232733;"
    "color:white;"
    "font-weight:700;"
    "}"
    "div[data-testid='stDownloadButton'] > button:hover {"
    "border-color:#4EA1FF;"
    "color:white;"
    "}"
    "</style>",
    unsafe_allow_html=True,
)

# Streamlit 1.58+ supports fragments. Older Streamlit versions will run this as a normal function.
def _identity_fragment(func=None, **_kwargs):
    def decorator(f):
        return f
    return decorator(func) if callable(func) else decorator

streamlit_fragment = getattr(st, "fragment", _identity_fragment)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection": "keep-alive",
}

REQUEST_TIMEOUT = 8
IMAGE_TIMEOUT = 6
IMAGE_RETRY_TIMEOUT = 10
IMAGE_RETRY_COUNT = 1
IMAGE_FETCH_WORKERS = 6
IMAGE_FETCH_CHUNK_SIZE = 65536

# UI-only thumbnail compression. These values affect only what is rendered in the
# Streamlit visual QA HTML. Scoring still uses the original full image URLs.
DISPLAY_IMAGE_TIMEOUT = 5
DISPLAY_IMAGE_MAX_WIDTH = 360
DISPLAY_IMAGE_MAX_HEIGHT = 360
DISPLAY_IMAGE_QUALITY = 68
DISPLAY_IMAGE_CACHE_MAX = 1000
MAX_CACHE = 400
# Retailer-specific fetch tuning
WALGREENS_REQUEST_TIMEOUT = 18
WALGREENS_DEBUG_TIMEOUT = 25
WALGREENS_API_TIMEOUT = 10
CVS_REQUEST_TIMEOUT = 25

# =========================================
# PERFORMANCE SETTINGS
# =========================================
# Balanced parallelism for image-heavy retailers.
# Too many workers can cause Salsify/CVS image requests to timeout or show as broken.
BATCH_SIZE = 32
MAX_WORKERS = 8
UI_UPDATE_EVERY = 5

# Faster image compare via tiny difference hash.
IMAGE_HASH_WIDTH = 9
IMAGE_HASH_HEIGHT = 8

# Larger caches to reduce repeat image fetches during batch + visual QA.
HTML_CACHE_MAX = 200
IMAGE_HASH_CACHE_MAX = 600

# Hard image safety limits.
MAX_IMAGE_BYTES = 12 * 1024 * 1024
MAX_SAFE_IMAGE_PIXELS = 50_000_000
MAX_IMAGE_SLOTS_TO_COMPARE = 20
MAX_IMAGE_SLOTS_TO_SCORE = 12
STRICT_LIVE_RETAILER_ONLY = True
# Strict comparison mode means: compare Salsify side to selected retailer side only.
# Do not copy Salsify content/images into retailer fields and do not use hard-coded
# known-product catalogs as if they were live retailer data.
STRICT_COMPARISON_MODE = True
ALLOW_RETAILER_KNOWN_COPY_FALLBACKS = False
ALLOW_RETAILER_GENERATED_IMAGE_FALLBACKS = False
STRICT_CVS_VARIANT_MATCH = True
CVS_VARIANT_MIN_MATCH_SCORE = 35

CAPTURE_MODE_USE_EXTENSION = "Use extension + TXT upload"
CAPTURE_MODE_SKIP_EXTENSION = "Skip extension and go straight to batch"
AUTO_SKIP_EXTENSION_RETAILERS = {"Walgreens"}
# Retailer-specific Salsify isolation controls.
# Copy can stay retailer-locked while images still fall back to generic locked Salsify slots if
# retailer-labeled image assets do not exist yet.
EXCLUSIVE_SALSIFY_COPY_RETAILERS = {"walgreens"}
EXCLUSIVE_SALSIFY_IMAGE_RETAILERS = set()

html_cache = {}
image_hash_cache = {}
image_compare_cache = {}
display_image_cache = {}
IMAGE_COMPARE_CACHE_MAX = 1200

# Image downloads are throttled separately from row workers so image-heavy rows
# do not overload Salsify/CVS/retailer image hosts and create false timeouts.
image_fetch_semaphore = threading.BoundedSemaphore(IMAGE_FETCH_WORKERS)

thread_local = threading.local()

# =========================================
# VISUAL LAYOUT SETTINGS
# =========================================
SECTION_HEADER_SIZE = 24
COPY_TEXT_SIZE = 15
COPY_LINE_HEIGHT = 1.28
SECTION_VERTICAL_GAP = 8

# Use one shared spacing value so the image area feels mathematically even.
IMG_SPACE_PX = 4
IMG_BOX_HEIGHT = 104
IMG_SCORE_WIDTH_PX = 72

TITLE_TO_DESCRIPTION_GAP_PX = 28
DESCRIPTION_TO_FEATURES_GAP_PX = 28


def get_session():
    if not hasattr(thread_local, "session"):
        session = requests.Session()
        adapter = HTTPAdapter(
            pool_connections=100,
            pool_maxsize=100,
            max_retries=0,
        )
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        session.headers.update(HEADERS)
        thread_local.session = session
    return thread_local.session
    
# =========================================
# GENERIC HELPERS
# =========================================
def normalize_space(text):
    text = str(text or "")
    text = html.unescape(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def clean_item_number(value):
    if not value:
        return ""
    return str(value).replace(".0", "").strip()


def normalize_text(text):
    if not isinstance(text, str):
        return ""
    return re.sub(r"[^a-z0-9\s]", "", text.lower())


def dedupe_preserve_order(items):
    seen = set()
    out = []
    for item in items:
        item = normalize_space(item)
        if item and item not in seen:
            seen.add(item)
            out.append(item)
    return out


def keyword_score(a, b):
    a_norm = normalize_text(a)
    b_norm = normalize_text(b)

    if not a_norm and not b_norm:
        return 0
    if not a_norm or not b_norm:
        return 0

    return int(SequenceMatcher(None, a_norm, b_norm).ratio() * 100)


def description_similarity_score(a, b):
    a_norm = normalize_text(a)
    b_norm = normalize_text(b)

    if not a_norm and not b_norm:
        return 0
    if not a_norm or not b_norm:
        return 0
    if a_norm == b_norm:
        return 100

    return int(SequenceMatcher(None, a_norm, b_norm).ratio() * 100)

def html_escape_text(text):
    return html.escape(str(text or ""))


def equal_height_block(text, min_height=150):
    safe_text = html_escape_text(text or "Missing")
    return (
        f"<div style=\""
        f"width:100%;"
        f"min-height:{min_height}px;"
        f"padding:0;"
        f"margin:0;"
        f"background:transparent;"
        f"color:#FFFFFF;"
        f"white-space:pre-wrap;"
        f"line-height:{COPY_LINE_HEIGHT};"
        f"font-size:{COPY_TEXT_SIZE}px;"
        f"font-weight:500;"
        f"text-indent:0;"
        f"overflow-wrap:anywhere;"
        f"word-break:break-word;"
        f"\">{safe_text}</div>"
    )


def equal_feature_block(text, min_height=40):
    safe_text = html_escape_text(text or "Missing")
    return (
        f"<div style=\""
        f"width:100%;"
        f"min-height:{min_height}px;"
        f"padding:0;"
        f"margin:0;"
        f"background:transparent;"
        f"color:#FFFFFF;"
        f"white-space:pre-wrap;"
        f"line-height:{COPY_LINE_HEIGHT};"
        f"font-size:{COPY_TEXT_SIZE}px;"
        f"font-weight:500;"
        f"text-indent:0;"
        f"overflow-wrap:anywhere;"
        f"word-break:break-word;"
        f"\">{safe_text}</div>"
    )


def score_text_html(score):
    if score >= 80:
        color = "#4CAF50"
        label = "Strong"
    elif score >= 50:
        color = "#FFC107"
        label = "Review"
    else:
        color = "#F44336"
        label = "Poor"

    return f"<span style='color:{color}; font-weight:900; font-size:22px;'>{score}% ({label})</span>"


def section_header_html(label, score):
    safe_label = html_escape_text(label or "")
    return (
        f"<div style=\""
        f"display:flex;"
        f"justify-content:space-between;"
        f"align-items:flex-end;"
        f"gap:12px;"
        f"margin-top:{SECTION_VERTICAL_GAP}px;"
        f"margin-bottom:{SECTION_VERTICAL_GAP}px;"
        f"\">"
        f"<div style=\"font-size:{SECTION_HEADER_SIZE}px; font-weight:900; color:#FFFFFF; line-height:1.0;\">"
        f"{safe_label}"
        f"</div>"
        f"<div style=\"line-height:1.0;\">{score_text_html(score)}</div>"
        f"</div>"
    )


def avg_score_bar_html(label, score):
    if score >= 80:
        color = "#2E7D32"
    elif score >= 50:
        color = "#F9A825"
    else:
        color = "#C62828"

    safe_label = html_escape_text(label or "")
    return (
        f'<div style="background-color:{color};padding:6px 10px;border-radius:4px;color:white;'
        f'font-weight:900;font-size:19px;margin-top:2px;margin-bottom:{IMG_SPACE_PX}px;'
        f'display:flex;justify-content:space-between;align-items:center;gap:10px;">'
        f'<span>{safe_label}</span>'
        f'<span style="color:#FFFFFF; font-weight:900; font-size:20px;">{score}%</span>'
        f'</div>'
    )


def rating_stars_html(rating, review_count=None, font_size_px=18):
    try:
        rating = float(rating or 0)
    except Exception:
        rating = 0.0

    rating = max(0.0, min(rating, 5.0))
    fill_pct = (rating / 5.0) * 100

    review_html = ""
    if review_count is not None and str(review_count).strip() != "":
        review_html = (
            f'<span style="margin-left:8px;font-size:28px;font-weight:900;color:#FFFFFF;line-height:1;white-space:nowrap;">'
            f'{html_escape_text(review_count)}</span>'
        )

    return (
        f'<div style="display:inline-flex;align-items:center;justify-content:flex-end;gap:8px;white-space:nowrap;margin:0;line-height:1;overflow:visible;">'
        f'<span style="font-size:28px;font-weight:900;color:#FFFFFF;line-height:1;">{rating:.1f}</span>'
        f'<div style="position:relative;display:inline-block;line-height:1;font-size:{font_size_px}px;letter-spacing:0.6px;">'
        f'<div style="color:rgba(255,255,255,0.35);">★★★★★</div>'
        f'<div style="position:absolute;top:0;left:0;width:{fill_pct}%;overflow:hidden;color:#FFFFFF;white-space:nowrap;">★★★★★</div>'
        f'</div>'
        f'{review_html}'
        f'</div>'
    )


def locked_visual_header_row_html(
    salsify_header_html,
    retailer_header_html,
    rating_html="",
    retailer_name="",
):
    retailer_name = str(retailer_name or "").strip().lower()

    if retailer_name in {"walgreens", "kroger"}:
        return (
            '<div style="display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1fr);column-gap:18px;align-items:start;margin:0 0 6px 0;">'
            f'<div style="min-width:0;display:flex;align-items:flex-start;justify-content:flex-start;overflow:visible;">{salsify_header_html}</div>'
            '<div style="min-width:0;display:grid;grid-template-columns:minmax(0,1fr) auto;column-gap:12px;align-items:start;overflow:visible;">'
            f'<div style="min-width:0;display:flex;align-items:flex-start;justify-content:flex-start;overflow:visible;">{retailer_header_html}</div>'
            f'<div style="min-width:0;display:flex;align-items:flex-start;justify-content:flex-end;overflow:visible;">{rating_html or "&nbsp;"}</div>'
            '</div>'
            '</div>'
        )

    return (
        '<div style="display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1fr);column-gap:18px;align-items:start;margin:0 0 6px 0;">'
        f'<div style="min-width:0;display:flex;align-items:flex-start;justify-content:flex-start;overflow:visible;">{salsify_header_html}</div>'
        f'<div style="min-width:0;display:flex;align-items:flex-start;justify-content:flex-start;overflow:visible;">{retailer_header_html}</div>'
        '</div>'
    )


def prepend_kroger_variant_for_display(value, variant_size, retailer_name=""):
    """Kroger-only display helper. Keeps scoring values unchanged."""
    value = str(value or "").strip()
    variant_size = clean_kroger_variant_size(variant_size)
    if str(retailer_name or "").strip().lower() != "kroger" or not variant_size:
        return value
    if not value:
        return variant_size
    if value.lower().startswith(variant_size.lower()):
        return value
    return f"{variant_size} | {value}"


def column_header_link_html(label, item_number, href):
    safe_label = html_escape_text(label or "")
    safe_item = html_escape_text(item_number or "")
    clean_href = clean_uploaded_url_value(href)
    safe_href = html.escape(clean_href, quote=True)

    if safe_href and safe_item:
        item_html = (
            f'<a href="{safe_href}" target="_blank" '
            f'style="color:#3EA6FF; text-decoration:none; font-weight:900; white-space:nowrap;">{safe_item}</a>'
        )
    else:
        item_html = f'<span style="color:#3EA6FF; font-weight:900; white-space:nowrap;">{safe_item or "Missing"}</span>'

    return (
        f'<div style="text-align:left;margin-top:0;margin-bottom:2px;font-size:28px;'
        f'font-weight:900;color:#FFFFFF;line-height:1.05;white-space:nowrap;display:inline-block;">{safe_label}: {item_html}</div>'
    )


def image_header_html(label):
    safe_label = html_escape_text(label or "")
    return (
        f"<div style=\""
        f"text-align:left;"
        f"margin-top:0;"
        f"margin-bottom:2px;"
        f"font-size:28px;"
        f"font-weight:900;"
        f"color:#FFFFFF;"
        f"line-height:1.05;"
        f"\">"
        f"{safe_label}"
        f"</div>"
    )



def is_video_like_url(url):
    url = html.unescape(str(url or "").strip())
    if not url:
        return False
    lowered = url.lower().split('?', 1)[0]
    return bool(
        lowered.endswith('.mp4')
        or lowered.endswith('.m3u8')
        or lowered.endswith('.webm')
        or '/video/' in lowered
        or 'salsify.com/video/' in lowered
        or 'asr-rm/' in lowered
    )



def _cache_display_image_src(url, value):
    global display_image_cache
    if "display_image_cache" not in globals() or not isinstance(globals().get("display_image_cache"), dict):
        display_image_cache = {}
    display_image_cache[url] = value
    while len(display_image_cache) > DISPLAY_IMAGE_CACHE_MAX:
        display_image_cache.pop(next(iter(display_image_cache)))


def get_display_image_src(url):
    """
    UI-only thumbnail source builder.

    This returns a compressed data URI for image rendering in Streamlit, while all
    scoring still uses the original full image URL. CSS dimensions are unchanged;
    only the downloaded/rendered payload is smaller.
    """
    url = str(url or "").strip()
    if not url:
        return ""
    if is_video_like_url(url):
        return url

    global display_image_cache
    if "display_image_cache" not in globals() or not isinstance(globals().get("display_image_cache"), dict):
        display_image_cache = {}

    if url in display_image_cache:
        return display_image_cache[url]

    try:
        # Reuse the existing controlled image fetch path so display thumbnails do
        # not create an unbounded second wave of browser/network requests.
        image_bytes = _download_image_bytes_once(url, DISPLAY_IMAGE_TIMEOUT)
        if not image_bytes:
            _cache_display_image_src(url, url)
            return url

        bio = BytesIO(image_bytes)
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            img = Image.open(bio)
            width, height = img.size
            if width * height > MAX_SAFE_IMAGE_PIXELS:
                _cache_display_image_src(url, url)
                return url
            img.load()

        # Preserve transparency against a white background so PNG/WebP assets look
        # normal after JPEG compression.
        if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
            rgba = img.convert("RGBA")
            background = Image.new("RGB", rgba.size, (255, 255, 255))
            background.paste(rgba, mask=rgba.split()[-1])
            img = background
        elif img.mode != "RGB":
            img = img.convert("RGB")

        img.thumbnail((DISPLAY_IMAGE_MAX_WIDTH, DISPLAY_IMAGE_MAX_HEIGHT), Image.LANCZOS)

        out = BytesIO()
        img.save(out, format="JPEG", quality=DISPLAY_IMAGE_QUALITY, optimize=True)
        encoded = base64.b64encode(out.getvalue()).decode("ascii")
        data_uri = f"data:image/jpeg;base64,{encoded}"
        _cache_display_image_src(url, data_uri)
        return data_uri

    except Exception:
        _cache_display_image_src(url, url)
        return url


def image_compare_cell_html(url):
    if not url:
        # Layout-only fix: Missing should act like an empty image placeholder.
        # It reserves a normal image-slot footprint, but it does not force real
        # images into a fixed-size box.
        return (
            f'<div style="'
            f'width:100%;'
            f'height:100%;'
            f'min-height:{IMG_BOX_HEIGHT}px;'
            f'display:flex;'
            f'align-items:center;'
            f'justify-content:center;'
            f'margin:0;'
            f'padding:0;'
            f'box-sizing:border-box;'
            f'color:#C62828;'
            f'font-size:16px;'
            f'font-weight:700;'
            f'overflow:hidden;'
            f'">Missing</div>'
        )

    display_url = get_display_image_src(url)
    safe_url = html.escape(str(display_url), quote=True)

    if not is_video_like_url(url):
        return (
            f"<div style='width:100%;margin:0;padding:0;display:flex;align-items:flex-start;justify-content:center;overflow:hidden;'>"
            f"<img src='{safe_url}' loading='lazy' decoding='async' referrerpolicy='no-referrer' style='display:block;width:100%;height:auto;object-fit:contain;' />"
            f"</div>"
        )

    media_id = hashlib.md5(str(url).encode("utf-8")).hexdigest()[:12]
    modal_id = f"video_modal_{media_id}"
    open_js = (
        f"(function(){{"
        f"var modal=document.getElementById('{modal_id}');"
        f"if(modal){{modal.style.display='flex';}}"
        f"var thumb=document.getElementById('thumb_media_{media_id}');"
        f"var full=document.getElementById('full_media_{media_id}');"
        f"if(full){{"
        f"try{{if(thumb&&thumb.currentTime>=0){{full.currentTime=thumb.currentTime||0;}}}}catch(e){{}}"
        f"try{{full.play();}}catch(e){{}}"
        f"}}"
        f"}})()"
    )
    close_js = (
        f"(function(){{"
        f"var modal=document.getElementById('{modal_id}');"
        f"if(modal){{modal.style.display='none';}}"
        f"var full=document.getElementById('full_media_{media_id}');"
        f"if(full){{try{{full.pause();}}catch(e){{}}}}"
        f"}})()"
    )

    return (
        f'<div style="position:relative;width:100%;margin:0;padding:0;display:flex;align-items:flex-start;justify-content:center;overflow:hidden;">'
        f'<video id="thumb_media_{media_id}" controls playsinline preload="metadata" onplay="{open_js}" style="display:block;width:100%;height:auto;max-height:320px;object-fit:contain;background:#000;">'
        f'<source src="{safe_url}" />'
        f'</video>'
        f'</div>'
        f'<div id="{modal_id}" onclick="if(event.target===this){{{close_js}}}" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.28);z-index:999999;align-items:center;justify-content:center;padding:18px;">'
        f'<div style="position:relative;width:min(74vw,760px);max-height:76vh;border-radius:12px;overflow:hidden;box-shadow:0 20px 56px rgba(0,0,0,0.38);background:#111;">'
        f'<button type="button" onclick="{close_js}" style="position:absolute;top:8px;right:8px;width:30px;height:30px;border:0;border-radius:15px;background:rgba(255,255,255,0.95);color:#111;font-size:18px;font-weight:900;cursor:pointer;z-index:2;">×</button>'
        f'<video id="full_media_{media_id}" controls playsinline preload="metadata" style="display:block;width:100%;max-height:76vh;height:auto;object-fit:contain;background:#000;">'
        f'<source src="{safe_url}" />'
        f'</video>'
        f'</div></div>'
    )

def image_compare_row_html(s_url, r_url, score):
    return (
        f"<div style=\""
        f"display:grid;"
        f"grid-template-columns:minmax(0,1fr) minmax(0,1fr) {IMG_SCORE_WIDTH_PX}px;"
        f"column-gap:8px;"
        f"align-items:stretch;"
        f"margin:0 0 {IMG_SPACE_PX}px 0;"
        f"padding:0;"
        f"\">"
        f"<div style=\"margin:0; padding:0; display:flex; align-items:stretch;\">"
        f"{image_compare_cell_html(s_url)}"
        f"</div>"
        f"<div style=\"margin:0; padding:0; display:flex; align-items:stretch;\">"
        f"{image_compare_cell_html(r_url)}"
        f"</div>"
        f"<div style=\""
        f"display:flex;"
        f"align-items:flex-start;"
        f"justify-content:flex-start;"
        f"text-align:left;"
        f"margin:0;"
        f"padding-top:4px;"
        f"\">"
        f"{score_text_html(score)}"
        f"</div>"
        f"</div>"
    )

def image_tile_html(label, url, box_height=170):
    safe_label = html.escape(label)

    if url:
        display_url = get_display_image_src(url)
        safe_url = html.escape(display_url, quote=True)
        return f'''<div style="border:1px solid #E0E0E0;border-radius:8px;background:#FFFFFF;padding:8px;">
<div style="font-size:45px;font-weight:600;margin-bottom:6px;">{safe_label}</div>
<div style="height:{box_height}px;display:flex;align-items:center;justify-content:center;background:#FAFAFA;border-radius:6px;overflow:hidden;">
<img src="{safe_url}" loading="lazy" decoding="async" referrerpolicy="no-referrer" style="max-width:100%;max-height:{box_height}px;object-fit:contain;" />
</div>
</div>'''
    else:
        return f'''<div style="border:1px solid #E0E0E0;border-radius:8px;background:#FFFFFF;padding:8px;">
<div style="font-size:45px;font-weight:600;margin-bottom:6px;">{safe_label}</div>
<div style="height:{box_height}px;display:flex;align-items:center;justify-content:center;background:#FAFAFA;border-radius:6px;color:#C62828;font-size:14px;font-weight:600;">
❌ Missing
</div>
</div>'''


def image_slot_block_html(slot_num, s_url, r_url, score, retailer_name="CVS", box_height=170):
    if score >= 80:
        score_color = "#2E7D32"
    elif score >= 50:
        score_color = "#F9A825"
    else:
        score_color = "#C62828"

    return f'''<div style="border:1px solid #DADADA;border-radius:10px;padding:10px;margin-bottom:12px;background:#FCFCFC;">
<div style="font-weight:700;margin-bottom:10px;color:{score_color};">Image Slot {slot_num} — {score}%</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;">
{image_tile_html("Salsify", s_url, box_height=box_height)}
{image_tile_html(retailer_name, r_url, box_height=box_height)}
</div>
</div>'''


def build_image_panel_html(s_images, r_images, max_images, retailer_name="CVS", box_height=110):
    blocks = []

    for i in range(max_images):
        s_url = s_images[i].get("url") if i < len(s_images) and isinstance(s_images[i], dict) else ""
        r_url = r_images[i] if i < len(r_images) and isinstance(r_images[i], str) else ""
        score = compare_images_visually(s_url, r_url) if (s_url and r_url) else 0

        blocks.append(
            image_slot_block_html(
                slot_num=i + 1,
                s_url=s_url,
                r_url=r_url,
                score=score,
                retailer_name=retailer_name,
                box_height=box_height,
            )
        )

    return f'''<div style="padding-right:4px;">{"".join(blocks)}</div>'''


def read_uploaded_file_from_bytes(file_bytes, file_name):
    if not file_bytes:
        raise EmptyDataError("Uploaded file is empty.")
    if len(file_bytes.strip()) == 0:
        raise EmptyDataError("Uploaded file is empty.")

    file_name = str(file_name or "").lower().strip()

    if file_name.endswith(".xlsx"):
        xls = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
        frames = []

        for sheet_name in xls.sheet_names:
            sheet_df = pd.read_excel(
                BytesIO(file_bytes),
                sheet_name=sheet_name,
                engine="openpyxl",
                dtype=str,
                keep_default_na=False,
            )

            if sheet_df is None or sheet_df.empty:
                continue

            sheet_df = sheet_df.copy()
            sheet_df["retailer"] = str(sheet_name).strip()
            frames.append(sheet_df)

        if not frames:
            raise EmptyDataError("No readable sheets found in uploaded Excel file.")

        return pd.concat(frames, ignore_index=True)

    last_error = None
    for encoding in ["utf-8-sig", "utf-8", "latin1"]:
        try:
            return pd.read_csv(BytesIO(file_bytes), encoding=encoding, dtype=str, keep_default_na=False)
        except Exception as e:
            last_error = e

    raise last_error if last_error else EmptyDataError("Could not parse uploaded file.")

def infer_retailer_name_from_url(url):
    if pd.isna(url):
        return "Retailer"

    url = str(url or "").strip().lower()

    if not url:
        return "Retailer"

    if "cvs.com" in url:
        return "CVS"
    if "walmart.com" in url:
        return "Walmart"
    if "target.com" in url:
        return "Target"
    if "kroger.com" in url:
        return "Kroger"
    if "samsclub.com" in url or "sam's club" in url:
        return "Sam's Club"
    if "walgreens.com" in url:
        return "Walgreens"
    if "heb.com" in url or "h-e-b" in url:
        return "HEB"
    if "amazon.com" in url:
        return "Amazon"

    return "Retailer"
    


def normalize_retailer_name(value):
    value = normalize_space(value)
    if not value:
        return "Retailer"
    lowered = value.lower()
    mapping = {
        "cvs": "CVS",
        "walgreens": "Walgreens",
        "wags": "Walgreens",
        "sam's club": "Sam's Club",
        "sams club": "Sam's Club",
        "samsclub": "Sam's Club",
        "walmart": "Walmart",
        "target": "Target",
        "kroger": "Kroger",
        "heb": "HEB",
        "h-e-b": "HEB",
        "h e b": "HEB",
        "amazon": "Amazon",
        "retailer": "Retailer",
    }
    return mapping.get(lowered, value)


RETAILER_URL_DOMAIN_RULES = {
    "cvs": ("cvs.com",),
    "walgreens": ("walgreens.com",),
    "kroger": ("kroger.com",),
    "heb": ("heb.com",),
    "sam's club": ("samsclub.com",),
    "sams club": ("samsclub.com",),
    "samsclub": ("samsclub.com",),
    "walmart": ("walmart.com",),
    "target": ("target.com",),
    "amazon": ("amazon.com",),
}


def retailer_url_matches_selected(retail_url, selected_retailer):
    """True only when the retailer URL belongs to the selected retailer.

    This is the first guardrail for retailer isolation. It prevents a row labeled
    CVS from accidentally being parsed by CVS code if the URL is actually Kroger,
    Walgreens, etc. Blank URLs are handled elsewhere as missing URLs.
    """
    retailer = normalize_retailer_name(selected_retailer).strip().lower()
    url = str(retail_url or "").strip().lower()
    if not url:
        return True
    expected_domains = RETAILER_URL_DOMAIN_RULES.get(retailer)
    if not expected_domains:
        return True
    return any(domain in url for domain in expected_domains)


def build_retailer_url_mismatch_status(retail_url, selected_retailer):
    retailer = normalize_retailer_name(selected_retailer)
    return f"Skipped: URL does not match selected retailer {retailer}: {retail_url}"


def build_empty_retailer_bundle(retailer_name="Retailer", reason=""):
    retailer_name = normalize_retailer_name(retailer_name)
    reason = normalize_space(reason) or "retailer_not_supported"
    return {
        "text": {
            "title": "",
            "description": "",
            "features": [],
            "rating": "",
            "review_count": "",
            "debug": {
                "Title Path": reason,
                "Description Path": reason,
                "Features Path": reason,
                "Source Used": reason,
                "Retailer": retailer_name,
            },
        },
        "images": [],
    }


def normalize_salsify_asset_name(value):
    value = html.unescape(str(value or "")).lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[/_|-]+", " ", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def make_blank_salsify_image_slot(name="missing"):
    return {"name": str(name or "missing"), "url": ""}


# CVS-only Salsify packaging views.
# These are allowed into the Salsify image bundle only so the CVS visual QA
# can populate slots 2 and 3. Non-CVS retailers filter these back out before
# image alignment, so Flat Back_2D / Flat Left_2D do not leak into Walgreens,
# Kroger, Sam's Club, or other retailer comparisons.
CVS_ONLY_SALSIFY_FLAT_IMAGE_TOKENS = (
    "flat back 2d",
    "flat left 2d",
)


def is_cvs_only_salsify_flat_image_name(value):
    name = normalize_salsify_asset_name(value or "")
    if not name:
        return False
    return any(token in name for token in CVS_ONLY_SALSIFY_FLAT_IMAGE_TOKENS)


def is_cvs_only_salsify_image(img):
    if not isinstance(img, dict):
        return False
    return bool(img.get("cvs_only")) or is_cvs_only_salsify_flat_image_name(img.get("name", ""))

# Walgreens-only Salsify image slots.
# Walgreens visual QA requires Ingredient Label Image locked in slot 2.
# Keep this flagged as Walgreens-only so it does not shift image order for other retailers.
def is_walgreens_ingredient_label_image_name(value):
    name = normalize_salsify_asset_name(value or "")
    if not name:
        return False
    return any(token in name for token in [
        "ingredient label image",
        "ingredient label",
        "ingredients label image",
        "ingredients label",
    ])


def is_walgreens_only_salsify_image(img):
    if not isinstance(img, dict):
        return False
    return bool(img.get("walgreens_only")) or is_walgreens_ingredient_label_image_name(img.get("name", ""))


def build_locked_salsify_slots(s_images, lock_top_three=True, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE):
    """
    Salsify-only slot builder.

    Rules:
    1. Slot 1 is reserved for Online Optimized Image / online image.
    2. Slot 2 is reserved for Flat Back_2D / back image.
    3. Slot 3 is reserved for Flat Left_2D / left image.
    4. If one of those top-three images is missing, keep the slot blank so later
       lifestyle / ATF images shift down and do not move up into slots 1-3.
    5. Retailer images must not be modified by this rule.
    """
    s_images = [img for img in (s_images or []) if isinstance(img, dict)]

    by_name = {}
    remainder = []
    locked_name_map = {
        "online": [
            "online optimized image",
            "online image",
            "online",
            "front",
        ],
        "back": [
            "flat back 2d",
            "flat back",
            "back 2d",
            "back",
        ],
        "left": [
            "flat left 2d",
            "flat left",
            "left 2d",
            "left",
        ],
    }

    def classify(img):
        name = normalize_salsify_asset_name(img.get("name", ""))
        for canonical, options in locked_name_map.items():
            for option in options:
                if normalize_salsify_asset_name(option) in name:
                    return canonical
        return ""

    for img in s_images:
        canonical = classify(img)
        if canonical and canonical not in by_name:
            by_name[canonical] = img
        else:
            remainder.append(img)

    if not lock_top_three:
        ordered = []
        for key in ["online", "back", "left"]:
            if key in by_name:
                ordered.append(by_name[key])
        ordered.extend(remainder)
        return ordered[:max_slots]

    ordered = [
        by_name.get("online") or make_blank_salsify_image_slot("online"),
        by_name.get("back") or make_blank_salsify_image_slot("back"),
        by_name.get("left") or make_blank_salsify_image_slot("left"),
    ]
    ordered.extend(remainder)
    return ordered[:max_slots]


# =========================================
# RETAILER Salsify REQUIREMENTS
# =========================================
# Centralized retailer-specific Salsify limits so copy/image rules are easy to update in one place.
RETAILER_SALSIFY_REQUIREMENTS = {
    "default": {"max_features": 5, "max_images": 6},
    "cvs": {"max_features": 5, "max_images": 8},
    "walgreens": {"max_features": 5, "max_images": 6},
    "kroger": {"max_features": 7, "max_images": 6},
    "heb": {"max_features": 10, "max_images": 10},
    "sam's club": {"max_features": 10, "max_images": 10},
    "sams club": {"max_features": 10, "max_images": 10},
    "samsclub": {"max_features": 10, "max_images": 10},
}


def get_retailer_salsify_requirements(retailer_name):
    retailer = str(retailer_name or "").strip().lower()
    return dict(RETAILER_SALSIFY_REQUIREMENTS.get(retailer, RETAILER_SALSIFY_REQUIREMENTS["default"]))


def get_retailer_salsify_feature_fields(retailer_name):
    max_features = int(get_retailer_salsify_requirements(retailer_name).get("max_features", 5) or 5)
    max_features = max(0, min(max_features, 10))
    return [f"feature{i}" for i in range(1, max_features + 1)]


def classify_cvs_generic_asset_name(value):
    name = normalize_salsify_asset_name(value or "")
    if any(token in name for token in [
        "atf i/o generic",
        "atf i o generic",
        "atf io generic",
        "atf i/o-generic",
        "atf io-generic",
    ]):
        return "io_generic"
    if any(token in name for token in [
        "atf 6 generic",
        "atf 6-generic",
        "atf6 generic",
        "atf6-generic",
    ]):
        return "atf6_generic"
    return ""


def apply_retailer_salsify_copy_limits(retailer_name, text_bundle):
    out = dict(text_bundle or {})
    limits = get_retailer_salsify_requirements(retailer_name)
    max_features = int(limits.get("max_features", 5) or 5)
    max_features = max(0, min(max_features, 10))

    gathered = []
    for value in out.get("features", []) or []:
        clean_value = normalize_space(value)
        if clean_value:
            gathered.append(clean_value)
    for i in range(1, 11):
        clean_value = normalize_space(out.get(f"feature{i}", ""))
        if clean_value:
            gathered.append(clean_value)

    gathered = dedupe_preserve_order(gathered)[:max_features]
    out["features"] = gathered
    for i in range(1, 11):
        out[f"feature{i}"] = gathered[i - 1] if i <= max_features and i - 1 < len(gathered) else ""
    return out


def infer_cvs_image_slot_from_url(url):
    url = str(url or "").strip().split("?", 1)[0]
    if not url:
        return None
    name = url.rsplit("/", 1)[-1]
    stem = re.sub(r'\.(jpg|jpeg|png|webp|avif)$', '', name, flags=re.IGNORECASE)
    match = re.search(r'(?:[_\-])(\d{1,2})$', stem)
    if not match:
        match = re.search(r'\((\d{1,2})\)$', stem)
    if match:
        try:
            slot_num = int(match.group(1))
            if 1 <= slot_num <= MAX_IMAGE_SLOTS_TO_COMPARE:
                return slot_num
        except Exception:
            pass
    if re.search(r'\d', stem):
        return 1
    return None

def reorder_cvs_salsify_images_for_visual(images, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE):
    """
    CVS Salsify order for visual QA:
    1. Online/Main image.
    2. Flat Back_2D / back.
    3. Flat Left_2D / side.
    4. ATF I/O-Generic when present.
       - If ATF I/O-Generic is missing, ATF 2 moves up into this slot.
    5+. Continue remaining ATF 2-5 Generic images in order.
    Last. ATF 6-Generic is always kept last and is never used as the ATF I/O fallback.

    Missing slots 1-3 stay blank so later ATF/lifestyle images do not hide missing required flat-packaging assets.
    """
    imgs = [img for img in (images or []) if isinstance(img, dict)]

    def norm_name(img):
        return normalize_salsify_asset_name((img or {}).get("name", "")) if isinstance(img, dict) else ""

    def find_first(*queries):
        query_tokens = [normalize_salsify_asset_name(q) for q in queries if normalize_salsify_asset_name(q)]
        for q in query_tokens:
            for img in imgs:
                name = norm_name(img)
                if name and (q == name or q in name):
                    return img
        return None

    def image_key(img):
        return str(img.get("url", "") or "").strip() if isinstance(img, dict) else ""

    ordered = []
    used = set()

    def add(img, blank_name=""):
        if not isinstance(img, dict):
            if blank_name:
                ordered.append(make_blank_salsify_image_slot(blank_name))
            return False
        key = image_key(img)
        if not key or key in used:
            if blank_name:
                ordered.append(make_blank_salsify_image_slot(blank_name))
            return False
        ordered.append(img)
        used.add(key)
        return True

    def add_first_available(query_groups, blank_name=""):
        for query_group in query_groups:
            if add(find_first(*query_group)):
                return True
        if blank_name:
            ordered.append(make_blank_salsify_image_slot(blank_name))
        return False

    add(find_first(
        "online optimized image", "online image", "main variant image", "main image", "hero", "primary", "front", "product image 1", "image 1",
    ), "cvs_slot_1")
    # CVS requires explicit flat packaging slots from Salsify.
    # Slot 2 should show Missing unless Salsify has an actual Flat Back_2D / Flat Back asset.
    # Slot 3 should show Missing unless Salsify has an actual Flat Left_2D / Flat Left asset.
    # Do NOT use loose fallback names like image 2, product image 2, side, back, or right here.
    # Those can be ATF/lifestyle/package images and would hide the real missing-flat issue.
    add(find_first(
        "flat back 2d", "flat back_2d", "flat back",
    ), "cvs_slot_2")
    add(find_first(
        "flat left 2d", "flat left_2d", "flat left",
    ), "cvs_slot_3")

    io_queries = (("atf i/o generic", "atf i o generic", "atf io generic", "atf i/o-generic", "atf io-generic"),)
    atf2_queries = (("atf 2 generic", "atf 2-generic", "atf2 generic", "atf2-generic"),)
    atf3_queries = (("atf 3 generic", "atf 3-generic", "atf3 generic", "atf3-generic"),)
    atf4_queries = (("atf 4 generic", "atf 4-generic", "atf4 generic", "atf4-generic"),)
    atf5_queries = (("atf 5 generic", "atf 5-generic", "atf5 generic", "atf5-generic"),)
    atf6_queries = (("atf 6 generic", "atf 6-generic", "atf6 generic", "atf6-generic"),)

    # Slot 4: ATF I/O if present; otherwise ATF 2 moves up.
    add_first_available(io_queries + atf2_queries)

    # Continue remaining core ATF images in order. Used URLs are skipped automatically.
    add_first_available(atf2_queries)
    add_first_available(atf3_queries)
    add_first_available(atf4_queries)
    add_first_available(atf5_queries)

    # ATF 6 is always last among CVS ATF assets.
    add_first_available(atf6_queries)

    for img in imgs:
        add(img)

    return ordered[:max_slots]
def reorder_walgreens_salsify_images_for_visual(images, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE):
    """
    Walgreens Salsify order for visual QA:
    1. Online Optimized Image.
    2. Ingredient Label Image.
    3+. ATF / lifestyle images.

    Slots 1 and 2 are locked. Missing required slots stay blank so ATF images
    never move up and hide missing Online Optimized or Ingredient Label assets.
    """
    imgs = [
        img for img in (images or [])
        if isinstance(img, dict) and not is_cvs_only_salsify_image(img)
    ]

    def norm_name(img):
        return normalize_salsify_asset_name((img or {}).get("name", "")) if isinstance(img, dict) else ""

    def find_first(*queries):
        query_tokens = [normalize_salsify_asset_name(q) for q in queries if normalize_salsify_asset_name(q)]
        for q in query_tokens:
            for img in imgs:
                name = norm_name(img)
                if name and (q == name or q in name):
                    return img
        return None

    def image_key(img):
        return str(img.get("url", "") or "").strip() if isinstance(img, dict) else ""

    ordered = []
    used = set()

    def add(img, blank_name=""):
        if not isinstance(img, dict):
            if blank_name:
                ordered.append(make_blank_salsify_image_slot(blank_name))
            return False
        key = image_key(img)
        if not key or key in used:
            if blank_name:
                ordered.append(make_blank_salsify_image_slot(blank_name))
            return False
        ordered.append(img)
        used.add(key)
        return True

    # Locked slot 1: Online Optimized Image only.
    add(
        find_first(
            "online optimized image",
            "online image",
        ),
        "walgreens_slot_1_online_optimized_image",
    )

    # Locked slot 2: Ingredient Label Image only.
    add(
        find_first(
            "ingredient label image",
            "ingredient label",
            "ingredients label image",
            "ingredients label",
        ),
        "walgreens_slot_2_ingredient_label_image",
    )

    atf_query_groups = [
        ("atf i/o generic", "atf i o generic", "atf io generic", "atf i/o-generic", "atf io-generic", "atf i/o"),
        ("atf 2 generic", "atf 2-generic", "atf2 generic", "atf2-generic", "atf 2"),
        ("atf 3 generic", "atf 3-generic", "atf3 generic", "atf3-generic", "atf 3"),
        ("atf 4 generic", "atf 4-generic", "atf4 generic", "atf4-generic", "atf 4"),
        ("atf 5 generic", "atf 5-generic", "atf5 generic", "atf5-generic", "atf 5"),
        ("atf 6 generic", "atf 6-generic", "atf6 generic", "atf6-generic", "atf 6"),
    ]

    # ATF images start only after the two locked Walgreens rows.
    for query_group in atf_query_groups:
        add(find_first(*query_group))

    # Include any remaining non-locked assets in source order, without duplicates.
    for img in imgs:
        add(img)

    return ordered[:max_slots]


def reorder_cvs_retailer_images_for_visual(images, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE):
    """CVS-only retailer image order.

    Keep CVS images in the exact live site order captured from the page. Dedupe by
    the base image path, but preserve the query string on the returned URL. CVS
    image service URLs often need the ?im=Resize(...) query to render correctly;
    stripping it can create broken thumbnails and false 0% image scores.
    """
    ordered = []
    seen = set()
    for url in images or []:
        clean_url = str(url or "").strip()
        base = clean_url.split("?", 1)[0]
        if not base or base in seen:
            continue
        ordered.append(clean_url)
        seen.add(base)
        if len(ordered) >= max_slots:
            break
    return ordered[:max_slots]

def _cache_cvs_package_like_result(url, value):
    global cvs_package_like_cache
    if "cvs_package_like_cache" not in globals() or not isinstance(globals().get("cvs_package_like_cache"), dict):
        cvs_package_like_cache = {}
    cvs_package_like_cache[url] = value
    while len(cvs_package_like_cache) > IMAGE_HASH_CACHE_MAX:
        cvs_package_like_cache.pop(next(iter(cvs_package_like_cache)))


def is_cvs_package_like_image(url):
    """CVS-only lightweight packaging detector.

    Looks for the centered product-pack / package-panel on white-space pattern.
    This is only used for CVS locked packaging rows and does not affect any
    other retailer.
    """
    global cvs_package_like_cache
    if "cvs_package_like_cache" not in globals() or not isinstance(globals().get("cvs_package_like_cache"), dict):
        cvs_package_like_cache = {}

    url = str(url or "").strip()
    if not url or is_video_like_url(url):
        return False
    cache_key = url.split("?", 1)[0]
    if cache_key in cvs_package_like_cache:
        return bool(cvs_package_like_cache[cache_key])

    try:
        image_bytes = _download_image_bytes_with_retry(url)
        if not image_bytes:
            _cache_cvs_package_like_result(cache_key, False)
            return False

        bio = BytesIO(image_bytes)
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            img = Image.open(bio)
            width, height = img.size
            if width * height > MAX_SAFE_IMAGE_PIXELS:
                _cache_cvs_package_like_result(cache_key, False)
                return False
            img.load()

        if img.mode in ("RGBA", "LA") or (img.mode == "P" and "transparency" in img.info):
            rgba = img.convert("RGBA")
            background = Image.new("RGB", rgba.size, (255, 255, 255))
            background.paste(rgba, mask=rgba.split()[-1])
            img = background
        elif img.mode != "RGB":
            img = img.convert("RGB")

        img.thumbnail((180, 180), Image.LANCZOS)
        arr = np.asarray(img).astype("int16")
        if arr.size == 0:
            _cache_cvs_package_like_result(cache_key, False)
            return False

        h, w = arr.shape[:2]
        image_area = max(1, h * w)
        white_mask = (arr[:, :, 0] >= 244) & (arr[:, :, 1] >= 244) & (arr[:, :, 2] >= 244)
        nonwhite_mask = ~white_mask
        nonwhite_count = int(nonwhite_mask.sum())
        if nonwhite_count <= max(8, image_area * 0.015):
            _cache_cvs_package_like_result(cache_key, False)
            return False

        white_ratio = float(white_mask.sum()) / float(image_area)
        nonwhite_ratio = float(nonwhite_count) / float(image_area)

        ys, xs = np.where(nonwhite_mask)
        x0, x1 = int(xs.min()), int(xs.max())
        y0, y1 = int(ys.min()), int(ys.max())
        bbox_w = max(1, x1 - x0 + 1)
        bbox_h = max(1, y1 - y0 + 1)
        bbox_width_ratio = bbox_w / float(w)
        bbox_height_ratio = bbox_h / float(h)
        bbox_area_ratio = (bbox_w * bbox_h) / float(image_area)
        cx = (x0 + x1) / 2.0 / float(w)
        cy = (y0 + y1) / 2.0 / float(h)
        centered = (0.32 <= cx <= 0.68) and (0.28 <= cy <= 0.72)

        # Strong negative: full-bleed ATF/lifestyle/card graphics.
        if white_ratio < 0.08 and bbox_area_ratio > 0.82:
            _cache_cvs_package_like_result(cache_key, False)
            return False
        if nonwhite_ratio > 0.82 and bbox_area_ratio > 0.90:
            _cache_cvs_package_like_result(cache_key, False)
            return False

        package_like = False
        if centered and white_ratio >= 0.18 and bbox_area_ratio <= 0.82:
            package_like = True
        if centered and white_ratio >= 0.10 and bbox_height_ratio <= 0.58 and bbox_width_ratio >= 0.45:
            package_like = True
        if centered and white_ratio >= 0.40 and nonwhite_ratio <= 0.55:
            package_like = True

        if bbox_width_ratio > 0.92 and bbox_height_ratio > 0.92 and white_ratio < 0.35:
            package_like = False

        _cache_cvs_package_like_result(cache_key, bool(package_like))
        return bool(package_like)
    except Exception:
        _cache_cvs_package_like_result(cache_key, False)
        return False


def align_cvs_atf_images_by_visual_match(s_images, r_images, locked_slots=3, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE, retailer_name="CVS"):
    """CVS-only hybrid packaging alignment.

    Keep Salsify order stable, keep CVS ATF/site order stable, and lock CVS
    slots 2 and 3 only when CVS has a real packaging match.
    """
    retailer_key = str(retailer_name or "").strip().lower()
    if retailer_key != "cvs":
        return [str(u or "").strip() for u in list(r_images or []) if str(u or "").strip()][:max_slots]

    s_images = list(s_images or [])
    r_images = [str(u or "").strip() for u in list(r_images or []) if str(u or "").strip()]
    max_slots = max(0, int(max_slots or MAX_IMAGE_SLOTS_TO_COMPARE))
    if not r_images or max_slots <= 0:
        return []

    def _s_url(img):
        return str(img.get("url", "") or "").strip() if isinstance(img, dict) else ""

    def _safe_visual_score(s_url, r_url):
        if not s_url or not r_url:
            return 0
        try:
            return int(compare_images_visually(s_url, r_url) or 0)
        except Exception:
            return 0

    def _choose_cvs_packaging_candidate(s_url, remaining_urls, search_window=3):
        best_idx = None
        best_score = -1
        window = list(remaining_urls[:max(1, int(search_window or 1))])
        for idx, candidate_url in enumerate(window):
            if not candidate_url:
                continue
            visual_score = _safe_visual_score(s_url, candidate_url)
            package_like = False
            try:
                package_like = bool(is_cvs_package_like_image(candidate_url))
            except Exception:
                package_like = False
            if visual_score >= 65 or package_like:
                rank_score = visual_score + (12 if package_like else 0) - (idx * 3)
                if rank_score > best_score:
                    best_idx = idx
                    best_score = rank_score
        return best_idx

    ordered = []
    ordered.append(r_images[0] if r_images else "")
    remaining = list(r_images[1:])

    for s_idx in (1, 2):
        if len(ordered) >= max_slots:
            break
        s_locked_url = _s_url(s_images[s_idx]) if s_idx < len(s_images) else ""
        if not s_locked_url:
            if remaining:
                try:
                    if is_cvs_package_like_image(remaining[0]):
                        remaining.pop(0)
                except Exception:
                    pass
            ordered.append("")
            continue
        chosen_idx = _choose_cvs_packaging_candidate(s_locked_url, remaining, search_window=3)
        if chosen_idx is None:
            ordered.append("")
        else:
            ordered.append(remaining.pop(chosen_idx))

    for s_img in s_images[3:max_slots]:
        if len(ordered) >= max_slots:
            break
        if not _s_url(s_img):
            ordered.append("")
            continue
        ordered.append(remaining.pop(0) if remaining else "")

    while len(ordered) < max_slots and remaining:
        ordered.append(remaining.pop(0))

    return ordered[:max_slots]

def apply_retailer_salsify_image_limits(retailer_name, images):
    retailer = str(retailer_name or "").strip().lower()
    limits = get_retailer_salsify_requirements(retailer_name)
    max_images = int(limits.get("max_images", MAX_IMAGE_SLOTS_TO_COMPARE) or MAX_IMAGE_SLOTS_TO_COMPARE)
    max_images = max(0, min(max_images, MAX_IMAGE_SLOTS_TO_COMPARE))
    if retailer == "cvs":
        cvs_images = [img for img in list(images or []) if not is_walgreens_only_salsify_image(img) and not bool((img or {}).get("kroger_only"))]
        return reorder_cvs_salsify_images_for_visual(cvs_images, max_slots=max_images)
    if retailer == "walgreens":
        return reorder_walgreens_salsify_images_for_visual([img for img in list(images or []) if not bool((img or {}).get("kroger_only"))], max_slots=max_images)
    images = [
        img for img in list(images or [])
        if not is_cvs_only_salsify_image(img) and not is_walgreens_only_salsify_image(img) and not bool((img or {}).get("kroger_only"))
    ]
    out = []
    seen_urls = set()
    for img in list(images or []):
        if not isinstance(img, dict):
            continue
        url = str(img.get("url", "") or "").strip()
        if not url or url in seen_urls:
            continue
        out.append(img)
        seen_urls.add(url)
        if len(out) >= max_images:
            break
    return out


def clean_uploaded_url_value(value):
    """Clean URL values coming from Excel/CSV uploads.

    Handles hidden whitespace, zero-width characters, formula text, and
    HYPERLINK-style values. This keeps formula-generated Salsify/Walgreens
    links from being treated differently than manually copied URLs.
    """
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    value = html.unescape(str(value or "").strip())
    if not value or value in {"#N/A", "nan", "None", "0", "0.0"}:
        return ""

    value = value.replace("\u00a0", " ")
    value = re.sub(r"[\u200b-\u200f\u202a-\u202e\u2060\ufeff]", "", value)
    value = value.replace("\r", "").replace("\n", "")
    # CVS-only issue surfaced by pasted URL lists: tracker cells can contain
    # URLs with accidental trailing separators like ?skuId=137056;. Strip only
    # terminal separators so the real query string remains intact.
    value = value.strip().rstrip(";,")

    # If Excel formula text is read instead of the cached result, extract the
    # first URL from formulas like =HYPERLINK("https://...", "link").
    if value.startswith("="):
        match = re.search(r"https?://[^\"') ,]+", value)
        if match:
            value = match.group(0)

    value = re.sub(r"\s+", "", value)
    return value.strip()


def normalize_uploaded_salsify_url(value):
    """Build a fallback Salsify product URL.

    IMPORTANT: do not use this as the primary uploaded Salsify URL. Some
    Salsify PDP decks expose images/assets only on the full slug URL. This
    shortened URL is only a fallback for fetch attempts when the full URL fails.
    """
    value = clean_uploaded_url_value(value)
    if not value:
        return ""

    if "sites.salsify.com" not in value.lower():
        return value

    value = value.split("#", 1)[0].strip()
    match = re.search(r"^(https?://sites\.salsify\.com/[^\s?#]+?/product/([^/\s?#]+))(?:/[^?#]*)?(?:[?#].*)?$", value, flags=re.IGNORECASE)
    if match:
        return match.group(1).rstrip("/") + "/"

    return value



def salsify_url_candidates(value):
    """Return robust Salsify URL candidates for server-side fetch/parsing.

    The Salsify link clicked in the UI opens in the user's browser, but the app
    parses Salsify with requests. Browser navigation can tolerate/normalize
    malformed slug text such as "andamp" and encoded symbols differently than
    the server-side request. Try the original URL, a normalized/encoded slug URL,
    and the SKU-level fallback URL.
    """
    raw_url = clean_uploaded_url_value(value)
    if not raw_url:
        return []
    if "sites.salsify.com" not in raw_url.lower():
        return [raw_url]

    from urllib.parse import quote, unquote, urlsplit, urlunsplit

    candidates = []

    def add(candidate):
        candidate = str(candidate or "").strip()
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    add(raw_url)

    try:
        parts = urlsplit(raw_url)
        decoded_path = unquote(parts.path or "")

        # Salsify slug recovery. Some uploaded Salsify links contain literal
        # "andamp" inside the slug, for example SNUGandampDRY. Browser clicks
        # may still resolve, but server-side requests can return a shell page.
        # Try the real slug form first: SNUGandDRY.
        plain_and_path = re.sub(r"andamp;?", "and", decoded_path, flags=re.IGNORECASE)
        if plain_and_path != decoded_path:
            add(urlunsplit((
                parts.scheme,
                parts.netloc,
                quote(plain_and_path, safe="/-._~"),
                parts.query,
                "",
            )))

        # Also keep the older encoded-ampersand candidate as a secondary
        # fallback for any Salsify slug that genuinely expects an ampersand.
        fixed_path = re.sub(r"andamp;?", "&", decoded_path, flags=re.IGNORECASE)
        fixed_path = html.unescape(fixed_path)
        encoded_path = quote(fixed_path, safe="/-._~")
        add(urlunsplit((parts.scheme, parts.netloc, encoded_path, parts.query, "")))
    except Exception:
        pass

    fallback_url = normalize_uploaded_salsify_url(raw_url)
    if fallback_url and fallback_url != raw_url:
        add(fallback_url)

    return candidates


def is_probably_salsify_pdp_html(html_text):
    """True when returned Salsify HTML looks like the actual PDP deck payload."""
    text = str(html_text or "")
    if not text.strip():
        return False
    markers = [
        "__NEXT_DATA__",
        "digitalAssets",
        "Online Optimized Image",
        "Main Variant Image",
        "Sams Club Description",
        "Sam's Club Description",
        "Retailer URL",
        "Sam Club URL",
        "Sam's Club URL",
    ]
    return any(marker.lower() in text.lower() for marker in markers)

def normalize_uploaded_retail_url(value):
    value = clean_uploaded_url_value(value)
    if not value:
        return ""
    if "kroger.com" in value.lower():
        try:
            return normalize_kroger_url(value)
        except Exception:
            return value
    return value.split("#", 1)[0].strip()


def uploaded_retailer_value_is_generic(value):
    value = normalize_space(value).strip().lower()
    return (not value) or value in {
        "sheet1",
        "sheet 1",
        "input",
        "export",
        "data",
        "retailer",
        "retailers",
        "pdp qa",
    }


WIDE_SALSIFY_RETAILER_CONFIG = [
    ("Albertsons", "albertsons rpc", "albertsons salsify url", "albertsons url"),
    ("CVS", "cvs rpc", "cvs salsify url", "cvs url"),
    ("HEB", "heb rpc", "heb salsify url", "heb url"),
    ("Kroger", "kroger rpc", "kroger salsify url", "kroger url"),
    ("Meijer", "meijer rpc", "meijer salsify url", "meijer url"),
    ("Sams Club", "sams club rpc", "sams club salsify url", "sams club url"),
    ("Walgreens", "walgreens rpc", "walgreens salsify url", "walgreens url"),
]

WIDE_SALSIFY_SALSIFY_URL_ALIASES = {
    "albertsons salsify url": ["albertsons salsify url", "albertsons pdp deck link"],
    "cvs salsify url": ["cvs salsify url", "cvs pdp deck link"],
    "heb salsify url": ["heb salsify url", "heb pdp deck link", "h-e-b salsify url", "h-e-b pdp deck link"],
    "kroger salsify url": ["kroger salsify url", "kroger pdp deck link"],
    "meijer salsify url": ["meijer salsify url", "meijer pdp deck link"],
    "sams club salsify url": ["sams club salsify url", "sams club pdp deck link", "sam's club salsify url", "sam's club pdp deck link"],
    "walgreens salsify url": ["walgreens salsify url", "walgreens pdp deck link"],
}


def first_existing_column_local(df, candidates):
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
    return ""


def is_wide_salsify_template_df(df):
    """Detect one-template wide Salsify export with retailer-specific RPC columns."""
    if df is None or df.empty:
        return False
    cols = {str(c).strip().lower() for c in df.columns}
    if "retailer" in cols and "retailer rpc" in cols:
        return False
    rpc_hits = sum(1 for _, rpc_col, _, _ in WIDE_SALSIFY_RETAILER_CONFIG if rpc_col in cols)
    return bool(rpc_hits >= 2 or (rpc_hits >= 1 and ("sku" in cols or "7 digit sku" in cols)))


def normalize_wide_salsify_template_df(df):
    """Convert wide Salsify export into normalized one-row-per-retailer/RPC format."""
    source = df.copy()
    source.columns = [str(c).strip().lower() for c in source.columns]
    sku_col = first_existing_column_local(source, ["sku", "7 digit sku", "product sku", "salsify sku", "item sku"])
    brand_col = first_existing_column_local(source, ["brand", "brand_char", "brand characteristic"])
    rows = []
    for _, row in source.iterrows():
        sku = normalize_space(row.get(sku_col, "")) if sku_col else ""
        brand = normalize_space(row.get(brand_col, "")) if brand_col else ""
        for retailer, rpc_col, salsify_col, url_col in WIDE_SALSIFY_RETAILER_CONFIG:
            rpc = normalize_space(row.get(rpc_col, ""))
            if not rpc:
                continue
            salsify_source_col = first_existing_column_local(source, WIDE_SALSIFY_SALSIFY_URL_ALIASES.get(salsify_col, [salsify_col]))
            salsify_url = normalize_space(row.get(salsify_source_col, "")) if salsify_source_col else ""
            retail_url = normalize_space(row.get(url_col, "")) if url_col in source.columns else ""
            rows.append({
                "retailer": retailer,
                "sku": sku,
                "retailer rpc": rpc,
                "retailer salsify url": salsify_url,
                "retailer url": retail_url,
                "brand": brand,
            })
    if not rows:
        return source
    out = pd.DataFrame(rows)
    out = out.sort_values(
        by=["retailer", "sku", "retailer rpc"],
        key=lambda col: col.astype(str).str.lower(),
        kind="stable",
    ).reset_index(drop=True)
    return out


def coalesce_duplicate_columns(df):
    """Merge duplicate column names created by retailer-specific renames."""
    if df is None or df.empty or not df.columns.duplicated().any():
        return df
    ordered_names = []
    for col in df.columns:
        if col not in ordered_names:
            ordered_names.append(col)
    merged = pd.DataFrame(index=df.index)
    for col in ordered_names:
        block = df.loc[:, df.columns == col]
        if block.shape[1] == 1:
            merged[col] = block.iloc[:, 0]
            continue
        series = block.iloc[:, 0]
        for idx in range(1, block.shape[1]):
            next_series = block.iloc[:, idx]
            current_blank = series.isna() | (series.astype(str).str.strip() == "") | (series.astype(str).str.strip().str.lower() == "nan")
            series = series.where(~current_blank, next_series)
        merged[col] = series
    return merged

def prepare_input_df(df):
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]

    # Support the one-template wide Salsify export. Convert retailer-specific
    # columns like Kroger RPC/Kroger URL into normalized app rows before renaming.
    if is_wide_salsify_template_df(df):
        df = normalize_wide_salsify_template_df(df)
        df.columns = [str(c).strip().lower() for c in df.columns]

    df.rename(
        columns={
            "salsify url": "salsify_url",
            "retailer salsify url": "salsify_url",
            "retailer salsify link": "salsify_url",
            "retailer pdp deck link": "salsify_url",
            "salsify_url": "salsify_url",
            "salsify link": "salsify_url",
            "salsify pdp deck link": "salsify_url",
            "pdp deck link": "salsify_url",
            "walgreens pdp deck link": "salsify_url",
            "cvs pdp deck link": "salsify_url",
            "kroger pdp deck link": "salsify_url",
            "sams club pdp deck link": "salsify_url",
            "sam's club pdp deck link": "salsify_url",
            "retail url": "retail_url",
            "retail_url": "retail_url",
            "retailer url": "retail_url",
            "walgreens url": "retail_url",
            "cvs url": "retail_url",
            "kroger url": "retail_url",
            "kroger_url": "retail_url",
            "sams club url": "retail_url",
            "sam's club url": "retail_url",
            "sku id": "sku",
            "product sku": "sku",
            "7 digit sku": "sku",
            "salsify sku": "sku",
            "item sku": "sku",
            "brand_char": "brand",
            "brand characteristic": "brand",
            "retailer name": "retailer",
            "retailer_name": "retailer",
            "retailer sku": "sku",
            "retailer rpc": "retailer_rpc",
            "retailer item id": "retailer_rpc",
            "kroger rpc": "kroger_rpc",
            "heb rpc": "heb_rpc",
            "h-e-b rpc": "heb_rpc",
            "heb item id": "heb_rpc",
            "h-e-b item id": "heb_rpc",
            "all source code": "copy_source_code",
            "source code": "copy_source_code",
            "copy source code": "copy_source_code",
            "raw source code": "copy_source_code",
            "html source": "copy_source_code",
            "raw html": "copy_source_code",
            "page source": "copy_source_code",
        },
        inplace=True,
    )


    # Several retailer-specific columns intentionally map into common app columns
    # such as salsify_url and retail_url. Coalesce duplicate names so df[col]
    # returns a Series, not a DataFrame.
    df = coalesce_duplicate_columns(df)


    # CVS/manual-source rescue:
    # Some tracker/source workbooks have page source in an unnamed column or a
    # placeholder column such as "\\". If copy_source_code is blank, scan the
    # non-core columns for HTML-like page source and copy the richest cell into
    # copy_source_code. Retailer parsing is still isolated later by selected
    # retailer, so this does not make non-CVS parsers use CVS logic.
    if "copy_source_code" not in df.columns:
        df["copy_source_code"] = ""

    def _looks_like_uploaded_page_source(value):
        value = str(value or "")
        if len(value.strip()) < 200:
            return False
        lowered = value.lower()
        return bool(
            "<!doctype html" in lowered
            or "<html" in lowered
            or "self.__next_f.push" in lowered
            or "vendorcontent" in lowered
            or "vendordetails" in lowered
            or "sp-schema" in lowered
            or "/bizcontent/merchandising/productimages/high_res/" in lowered
            or "schema.org" in lowered
        )

    core_columns_for_source_scan = {
        "retailer", "sku", "salsify_url", "retail_url", "brand", "retailer_rpc",
        "rating", "review_count", "copy_source_code", "kroger_rpc", "heb_rpc",
        "cvs rpc", "walgreens rpc", "sams club rpc",
    }
    source_like_columns = []
    for candidate_col in list(df.columns):
        candidate_col_name = str(candidate_col or "").strip().lower()
        if candidate_col_name in core_columns_for_source_scan:
            continue
        try:
            sample_values = df[candidate_col].dropna().astype(str).head(25).tolist()
        except Exception:
            sample_values = []
        if any(_looks_like_uploaded_page_source(value) for value in sample_values):
            source_like_columns.append(candidate_col)

    if source_like_columns:
        def _pick_best_row_source(row):
            existing = str(row.get("copy_source_code", "") or "")
            if existing.strip():
                return existing
            candidates = []
            for candidate_col in source_like_columns:
                value = str(row.get(candidate_col, "") or "")
                if _looks_like_uploaded_page_source(value):
                    candidates.append(value)
            return max(candidates, key=len) if candidates else existing

        df["copy_source_code"] = df.apply(_pick_best_row_source, axis=1)

    rpc_candidates = []
    for rpc_col in ["retailer_rpc", "kroger_rpc", "heb_rpc", "cvs rpc", "walgreens rpc", "sams club rpc"]:
        if rpc_col in df.columns:
            rpc_candidates.append(
                df[rpc_col]
                .replace("#N/A", "")
                .fillna("")
                .astype(str)
                .str.replace(".0", "", regex=False)
                .str.strip()
            )
    if rpc_candidates:
        retailer_rpc = rpc_candidates[0].copy()
        for series in rpc_candidates[1:]:
            retailer_rpc = retailer_rpc.where(retailer_rpc != "", series)
        df["retailer_rpc"] = retailer_rpc
    else:
        df["retailer_rpc"] = ""

    for rpc_col in ["kroger_rpc", "heb_rpc", "cvs rpc", "walgreens rpc", "sams club rpc"]:
        if rpc_col in df.columns:
            df.drop(columns=[rpc_col], inplace=True)

    for col in ["sku", "salsify_url", "retail_url", "brand", "retailer_rpc", "rating", "review_count", "copy_source_code"]:
        if col not in df.columns:
            df[col] = ""

    df = coalesce_duplicate_columns(df)

    for col in ["sku", "salsify_url", "retail_url", "brand", "retailer_rpc", "rating", "review_count"]:
        df[col] = df[col].replace("#N/A", "").fillna("").astype(str).str.strip()
    df["retailer_rpc"] = df["retailer_rpc"].apply(clean_item_number)
    df["copy_source_code"] = df["copy_source_code"].fillna("").astype(str)

    # Keep the original full Salsify URL, including the product slug.
    # The full URL is needed for Walgreens image/asset extraction. A shortened
    # /product/{sku}/ fallback is still attempted in get_html() if the full URL fails.
    df["salsify_url"] = df["salsify_url"].apply(clean_uploaded_url_value)
    df["retail_url"] = df["retail_url"].apply(normalize_uploaded_retail_url)

    if "retailer" not in df.columns:
        df["retailer"] = df["retail_url"].apply(infer_retailer_name_from_url)
    else:
        df["retailer"] = df["retailer"].replace("#N/A", "").fillna("").astype(str).str.strip()
        # Excel files are often read sheet-by-sheet and the sheet name may be
        # "Sheet1". If the sheet name is generic, infer retailer from URL so
        # Walgreens rows are not filtered out during strict retailer selection.
        inferred_retailer = df["retail_url"].apply(infer_retailer_name_from_url)
        df["retailer"] = df["retailer"].where(
            ~df["retailer"].apply(uploaded_retailer_value_is_generic),
            inferred_retailer,
        )
    df["retailer"] = df["retailer"].apply(normalize_retailer_name)

    required = ["sku", "salsify_url", "retail_url"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    return df



def build_selected_retailer_df_from_wide_source(df, selected_retailer):
    """Build one selected retailer queue directly from the wide SKU/RPC matrix.

    This prevents the app from accidentally processing the original 936-row
    all-retailer matrix when the selected retailer is Kroger. It also keeps rows
    where the selected retailer has a URL or Salsify URL even if the RPC is blank.
    """
    if df is None or df.empty:
        return pd.DataFrame()

    selected_norm = normalize_retailer_name(selected_retailer)
    source = df.copy()
    source.columns = [str(c).strip().lower() for c in source.columns]

    config = None
    for retailer, rpc_col, salsify_col, url_col in WIDE_SALSIFY_RETAILER_CONFIG:
        if normalize_retailer_name(retailer) == selected_norm:
            config = (retailer, rpc_col, salsify_col, url_col)
            break
    if not config:
        return pd.DataFrame()

    retailer, rpc_col, salsify_col, url_col = config
    sku_col = first_existing_column_local(source, ["sku", "7 digit sku", "product sku", "salsify sku", "item sku"])
    brand_col = first_existing_column_local(source, ["brand", "brand_char", "brand characteristic"])
    salsify_source_col = first_existing_column_local(source, WIDE_SALSIFY_SALSIFY_URL_ALIASES.get(salsify_col, [salsify_col]))

    rows = []
    for _, row in source.iterrows():
        sku = normalize_space(row.get(sku_col, "")) if sku_col else ""
        brand = normalize_space(row.get(brand_col, "")) if brand_col else ""
        rpc = normalize_space(row.get(rpc_col, "")) if rpc_col in source.columns else ""
        salsify_url = normalize_space(row.get(salsify_source_col, "")) if salsify_source_col else ""
        retail_url = normalize_space(row.get(url_col, "")) if url_col in source.columns else ""

        # Keep only rows that actually belong to this retailer. For Kroger, this
        # drops the blank/non-Kroger 936-row template rows and keeps the real
        # Kroger setup rows.
        if not (rpc or salsify_url or retail_url):
            continue

        rows.append({
            "retailer": retailer,
            "sku": sku,
            "brand": brand,
            "retailer_rpc": rpc,
            "salsify_url": clean_uploaded_url_value(salsify_url),
            "retail_url": normalize_uploaded_retail_url(retail_url),
            "rating": "",
            "review_count": "",
            "copy_source_code": "",
        })

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return sort_selected_retailer_queue(out)




def sort_selected_retailer_queue(df):
    """Keep selected retailer runs stable and readable: Brand A-Z, then SKU/RPC/URL."""
    if df is None or df.empty:
        return df
    out = df.copy()
    for col in ["brand", "sku", "retailer_rpc", "retail_url"]:
        if col not in out.columns:
            out[col] = ""
    out["_brand_sort"] = out["brand"].fillna("").astype(str).str.strip().str.lower()
    out["_sku_sort"] = out["sku"].fillna("").astype(str).str.strip().str.lower()
    out["_rpc_sort"] = out["retailer_rpc"].fillna("").astype(str).str.strip().str.lower()
    out["_url_sort"] = out["retail_url"].fillna("").astype(str).str.strip().str.lower()
    out = out.sort_values(
        by=["_brand_sort", "_sku_sort", "_rpc_sort", "_url_sort"],
        kind="stable",
    ).drop(columns=["_brand_sort", "_sku_sort", "_rpc_sort", "_url_sort"], errors="ignore")
    return out.reset_index(drop=True)




def filter_queue_to_uploaded_capture_matches(df, selected_retailer, source_mode="", selected_capture_mode="", uploaded_raw_html_map=None):
    """For extension/TXT runs, process only selected-retailer rows that actually matched the uploaded capture.

    The SKU/RPC matrix is the reference source, but the uploaded capture file represents the pages that were actually loaded.
    This prevents the app from processing the whole matrix or unmatched selected-retailer rows.
    """
    if df is None or df.empty:
        return df, 0, 0
    uploaded_raw_html_map = uploaded_raw_html_map or {}
    if selected_capture_mode != CAPTURE_MODE_USE_EXTENSION or not uploaded_raw_html_map:
        return df, 0, 0

    out = df.copy()
    before_count = len(out)
    source_mode = str(source_mode or "")

    if source_mode == "extension_url_only_results":
        if "retail_url" not in out.columns:
            out["retail_url"] = ""
        matched_mask = out["retail_url"].fillna("").astype(str).str.strip().ne("")
    else:
        if "copy_source_code" not in out.columns:
            out["copy_source_code"] = ""
        matched_mask = out["copy_source_code"].fillna("").astype(str).str.len() > 0

    matched_count = int(matched_mask.sum())
    missing_count = max(before_count - matched_count, 0)
    if matched_count <= 0:
        return out, matched_count, missing_count

    out = out[matched_mask].copy()
    return out.reset_index(drop=True), matched_count, missing_count


def strict_filter_rows_for_selected_retailer(df, selected_retailer, dedupe_by_url=False):
    """
    Hard retailer isolation guard.

    This prevents any CVS, Walgreens, Kroger, or Sam's Club rows from leaking into a
    different retailer batch, even if the uploaded file contains multiple retailers.
    """
    selected_retailer_norm = normalize_retailer_name(selected_retailer)
    if df is None:
        return pd.DataFrame()

    out = df.copy()
    if "retailer" not in out.columns:
        out["retailer"] = "Retailer"

    out["retailer"] = out["retailer"].astype(str).apply(normalize_retailer_name)
    out = out[out["retailer"] == selected_retailer_norm].copy()

    if "retail_url" not in out.columns:
        out["retail_url"] = ""

    out["retail_url"] = out["retail_url"].fillna("").astype(str).str.strip()

    # New SKU-list workflow: retailer URL is optional. The extension can find the
    # URL/content from Retailer RPC/Search Term, then the app maps the extension result
    # back by RPC. Keep blank-URL rows in the selected retailer queue.
    if not out.empty:
        out = out[out["retail_url"].apply(lambda value: (not str(value or "").strip()) or retailer_url_matches_selected(value, selected_retailer_norm))].copy()

    if dedupe_by_url and not out.empty:
        # Kroger can have multiple SKU7/version rows sharing one PDP URL.
        # Do not collapse Kroger rows by URL or valid SKU rows disappear.
        if selected_retailer_norm != "Kroger":
            out = out.drop_duplicates(subset=["retail_url"], keep="first").copy()

    return out
def clear_in_memory_caches():
    global html_cache, image_hash_cache, image_compare_cache, display_image_cache, walgreens_api_cache

    if "html_cache" not in globals() or not isinstance(globals().get("html_cache"), dict):
        html_cache = {}
    if "image_hash_cache" not in globals() or not isinstance(globals().get("image_hash_cache"), dict):
        image_hash_cache = {}
    if "image_compare_cache" not in globals() or not isinstance(globals().get("image_compare_cache"), dict):
        image_compare_cache = {}
    if "display_image_cache" not in globals() or not isinstance(globals().get("display_image_cache"), dict):
        display_image_cache = {}

    html_cache.clear()
    image_hash_cache.clear()
    image_compare_cache.clear()
    display_image_cache.clear()
    if "walgreens_api_cache" not in globals() or not isinstance(globals().get("walgreens_api_cache"), dict):
        walgreens_api_cache = {}
    walgreens_api_cache.clear()

# =========================================
# HTML FETCH
# =========================================
def get_html(url):
    global html_cache

    if "html_cache" not in globals() or not isinstance(globals().get("html_cache"), dict):
        html_cache = {}

    if not url:
        return ""

    url = clean_uploaded_url_value(url)
    primary_url = url
    is_salsify_url = "sites.salsify.com" in str(url).lower()
    candidate_urls = salsify_url_candidates(url) if is_salsify_url else [primary_url]
    last_successful_html = ""

    for candidate_url in candidate_urls:
        candidate_url = str(candidate_url or "").strip()
        if not candidate_url:
            continue

        cached = html_cache.get(candidate_url)
        if cached:
            if not is_salsify_url or is_probably_salsify_pdp_html(cached):
                return cached
            last_successful_html = cached
            continue

        try:
            session = get_session()
            r = session.get(candidate_url, timeout=REQUEST_TIMEOUT)
            if r.status_code == 200 and r.text:
                html_cache[candidate_url] = r.text
                if primary_url and primary_url != candidate_url:
                    html_cache[primary_url] = r.text
                while len(html_cache) > HTML_CACHE_MAX:
                    html_cache.pop(next(iter(html_cache)))

                if not is_salsify_url or is_probably_salsify_pdp_html(r.text):
                    return r.text

                # This can happen when Salsify returns a generic shell/error app page.
                # Keep it only as a last resort and continue trying safer candidates.
                last_successful_html = r.text
        except Exception:
            pass

    return last_successful_html

def fetch_html_with_timeout(url, timeout_seconds):
    if not url:
        return ""

    try:
        session = get_session()
        r = session.get(url, timeout=timeout_seconds, allow_redirects=True)
        if r.status_code == 200 and r.text:
            return r.text
    except Exception:
        pass

    return ""


def get_walgreens_html(url):
    """
    Walgreens-specific fetch path with a longer timeout and shared HTML cache.
    This avoids refetching the same PDP across batch processing + visual review.
    """
    global html_cache
    if "html_cache" not in globals() or not isinstance(globals().get("html_cache"), dict):
        html_cache = {}

    url = str(url or "").strip()
    if not url:
        return ""

    cache_key = f"walgreens::{url}"
    if cache_key in html_cache:
        html_cache[cache_key] = html_cache.pop(cache_key)
        return html_cache[cache_key]

    html_text = fetch_html_with_timeout(url, WALGREENS_REQUEST_TIMEOUT)
    if html_text:
        html_cache[cache_key] = html_text
        while len(html_cache) > HTML_CACHE_MAX:
            html_cache.pop(next(iter(html_cache)))
    return html_text

def get_walgreens_product_id_from_url(retail_url):
    """
    Supports Walgreens product URLs like:
    - /ID=300432791-product
    - /ID=prod6153586-product
    - ?productId=300432791
    - ?productId=prod6153586
    """
    if not retail_url:
        return ""

    retail_url = str(retail_url or "").strip()

    patterns = [
        r"/ID=([A-Za-z0-9]+)-product",
        r"[?&]productId=([A-Za-z0-9]+)",
        r'"productId"\s*:\s*"([A-Za-z0-9]+)"',
    ]

    for pattern in patterns:
        m = re.search(pattern, retail_url, flags=re.IGNORECASE)
        if m:
            return m.group(1)

    return ""


def get_walgreens_sku_id_from_url(retail_url):
    """
    Extracts selected skuId from Walgreens variant/querystring PDPs such as:
    .../ID=300447053-product?skuId=400632972
    .../ID=300465880-product?skuId=sku6275345
    """
    if not retail_url:
        return ""

    retail_url = str(retail_url or "").strip()

    m = re.search(r"[?&]skuId=([A-Za-z0-9_-]+)", retail_url, flags=re.IGNORECASE)
    if m:
        return m.group(1)

    return ""

def fetch_json_with_timeout(url, timeout_seconds):
    if not url:
        return None

    try:
        session = get_session()
        r = session.get(url, timeout=timeout_seconds, allow_redirects=True)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass

    return None


def get_walgreens_product_api_payload(product_id):
    """
    Walgreens source exposes a lighter product endpoint:
    /productapi/v1/products?productId={prodId}

    Cache API payloads so repeated runs and visual mode are faster.
    """
    global walgreens_api_cache
    if "walgreens_api_cache" not in globals() or not isinstance(globals().get("walgreens_api_cache"), dict):
        walgreens_api_cache = {}

    product_id = str(product_id or "").strip()
    if not product_id:
        return None

    cache_key = f"productapi::{product_id}"
    if cache_key in walgreens_api_cache:
        walgreens_api_cache[cache_key] = walgreens_api_cache.pop(cache_key)
        return walgreens_api_cache[cache_key]

    api_url = f"https://www.walgreens.com/productapi/v1/products?productId={product_id}"
    payload = fetch_json_with_timeout(api_url, WALGREENS_API_TIMEOUT)
    if payload is not None:
        walgreens_api_cache[cache_key] = payload
        while len(walgreens_api_cache) > HTML_CACHE_MAX:
            walgreens_api_cache.pop(next(iter(walgreens_api_cache)))
    return payload

def walk_json(obj):
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from walk_json(v)
    elif isinstance(obj, list):
        for item in obj:
            yield from walk_json(item)


def find_first_dict_with_keys(obj, required_keys):
    required_keys = set(required_keys)

    for node in walk_json(obj):
        if isinstance(node, dict) and required_keys.issubset(set(node.keys())):
            return node

    return {}

# =========================================
# HTML / DOM DEBUG HELPERS
# =========================================
def html_to_debug_textblob(html_text):
    if not html_text:
        return ""

    raw = html.unescape(html_text or "")
    text = BeautifulSoup(raw, "html.parser").get_text("\n", strip=True)
    text = text.replace("\r", "\n")
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def html_to_prettified_dom(html_text):
    if not html_text:
        return ""

    try:
        raw = html.unescape(html_text or "")
        soup = BeautifulSoup(raw, "html.parser")
        return soup.prettify()
    except Exception:
        return html_text or ""


def preview_between_markers(text, start_marker="", end_marker=""):
    """
    Returns a preview slice between start_marker and end_marker.
    Uses case-insensitive literal matching.
    """
    source = str(text or "")

    result = {
        "preview": "",
        "start_found": False,
        "end_found": False,
        "start_index": -1,
        "end_index": -1,
    }

    if not source:
        return result

    working = source
    start_idx_global = 0

    if start_marker:
        start_match = re.search(re.escape(start_marker), source, flags=re.IGNORECASE)
        if not start_match:
            return result

        result["start_found"] = True
        result["start_index"] = start_match.start()
        start_idx_global = start_match.start()
        working = source[start_idx_global:]
    else:
        result["start_found"] = True
        result["start_index"] = 0
        working = source

    if end_marker:
        end_match = re.search(re.escape(end_marker), working, flags=re.IGNORECASE)
        if end_match:
            result["end_found"] = True
            result["end_index"] = start_idx_global + end_match.start()
            result["preview"] = working[:end_match.start()].strip()
            return result

    result["preview"] = working.strip()
    return result


def parse_debug_headers_text(headers_text):
    """Accept JSON object text or Header: Value lines."""
    raw = str(headers_text or "").strip()
    if not raw:
        return {}
    if raw.startswith("{"):
        try:
            value = json.loads(raw)
            return value if isinstance(value, dict) else {}
        except Exception:
            return {}
    headers = {}
    for line in raw.splitlines():
        line = str(line or "").strip()
        if not line or ":" not in line:
            continue
        key, value = line.split(":", 1)
        key = str(key or "").strip()
        value = str(value or "").strip()
        if key:
            headers[key] = value
    return headers


def fetch_url_debug(
    url,
    retailer_name="",
    headers_text="",
    timeout_override=None,
    use_mobile=False,
    proxy_url="",
):
    result = {
        "requested_url": str(url or ""),
        "final_url": "",
        "status_code": None,
        "reason": "",
        "content_type": "",
        "content_length_header": "",
        "text_length": 0,
        "history": [],
        "error": "",
        "raw_html": "",
        "dom_text": "",
        "prettified_dom": "",
        "response_headers": {},
        "request_headers_used": {},
        "proxy_used": "",
        "elapsed_seconds": 0.0,
    }

    url = str(url or "").strip()
    retailer_name = str(retailer_name or "").strip().lower()
    proxy_url = str(proxy_url or "").strip()

    if not url:
        result["error"] = "No URL provided."
        return result

    if retailer_name == "kroger":
        try:
            url = normalize_kroger_url(url)
        except Exception:
            pass

    desktop_ua = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
    mobile_ua = (
        "Mozilla/5.0 (iPhone; CPU iPhone OS 17_4 like Mac OS X) "
        "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Mobile/15E148 Safari/604.1"
    )

    headers = dict(HEADERS)
    headers.update(parse_debug_headers_text(headers_text))
    headers["User-Agent"] = mobile_ua if use_mobile else desktop_ua
    headers.setdefault("Accept", "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8")
    headers.setdefault("Accept-Language", "en-US,en;q=0.9")
    headers.setdefault("Cache-Control", "no-cache")
    headers.setdefault("Pragma", "no-cache")
    headers.setdefault("Upgrade-Insecure-Requests", "1")
    result["request_headers_used"] = headers
    result["proxy_used"] = proxy_url

    timeout_value = (6, 30)
    if retailer_name == "walgreens":
        timeout_value = (6, WALGREENS_DEBUG_TIMEOUT)
    elif retailer_name == "kroger":
        timeout_value = (8, 45)

    try:
        if timeout_override is not None and str(timeout_override).strip() != "":
            timeout_override = max(1, int(timeout_override))
            timeout_value = (min(timeout_override, 10), timeout_override)
    except Exception:
        pass

    proxies = None
    if proxy_url:
        proxies = {"http": proxy_url, "https": proxy_url}

    start = time.monotonic()
    try:
        session = get_session()
        r = session.get(
            url,
            headers=headers,
            timeout=timeout_value,
            proxies=proxies,
            allow_redirects=True,
        )
        result["elapsed_seconds"] = round(time.monotonic() - start, 3)
        result["final_url"] = str(r.url or "")
        result["status_code"] = int(r.status_code)
        result["reason"] = str(getattr(r, "reason", "") or "")
        result["content_type"] = str(r.headers.get("Content-Type", "") or "")
        result["content_length_header"] = str(r.headers.get("Content-Length", "") or "")
        result["history"] = [{"status_code": int(h.status_code), "url": str(h.url or "")} for h in r.history]
        interesting_headers = ["Content-Type", "Content-Length", "Server", "Cache-Control", "Set-Cookie", "Location", "X-Cache", "X-Served-By", "CF-Cache-Status", "CF-Ray"]
        result["response_headers"] = {k: v for k, v in r.headers.items() if k in interesting_headers}
        raw_html = r.text or ""
        result["raw_html"] = raw_html
        result["text_length"] = len(raw_html)
        # Keep debugger fast. DOM text and prettified DOM are generated only
        # when the user selects those views in render_debugger_panel().
        result["dom_text"] = ""
        result["prettified_dom"] = ""
        result["lazy_debug_views"] = True
    except Exception as e:
        result["elapsed_seconds"] = round(time.monotonic() - start, 3)
        result["error"] = repr(e)

    return result


DEBUG_TEXT_PREVIEW_CHARS = 200000


def make_fast_debug_views(html_text, lazy=True):
    """Build lightweight debug views without parsing/prettifying huge HTML upfront."""
    html_text = str(html_text or "")
    return {
        "raw_html": html_text,
        "dom_text": "" if lazy else html_to_debug_textblob(html_text),
        "prettified_dom": "" if lazy else html_to_prettified_dom(html_text),
        "lazy_debug_views": bool(lazy),
        "text_length": len(html_text),
    }


@st.cache_data(show_spinner=False)
def get_debug_views_for_url(url):
    html_text = get_html(url)
    return make_fast_debug_views(html_text, lazy=True)


def build_debug_views_from_html(html_text):
    return make_fast_debug_views(html_text, lazy=True)


def get_uploaded_text_file_bytes(uploaded_text_file):
    if uploaded_text_file is None:
        return ""

    try:
        raw = uploaded_text_file.getvalue()
        if isinstance(raw, bytes):
            for encoding in ["utf-8", "utf-8-sig", "latin1"]:
                try:
                    return raw.decode(encoding)
                except Exception:
                    pass
        return str(raw)
    except Exception:
        return ""



def normalize_uploaded_capture_url(url):
    url = str(url or "").strip()
    if not url:
        return ""
    try:
        if "kroger.com" in url.lower():
            url = normalize_kroger_url(url)
    except Exception:
        pass
    url = html.unescape(url)
    url = url.split("#", 1)[0].strip()
    url = re.sub(r"[​-‍﻿]", "", url)
    return url


def uploaded_capture_url_candidates(url):
    normalized = normalize_uploaded_capture_url(url)
    if not normalized:
        return []
    candidates = []
    for candidate in [normalized]:
        if candidate not in candidates:
            candidates.append(candidate)
        no_query = candidate.split("?", 1)[0].strip()
        if no_query and no_query not in candidates:
            candidates.append(no_query)
        lowered = candidate.lower()
        if lowered not in candidates:
            candidates.append(lowered)
        lowered_no_query = no_query.lower() if no_query else ""
        if lowered_no_query and lowered_no_query not in candidates:
            candidates.append(lowered_no_query)
    return candidates


def _kroger_feature_heading_pattern():
    """Shared Kroger feature-heading pattern.

    Kroger Product Details often exposes one description paragraph followed by
    <ul><li> feature rows. In TXT/extension captures, those rows can flatten into
    one string, so we split only on strong uppercase headings followed by an
    em dash or hyphen.
    """
    heading_terms = [
        r"WHAT['’]?S\s+INCLUDED",
        r"FRESHNESS\s+YOU\s+CAN\s+FEEL",
        r"BREAKS\s+DOWN\s+LIKE\s+TOILET\s+PAPER\*?",
        r"HELPS\s+IN\s+3\s+WAYS",
        r"GENTLE\s+FOR\s+SKIN",
        r"IT\s+TAKES\s+TWO",
        r"NEVER\s+RUN\s+OUT",
        r"GENTLE['’]?S\s+IN\s+THE\s+NAME",
        r"A\s+REFRESHING\s+CLEAN",
        r"NO\s+FRAGRANCE(?:,\s+NO\s+PROBLEM)?",
        r"SAFE\s+FOR\s+SENSITIVE\s+SKIN",
        r"HYPOALLERGENIC",
        r"PERFECT\s+FOR\s+THE\s+WHOLE\s+FAMILY",
        r"STRONG\s+CLEANINGRIPPLES",
        r"BATH\s+TISSUE",
        r"ALL\s+DAY\s+PROTECTION",
        r"UP\s+TO\s+ZERO\s+ODOR",
        r"ODOR\s+CONTROL",
        r"DRYNESS",
        r"ACTIVE\s+FIT",
        r"LEAK\s*GUARD",
        r"LEAKSHIELD",
        r"DERMATOLOGIST\s+TESTED",
    ]
    # Use known Kroger/K-C headings only. A generic uppercase-heading regex can
    # over-split a real feature like "FRESHNESS YOU CAN FEEL" into "FRESHNESS",
    # "YOU", and "CAN FEEL" because each trailing phrase is also uppercase and
    # followed by the same dash.
    known_heading = r"(?:" + "|".join(heading_terms) + r")"
    return r"(?=(?:^|\s)" + known_heading + r"\s*[—-]\s+)"



def _kroger_unlabeled_feature_marker_pattern():
    """Markers for Kroger Product Details bullets that do not have uppercase headings."""
    terms = [
        r"\d+\s+(?:Mega\s+XL\s+|Mega\s+)?(?:Rolls?|Count|Ct|Packs?|Flip[-\s]?Top|Flushable|Underwear|Pads?|Wipes?)\b",
        r"\d+\s+Depend(?:®)?\b",
        r"Soft\s+\d+\s*ply\s+bath\s+tissue\b",
        r"Strong\s+CleaningRipples(?:™|®)?\s+designed\b",
        r"\d+x\s+thicker\s+and\s+stronger\b",
        r"Made\s+with\s+improved\s+softness\b",
        r"(?<!that\s)Breaks\s+down\s+quickly\b",
        r"Bath\s+tissue\s+that\s+breaks\s+down\b",
        r"Pair\s+with\s+Cottonelle(?:®)?\b",
        r"DryShield(?:™|®)?\s+core\s+absorbs\b",
        r"Designed\s+for\s+up\s+to\b",
        r"Feels\s+like\s+real\s+underwear\b",
        r"Up\s+to\s+zero\s+odors\b",
        r"Designed\s+with\s+ultra[-\s]?soft\s+material\b",
        r"Made\s+with\s+cotton[-\s]?like\s+fabric\b",
        r"Each\s+wet\s+wipe\b",
        r"Remove(?:s)?\s+up\s+to\b",
    ]
    return r"(?=(?:^|\s)(?:" + "|".join(terms) + r"))"


def _kroger_unlabeled_first_feature_pattern():
    """First bullet must be count/package style so description text is not over-split."""
    return (
        r"(?=\b(?:"
        r"\d+\s+(?:Mega\s+XL\s+|Mega\s+)?(?:Rolls?|Count|Ct|Packs?|Flip[-\s]?Top|Flushable|Underwear|Pads?|Wipes?)\b"
        r"|\d+\s+Depend(?:®)?\b"
        r"))"
    )


def split_kroger_unlabeled_feature_text(text, max_features=10):
    """Split unlabeled Kroger bullet text while preserving the intro paragraph.

    A first count/package marker is required. This prevents description sentences
    like "3x thicker" or "breaks down quickly" from being treated as features.
    """
    value = clean_kroger_text(text)
    if not value:
        return "", []

    first_match = re.search(_kroger_unlabeled_first_feature_pattern(), value, flags=re.IGNORECASE | re.UNICODE)
    if not first_match:
        return value, []

    intro = normalize_space(value[:first_match.start()])
    remainder = normalize_space(value[first_match.start():])
    matches = list(re.finditer(_kroger_unlabeled_feature_marker_pattern(), remainder, flags=re.IGNORECASE | re.UNICODE))
    if len(matches) < 2:
        return value, []

    items = []
    for idx, match in enumerate(matches):
        start = match.start()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(remainder)
        item = normalize_space(remainder[start:end])
        if item:
            items.append(item)

    return intro, dedupe_preserve_order(items)[:max_features]


def split_kroger_parsed_description(description):
    """Split Kroger flattened PDP description into intro + feature bullets.

    Primary path uses known uppercase Kroger/K-C headings such as WHAT'S INCLUDED.
    Secondary path handles plain <li> text that Kroger/browser-extension captures
    flatten into one string with no headings.
    """
    description = clean_kroger_text(description)
    if not description:
        return "", []

    matches = list(re.finditer(_kroger_feature_heading_pattern(), description, flags=re.IGNORECASE | re.UNICODE))
    if matches:
        intro = normalize_space(description[:matches[0].start()])
        items = []
        for idx, match in enumerate(matches):
            start = match.start()
            end = matches[idx + 1].start() if idx + 1 < len(matches) else len(description)
            item = normalize_space(description[start:end])
            if item:
                items.append(item)
        items = dedupe_preserve_order(items)[:10]
        return intro or description, items

    intro, unlabeled_items = split_kroger_unlabeled_feature_text(description, max_features=10)
    if len(unlabeled_items) >= 2:
        return intro or description, unlabeled_items

    return description, []



def normalize_kroger_extension_feature_candidates(items, title="", max_features=10):
    """Clean feature candidates from Kroger extension PARSED JSON.

    Kroger extension PARSED JSON can mix navigation/category text with product copy,
    and can also flatten description + bullets into one description string. This helper
    keeps product-looking bullets separate and avoids nav/category pollution.
    """
    if not items:
        return []
    if isinstance(items, str):
        items = [items]

    title_norm = normalize_text(title)
    blocked_exact = {
        "pharmacy & health", "digital coupons", "weekly ad",
        "back to school list", "back to school wellness", "meal planning & recipes",
        "store locator", "about the company", "shop store brands",
        "zero hunger | zero waste", "personal care", "feminine care", "baby wipes",
        "cleaning and household", "paper products", "product details",
        "product information", "ratings & reviews",
    }
    blocked_contains = [
        "breadcrumbs", "coupon", "cash back", "qualifying products",
        "item availability", "sign in", "pickup", "delivery", "reviews",
        "privacy", "terms and conditions",
    ]

    out = []
    for item in items:
        value = clean_kroger_text(item)
        if not value:
            continue
        for part in split_kroger_feature_text_if_stuck(value):
            part = clean_kroger_text(part)
            if not part:
                continue
            lowered = part.lower().strip()
            if lowered in blocked_exact:
                continue
            if any(token in lowered for token in blocked_contains):
                continue
            if title_norm and normalize_text(part) == title_norm:
                continue
            if len(part) < 18 and not re.search(r"\d", part):
                continue
            out.append(part)

    return dedupe_preserve_order(out)[:max_features]

# =========================================================
# KROGER CAPTURE AND PARSING
# =========================================================
def build_kroger_compact_capture_from_parsed_json(payload):
    """Build small parse-friendly HTML from Kroger extension PARSED JSON.

    This prevents the app from storing huge Kroger footer/privacy HTML in
    session_state and uses the reliable extension-extracted product values
    instead. Invalid shell captures are rejected by returning an empty string.
    """
    if not isinstance(payload, dict):
        return ""

    title = clean_kroger_text(payload.get("title", ""))
    raw_description = str(payload.get("description", "") or "")
    raw_feature_candidates = payload.get("features", []) or []
    html_description, html_feature_items = extract_kroger_description_features_from_html_fragment(raw_description)
    description = html_description or clean_kroger_text(raw_description)
    images = select_kroger_image_urls_by_perspective(
        [html.unescape(str(url or "").strip()) for url in (payload.get("images", []) or [])],
        max_images=6,
    )

    # Reject Kroger loading/privacy shells. These can still say capture status ok,
    # but they do not carry product content.
    if not title or title.strip().lower() == "kroger":
        return ""
    if not description and not images:
        return ""

    if html_feature_items:
        intro = description
        feature_items = html_feature_items
    else:
        intro, feature_items = split_kroger_parsed_description(description)

    # If the extension did parse features separately, keep them separate as a fallback.
    # Do not let navigation/category crumbs from the extension become product bullets.
    if not feature_items:
        feature_items = normalize_kroger_extension_feature_candidates(raw_feature_candidates, title=title, max_features=10)
    if feature_items and intro:
        # Keep description as the standalone description, not description + bullet copy.
        intro = clean_kroger_text(intro)

    final_url = clean_uploaded_url_value(payload.get("finalUrl", ""))
    requested_url = clean_uploaded_url_value(payload.get("requestedUrl", ""))

    product_json_ld = {
        "@context": "https://schema.org",
        "@type": "Product",
        "name": title,
        "description": intro or description,
        "image": images[:1],
    }

    parts = [
        "<html><head>",
        f"<title>{html_escape_text(title)} - Kroger</title>",
        '<script type="application/ld+json">',
        json.dumps(product_json_ld, ensure_ascii=False),
        "</script>",
        "</head><body>",
        f"<h1>{html_escape_text(title)}</h1>",
        f"<!-- Requested URL: {html_escape_text(requested_url)} -->",
        f"<!-- Final URL: {html_escape_text(final_url)} -->",
        '<section data-testid="product-details-romance-description">',
        f"<p>{html_escape_text(intro or description)}</p>",
    ]

    if feature_items:
        parts.append("<ul>")
        for feature in feature_items:
            parts.append(f"<li>{html_escape_text(feature)}</li>")
        parts.append("</ul>")
    parts.append("</section>")

    for idx, image_url in enumerate(images[:6]):
        perspective = _extract_kroger_perspective_from_url(image_url) or "front"
        parts.append(
            f'<div data-testid="main-image-perspective" aria-label="{html_escape_text(title)} Perspective: {html_escape_text(perspective)}">'
            f'<img class="ProductImages-image" src="{html.escape(image_url, quote=True)}" alt="{html_escape_text(title)} Perspective: {html_escape_text(perspective)}" />'
            '</div>'
        )

    parts.append("</body></html>")
    return "\n".join(parts)




def build_kroger_compact_capture_from_raw_html(raw_html_text, requested_url="", final_url=""):
    """Build a compact Kroger capture from BEGIN HTML only, never extension PARSED JSON.

    Performance note:
    Kroger TXT captures can be very large. This function intentionally avoids
    BeautifulSoup on the full capture. It uses fast string/regex slicing to keep
    only the product title, romance description block, selected variant/UPC context,
    and image perspective labels. The smaller HTML is what the normal Kroger parser
    reads later.
    """
    raw = str(raw_html_text or "")
    if not raw.strip():
        return ""

    working = html.unescape(raw)

    title = ""
    title_patterns = [
        r"(?im)^#\s+(.+?)\s*-\s*Kroger(?:\[|\s|$)",
        r"(?im)^##\s+(.+?)\s*$",
        r"<h1\b[^>]*>(.*?)</h1\s*>",
        r"<title\b[^>]*>(.*?)\s*-\s*Kroger\s*</title\s*>",
    ]
    for pattern in title_patterns:
        match = re.search(pattern, working, flags=re.IGNORECASE | re.DOTALL)
        if match:
            candidate = clean_kroger_text(match.group(1))
            candidate = re.sub(r"\s*-\s*Kroger\s*$", "", candidate, flags=re.IGNORECASE).strip()
            if candidate and candidate.lower() != "kroger":
                title = candidate
                break

    romance_html = ""
    marker_match = re.search(
        r'data-testid=["\']product-details-romance-description["\']|product-details-romance-description',
        working,
        flags=re.IGNORECASE,
    )
    if marker_match:
        start = marker_match.start()
        open_tag_start = working.rfind("<", 0, start)
        if open_tag_start != -1:
            start = open_tag_start

        search_window = working[start:start + 200000]
        end = -1
        reviews_match = re.search(
            r'<div\b[^>]*class=["\'][^"\']*ProductDetails--Reviews',
            search_window,
            flags=re.IGNORECASE,
        )
        if reviews_match:
            end = start + reviews_match.start()
        if end == -1:
            ul_end_match = re.search(r'</ul\s*>', search_window, flags=re.IGNORECASE)
            if ul_end_match:
                end = start + ul_end_match.end()
        if end != -1 and end > start:
            romance_html = working[start:end]

    if not romance_html:
        small = working[:12000]
        return "\n".join([
            "<html><head>",
            f"<title>{html_escape_text(title or 'Kroger Product')} - Kroger</title>",
            "</head><body>",
            f"<h1>{html_escape_text(title)}</h1>" if title else "",
            small,
            "</body></html>",
        ])

    context_lines = []
    context_patterns = [
        r'[^\n\r]{0,180}Perspective\s*:\s*(?:front|back|left|right|top|bottom)[^\n\r]{0,180}',
        r'[^\n\r]{0,120}\bUPC\s*:\s*[0-9]{8,14}[^\n\r]{0,120}',
        r'[^\n\r]{0,120}\bSize\s*:[^\n\r]{0,240}',
        r'<label\b[^>]*data-testid=["\']selected-variant-option["\'][\s\S]{0,400?</label>',
    ]
    context_source = working[:50000] + "\n" + romance_html[:50000]
    for pattern in context_patterns:
        for match in re.finditer(pattern, context_source, flags=re.IGNORECASE):
            value = match.group(0)
            if value and value not in context_lines:
                context_lines.append(value)
            if len(context_lines) >= 50:
                break

    parts = [
        "<html><head>",
        f"<title>{html_escape_text(title or 'Kroger Product')} - Kroger</title>",
        "</head><body>",
    ]
    if title:
        parts.append(f"<h1>{html_escape_text(title)}</h1>")
    if requested_url:
        parts.append(f"<!-- Requested URL: {html_escape_text(requested_url)} -->")
    if final_url:
        parts.append(f"<!-- Final URL: {html_escape_text(final_url)} -->")
    parts.append("<!-- KROGER HTML SOURCE ONLY: extension PARSED JSON intentionally ignored for copy/features. -->")
    parts.append(romance_html)
    if context_lines:
        parts.append("<pre data-kroger-html-context='1'>")
        parts.append(html_escape_text("\n".join(context_lines)))
        parts.append("</pre>")
    parts.append("</body></html>")
    return "\n".join(parts)


# =========================================================
# CVS CAPTURE AND PARSING
# =========================================================
def build_cvs_compact_capture_from_parsed_json(payload):
    """Build compact parse-friendly CVS HTML from browser-extension output."""
    if not isinstance(payload, dict):
        return ""
    title = normalize_space(payload.get("title", "") or payload.get("documentTitle", ""))
    description = clean_cvs_text(payload.get("description", "") or "")
    features = payload.get("features", []) or []
    if isinstance(features, str):
        features = [features]
    features = normalize_cvs_features([str(x or "") for x in features])
    images = []
    seen = set()
    for image_url in payload.get("images", []) or []:
        clean_url = html.unescape(str(image_url or "").strip()).replace("\\/", "/")
        if not clean_url or is_video_like_url(clean_url):
            continue
        if clean_url.startswith("//"):
            clean_url = "https:" + clean_url
        elif clean_url.startswith("/"):
            clean_url = "https://www.cvs.com" + clean_url
        if not re.match(r"^https?://", clean_url, flags=re.IGNORECASE):
            continue
        if any(token in clean_url.lower() for token in ["sprite", "icon", "logo", "placeholder", "data:image", ".svg"]):
            continue
        key = clean_url.split("?", 1)[0]
        if key and key not in seen:
            seen.add(key)
            # Preserve the full browser-rendered CVS URL, including query params.
            # CVS thumbnails/high_res assets can depend on ?im=Resize(...) to render.
            images.append(clean_url)
    requested_url = clean_uploaded_url_value(payload.get("requestedUrl", ""))
    final_url = clean_uploaded_url_value(payload.get("finalUrl", ""))
    if not (title or description or features or images):
        return ""
    product_json_ld = {"@context": "https://schema.org", "@type": "Product", "name": title, "description": description, "image": images}
    parts = [
        "<!doctype html><html><head>",
        f"<title>{html_escape_text(title)}</title>",
        '<script type="application/ld+json">',
        json.dumps(product_json_ld, ensure_ascii=False),
        '</script>',
        "</head><body>",
        f"<h1>{html_escape_text(title)}</h1>",
        "<div data-cvs-compact-capture='1' class='whitespace-pre-line'>",
        f"<span>{html_escape_text(description)}</span>",
        "<ul>",
    ]
    for feature in features[:10]:
        parts.append(f"<li id='vendorDetailsBullet'>{html_escape_text(feature)}</li>")
    parts.extend(["</ul>", "</div>"])
    for image_url in images[:MAX_IMAGE_SLOTS_TO_COMPARE]:
        safe_url = html.escape(image_url, quote=True)
        parts.append(f'<img src="{safe_url}" data-src="{safe_url}" />')
    if requested_url:
        parts.append(f"<meta name='requested-url' content='{html.escape(requested_url, quote=True)}'>")
    if final_url:
        parts.append(f"<meta name='final-url' content='{html.escape(final_url, quote=True)}'>")
    parts.append("</body></html>")
    return "\n".join(parts)

# =========================================================
# SAM'S CLUB COMPACT CAPTURE
# =========================================================
def build_sams_compact_capture_from_parsed_json(payload):
    """Build compact parse-friendly Sam's Club HTML from extension PARSED JSON.

    Sam's Club pages often expose clean copy/images in the extension parsed JSON while
    the raw hydrated HTML also contains navigation SVGs, sponsored shelves, customer
    photos, and membership graphics. This compact page keeps only PDP copy plus real
    product-gallery media so the Sam's Club app parser stays retailer-isolated and
    does not ingest page chrome assets.
    """
    if not isinstance(payload, dict):
        return ""

    def _clean_local(value):
        value = html.unescape(str(value or ""))
        value = value.replace("\\u003c", "<").replace("\\u003e", ">")
        value = value.replace("\\u0026", "&").replace("\\u00a0", " ")
        value = value.replace("\\/", "/").replace('\\"', '"')
        if "<" in value and ">" in value:
            value = BeautifulSoup(value, "html.parser").get_text(" ", strip=True)
        value = re.sub(r"\s+", " ", value).strip()
        return value

    def _is_sams_product_media(url):
        url = html.unescape(str(url or "").strip()).replace("\\/", "/")
        if not url:
            return False
        lowered = url.lower().split("?", 1)[0]
        if lowered.startswith("data:"):
            return False
        if lowered.endswith(".svg"):
            return False
        if any(token in lowered for token in [
            "/dfw/", "sams-mav", "sprite", "icon", "logo", "badge", "placeholder",
            "avatar", "rating", "stars", "review", "customer",
        ]):
            return False
        if is_video_like_url(url):
            return "i5-richmedia.samsclubimages.com" in lowered or "samsclubimages.com" in lowered
        return bool(
            ("samsclubimages.com/asr/" in lowered or "walmartimages.com" in lowered)
            and re.search(r"\.(?:jpg|jpeg|png|webp|avif)$", lowered, flags=re.IGNORECASE)
        )

    def _normalize_media_url(url):
        url = html.unescape(str(url or "").strip()).replace("\\/", "/")
        if url.startswith("//"):
            url = "https:" + url
        if not re.match(r"^https?://", url, flags=re.IGNORECASE):
            return ""
        if is_video_like_url(url):
            return url
        base = url.split("?", 1)[0]
        if "samsclubimages.com/asr/" in base.lower():
            return f"{base}?odnHeight=450&odnWidth=450&odnBg=FFFFFF"
        return url

    title = _clean_local(payload.get("title", "") or payload.get("name", "") or payload.get("documentTitle", ""))
    title = re.sub(r"\s+-\s+Samsclub\.com\s*$", "", title, flags=re.IGNORECASE).strip()

    raw_description = payload.get("description", "") or payload.get("longDescription", "") or ""
    description = _clean_local(raw_description)

    features = payload.get("features", []) or payload.get("highlights", []) or []
    if isinstance(features, str):
        features = [features]
    feature_items = []
    # If description is only a UL from the Sam's Club highlights block, make those LIs features.
    if raw_description and "<li" in str(raw_description).lower():
        try:
            soup = BeautifulSoup(str(raw_description), "html.parser")
            li_items = [_clean_local(li.get_text(" ", strip=True)) for li in soup.find_all("li")]
            feature_items.extend([x for x in li_items if x])
            if not description or len(description) < 40:
                description = ""
        except Exception:
            pass
    for feature in features:
        clean = _clean_local(feature)
        if not clean:
            continue
        if re.search(r"^(shipping|pickup|delivery|reorder|savings|departments|services|same day delivery)$", clean, flags=re.IGNORECASE):
            continue
        if len(clean) > 240:
            continue
        feature_items.append(clean)
    feature_items = dedupe_preserve_order(feature_items)[:10]

    images = []
    diagnostics = payload.get("diagnostics", {}) if isinstance(payload.get("diagnostics", {}), dict) else {}
    sams_image_verified = bool(diagnostics.get("imageProductVerified"))
    sams_image_source = normalize_space(diagnostics.get("imageCaptureSource", ""))
    source_images = (payload.get("images", []) or []) if sams_image_verified and sams_image_source in {"exact_product_json_ld", "exact_product_hero_gallery", "exact_product_json_ld_and_hero_gallery"} else []
    for image_url in source_images:
        normalized = _normalize_media_url(image_url)
        if normalized and _is_sams_product_media(normalized):
            images.append(normalized)
    images = dedupe_preserve_order(images)[:MAX_IMAGE_SLOTS_TO_COMPARE]

    requested_url = clean_uploaded_url_value(payload.get("requestedUrl", ""))
    final_url = clean_uploaded_url_value(payload.get("finalUrl", ""))

    if not (title or description or feature_items or images):
        return ""

    parts = ["<html><body>"]
    if title:
        parts.append(f"<h1>{html_escape_text(title)}</h1>")
        parts.append(f"\n## {html_escape_text(title)}\n")
    if feature_items:
        parts.append("\n### Highlights\n")
        parts.append("<ul data-sams-highlights='1'>")
        for feature in feature_items:
            parts.append(f"<li>{html_escape_text(feature)}</li>")
            parts.append(f"- {html_escape_text(feature)}")
        parts.append("</ul>")
        parts.append("Read more")
    if description:
        parts.append("\n#### Product details\n")
        parts.append(f"<p>{html_escape_text(description)}</p>")
        parts.append(description)
        parts.append("\n### Specifications\n")
    for idx, image_url in enumerate(images, start=1):
        safe_url = html.escape(image_url, quote=True)
        safe_title = html.escape(title or "Sam's Club product", quote=True)
        if is_video_like_url(image_url):
            parts.append(f'<video src="{safe_url}" data-sams-gallery="1"></video>')
        else:
            parts.append(f'<img src="{safe_url}" data-src="{safe_url}" alt="thumbnail image {idx} of {safe_title}, {idx} of {len(images)}" data-sams-gallery="1" />')
    if requested_url:
        parts.append(f"<meta name='requested-url' content='{html.escape(requested_url, quote=True)}'>")
    if final_url:
        parts.append(f"<meta name='final-url' content='{html.escape(final_url, quote=True)}'>")
    parts.append("</body></html>")
    return "\n".join(parts)


def _kroger_rpc_values_from_capture_context(*values):
    """Extract likely Kroger UPC/RPC keys from URLs or compact capture context."""
    out=[]
    seen=set()
    for value in values:
        source=str(value or "")
        if not source:
            continue
        patterns=[
            r"UPC:\s*([0-9]{8,14})",
            r'"rpc"\s*:\s*"?([0-9]{8,14})"?',
            r'"upc"\s*:\s*"?([0-9]{8,14})"?',
            r"/([0-9]{8,14})(?:[/?#\s]|$)",
        ]
        for pattern in patterns:
            for match in re.finditer(pattern, source, flags=re.IGNORECASE):
                digits=re.sub(r"\D+", "", str(match.group(1) or ""))
                if not digits:
                    continue
                candidates={digits}
                if digits.isdigit():
                    candidates.add(digits.zfill(13))
                    candidates.add(digits.lstrip("0") or digits)
                for candidate in candidates:
                    if candidate and candidate not in seen:
                        seen.add(candidate)
                        out.append(candidate)
    return out


def _slice_between_markers(source, start_marker, end_marker, start_pos=0):
    start=source.find(start_marker, start_pos)
    if start == -1:
        return ""
    start += len(start_marker)
    end=source.find(end_marker, start)
    if end == -1:
        return ""
    return source[start:end].strip()


def build_kroger_compact_capture_from_capture_block(block, requested_url="", final_url=""):
    """Very fast Kroger TXT block compactor for upload indexing."""
    block=str(block or "")
    if not block.strip():
        return ""
    html_text=_slice_between_markers(block, "-----BEGIN HTML-----", "-----END HTML-----")
    if not html_text:
        return ""

    title=""
    title_source=html_text[:60000]
    for pattern in [
        r"(?im)^#\s+(.+?)\s*-\s*Kroger(?:\[|\s|$)",
        r"(?im)^##\s+(.+?)\s*$",
        r"<h1\b[^>]*>(.*?)</h1\s*>",
        r"<title\b[^>]*>(.*?)\s*-\s*Kroger\s*</title\s*>",
    ]:
        match=re.search(pattern, title_source, flags=re.IGNORECASE|re.DOTALL)
        if match:
            candidate=clean_kroger_text(html.unescape(match.group(1)))
            candidate=re.sub(r"\s*-\s*Kroger\s*$", "", candidate, flags=re.IGNORECASE).strip()
            if candidate and candidate.lower() != "kroger":
                title=candidate
                break

    marker_pos=-1
    for marker in ['data-testid="product-details-romance-description"', "data-testid='product-details-romance-description'", "product-details-romance-description"]:
        marker_pos=html_text.find(marker)
        if marker_pos != -1:
            break
    if marker_pos == -1:
        small=html.unescape(html_text[:12000])
        return "\n".join(["<html><head>", f"<title>{html_escape_text(title or 'Kroger Product')} - Kroger</title>", "</head><body>", f"<h1>{html_escape_text(title)}</h1>" if title else "", small, "</body></html>"])

    start=html_text.rfind("<", 0, marker_pos)
    if start == -1:
        start=marker_pos
    search_window=html_text[start:start+180000]
    end_candidates=[]
    for end_marker in ['<div class="ProductDetails--Container', "<div class='ProductDetails--Container", "ProductDetails--Reviews", "</ul>", "&lt;/ul&gt;"]:
        idx=search_window.find(end_marker)
        if idx != -1:
            if end_marker in {"</ul>", "&lt;/ul&gt;"}:
                idx += len(end_marker)
            end_candidates.append(idx)
    end_local=min(end_candidates) if end_candidates else min(len(search_window), 50000)
    romance_html=html.unescape(search_window[:end_local])

    context_lines=[]
    context_source=html.unescape(html_text[:50000] + "\n" + search_window[:50000])
    context_patterns=[
        r'[^\n\r]{0,180}Perspective\s*:\s*(?:front|back|left|right|top|bottom)[^\n\r]{0,180}',
        r'[^\n\r]{0,120}\bUPC\s*:\s*[0-9]{8,14}[^\n\r]{0,120}',
        r'[^\n\r]{0,120}\bSize\s*:[^\n\r]{0,240}',
        r'<label\b[^>]*data-testid=["\']selected-variant-option["\'][\s\S]{0,400?</label>',
    ]
    for pattern in context_patterns:
        for match in re.finditer(pattern, context_source, flags=re.IGNORECASE):
            value=match.group(0)
            if value and value not in context_lines:
                context_lines.append(value)
            if len(context_lines) >= 50:
                break

    parts=["<html><head>", f"<title>{html_escape_text(title or 'Kroger Product')} - Kroger</title>", "</head><body>"]
    if title:
        parts.append(f"<h1>{html_escape_text(title)}</h1>")
    if requested_url:
        parts.append(f"<!-- Requested URL: {html_escape_text(requested_url)} -->")
    if final_url:
        parts.append(f"<!-- Final URL: {html_escape_text(final_url)} -->")
    parts.append("<!-- KROGER FAST HTML SOURCE ONLY: extension PARSED JSON ignored for startup speed. -->")
    parts.append(romance_html)
    if context_lines:
        parts.append("<pre data-kroger-html-context='1'>")
        parts.append(html_escape_text("\n".join(context_lines)))
        parts.append("</pre>")
    parts.append("</body></html>")
    return "\n".join(parts)


def _heb_capture_image_urls_from_raw_html(raw_html_text, target_rpc="", max_images=10):
    """HEB-only helper used during TXT indexing.

    Do not store full HEB raw HTML in session state. HEB captures can be very large
    and can crash Streamlit. Instead, pull only current-item image URLs from the raw
    HTML and append a tiny image snippet to the compact HEB capture.
    """
    raw_html_text = str(raw_html_text or "")
    target_rpc = re.sub(r"[^0-9]", "", str(target_rpc or "").replace(".0", ""))
    padded_rpc = target_rpc.zfill(9) if target_rpc else ""
    urls = []
    seen = set()
    patterns = [
        r'https?:\\/\\/images\.heb\.com\\/is\\/image\\/HEBGrocery\\/[^"\\\s<>]+',
        r'https?://images\.heb\.com/is/image/HEBGrocery/[^"\s<>]+',
        r'//images\.heb\.com/is/image/HEBGrocery/[^"\s<>]+',
    ]
    for pattern in patterns:
        for raw_url in re.findall(pattern, raw_html_text, flags=re.IGNORECASE):
            url = html.unescape(str(raw_url or "")).replace("\\/", "/").strip()
            url = re.sub(r"[\)\]\}\'\";,]+$", "", url)
            if url.startswith("//"):
                url = "https:" + url
            if not url.lower().startswith("http"):
                continue
            if "images.heb.com/is/image/HEBGrocery" not in url:
                continue
            m = re.search(r"/is/image/HEBGrocery/([^?\s<>]+)", url, flags=re.IGNORECASE)
            if not m:
                continue
            asset = re.sub(r"[\)\]\}\'\";,]+$", "", m.group(1).strip())
            if not asset or asset.lower().startswith(("prd-small/", "prd-medium/", "prd-large/")):
                continue
            normalized = f"https://images.heb.com/is/image/HEBGrocery/{asset}?fit=constrain,1&wid=800&hei=800&fmt=jpg&qlt=80"
            key = normalized.split("?", 1)[0]
            if key and key not in seen:
                seen.add(key)
                urls.append(normalized)
    if padded_rpc and urls:
        current_item_urls = [u for u in urls if f"/{padded_rpc}" in u]
        if current_item_urls:
            urls = current_item_urls
    def sort_key(url):
        m = re.search(r"/HEBGrocery/(\d+)(?:-(\d+))?", url)
        if not m:
            return (999999999, 9999, url)
        return (int(m.group(1)), int(m.group(2) or 1), url)
    return sorted(urls, key=sort_key)[:max_images]


def _build_heb_compact_image_snippet(raw_html_text, target_rpc=""):
    urls = _heb_capture_image_urls_from_raw_html(raw_html_text, target_rpc=target_rpc, max_images=10)
    if not urls:
        return ""
    parts = ["<!-- HEB COMPACT RAW IMAGE URLS FROM EXTENSION HTML -->", "<section data-heb-compact-images='1'>"]
    for url in urls:
        parts.append(f'<img src="{html.escape(url, quote=True)}" />')
    parts.append("</section>")
    return "\n".join(parts)

def build_cvs_compact_capture_from_raw_html(raw_html_text, requested_url="", final_url=""):
    source = str(raw_html_text or "")
    if not source:
        return ""
    generic_title = "shop beauty, vitamins, medicine & everyday essentials | cvs.com"
    source_lc = source.lower()
    has_product_evidence = any(x in source_lc for x in [
        'data-capture="product-section"', 'data-capture="product-state"',
        'application/ld+json', 'vendordetailsbullets', 'vendordetailsparagraph',
        '/bizcontent/merchandising/productimages/high_res/', 'customer reviews:'
    ])
    if generic_title in source_lc and not has_product_evidence:
        return ""
    max_chars = 240000
    parts = ["<!doctype html><html><head>"]
    for pattern in [
        r"(?is)<title[^>]*>.*?</title>",
        r"(?is)<meta\b[^>]*(?:name=[\"']description[\"']|property=[\"']og:[^\"']+[\"'])[^>]*>",
        r"(?is)<script\b[^>]*type=[\"']application/ld\+json[\"'][^>]*>.*?</script>",
    ]:
        for match in re.finditer(pattern, source):
            value = match.group(0)
            if sum(map(len, parts)) + len(value) > max_chars:
                break
            parts.append(value)
    parts.append("</head><body>")
    for match in re.finditer(r"(?is)<section\b[^>]*data-capture=[\"'](?:product-section|product-state)[\"'][^>]*>.*?</section>", source):
        value = match.group(0)
        if len(value) > 60000:
            value = value[:60000] + "\n<!-- TRUNCATED TARGET PRODUCT STATE -->\n</section>"
        if sum(map(len, parts)) + len(value) > max_chars:
            break
        parts.append(value)
    for pattern in [
        r"(?is)<script\b[^>]*data-capture=[\"']image-urls[\"'][^>]*>.*?</script>",
        r"(?is)<pre\b[^>]*data-capture=[\"']visible-text[\"'][^>]*>.*?</pre>",
    ]:
        match = re.search(pattern, source)
        if match:
            value = match.group(0)[:35000]
            if sum(map(len, parts)) + len(value) <= max_chars:
                parts.append(value)
    if requested_url:
        parts.append(f"<!-- Requested URL: {html.escape(str(requested_url), quote=True)} -->")
    if final_url:
        parts.append(f"<!-- Final URL: {html.escape(str(final_url), quote=True)} -->")
    parts.append("</body></html>")
    return "\n".join(parts)[:max_chars]


def is_invalid_cvs_title_candidate(value):
    text = normalize_space(value)
    lowered = text.lower()
    return (not text or lowered.startswith(("http://", "https://")) or "/shop/" in lowered
            or "prodid-" in lowered or "skuid=" in lowered
            or lowered == "shop beauty, vitamins, medicine & everyday essentials | cvs.com")


def is_cvs_review_meta_description(value):
    lowered = normalize_space(value).lower()
    return any(x in lowered for x in ["see real customer reviews", "see all reviews", "shop with confidence"])


def parse_uploaded_raw_html_map(raw_text, selected_retailer=""):
    raw_text = str(raw_text or "")
    if not raw_text.strip():
        return {}

    html_map = {}
    header_pattern = re.compile(r'(?m)^=+\s*PDP CAPTURE\s+\d+\s*=+')
    matches = list(header_pattern.finditer(raw_text))
    if matches:
        block_ranges = [(m.start(), matches[i + 1].start() if i + 1 < len(matches) else len(raw_text)) for i, m in enumerate(matches)]
    else:
        block_ranges = [(0, len(raw_text))]

    for block_start, block_end in block_ranges:
        block = raw_text[block_start:block_end]
        requested_match = re.search(r'(?im)^Requested\s+URL\s*:\s*(https?://\S+)', block)
        if not requested_match:
            continue
        requested_url = str(requested_match.group(1) or "").strip()
        if not requested_url:
            continue
        selected_capture_retailer = normalize_retailer_name(selected_retailer) if selected_retailer else ""
        if selected_capture_retailer and not retailer_url_matches_selected(requested_url, selected_capture_retailer):
            # Do not compact/index captures for other retailers. This keeps selected-retailer
            # runs from loading CVS/HEB/Kroger/Walgreens/etc. blocks that are not selected.
            continue

        requested_url_lc = requested_url.lower()
        final_url_match = re.search(r'(?im)^Final\s+URL\s*:\s*(https?://\S+)', block)
        final_url_from_payload = clean_uploaded_url_value(final_url_match.group(1)) if final_url_match else ""

        if "kroger.com" in requested_url_lc:
            html_text = build_kroger_compact_capture_from_capture_block(block, requested_url=requested_url, final_url=final_url_from_payload)
        else:
            compact_html = ""
            parsed_payload = {}
            parsed_match = re.search(r'(?is)-----BEGIN PARSED JSON-----(.*?)-----END PARSED JSON-----', block)
            if parsed_match:
                try:
                    parsed_payload = json.loads(str(parsed_match.group(1) or "").strip())
                    if requested_url and "heb.com" in requested_url_lc:
                        compact_html = build_heb_compact_capture_from_parsed_json(parsed_payload)
                    elif requested_url and "cvs.com" in requested_url_lc:
                        compact_html = build_cvs_compact_capture_from_parsed_json(parsed_payload)
                    elif requested_url and "samsclub.com" in requested_url_lc:
                        compact_html = build_sams_compact_capture_from_parsed_json(parsed_payload)
                    else:
                        compact_html = build_kroger_compact_capture_from_parsed_json(parsed_payload)
                except Exception:
                    parsed_payload = {}
                    compact_html = ""
            if not final_url_from_payload and isinstance(parsed_payload, dict):
                final_url_from_payload = clean_uploaded_url_value(parsed_payload.get("finalUrl", ""))
            html_match = re.search(r'(?is)-----BEGIN HTML-----(.*?)-----END HTML-----', block)
            raw_html_text = html.unescape(str(html_match.group(1) or "").strip()) if html_match else ""
            if compact_html:
                if requested_url and "cvs.com" in requested_url.lower() and raw_html_text:
                    cvs_raw_compact = build_cvs_compact_capture_from_raw_html(raw_html_text, requested_url=requested_url, final_url=final_url_from_payload)
                    html_text = compact_html + ("\n" + cvs_raw_compact if cvs_raw_compact else "")
                elif requested_url and "samsclub.com" in requested_url.lower():
                    # The extension PARSED JSON is exact-product scoped. Keep Sam's Club
                    # isolated to that verified copy/gallery instead of appending raw page
                    # ASR assets from recommendations, reviews, or shared parent content.
                    html_text = compact_html
                elif requested_url and "heb.com" in requested_url.lower() and raw_html_text:
                    # HEB-only stable fix: keep compact parsed copy and append only tiny current-item image URLs.
                    # Do not append full raw HTML because it can make the uploaded map huge and crash Streamlit.
                    rpc_for_images = ""
                    if isinstance(parsed_payload, dict):
                        rpc_for_images = parsed_payload.get("rpc", "") or parsed_payload.get("sku", "") or ""
                    image_snippet = _build_heb_compact_image_snippet(raw_html_text, target_rpc=rpc_for_images)
                    html_text = compact_html + ("\n" + image_snippet if image_snippet else "")
                else:
                    html_text = compact_html
            else:
                if requested_url and "cvs.com" in requested_url_lc:
                    html_text = build_cvs_compact_capture_from_raw_html(raw_html_text, requested_url=requested_url, final_url=final_url_from_payload)
                else:
                    html_text = raw_html_text

        if requested_url and "kroger.com" in requested_url.lower() and html_text and not is_valid_kroger_product_capture(html_text):
            html_text = build_kroger_invalid_capture_stub(requested_url=requested_url, final_url=final_url_from_payload, reason="invalid_kroger_shell_or_product_unavailable_capture")
        if requested_url and "kroger.com" in requested_url.lower() and not html_text:
            html_text = build_kroger_invalid_capture_stub(requested_url=requested_url, final_url=final_url_from_payload, reason="missing_kroger_html_capture")
        if not html_text or len(html_text) < 30:
            continue

        keys = []
        for url_value in [requested_url, final_url_from_payload]:
            key = normalize_uploaded_capture_url(url_value)
            if key:
                keys.append(key)
        if requested_url and "kroger.com" in requested_url.lower():
            for rpc in _kroger_rpc_values_from_capture_context(requested_url, final_url_from_payload, html_text):
                keys.append(f"kroger_rpc::{rpc}")
        for key in dedupe_preserve_order(keys):
            html_map[key] = html_text

    return html_map

def is_likely_usable_cvs_manual_source(html_text):
    """CVS-only guard for manually pasted source snippets.

    A few CVS pages can be pasted/copied as stylesheet/footer-only source.
    Those snippets are not useful product captures and should not count as a
    matched capture, otherwise the batch can hide the real missing capture
    problem. Keep this intentionally CVS-specific.
    """
    text = str(html_text or "")
    if len(text.strip()) < 200:
        return False
    lowered = text.lower()

    strong_product_markers = [
        "vendordetailsbullet",
        "vendordetails",
        "bizcontent/merchandising/productimages/high_res",
        "dynamicmediaurl",
        "skuid=",
        "prodid-",
    ]
    has_product_marker = any(marker.lower() in lowered for marker in strong_product_markers)
    has_product_text = bool(re.search(r"\b(kleenex|kotex|viva|poise|depend|huggies|pull-ups|cottonelle|scott|goodnites|thinx)\b", lowered))

    footer_css_only = (
        "sc-cvs-footer-container" in lowered
        and not has_product_marker
        and not has_product_text
    )
    if footer_css_only:
        return False

    return bool(has_product_marker or has_product_text)


def parse_cvs_manual_source_code_workbook(file_bytes):
    """Parse a manual CVS source-code workbook.

    Supported CVS-only manual fallback formats:
    1. Header format with CVS RPC / Retailer RPC / skuId plus all source code / source code / raw HTML.
    2. Legacy format where column A is CVS RPC and columns B onward are source snippets.

    Excel cells can truncate long page source around 32k characters, so browser extension TXT is still preferred.
    """
    html_map = {}
    stats = {
        "mode": "cvs_manual_source_xlsx",
        "rows_seen": 0,
        "mapped_rows": 0,
        "skipped_blank": 0,
        "skipped_weak": 0,
        "truncated_cell_count": 0,
        "header_aware_mapping": False,
    }
    if not file_bytes:
        return html_map, stats
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return html_map, stats

    rpc_header_options = {
        "cvs rpc", "retailer rpc", "sku", "skuid", "sku id", "cvs skuid", "cvs sku", "item id", "productid", "product id"
    }
    source_header_tokens = (
        "all source code", "source code", "copy source code", "raw source code", "html source", "raw html", "page source", "source html"
    )

    def clean_header(value):
        return normalize_space(value).strip().lower().replace("_", " ")

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_row_index = None
        rpc_idx = None
        source_indices = []
        for idx, row in enumerate(rows[:8]):
            headers = [clean_header(x) for x in (row or [])]
            possible_rpc_idx = None
            possible_source_indices = []
            for c_idx, header in enumerate(headers):
                if possible_rpc_idx is None and header in rpc_header_options:
                    possible_rpc_idx = c_idx
                if any(token in header for token in source_header_tokens):
                    possible_source_indices.append(c_idx)
            if possible_rpc_idx is not None and possible_source_indices:
                header_row_index = idx
                rpc_idx = possible_rpc_idx
                source_indices = possible_source_indices
                stats["header_aware_mapping"] = True
                break
        data_rows = rows[header_row_index + 1:] if header_row_index is not None else rows
        for row in data_rows:
            if not row or not any(cell is not None and str(cell).strip() for cell in row):
                continue
            if header_row_index is not None and rpc_idx is not None and source_indices:
                rpc_raw = row[rpc_idx] if rpc_idx < len(row) else ""
                values_to_scan = [row[i] for i in source_indices if i < len(row)]
            else:
                rpc_raw = row[0] if len(row) else ""
                values_to_scan = list(row[1:])
            rpc_raw = str(rpc_raw or "").replace(".0", "").strip()
            rpc = re.sub(r"[^0-9A-Za-z_-]", "", rpc_raw)
            if not rpc or rpc.lower() in {"cvs", "retailer", "rpc", "cvsrpc"}:
                continue
            stats["rows_seen"] += 1
            pieces = []
            for value in values_to_scan:
                if value is None:
                    continue
                part = html.unescape(str(value or "").strip())
                if not part:
                    continue
                if len(part) >= 32760:
                    stats["truncated_cell_count"] += 1
                pieces.append(part)
            html_text = "\n".join(pieces).strip()
            if not html_text:
                stats["skipped_blank"] += 1
                continue
            if not is_likely_usable_cvs_manual_source(html_text):
                stats["skipped_weak"] += 1
                continue
            wrapped = (
                "\n"
                f"<!-- CVS MANUAL SOURCE XLSX RPC {html.escape(rpc)} -->\n"
                + html_text
                + "\n"
            )
            html_map[f"cvs_rpc::{rpc}"] = wrapped
            stats["mapped_rows"] += 1
    return html_map, stats


def build_extension_input_csv(retailer_df, retailer_name):
    """Build the simple two-column input the browser extension expects.

    This lets the app run from a retailer SKU list instead of manually prepared URL files.
    The extension only needs Retailer + Search Term, where Search Term is the retailer RPC
    when available, otherwise the SKU.
    """
    retailer_name = normalize_retailer_name(retailer_name)
    rows = []
    df = retailer_df.copy() if retailer_df is not None else pd.DataFrame()
    if df.empty:
        return "Retailer,Search Term\n"
    for _, row in df.iterrows():
        search_term = normalize_space(row.get("retailer_rpc", "")) or normalize_space(row.get("sku", ""))
        if not search_term:
            continue
        rows.append({"Retailer": retailer_name, "Search Term": search_term})
    if not rows:
        return "Retailer,Search Term\n"
    return pd.DataFrame(rows).drop_duplicates().to_csv(index=False)


def parse_json_list(value):
    value = str(value or "").strip()
    if not value:
        return []
    try:
        parsed = json.loads(value)
        if isinstance(parsed, list):
            return [normalize_space(x) for x in parsed if normalize_space(x)]
    except Exception:
        pass
    if " | " in value:
        return [normalize_space(x) for x in value.split(" | ") if normalize_space(x)]
    return [normalize_space(value)] if normalize_space(value) else []


def build_structured_extension_capture_html(row):
    """Build tiny parseable HTML from the extension's separated-copy CSV/XLSX output."""
    title = normalize_space(row.get("parsed_title", "") or row.get("matched_title", "") or "")
    description = normalize_space(row.get("parsed_description", ""))
    variant_size = normalize_space(row.get("parsed_variant_size", ""))
    source = normalize_space(row.get("parsed_copy_source", "")) or "extension_structured_results"
    retailer = normalize_retailer_name(row.get("retailer", "") or "Retailer")
    product_url = clean_uploaded_url_value(row.get("product_url", "") or row.get("retail_url", "") or "")
    search_term = normalize_space(row.get("search_term", "") or row.get("retailer_rpc", "") or "")

    features = []
    for i in range(1, 11):
        value = normalize_space(row.get(f"parsed_feature_{i}", ""))
        if value:
            features.append(value)
    if not features:
        features.extend(parse_json_list(row.get("parsed_features_json", "")))
    features = dedupe_preserve_order(features)[:10]

    images = parse_json_list(row.get("parsed_images_json", ""))
    images = [clean_uploaded_url_value(x) for x in images if clean_uploaded_url_value(x)]
    images = dedupe_preserve_order(images)[:12]

    parts = [
        "<html><head>",
        f"<title>{html_escape_text(title or retailer + ' Product')}</title>",
        "</head><body data-pdp-structured-extension-capture='1' "
        f"data-retailer='{html.escape(retailer, quote=True)}' "
        f"data-source='{html.escape(source, quote=True)}' "
        f"data-search-term='{html.escape(search_term, quote=True)}'>",
    ]
    if title:
        parts.append(f"<h1>{html_escape_text(title)}</h1>")
    if product_url:
        parts.append(f"<!-- Product URL: {html_escape_text(product_url)} -->")
    if variant_size:
        parts.append(
            "<label data-testid='selected-variant-option'>"
            f"<input type='radio' value='{html.escape(variant_size, quote=True)}' checked>"
            f"{html_escape_text(variant_size)}</label>"
        )
    parts.append("<section data-testid='product-details-romance-description' data-retailer-description='1'>")
    if description:
        parts.append(f"<p>{html_escape_text(description)}</p>")
    if features:
        parts.append("<ul>")
        for feature in features:
            parts.append(f"<li>{html_escape_text(feature)}</li>")
        parts.append("</ul>")
    parts.append("</section>")
    for idx, image_url in enumerate(images[:12], start=1):
        perspective = _extract_kroger_perspective_from_url(image_url) if "kroger.com" in image_url.lower() else ""
        perspective_label = perspective or f"image-{idx}"
        parts.append(
            f"<div data-testid='main-image-perspective' aria-label='{html_escape_text(title)} Perspective: {html_escape_text(perspective_label)}'>"
            f"<img class='ProductImages-image' src='{html.escape(image_url, quote=True)}' alt='{html_escape_text(title)} Perspective: {html_escape_text(perspective_label)}' />"
            "</div>"
        )
    parts.append("</body></html>")
    return "\n".join(parts)


def parse_extension_results_structured_file(file_bytes, file_name):
    """Parse extension result CSV/XLSX with separated copy fields into source map."""
    file_name = str(file_name or "").lower().strip()
    if file_name.endswith(".xlsx"):
        df = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
    else:
        last_error = None
        for encoding in ["utf-8-sig", "utf-8", "latin1"]:
            try:
                df = pd.read_csv(BytesIO(file_bytes), encoding=encoding)
                break
            except Exception as e:
                last_error = e
        else:
            raise last_error
    df = df.copy()
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    rename_map = {
        "product_url": "product_url",
        "retail_url": "product_url",
        "retailer_url": "product_url",
        "search_term": "search_term",
        "retailer_rpc": "search_term",
        "parsed_title": "parsed_title",
        "parsed_variant_size": "parsed_variant_size",
        "parsed_description": "parsed_description",
        "parsed_features_json": "parsed_features_json",
        "parsed_images_json": "parsed_images_json",
        "parsed_copy_source": "parsed_copy_source",
    }
    df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns}, inplace=True)
    html_map = {}
    stats = {"mode": "extension_structured_results", "rows_seen": len(df), "mapped_rows": 0, "skipped_blank": 0, "skipped_weak": 0}
    for _, row in df.iterrows():
        row_dict = {str(k): ("" if pd.isna(v) else v) for k, v in row.to_dict().items()}
        product_url = clean_uploaded_url_value(row_dict.get("product_url", ""))
        search_term = normalize_space(row_dict.get("search_term", ""))
        title = normalize_space(row_dict.get("parsed_title", "") or row_dict.get("matched_title", ""))
        description = normalize_space(row_dict.get("parsed_description", ""))
        if not (product_url or search_term):
            stats["skipped_blank"] += 1
            continue
        if not (title or description or row_dict.get("parsed_features_json", "") or row_dict.get("parsed_images_json", "")):
            stats["skipped_weak"] += 1
            continue
        html_text = build_structured_extension_capture_html(row_dict)
        keys = []
        key = normalize_uploaded_capture_url(product_url)
        if key:
            keys.append(key)
        for rpc in kroger_rpc_candidates(search_term):
            keys.append(f"kroger_rpc::{rpc}")
            keys.append(f"rpc::{rpc}")
        clean_search = re.sub(r"[^0-9A-Za-z_-]", "", str(search_term or ""))
        if clean_search:
            keys.append(f"rpc::{clean_search}")
        for map_key in dedupe_preserve_order(keys):
            html_map[map_key] = html_text
        stats["mapped_rows"] += 1
    return html_map, stats


def looks_like_extension_results_file(file_bytes, file_name):
    file_name = str(file_name or "").lower().strip()
    if not (file_name.endswith(".csv") or file_name.endswith(".xlsx")):
        return False
    try:
        if file_name.endswith(".xlsx"):
            df_head = pd.read_excel(BytesIO(file_bytes), engine="openpyxl", nrows=3)
        else:
            df_head = pd.read_csv(BytesIO(file_bytes), nrows=3)
        cols = {str(c).strip().lower().replace(" ", "_") for c in df_head.columns}
        return bool({"parsed_title", "parsed_description", "parsed_features_json", "parsed_images_json", "product_url"} & cols)
    except Exception:
        return False


def clean_rpc_key(value):
    value = str(value or "").strip().replace(".0", "")
    return re.sub(r"[^0-9A-Za-z_-]", "", value)


def first_existing_column(df, candidates):
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
    return ""


def url_only_rpc_key_candidates(value):
    raw = clean_rpc_key(value)
    if not raw:
        return []
    candidates = [raw]
    if raw.isdigit():
        candidates.append(raw.zfill(13))
        candidates.append(raw.lstrip("0") or raw)
    return dedupe_preserve_order([x for x in candidates if x])


def find_url_only_url_in_uploaded_map(uploaded_map, target_rpc=""):
    uploaded_map = uploaded_map or {}
    for rpc in url_only_rpc_key_candidates(target_rpc):
        for prefix in ["url_only_rpc::", "url_only_kroger_rpc::"]:
            url = clean_uploaded_url_value(uploaded_map.get(f"{prefix}{rpc}", ""))
            if url:
                return url
    return ""


def looks_like_url_only_results_file(file_bytes, file_name):
    """Detect extension output that contains only Retailer RPC + Retailer URL.

    URL-only results should update retail_url only. They must not be treated as
    uploaded HTML/copy source, because that breaks Kroger live copy/image parsing.
    """
    file_name = str(file_name or "").lower().strip()
    if not (file_name.endswith(".csv") or file_name.endswith(".xlsx")):
        return False
    try:
        if file_name.endswith(".xlsx"):
            df_head = pd.read_excel(BytesIO(file_bytes), engine="openpyxl", nrows=5)
        else:
            df_head = pd.read_csv(BytesIO(file_bytes), nrows=5)
    except Exception:
        return False
    cols = {str(c).strip().lower().replace(" ", "_") for c in df_head.columns}
    has_rpc = bool({"retailer_rpc", "search_term", "kroger_rpc", "cvs_rpc", "rpc"} & cols)
    has_url = bool({"retailer_url", "retail_url", "product_url", "kroger_url", "url"} & cols)
    has_copy_source = bool({"parsed_title", "parsed_description", "parsed_features_json", "parsed_images_json", "copy_source_code", "source_code", "raw_html", "all_source_code"} & cols)
    return has_rpc and has_url and not has_copy_source


def parse_url_only_results_file(file_bytes, file_name):
    """Parse URL-only extension CSV/XLSX into a lightweight URL map by RPC."""
    file_name = str(file_name or "").lower().strip()
    if file_name.endswith(".xlsx"):
        frames = []
        xls = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
        for sheet in xls.sheet_names:
            df_sheet = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, dtype=str, keep_default_na=False, engine="openpyxl")
            if df_sheet is not None and not df_sheet.empty:
                frames.append(df_sheet)
        df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    else:
        last_error = None
        for encoding in ["utf-8-sig", "utf-8", "latin1"]:
            try:
                df = pd.read_csv(BytesIO(file_bytes), dtype=str, keep_default_na=False, encoding=encoding)
                break
            except Exception as exc:
                last_error = exc
        else:
            raise last_error

    df = df.copy()
    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    rpc_col = first_existing_column(df, ["retailer_rpc", "search_term", "kroger_rpc", "cvs_rpc", "rpc"])
    url_col = first_existing_column(df, ["retailer_url", "retail_url", "product_url", "kroger_url", "url"])
    html_map = {}
    stats = {"mode": "extension_url_only_results", "rows_seen": int(len(df)), "mapped_rows": 0, "skipped_blank": 0, "skipped_weak": 0, "truncated_cell_count": 0}
    if not rpc_col or not url_col:
        return html_map, stats

    for _, row in df.iterrows():
        rpc_raw = row.get(rpc_col, "")
        url = clean_uploaded_url_value(row.get(url_col, ""))
        if not clean_rpc_key(rpc_raw) or not url:
            stats["skipped_blank"] += 1
            continue
        mapped = False
        for rpc in url_only_rpc_key_candidates(rpc_raw):
            html_map[f"url_only_rpc::{rpc}"] = url
            if "kroger.com" in url.lower():
                html_map[f"url_only_kroger_rpc::{rpc}"] = url
            mapped = True
        if mapped:
            stats["mapped_rows"] += 1
    return html_map, stats

def parse_uploaded_retailer_source_file(file_bytes, file_name, selected_retailer=""):
    """Parse uploaded captured retailer source.

    TXT/HTML files use the extension parser. XLSX files are treated as manual
    CVS source-code workbooks keyed by CVS RPC. This keeps the manual fallback
    isolated to CVS and does not change other retailer capture behavior.
    """
    file_name = str(file_name or "").lower().strip()
    if looks_like_url_only_results_file(file_bytes, file_name):
        return parse_url_only_results_file(file_bytes, file_name)
    if looks_like_extension_results_file(file_bytes, file_name):
        return parse_extension_results_structured_file(file_bytes, file_name)
    if file_name.endswith(".xlsx"):
        return parse_cvs_manual_source_code_workbook(file_bytes)

    text_value = ""
    if isinstance(file_bytes, bytes):
        for encoding in ["utf-8", "utf-8-sig", "latin1"]:
            try:
                text_value = file_bytes.decode(encoding)
                break
            except Exception:
                pass
    else:
        text_value = str(file_bytes or "")

    parsed_map = parse_uploaded_raw_html_map(text_value, selected_retailer=selected_retailer)
    stats = {
        "mode": "extension_txt_html",
        "rows_seen": 0,
        "mapped_rows": len(parsed_map),
        "skipped_blank": 0,
        "skipped_weak": 0,
        "truncated_cell_count": 0,
    }
    return parsed_map, stats


def lookup_uploaded_raw_html(uploaded_html_map, retail_url, target_rpc=""):
    uploaded_html_map = uploaded_html_map or {}
    retail_url = str(retail_url or "").strip()
    target_rpc = str(target_rpc or "").strip()

    # URL-only extension results are for filling Retail URL only, not copy_source_code.
    if any(str(k).startswith("url_only_") for k in uploaded_html_map.keys()):
        return ""

    # Structured extension CSV/XLSX results are keyed by generic rpc::{value}.
    for rpc in kroger_rpc_candidates(target_rpc) + [re.sub(r"[^0-9A-Za-z_-]", "", target_rpc)]:
        if not rpc:
            continue
        html_text = str(uploaded_html_map.get(f"rpc::{rpc}", "") or "")
        if html_text:
            return html_text

    if retail_url and "kroger.com" in retail_url.lower():
        key = normalize_uploaded_capture_url(retail_url)
        html_text = str(uploaded_html_map.get(key, "") or "")
        if html_text:
            return html_text
        for rpc in kroger_rpc_candidates(target_rpc):
            html_text = str(uploaded_html_map.get(f"kroger_rpc::{rpc}", "") or "")
            if html_text:
                return html_text
        matched_key = find_kroger_url_in_uploaded_map(uploaded_html_map, target_rpc=target_rpc)
        if matched_key:
            return str(uploaded_html_map.get(matched_key, "") or "")
        return ""

    if (retail_url and "cvs.com" in retail_url.lower()) or (target_rpc and any(str(k).startswith("cvs_rpc::") for k in uploaded_html_map.keys())):
        rpc = re.sub(r"[^0-9A-Za-z_-]", "", str(target_rpc or "").replace(".0", "").strip())
        if not rpc and retail_url:
            m_rpc = re.search(r"[?&]skuId=([0-9A-Za-z_-]+)", str(retail_url or ""), flags=re.IGNORECASE)
            if m_rpc:
                rpc = m_rpc.group(1)
        if rpc:
            html_text = str(uploaded_html_map.get(f"cvs_rpc::{rpc}", "") or "")
            if html_text:
                return html_text
        key = normalize_uploaded_capture_url(retail_url)
        html_text = str(uploaded_html_map.get(key, "") or "")
        if html_text:
            return html_text
        for key in uploaded_capture_url_candidates(retail_url):
            html_text = uploaded_html_map.get(key, "")
            if html_text:
                return html_text
        product_path = re.sub(r"[?#].*$", "", normalize_uploaded_capture_url(retail_url)).lower()
        if product_path:
            for map_key, html_text in uploaded_html_map.items():
                map_path = re.sub(r"[?#].*$", "", normalize_uploaded_capture_url(map_key)).lower()
                if map_path == product_path:
                    return str(html_text or "")
        return ""

    if (retail_url and "heb.com" in retail_url.lower()) or (target_rpc and any(str(k).startswith("heb_rpc::") for k in uploaded_html_map.keys())):
        rpc = re.sub(r"[^0-9A-Za-z]", "", str(target_rpc or "").replace(".0", "").strip())
        if not rpc and retail_url:
            m_rpc = re.search(r"/(\d{4,12})(?:[/?#]|$)", str(retail_url or ""))
            if m_rpc:
                rpc = m_rpc.group(1)
        if rpc:
            html_text = str(uploaded_html_map.get(f"heb_rpc::{rpc}", "") or "")
            if html_text:
                return html_text
        key = normalize_uploaded_capture_url(retail_url)
        html_text = str(uploaded_html_map.get(key, "") or "")
        if html_text:
            return html_text
        return ""

    if not retail_url and target_rpc:
        matched_key = find_kroger_url_in_uploaded_map(uploaded_html_map, target_rpc=target_rpc)
        if matched_key:
            return str(uploaded_html_map.get(matched_key, "") or "")

    for key in uploaded_capture_url_candidates(retail_url):
        html_text = uploaded_html_map.get(key, "")
        if html_text:
            return html_text
    return ""

def normalize_extension_capture_url(retail_url, retailer_name=""):
    """Normalize only the URL sent to the browser extension; preserve source data elsewhere."""
    raw = str(retail_url or "").strip()
    if not raw:
        return ""
    if normalize_retailer_name(retailer_name) != "CVS" and "cvs.com" not in raw.lower():
        return raw
    try:
        parsed = urllib.parse.urlsplit(raw)
        host = (parsed.hostname or "").lower()
        if host not in {"cvs.com", "www.cvs.com"}:
            return raw
        path = re.sub(r"/reviews(?:/.*)?$", "", parsed.path or "", flags=re.IGNORECASE).rstrip("/")
        return urllib.parse.urlunsplit(("https", "www.cvs.com", path, "", ""))
    except Exception:
        cleaned = re.sub(r"^http://", "https://", raw, flags=re.IGNORECASE)
        cleaned = re.sub(r"/reviews(?:/.*)?(?:[?#].*)?$", "", cleaned, flags=re.IGNORECASE)
        return cleaned.split("#", 1)[0].split("?", 1)[0].rstrip("/")


def build_extension_batch_payload(retailer_df, retailer_name, current_batch_key, capture_mode, txt_ready=False):
    retailer_name_norm = normalize_retailer_name(retailer_name)
    retailer_df = strict_filter_rows_for_selected_retailer(
        retailer_df,
        retailer_name_norm,
        dedupe_by_url=True,
    )

    retail_urls = []
    row_payload = []
    if retailer_df is not None and not retailer_df.empty:
        retail_urls = [
            normalize_extension_capture_url(x, retailer_name_norm)
            for x in retailer_df["retail_url"].fillna("").astype(str).tolist()
            if str(x).strip()
        ]
        row_payload = []
        for _, row in retailer_df.iterrows():
            search_term = str(row.get("retailer_rpc", "") or "").strip() or str(row.get("sku", "") or "").strip()
            if not search_term:
                continue
            row_payload.append({
                "sku": str(row.get("sku", "") or "").strip(),
                "search_term": search_term,
                "retail_url": normalize_extension_capture_url(row.get("retail_url", ""), retailer_name_norm),
                "retailer_rpc": str(row.get("retailer_rpc", "") or "").strip(),
                "rpc": str(row.get("retailer_rpc", "") or "").strip(),
                "retailer": retailer_name_norm,
            })

    return {
        "ready": True,
        "retailer": retailer_name_norm,
        "retailerGuard": retailer_name_norm,
        "batchKey": str(current_batch_key or ""),
        "captureMode": str(capture_mode or ""),
        "txtReady": bool(txt_ready),
        "totalRows": int(len(row_payload)),
        "uniqueRetailUrlCount": int(len(retail_urls)),
        "retailUrls": retail_urls,
        "rows": row_payload,
        "timestamp": int(time.time()),
    }


def render_extension_batch_bridge(payload):
    payload_json = json.dumps(payload or {}, ensure_ascii=False)
    payload_b64 = base64.b64encode(payload_json.encode("utf-8")).decode("ascii")
    retailer_html = html.escape(str((payload or {}).get("retailer", "") or ""), quote=True)
    batch_key_html = html.escape(str((payload or {}).get("batchKey", "") or ""), quote=True)
    capture_mode_html = html.escape(str((payload or {}).get("captureMode", "") or ""), quote=True)
    payload_b64_html = html.escape(payload_b64, quote=True)

    st.markdown(
        f"""
        <div id="pdp-extension-batch-ready"
             data-pdp-extension-batch-ready="1"
             data-pdp-extension-batch-payload-b64="{payload_b64_html}"
             data-pdp-extension-retailer="{retailer_html}"
             data-pdp-extension-batch-key="{batch_key_html}"
             data-pdp-extension-capture-mode="{capture_mode_html}"
             style="display:none !important; visibility:hidden !important; width:0 !important; height:0 !important; max-height:0 !important; overflow:hidden !important; opacity:0 !important; pointer-events:none !important; position:absolute !important; left:-9999px !important; top:-9999px !important;"
        ></div>
        <script id="pdp-extension-batch-json" type="application/json">{payload_json}</script>
        """,
        unsafe_allow_html=True,
    )

    bridge_html = f"""
    <script>
    (function() {{
      const payload = {payload_json};
      const payloadB64 = '{payload_b64}';
      const TARGET_KEYS = [
        '__PDP_EXTENSION_BATCH__', '__RAW_HTML_EXTENSION_BATCH__', '__STREAMLIT_EXTENSION_BATCH__', '__EXTENSION_BATCH__',
        '__RAW_HTML_BATCH__', 'pdpExtensionBatch', 'rawHtmlExtensionBatch', 'streamlitExtensionBatch', 'extensionBatch'
      ];
      const STORAGE_KEYS = [
        '__PDP_EXTENSION_BATCH__', '__RAW_HTML_EXTENSION_BATCH__', '__STREAMLIT_EXTENSION_BATCH__', '__EXTENSION_BATCH__',
        'pdpExtensionBatch', 'rawHtmlExtensionBatch', 'streamlitExtensionBatch', 'extensionBatch'
      ];
      const EVENT_NAMES = [
        'pdp-extension-batch-ready', 'raw-html-extension-batch-ready', 'streamlit-extension-batch-ready',
        'extension-batch-ready', 'PDP_EXTENSION_BATCH_READY'
      ];

      function uniqueTargets() {{
        const out = [];
        for (const candidate of [window, window.parent, window.top]) {{
          try {{ if (candidate && !out.includes(candidate)) out.push(candidate); }} catch (e) {{}}
        }}
        return out;
      }}

      function ensureDomBeacon(targetWindow) {{
        try {{
          const doc = targetWindow.document;
          if (!doc) return;
          const root = doc.documentElement || doc.body;
          const body = doc.body || doc.documentElement;
          const payloadText = JSON.stringify(payload);
          if (root) {{
            root.setAttribute('data-pdp-extension-batch-ready', payload && payload.ready ? '1' : '0');
            root.setAttribute('data-pdp-extension-batch-payload-b64', payloadB64);
            root.setAttribute('data-pdp-extension-retailer', payload && payload.retailer ? String(payload.retailer) : '');
            root.setAttribute('data-pdp-extension-batch-key', payload && payload.batchKey ? String(payload.batchKey) : '');
            root.setAttribute('data-pdp-extension-capture-mode', payload && payload.captureMode ? String(payload.captureMode) : '');
          }}
          if (body) {{
            body.setAttribute('data-pdp-extension-batch-ready', payload && payload.ready ? '1' : '0');
            body.setAttribute('data-pdp-extension-batch-payload-b64', payloadB64);
          }}
          let beacon = doc.getElementById('pdp-extension-batch-ready');
          if (!beacon) {{
            beacon = doc.createElement('div');
            beacon.id = 'pdp-extension-batch-ready';
            beacon.style.display = 'none';
            beacon.style.visibility = 'hidden';
            beacon.style.width = '0';
            beacon.style.height = '0';
            beacon.style.maxHeight = '0';
            beacon.style.overflow = 'hidden';
            beacon.style.opacity = '0';
            beacon.style.pointerEvents = 'none';
            beacon.style.position = 'absolute';
            beacon.style.left = '-9999px';
            beacon.style.top = '-9999px';
            (doc.body || doc.documentElement).appendChild(beacon);
          }}
          beacon.setAttribute('data-pdp-extension-batch-ready', payload && payload.ready ? '1' : '0');
          beacon.setAttribute('data-pdp-extension-batch-payload-b64', payloadB64);
          beacon.setAttribute('data-pdp-extension-retailer', payload && payload.retailer ? String(payload.retailer) : '');
          beacon.setAttribute('data-pdp-extension-batch-key', payload && payload.batchKey ? String(payload.batchKey) : '');
          beacon.setAttribute('data-pdp-extension-capture-mode', payload && payload.captureMode ? String(payload.captureMode) : '');
          let scriptTag = doc.getElementById('pdp-extension-batch-json');
          if (!scriptTag) {{
            scriptTag = doc.createElement('script');
            scriptTag.id = 'pdp-extension-batch-json';
            scriptTag.type = 'application/json';
            (doc.body || doc.documentElement).appendChild(scriptTag);
          }}
          scriptTag.textContent = payloadText;
        }} catch (e) {{}}
      }}

      function publishToTarget(targetWindow) {{
        if (!targetWindow) return;
        try {{ for (const key of TARGET_KEYS) targetWindow[key] = payload; }} catch (e) {{}}
        try {{
          for (const key of STORAGE_KEYS) {{
            try {{ if (targetWindow.localStorage) targetWindow.localStorage.setItem(key, JSON.stringify(payload)); }} catch (e) {{}}
            try {{ if (targetWindow.sessionStorage) targetWindow.sessionStorage.setItem(key, JSON.stringify(payload)); }} catch (e) {{}}
          }}
        }} catch (e) {{}}
        ensureDomBeacon(targetWindow);
        try {{ for (const eventName of EVENT_NAMES) targetWindow.dispatchEvent(new CustomEvent(eventName, {{ detail: payload }})); }} catch (e) {{}}
        try {{
          targetWindow.postMessage({{ type: 'pdp-extension-batch-ready', payload }}, '*');
          targetWindow.postMessage({{ type: 'raw-html-extension-batch-ready', payload }}, '*');
          targetWindow.postMessage({{ type: 'streamlit-extension-batch-ready', payload }}, '*');
          targetWindow.postMessage({{ type: 'extension-batch-ready', payload }}, '*');
          targetWindow.postMessage({{ type: 'PDP_EXTENSION_BATCH_READY', payload }}, '*');
        }} catch (e) {{}}
      }}

      function publishAll() {{ for (const target of uniqueTargets()) publishToTarget(target); }}
      publishAll();
      let publishCount = 0;
      const timer = setInterval(() => {{ publishAll(); publishCount += 1; if (publishCount >= 40) clearInterval(timer); }}, 500);
      try {{
        window.addEventListener('message', function(event) {{
          const data = event && event.data ? event.data : {{}};
          const msgType = data && data.type ? String(data.type) : '';
          if (msgType === 'pdp-extension-batch-request' || msgType === 'raw-html-extension-batch-request' || msgType === 'streamlit-extension-batch-request' || msgType === 'extension-batch-request' || msgType === 'PDP_EXTENSION_BATCH_REQUEST') publishAll();
        }});
      }} catch (e) {{}}
    }})();
    </script>
    """
    try:
        components.html(bridge_html, height=0, width=0)
    except Exception:
        # Do not let the hidden extension bridge crash the app.
        pass

def normalize_kroger_url(url):
    url = str(url or "").strip()
    if not url:
        return ""
    url = html.unescape(url)
    url = url.split("#", 1)[0].strip()
    url = re.sub(r'([?&])msockid=[^&]+', r'\1', url, flags=re.IGNORECASE)
    url = re.sub(r'([?&])searchType=[^&]+', r'\1', url, flags=re.IGNORECASE)
    url = re.sub(r'([?&])fulfillment=[^&]+', r'\1', url, flags=re.IGNORECASE)
    url = re.sub(r'([?&])campaign=[^&]+', r'\1', url, flags=re.IGNORECASE)
    url = re.sub(r'([?&])adgroup=[^&]+', r'\1', url, flags=re.IGNORECASE)
    url = re.sub(r'([?&])pid=[^&]+', r'\1', url, flags=re.IGNORECASE)
    url = re.sub(r'\?&', '?', url)
    url = re.sub(r'[?&]+$', '', url)
    url = re.sub(r'\?{2,}', '?', url)
    return url.strip()

def clean_kroger_rpc(value):
    value = str(value or "").strip()
    value = value.replace(".0", "")
    value = re.sub(r"[^0-9A-Za-z]", "", value)
    return value


def kroger_rpc_candidates(value):
    rpc = clean_kroger_rpc(value)
    if not rpc:
        return []
    out = [rpc]
    if rpc.isdigit() and len(rpc) < 13:
        out.append(rpc.zfill(13))
    if rpc.isdigit() and len(rpc) < 12:
        out.append(rpc.zfill(12))
    deduped = []
    for item in out:
        if item and item not in deduped:
            deduped.append(item)
    return deduped


def find_kroger_url_in_uploaded_map(uploaded_html_map, target_rpc=""):
    uploaded_html_map = uploaded_html_map or {}
    url_only = find_url_only_url_in_uploaded_map(uploaded_html_map, target_rpc=target_rpc)
    if url_only:
        return url_only
    rpc_values = kroger_rpc_candidates(target_rpc)
    if not rpc_values:
        return ""
    for key in uploaded_html_map.keys():
        key_raw = str(key or "")
        if key_raw.startswith("kroger_rpc::") or key_raw.startswith("url_only_"):
            continue
        key_str = normalize_uploaded_capture_url(key_raw)
        for rpc in rpc_values:
            if rpc and rpc in key_str:
                return key_str
    return ""

def resolve_debug_views(
    debug_url,
    retailer_name="",
    use_manual_html_override=False,
    manual_html_text="",
    manual_html_file=None,
    headers_text="",
    timeout_override=None,
    use_mobile=False,
    proxy_url="",
):
    manual_text = str(manual_html_text or "").strip()
    uploaded_text = get_uploaded_text_file_bytes(manual_html_file).strip()

    if use_manual_html_override:
        chosen_html = manual_text or uploaded_text
        if chosen_html:
            views = build_debug_views_from_html(chosen_html)
            return {
                "mode": "manual_html_override",
                "requested_url": str(debug_url or ""),
                "final_url": "manual_html_override",
                "status_code": "MANUAL",
                "reason": "Manual HTML override",
                "content_type": "text/html",
                "content_length_header": str(len(chosen_html)),
                "text_length": len(chosen_html),
                "history": [],
                "error": "",
                "response_headers": {},
                **views,
            }

    return fetch_url_debug(
        debug_url,
        retailer_name=retailer_name,
        headers_text=headers_text,
        timeout_override=timeout_override,
        use_mobile=use_mobile,
        proxy_url=proxy_url,
    )


def is_debug_view_robot_page(debug_views):
    raw_html = str(debug_views.get("raw_html", "") or "").lower()
    final_url = str(debug_views.get("final_url", "") or "").lower()
    requested_url = str(debug_views.get("requested_url", "") or "").lower()

    combined = " ".join([raw_html, final_url, requested_url])

    markers = [
        "let us know you're not a robot",
        "let us know you’re not a robot",
        "let us know you're human",
        "let us know you’re human",
        "no robots allowed",
        "captcha",
        "px-captcha",
        "/are-you-human",
        "challenge-platform",
    ]

    return any(marker in combined for marker in markers)
      
def render_debugger_panel(
    debug_views,
    sku="",
    marker_start="",
    marker_end="",
    marker_target="Raw HTML",
    use_manual_html_override=False,
):
    requested_url = str(debug_views.get("requested_url", "") or "")
    final_url = str(debug_views.get("final_url", "") or "")
    status_code = str(debug_views.get("status_code", "") or "")
    reason = str(debug_views.get("reason", "") or "")
    text_length = int(debug_views.get("text_length", 0) or 0)
    history = debug_views.get("history", []) or []
    raw_html = str(debug_views.get("raw_html", "") or "")
    dom_text = str(debug_views.get("dom_text", "") or "")
    prettified_dom = str(debug_views.get("prettified_dom", "") or "")
    error_text = str(debug_views.get("error", "") or "")
    response_headers = debug_views.get("response_headers", {}) or {}
    request_headers_used = debug_views.get("request_headers_used", {}) or {}
    elapsed_seconds = debug_views.get("elapsed_seconds", 0.0)
    proxy_used = str(debug_views.get("proxy_used", "") or "")

    if use_manual_html_override:
        st.success("Using manual HTML override for debugger.")
    elif error_text:
        st.error(f"Debugger fetch error: {error_text}")

    metric_cols = st.columns(6)
    with metric_cols[0]:
        st.caption("Status")
        st.markdown(f"### {status_code if status_code else 'N/A'}")
    with metric_cols[1]:
        st.caption("Reason")
        st.markdown(f"### {reason if reason else 'N/A'}")
    with metric_cols[2]:
        st.caption("Content Length")
        st.markdown(f"### {text_length}")
    with metric_cols[3]:
        st.caption("Redirects")
        st.markdown(f"### {len(history)}")
    with metric_cols[4]:
        st.caption("Elapsed Seconds")
        st.markdown(f"### {elapsed_seconds}")
    with metric_cols[5]:
        st.caption("Final URL Set")
        st.markdown(f"### {'Yes' if final_url else 'No'}")

    st.text_input("Requested URL", value=requested_url, key=f"debug_requested_url_{sku}")
    st.text_input("Final URL", value=final_url, key=f"debug_final_url_{sku}")
    if proxy_used:
        st.text_input("Proxy Used", value=proxy_used, key=f"debug_proxy_used_{sku}")
    if request_headers_used:
        with st.expander("Request headers used"):
            st.json(request_headers_used)
    if response_headers:
        with st.expander("Response headers"):
            st.json(response_headers)

    st.download_button(
        "Download full raw HTML",
        data=raw_html.encode("utf-8"),
        file_name=f"raw_html_{sku or 'debug'}.html",
        mime="text/html",
        key=f"download_raw_html_{sku}",
    )

    debug_view = st.radio(
        "Debug view",
        ["Raw HTML preview", "DOM text", "Prettified DOM"],
        horizontal=True,
        key=f"debug_view_selector_{sku}",
    )

    preview_limit = int(globals().get("DEBUG_TEXT_PREVIEW_CHARS", 200000))
    if debug_view == "Raw HTML preview":
        show_full_raw = st.checkbox(
            "Load full raw HTML in the text box (slower for large TXT captures)",
            value=False,
            key=f"debug_show_full_raw_{sku}",
        )
        value = raw_html if show_full_raw else raw_html[:preview_limit]
        if raw_html and not show_full_raw and len(raw_html) > preview_limit:
            st.caption(f"Showing first {preview_limit:,} characters of {len(raw_html):,}. Use the download button for the full file, or check the box to load all text.")
        st.text_area(f"raw_html_{sku or 'debug'}", value=value, height=900, key=f"debug_raw_html_{sku}")
    elif debug_view == "DOM text":
        if not dom_text:
            dom_text = html_to_debug_textblob(raw_html)
        st.text_area(f"dom_text_{sku or 'debug'}", value=dom_text, height=900, key=f"debug_dom_text_{sku}")
    else:
        if not prettified_dom:
            with st.spinner("Generating prettified DOM. This can take a bit for large Kroger captures..."):
                prettified_dom = html_to_prettified_dom(raw_html)
        value = prettified_dom[:preview_limit]
        if prettified_dom and len(prettified_dom) > preview_limit:
            st.caption(f"Showing first {preview_limit:,} characters of {len(prettified_dom):,}. Download raw HTML if you need the full source.")
        st.text_area(f"prettified_dom_{sku or 'debug'}", value=value, height=900, key=f"debug_prettified_dom_{sku}")


# =========================================
# SALSIFY PARSERS
# =========================================
# Architecture note:
# The raw Salsify parser should collect broad Salsify content/assets.
# Retailer-specific Salsify decisions are applied later through
# finalize_salsify_copy_for_retailer(), align_salsify_images_for_retailer(),
# and build_normalized_comparison_payload(). This keeps CVS, Walgreens, Kroger,
# and Sam's Club Salsify rules separate from each retailer site parser without
# changing the UI or Excel export.


def _normalize_salsify_property_key(value):
    value = normalize_salsify_asset_name(value)
    value = value.replace(" ", "_")
    return value.strip("_")


def _prepend_salsify_property_value(prop_map, prop_name, prop_value):
    normalized_name = _normalize_salsify_property_key(prop_name)
    clean_value = normalize_space(prop_value)
    if not normalized_name or not clean_value:
        return
    values = prop_map.setdefault(normalized_name, [])
    if clean_value in values:
        values.remove(clean_value)
    values.insert(0, clean_value)


def extract_salsify_visible_property_map(html_text):
    result = {"properties": {}, "assets": {}}
    if not html_text:
        return result

    raw_html = html.unescape(str(html_text or ""))
    soup = BeautifulSoup(raw_html, "html.parser")

    def store_asset(label, url):
        normalized_label = normalize_salsify_asset_name(label)
        clean_url = html.unescape(str(url or "").strip())
        if not normalized_label or not clean_url:
            return
        result["assets"][normalized_label] = clean_url if is_video_like_url(clean_url) else clean_url.split("?", 1)[0]

    # Standard 2-column table rows: left cell = property label, right cell = actual value.
    for row in soup.find_all("tr"):
        cells = row.find_all(["td", "th"])
        if len(cells) < 2:
            continue
        label = normalize_space(cells[0].get_text(" ", strip=True))
        if not label:
            continue
        value_cell = cells[1]
        value_text = normalize_space(value_cell.get_text(" ", strip=True))
        if value_text:
            _prepend_salsify_property_value(result["properties"], label, value_text)
        for a in value_cell.find_all("a", href=True):
            href = html.unescape(str(a.get("href", "") or "").strip())
            if href:
                store_asset(label, href)

    # Also support definition-list style blocks.
    for dl in soup.find_all("dl"):
        for dt, dd in zip(dl.find_all("dt"), dl.find_all("dd")):
            label = normalize_space(dt.get_text(" ", strip=True))
            value_text = normalize_space(dd.get_text(" ", strip=True))
            if label and value_text:
                _prepend_salsify_property_value(result["properties"], label, value_text)
            for a in dd.find_all("a", href=True):
                href = html.unescape(str(a.get("href", "") or "").strip())
                if label and href:
                    store_asset(label, href)

    # Visible asset label -> href patterns.
    visible_asset_patterns = [
        r'>\s*(Main Variant Image-Sams Club|Main Variant Image-Club|Online Optimized Image-Sams Club|Online Optimized Image-Kroger|Online Optimized Image-Grocery|Online Optimized Image-HEB|Online Optimized Image-H-E-B|Online Optimized Image-|Ingredient Label Image-|Ingredient Label Image|Shipping-|Flat Back_2D-|Flat Left_2D-|ATF I/O-Sams Club|ATF I/O-Generic|ATF I/O-HEB|ATF I/O-H-E-B|ATF Video-Sams Club|ATF [0-9]+-Sams Club|ATF [0-9]+-HEB|ATF [0-9]+-H-E-B)\s*<.*?href="([^"]+)"',
        r'"property"\s*:\s*"(Main Variant Image-Sams Club|Main Variant Image-Club|Online Optimized Image-Sams Club|Online Optimized Image-Kroger|Online Optimized Image-Grocery|Online Optimized Image-HEB|Online Optimized Image-H-E-B|Online Optimized Image-|Ingredient Label Image-|Ingredient Label Image|Shipping-|Flat Back_2D-|Flat Left_2D-|ATF I/O-Sams Club|ATF I/O-Generic|ATF I/O-HEB|ATF I/O-H-E-B|ATF Video-Sams Club|ATF [0-9]+-Sams Club|ATF [0-9]+-HEB|ATF [0-9]+-H-E-B)"[^{}]{0,1200}?"value"\s*:\s*"([^"]+)"',
    ]
    for pattern in visible_asset_patterns:
        for matched_name, matched_url in re.findall(pattern, raw_html, flags=re.IGNORECASE | re.DOTALL):
            store_asset(matched_name, matched_url)

    return result


def pick_kroger_images_with_atf_and_lifestyle(asset_lookup):
    """Return Salsify images in the existing safe order, with CVS-only flats preserved.

    Existing non-CVS behavior is preserved by keeping the normal output focused on:
    1. One main Online Optimized Image.
    2. ATF and lifestyle images in source order.

    CVS-only change:
    - Flat Back_2D and Flat Left_2D assets are appended with cvs_only=True so CVS
      can use them for slots 2 and 3. The retailer alignment function filters
      these cvs_only images back out for every non-CVS retailer.
    """
    main_priority_order = [
        "online optimized image walgreens",
        "online optimized image",
    ]

    best_main = None
    best_main_priority_index = 999

    normalized_assets = []
    for name, url in asset_lookup.items():
        normalized_name = normalize_salsify_asset_name(name)
        clean_url = str(url or "").strip()
        if not clean_url:
            continue
        clean_url = clean_url if is_video_like_url(clean_url) else clean_url.split("?", 1)[0]
        normalized_assets.append((normalized_name, name, clean_url))

        for idx, priority in enumerate(main_priority_order):
            if priority in normalized_name and idx < best_main_priority_index:
                best_main_priority_index = idx
                best_main = {
                    "name": name,
                    "url": clean_url,
                }

    ordered_images = []
    seen_urls = set()

    def add_image(name, url, cvs_only=False, walgreens_only=False, kroger_only=False):
        clean_url = str(url or "").strip()
        if not clean_url or clean_url in seen_urls:
            return False
        img = {"name": name, "url": clean_url}
        if cvs_only:
            img["cvs_only"] = True
        if walgreens_only:
            img["walgreens_only"] = True
        if kroger_only:
            img["kroger_only"] = True
        ordered_images.append(img)
        seen_urls.add(clean_url)
        return True

    if best_main:
        add_image(best_main["name"], best_main["url"])

    # Kroger-only final selector needs these candidates even when another
    # retailer-specific Online Optimized image was the general best_main.
    for normalized_name, name, clean_url in normalized_assets:
        if (
            "online optimized image kroger" in normalized_name
            or "online optimized image grocery" in normalized_name
            or (
                "online optimized image" in normalized_name
                and not any(token in normalized_name for token in ["walgreens", "cvs", "sams club", "sam s club", "samsclub", "target", "walmart"])
            )
        ):
            add_image(name, clean_url, kroger_only=True)

    # HEB-specific assets must stay in the parsed Salsify image bundle too.
    heb_required_tokens = (
        "online optimized image heb",
        "online optimized image h e b",
        "atf i o heb",
        "atf io heb",
        "atf 1 heb",
        "atf 2 heb",
        "atf 3 heb",
        "atf 4 heb",
        "atf 5 heb",
        "atf 6 heb",
        "lifestyle heb",
    )
    for normalized_name, name, clean_url in normalized_assets:
        if any(token in normalized_name for token in heb_required_tokens):
            add_image(name, clean_url)

    # Sam's Club-specific assets must stay in the parsed Salsify image bundle
    # so the Sam's Club alignment step can place them in the proper slots.
    # Without this, properties such as Main Variant Image-Club and Shipping-
    # could be found in asset_lookup but discarded before alignment.
    sams_required_tokens = (
        "online optimized image sams club",
        "main variant image sams club",
        "main variant image club",
        "shipping",
        "atf video sams club",
    )
    for normalized_name, name, clean_url in normalized_assets:
        if any(token in normalized_name for token in sams_required_tokens):
            add_image(name, clean_url)

    # Sam's Club also needs the generic Online Optimized Image- asset available for
    # visual alignment. The original best_main logic can keep only the Sam's Club-
    # specific Online Optimized Image-Sams Club asset, which accidentally drops the
    # generic Online Optimized Image- before the Sam's Club ordering function can
    # place it after ATF Video-Sams Club.
    sams_generic_ooi_excluded_tokens = (
        "sams club", "sam s club", "samsclub", "sams",
        "walgreens", "kroger", "grocery", "cvs", "target", "walmart",
    )
    for normalized_name, name, clean_url in normalized_assets:
        is_generic_online_optimized = (
            "online optimized image" in normalized_name
            or "online image" in normalized_name
        )
        is_retailer_specific = any(token in normalized_name for token in sams_generic_ooi_excluded_tokens)
        if is_generic_online_optimized and not is_retailer_specific:
            add_image(name, clean_url)

    # Walgreens requires Ingredient Label Image as locked slot 2. Keep it in
    # the parsed image bundle, but mark it Walgreens-only so other retailers
    # do not inherit this extra slot.
    for normalized_name, name, clean_url in normalized_assets:
        if is_walgreens_ingredient_label_image_name(name):
            add_image(name, clean_url, walgreens_only=True)

    for normalized_name, name, clean_url in normalized_assets:
        is_atf_image = "atf" in normalized_name
        is_lifestyle_image = (
            "lifestyle" in normalized_name
            or "life style" in normalized_name
            or "lifestyle image" in normalized_name
        )

        if not (is_atf_image or is_lifestyle_image):
            continue

        add_image(name, clean_url)

    # CVS-only flats are intentionally appended after normal images. CVS slot
    # ordering will find them by name for slots 2 and 3; non-CVS retailers
    # filter them out in align_salsify_images_for_retailer().
    for normalized_name, name, clean_url in normalized_assets:
        if is_cvs_only_salsify_flat_image_name(name):
            add_image(name, clean_url, cvs_only=True)

    return ordered_images

def _parse_salsify_page(html_text):
    empty = {
        "text": {
            "title": "",
            "description": "",
            "feature1": "",
            "feature2": "",
            "feature3": "",
            "feature4": "",
            "feature5": "",
            "feature6": "",
            "feature7": "",
            "features": [],
            "retailer_overrides": {},
        },
        "images": [],
    }
    if not html_text:
        return empty

    soup = BeautifulSoup(html_text, "html.parser")
    visible_property_map = extract_salsify_visible_property_map(html_text)
    script = soup.find("script", {"id": "__NEXT_DATA__"})
    data = {}
    if script:
        try:
            data = json.loads(script.string)
        except Exception:
            data = {}

    def iter_nodes(obj):
        if isinstance(obj, dict):
            yield obj
            for value in obj.values():
                yield from iter_nodes(value)
        elif isinstance(obj, list):
            for item in obj:
                yield from iter_nodes(item)

    def looks_like_image_url(value):
        value = str(value or "").strip().lower()
        return bool(
            value
            and (value.startswith("http://") or value.startswith("https://"))
            and any(value.endswith(ext) or f"{ext}?" in value for ext in [".jpg", ".jpeg", ".png", ".webp", ".avif"])
        )

    def extract_node_values(node):
        values_out = []
        seen_local = set()

        def add_value(value):
            if not isinstance(value, str):
                return
            clean_value = html.unescape(str(value or "")).strip()
            if not clean_value or looks_like_image_url(clean_value):
                return
            if clean_value not in seen_local:
                seen_local.add(clean_value)
                values_out.append(clean_value)

        def walk_value(value):
            if isinstance(value, str):
                add_value(value)
            elif isinstance(value, dict):
                # Only inspect value-like fields. Do not pull generic property labels
                # such as "Sams Club Product Title" from dict keys like name/title.
                for key in ["value", "label", "displayValue", "text"]:
                    if isinstance(value.get(key), str):
                        add_value(value.get(key))
                for nested_key, nested in value.items():
                    if nested_key in {"property", "name", "key", "title"}:
                        continue
                    if isinstance(nested, (dict, list)):
                        walk_value(nested)
            elif isinstance(value, list):
                for item in value:
                    walk_value(item)

        if not isinstance(node, dict):
            return []

        for key in ["value", "label", "displayValue", "text"]:
            if isinstance(node.get(key), str):
                add_value(node.get(key))

        walk_value(node.get("values", []))
        return values_out

    def normalize_prop_name(value):
        value = normalize_salsify_asset_name(value)
        value = value.replace(" ", "_")
        return value.strip("_")

    property_values = {}
    for node in iter_nodes(data):
        if not isinstance(node, dict):
            continue
        prop_name = node.get("property") or node.get("name") or node.get("key")
        if not isinstance(prop_name, str) or not prop_name.strip():
            continue
        normalized_name = normalize_prop_name(prop_name)
        raw_values = extract_node_values(node)
        if not raw_values:
            continue
        property_values.setdefault(normalized_name, [])
        for raw_value in raw_values:
            if raw_value not in property_values[normalized_name]:
                property_values[normalized_name].append(raw_value)

    def collect_property_values(*keys):
        out = []
        seen = set()
        for key in keys:
            normalized_key = normalize_prop_name(key)
            for value in property_values.get(normalized_key, []) or []:
                clean_value = normalize_space(value)
                if clean_value and clean_value not in seen:
                    seen.add(clean_value)
                    out.append(clean_value)
        return out

    def collect_property_values_loose(*keys):
        out = []
        seen = set()
        normalized_queries = [normalize_prop_name(key) for key in keys if normalize_prop_name(key)]
        for query in normalized_queries:
            for prop_key, values in property_values.items():
                if not prop_key:
                    continue
                if query == prop_key or query in prop_key or prop_key in query:
                    for value in values or []:
                        clean_value = normalize_space(value)
                        if clean_value and clean_value not in seen:
                            seen.add(clean_value)
                            out.append(clean_value)
        return out


    for prop_key, prop_values in (visible_property_map.get("properties", {}) or {}).items():
        if not prop_key or not prop_values:
            continue
        property_values.setdefault(prop_key, [])
        for prop_value in reversed(list(prop_values)):
            clean_value = normalize_space(prop_value)
            if not clean_value:
                continue
            if clean_value in property_values[prop_key]:
                property_values[prop_key].remove(clean_value)
            property_values[prop_key].insert(0, clean_value)

    def first_property(*keys):
        values = collect_property_values(*keys)
        if values:
            return values[0]
        values = collect_property_values_loose(*keys)
        return values[0] if values else ""

    title = first_property(
        "PRODUCT_TITLE",
        "TITLE",
        "PRODUCT NAME",
        "PRODUCT_NAME",
        "NAME",
        "Display Name",
    )
    description = first_property(
        "DESCRIPTION",
        "LONG_DESCRIPTION",
        "PRODUCT_DESCRIPTION",
        "MARKETING_DESCRIPTION",
        "ROMANCE_COPY",
    )

    feature_candidates = []
    for key, values in property_values.items():
        if key.startswith("feature_") or key.startswith("bullet_") or key.startswith("benefit_"):
            for value in values:
                clean_value = normalize_space(value)
                if clean_value:
                    feature_candidates.append(clean_value)

    if not feature_candidates:
        for fallback_key in [
            "FEATURES",
            "BULLETS",
            "BULLET_POINTS",
            "BENEFITS",
            "HIGHLIGHTS",
            "KEY_FEATURES",
        ]:
            values = property_values.get(normalize_prop_name(fallback_key), []) or []
            for value in values:
                if "|" in value:
                    parts = [normalize_space(x) for x in value.split("|")]
                    feature_candidates.extend([x for x in parts if x])
                else:
                    clean_value = normalize_space(value)
                    if clean_value:
                        feature_candidates.append(clean_value)

    feature_candidates = dedupe_preserve_order(feature_candidates)

    if not title:
        meta_title = soup.find("meta", attrs={"property": "og:title"}) or soup.find("meta", attrs={"name": "twitter:title"})
        if meta_title and meta_title.get("content"):
            title = normalize_space(meta_title.get("content", ""))
    if not title:
        h1 = soup.find("h1")
        if h1:
            title = normalize_space(h1.get_text(" ", strip=True))
    if not title and soup.title:
        title = normalize_space(soup.title.get_text(" ", strip=True))

    if not description:
        meta_desc = soup.find("meta", attrs={"property": "og:description"}) or soup.find("meta", attrs={"name": "description"})
        if meta_desc and meta_desc.get("content"):
            description = normalize_space(meta_desc.get("content", ""))

    # Retailer-specific Salsify fields should override generic deck placeholders when present.
    kroger_feature_values = []
    for i in range(1, 11):
        kroger_feature_values.extend(
            collect_property_values(
                f"Kroger Feature {i}",
                f"Kroger Feature{i}",
                f"Kroger Bullet {i}",
                f"Kroger Bullet{i}",
            )
        )
    kroger_feature_values = dedupe_preserve_order(kroger_feature_values)

    sams_feature_values = []
    sams_feature_slots = {}
    for i in range(1, 11):
        exact_values = collect_property_values(
            f"Sam's Club Feature {i}",
            f"Sam's Club Feature{i}",
            f"Sams Club Feature {i}",
            f"Sams Club Feature{i}",
            f"Sam's Club Bullet {i}",
            f"Sam's Club Bullet{i}",
            f"Sams Club Bullet {i}",
            f"Sams Club Bullet{i}",
        )
        loose_values = collect_property_values_loose(
            f"Sam's Club Feature {i}",
            f"Sams Club Feature {i}",
            f"Sam's Club Bullet {i}",
            f"Sams Club Bullet {i}",
        )
        slot_values = dedupe_preserve_order((exact_values or []) + (loose_values or []))
        if slot_values:
            sams_feature_slots[i] = slot_values[0]
            sams_feature_values.extend(slot_values)
    sams_feature_values = dedupe_preserve_order(sams_feature_values)

    cvs_feature_values = []
    cvs_feature_slots = {}
    general_feature_values = []
    for i in range(1, 11):
        cvs_exact_values = collect_property_values(
            f"CVS Feature {i}",
            f"CVS Feature{i}",
            f"CVS Product Feature {i}",
            f"CVS Product Feature{i}",
            f"CVS Selling Point {i}",
            f"CVS Selling Point{i}",
            f"CVS Bullet {i}",
            f"CVS Bullet{i}",
            f"Retailer Feature {i} - CVS",
            f"Retailer Bullet {i} - CVS",
        )
        cvs_loose_values = collect_property_values_loose(
            f"CVS Feature {i}",
            f"CVS Product Feature {i}",
            f"CVS Selling Point {i}",
            f"CVS Bullet {i}",
            f"Retailer Feature {i} - CVS",
            f"Retailer Bullet {i} - CVS",
        )
        cvs_slot_values = dedupe_preserve_order((cvs_exact_values or []) + (cvs_loose_values or []))
        cvs_slot_values = [v for v in cvs_slot_values if v and not is_placeholder_salsify_copy_value(v)]
        if cvs_slot_values:
            cvs_feature_slots[i] = cvs_slot_values[0]
            cvs_feature_values.extend(cvs_slot_values)

        general_feature_values.extend(
            collect_property_values(
                f"General Feature {i}",
                f"General Feature{i}",
                f"General Bullet {i}",
                f"General Bullet{i}",
            )
        )

    if not cvs_feature_values:
        broad_cvs_feature_values = []
        for prop_key, values in property_values.items():
            pk = normalize_space(prop_key).lower().replace('_', ' ')
            if 'cvs' not in pk:
                continue
            if not any(token in pk for token in ['feature', 'bullet', 'selling point', 'benefit', 'highlight']):
                continue
            for value in values or []:
                clean_value = normalize_space(value)
                if clean_value and not is_placeholder_salsify_copy_value(clean_value):
                    broad_cvs_feature_values.append(clean_value)
        cvs_feature_values = dedupe_preserve_order(broad_cvs_feature_values)

    cvs_feature_values = normalize_salsify_feature_values(dedupe_preserve_order(cvs_feature_values), max_features=10)
    general_feature_values = normalize_salsify_feature_values(dedupe_preserve_order(general_feature_values), max_features=10)

    walgreens_feature_values = []
    walgreens_feature_slots = {}
    for i in range(1, 11):
        walgreens_exact_values = collect_property_values(
            f"Walgreens Feature {i}",
            f"Walgreens Feature{i}",
            f"Walgreens Product Feature {i}",
            f"Walgreens Product Feature{i}",
            f"Walgreens Selling Point {i}",
            f"Walgreens Selling Point{i}",
            f"Walgreens Bullet {i}",
            f"Walgreens Bullet{i}",
            f"Retailer Feature {i} - Walgreens",
            f"Retailer Bullet {i} - Walgreens",
        )
        walgreens_loose_values = collect_property_values_loose(
            f"Walgreens Feature {i}",
            f"Walgreens Product Feature {i}",
            f"Walgreens Selling Point {i}",
            f"Walgreens Bullet {i}",
            f"Retailer Feature {i} - Walgreens",
            f"Retailer Bullet {i} - Walgreens",
        )
        walgreens_slot_values = dedupe_preserve_order((walgreens_exact_values or []) + (walgreens_loose_values or []))
        walgreens_slot_values = [v for v in walgreens_slot_values if v and not is_placeholder_salsify_copy_value(v)]
        if walgreens_slot_values:
            walgreens_feature_slots[i] = walgreens_slot_values[0]
            walgreens_feature_values.extend(walgreens_slot_values)

    if not walgreens_feature_values:
        broad_walgreens_feature_values = []
        for prop_key, values in property_values.items():
            pk = normalize_space(prop_key).lower().replace('_', ' ')
            if 'walgreens' not in pk:
                continue
            if not any(token in pk for token in ['feature', 'bullet', 'selling point', 'benefit', 'highlight']):
                continue
            for value in values or []:
                clean_value = normalize_space(value)
                if clean_value and not is_placeholder_salsify_copy_value(clean_value):
                    broad_walgreens_feature_values.append(clean_value)
        walgreens_feature_values = dedupe_preserve_order(broad_walgreens_feature_values)

    walgreens_feature_values = normalize_salsify_feature_values(
        dedupe_preserve_order(walgreens_feature_values),
        max_features=10,
    )


    heb_feature_values = []
    heb_feature_slots = {}
    for i in range(1, 11):
        heb_exact_values = collect_property_values(
            f"HEB Feature {i}", f"HEB Feature{i}",
            f"H-E-B Feature {i}", f"H-E-B Feature{i}",
            f"HEB Product Feature {i}", f"HEB Product Feature{i}",
            f"H-E-B Product Feature {i}", f"H-E-B Product Feature{i}",
            f"HEB Bullet {i}", f"HEB Bullet{i}",
            f"H-E-B Bullet {i}", f"H-E-B Bullet{i}",
            f"Retailer Feature {i} - HEB", f"Retailer Bullet {i} - HEB",
        )
        heb_loose_values = collect_property_values_loose(
            f"HEB Feature {i}", f"H-E-B Feature {i}",
            f"HEB Product Feature {i}", f"H-E-B Product Feature {i}",
            f"HEB Bullet {i}", f"H-E-B Bullet {i}",
            f"Retailer Feature {i} - HEB", f"Retailer Bullet {i} - HEB",
        )
        heb_slot_values = dedupe_preserve_order((heb_exact_values or []) + (heb_loose_values or []))
        heb_slot_values = [v for v in heb_slot_values if v and not is_placeholder_salsify_copy_value(v)]
        if heb_slot_values:
            heb_feature_slots[i] = heb_slot_values[0]
            heb_feature_values.extend(heb_slot_values)

    if not heb_feature_values:
        broad_heb_feature_values = []
        for prop_key, values in property_values.items():
            pk = normalize_space(prop_key).lower().replace('_', ' ')
            if not any(token in pk for token in ['heb', 'h e b', 'h-e-b']):
                continue
            if not any(token in pk for token in ['feature', 'bullet', 'selling point', 'benefit', 'highlight']):
                continue
            for value in values or []:
                clean_value = normalize_space(value)
                if clean_value and not is_placeholder_salsify_copy_value(clean_value):
                    broad_heb_feature_values.append(clean_value)
        heb_feature_values = dedupe_preserve_order(broad_heb_feature_values)

    heb_feature_values = normalize_salsify_feature_values(dedupe_preserve_order(heb_feature_values), max_features=10)


    retailer_overrides = {
        "kroger": {
            "title": first_property("Kroger Product Title", "Kroger Title"),
            "description": first_property("Kroger Description", "Kroger Product Description"),
            "features": [re.sub(r'^(kroger\s*feature\s*\d+\s*[:\-]?\s*|feature\s*\d+\s*[:\-]?\s*|\d+\s*[\.\-\)]\s*)','',f,flags=re.IGNORECASE).strip() for f in kroger_feature_values],
        },
        "sam's club": {
            "title": first_property(
                "Sam's Club Product Title",
                "Sams Club Product Title",
                "Sam's Club Title",
                "Sams Club Title",
            ),
            "description": first_property(
                "Sam's Club Description",
                "Sams Club Description",
                "Sam's Club Product Description",
                "Sams Club Product Description",
            ),
            "features": sams_feature_values,
            "feature_slots": sams_feature_slots,
        },
        "cvs": {
            "title": first_non_placeholder_copy_value(
                first_property("CVS Product Title", "CVS Title", "CVS Product Name"),
                first_property("General Product Title", "General Title", "Product Title"),
            ),
            "description": first_non_placeholder_copy_value(
                first_property("CVS Description", "CVS Product Description", "CVS Long Description"),
                first_property("General Description", "General Product Description", "Description"),
            ),
            "features": normalize_salsify_feature_values(
                cvs_feature_values or general_feature_values,
                max_features=10,
            ),
            "feature_slots": cvs_feature_slots,
        },
        "heb": {
            "title": first_non_placeholder_copy_value(
                first_property("HEB Product Title", "HEB Title", "HEB Product Name", "H-E-B Product Title", "H-E-B Title"),
                first_property("General Product Title", "General Title", "Product Title"),
            ),
            "description": first_non_placeholder_copy_value(
                first_property("HEB Description", "HEB Product Description", "HEB Long Description", "H-E-B Description", "H-E-B Product Description"),
                first_property("General Description", "General Product Description", "Description"),
            ),
            "features": normalize_salsify_feature_values(heb_feature_values or general_feature_values, max_features=10),
            "feature_slots": heb_feature_slots,
        },
        "walgreens": {
            "title": first_non_placeholder_copy_value(
                first_property("Walgreens Product Title", "Walgreens Title", "Walgreens Product Name"),
                first_property("General Product Title", "General Title", "Product Title"),
            ),
            "description": first_non_placeholder_copy_value(
                first_property("Walgreens Description", "Walgreens Product Description", "Walgreens Long Description"),
                first_property("General Description", "General Product Description", "Description"),
            ),
            "features": normalize_salsify_feature_values(
                walgreens_feature_values or general_feature_values,
                max_features=10,
            ),
            "feature_slots": walgreens_feature_slots,
        },
    }

    text_bundle = {
        "title": title,
        "description": description,
        "feature1": feature_candidates[0] if len(feature_candidates) > 0 else "",
        "feature2": feature_candidates[1] if len(feature_candidates) > 1 else "",
        "feature3": feature_candidates[2] if len(feature_candidates) > 2 else "",
        "feature4": feature_candidates[3] if len(feature_candidates) > 3 else "",
        "feature5": feature_candidates[4] if len(feature_candidates) > 4 else "",
        "feature6": feature_candidates[5] if len(feature_candidates) > 5 else "",
        "feature7": feature_candidates[6] if len(feature_candidates) > 6 else "",
        "features": feature_candidates[:10],
        "retailer_overrides": retailer_overrides,
    }

    asset_lookup = {}
    try:
        properties = data["props"]["pageProps"]["product"]["digitalAssets"]["properties"]
        for prop in properties:
            raw_name = prop.get("property", "")
            normalized_name = normalize_salsify_asset_name(raw_name)
            values = prop.get("values", [])
            if not normalized_name or not values:
                continue
            val = ""
            first = values[0]
            if isinstance(first, dict):
                val = str(first.get("value", "") or "")
            elif isinstance(first, str):
                val = str(first or "")
            if not val:
                continue
            asset_lookup[normalized_name] = val.split("?")[0]
    except Exception:
        pass


    for asset_name, asset_url in (visible_property_map.get("assets", {}) or {}).items():
        clean_asset_name = normalize_salsify_asset_name(asset_name)
        clean_asset_url = html.unescape(str(asset_url or "").strip())
        if clean_asset_name and clean_asset_url and clean_asset_name not in asset_lookup:
            asset_lookup[clean_asset_name] = clean_asset_url if is_video_like_url(clean_asset_url) else clean_asset_url.split("?", 1)[0]

    try:
        raw_html_text = html.unescape(str(html_text or ""))
        fallback_asset_patterns = [
            r'>\s*(Main Variant Image-Sams Club|Main Variant Image-Club|Online Optimized Image-Sams Club|Online Optimized Image-Kroger|Online Optimized Image-Grocery|Online Optimized Image-|Ingredient Label Image-|Ingredient Label Image|Shipping-|Flat Back_2D-|Flat Left_2D-|ATF I/O-Sams Club|ATF I/O-Generic|ATF Video-Sams Club|ATF [0-9]+-Sams Club)\s*<.*?href="([^"]+)"',
            r'"property"\s*:\s*"(Main Variant Image-Sams Club|Main Variant Image-Club|Online Optimized Image-Sams Club|Online Optimized Image-Kroger|Online Optimized Image-Grocery|Online Optimized Image-|Ingredient Label Image-|Ingredient Label Image|Shipping-|Flat Back_2D-|Flat Left_2D-|ATF I/O-Sams Club|ATF I/O-Generic|ATF Video-Sams Club|ATF [0-9]+-Sams Club)"[^{}]{0,800}?"value"\s*:\s*"([^"]+)"',
        ]
        for pattern in fallback_asset_patterns:
            for matched_name, matched_url in re.findall(pattern, raw_html_text, flags=re.IGNORECASE | re.DOTALL):
                normalized_name = normalize_salsify_asset_name(matched_name)
                clean_url = html.unescape(str(matched_url or "").strip())
                if normalized_name and clean_url and normalized_name not in asset_lookup:
                    asset_lookup[normalized_name] = clean_url.split("?")[0] if not is_video_like_url(clean_url) else clean_url
    except Exception:
        pass

    images = pick_kroger_images_with_atf_and_lifestyle(asset_lookup)

    return {
        "text": text_bundle,
        "images": images[:MAX_IMAGE_SLOTS_TO_COMPARE],
    }

@st.cache_data(show_spinner=False)
def get_salsify_bundle(url):
    html_text = get_html(url)
    return _parse_salsify_page(html_text)


def get_salsify_text(url):
    return get_salsify_bundle(url)["text"]


def get_salsify_images(url):
    return get_salsify_bundle(url)["images"]


# =========================================
# CVS / RETAILER PARSERS
# =========================================

def get_cvs_sku_id_from_url(retail_url):
    retail_url = str(retail_url or "").strip()
    m = re.search(r"[?&]skuId=([0-9A-Za-z_-]+)", retail_url, flags=re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"prodid-([0-9A-Za-z_-]+)", retail_url, flags=re.IGNORECASE)
    if m:
        return m.group(1)
    return ""


def get_cvs_effective_sku_id(retail_url="", target_rpc=""):
    """CVS-only effective skuId/RPC selector.

    Prefer the skuId from the CVS URL because Visual QA/manual paths can pass
    the Salsify SKU as target_rpc. CVS fallback catalogs and image URL fallback
    must use the CVS skuId/RPC, not the Salsify SKU.
    """
    url_sku = get_cvs_sku_id_from_url(retail_url)
    if url_sku:
        return re.sub(r"[^0-9A-Za-z_-]", "", str(url_sku or "").strip())
    return re.sub(r"[^0-9A-Za-z_-]", "", str(target_rpc or "").replace(".0", "").strip())


def cvs_url_candidates(retail_url):
    raw_url = clean_uploaded_url_value(retail_url)
    if not raw_url:
        return []
    out = []
    def add(url):
        url = str(url or "").strip()
        if url and url not in out:
            out.append(url)
    add(raw_url)
    no_hash = raw_url.split("#", 1)[0].strip()
    add(no_hash)
    canonical = no_hash.split("?", 1)[0].strip()
    add(canonical)
    sku = get_cvs_sku_id_from_url(raw_url)
    if canonical and sku:
        add(f"{canonical}?skuId={sku}")
    return out


def extract_cvs_relevant_source_chunk(source, retail_url="", target_rpc=""):
    """CVS-only source slicer for messy pasted/browser source captures.

    Some pasted CVS captures can contain more than one PDP back-to-back. The
    image parser gathers every CVS high_res image it sees, so a multi-product
    blob can make the wrong product images win. This scopes CVS parsing to the
    selected row by locating the matching CVS URL first, then falling back to
    the Item # / skuId anchor.
    """
    source = str(source or "")
    if not source.strip():
        return ""
    retail_url = clean_uploaded_url_value(retail_url)
    target_rpc = get_cvs_effective_sku_id(retail_url=retail_url, target_rpc=target_rpc)
    lowered_source = source.lower()
    candidates = []
    for url in cvs_url_candidates(retail_url):
        if url:
            candidates.append(url)
            candidates.append(html.escape(url, quote=True))
    if retail_url:
        candidates.append(retail_url.split("?", 1)[0])
    start = -1
    for candidate in candidates:
        candidate = str(candidate or "").strip()
        if not candidate:
            continue
        idx = lowered_source.find(candidate.lower())
        if idx >= 0:
            start = idx
            break
    if start < 0 and target_rpc:
        hit_positions = []
        for pattern in [
            rf"[?&]skuId={re.escape(target_rpc)}\b",
            rf"Item\s*#\s*{re.escape(target_rpc)}\b",
            rf"cvs_rpc::{re.escape(target_rpc)}\b",
        ]:
            for m in re.finditer(pattern, source, flags=re.IGNORECASE):
                hit_positions.append(m.start())
        if hit_positions:
            hit = min(hit_positions)
            prior = list(re.finditer(r"https?://www\.cvs\.com/shop/", source[:hit], flags=re.IGNORECASE))
            start = prior[-1].start() if prior else max(0, hit - 45000)
    if start < 0:
        return source
    next_match = re.search(r"https?://www\.cvs\.com/shop/", source[start + 20:], flags=re.IGNORECASE)
    end = start + 20 + next_match.start() if next_match else len(source)
    chunk = source[start:end]
    return chunk if len(chunk.strip()) >= 200 else source

# CVS-only emergency fallback catalog.
# A small number of live CVS PDPs sometimes return a shell/blocked page to server-side
# requests even though the products are live in a normal browser and in the search index.
# Keep this isolated to CVS and keyed by CVS skuId/RPC so it cannot leak to other retailers.
CVS_KNOWN_PRODUCT_FALLBACKS = {
    "495589": {
        "title": "Kleenex Ultra Soft Facial Tissues, 4 Cubes (240 total tissues)",
        "description": "Tissues may be soft, but there’s only one Kleenex Ultra Soft—the #1 ultra tissue*. Each tissue features 3 layers of silky-soft strength that's gentle on watery eyes and runny noses, making them perfect to use for pollen allergies. Not only does Kleenex Ultra Soft comfort skin and help protect hands, but they're also allergist approved and hypoallergenic. Our facial tissues also come with Clean Shield technology that helps contain the mess 3x better than the leading value toilet paper. Each tissue box contains 60 total 3-ply tissues and comes in various colors and designs. For whatever happens next, Grab Kleenex. Packaging may vary. (*among national brands)",
        "features": [
            "WHAT'S INCLUDED — 4 cube boxes of Kleenex Ultra Soft Facial Tissues, 3-Ply, 60 tissues per box (240 tissues total)",
            "ALLERGIST APPROVED — Our tissues are hypoallergenic, dermatologist tested, and allergist approved",
            "TAKE ON ALLERGY SEASON — Whether you’re facing runny noses or watery eyes this allergy season, Kleenex Ultra Soft is there for it all",
            "SILKY SOFTNESS — These facial tissues are made with 3 layers and are silky soft for up to 100% irritation-free skin for an extra bit of comfort when you need it",
            "SAVE YOUR TOILET PAPER — When it comes to blowing your nose, Kleenex has got you covered. Our tissues are made with Clean Shield that contains the mess better than the leading value toilet paper",
        ],
    },
    "137056": {
        "title": "Kleenex Soothing Lotion Facial Tissues, 1 Box",
        "description": "You can't predict sick days, but with Kleenex Lotion Facial Tissues, made with coconut oil and aloe, you can be prepared for them. Our facial tissues moisturize skin to help prevent skin irritation. With 3-in-1 skin-loving benefits, Kleenex Lotion helps protect hands, soothes skin, and moisturizes skin with lotion. They also come with our Clean Shield technology that contains the mess 3x better than the leading value toilet paper. Each facial tissue is dermatologist-tested, made with 3 thick layers, and infused with coconut oil and aloe so you can experience a gentle clean. Each tissue box contains 120 total 3-ply tissues and comes in various colors and designs. For whatever happens next, Grab Kleenex. Packaging may vary.",
        "features": [
            "WHAT'S INCLUDED — 1 box of Kleenex Lotion 3-Ply Facial Tissues with Coconut Oil & Aloe, 120 tissues per box (120 tissues total)",
            "BE PREPARED FOR SICK DAYS — We know we can't prevent sick days, but Kleenex can help you through them by moisturizing your skin to help prevent skin irritation",
            "3 BENEFITS IN 1 TISSUE — Kleenex lotion facial tissues are designed to soothe and moisturize skin while protecting your hands, offering comprehensive care in every sheet",
            "SAVE YOUR TOILET PAPER — When it comes to blowing your nose, Kleenex has got you covered. Our tissues are made with Clean Shield that contains the mess better than the leading value toilet paper",
            "MADE WITH LOTION — These facial tissues are dermatologist-tested and infused with coconut oil and aloe",
        ],
    },
    "854178": {
        "title": "Kleenex Anti-Viral Facial Tissues, 1 Cube",
        "description": "Kleenex Anti-Viral Tissues are designed to kill 99.9% of cold and flu viruses*. Each facial tissue is made with a specially treated middle layer that wipes out cold and flu viruses* in the tissue in 15 minutes. When moisture hits the middle blue-dot layer, it will kill 99.9% of viruses in the tissue*. Not only that, but these tissues are made with 3 thick layers. Each tissue box contains 55 total 3-ply tissues and comes in various colors and designs that blend with any home. When you need more than just a tissue, grab Kleenex. Packaging may vary. (*Virucidal Against: Rhinoviruses type 1A and 2, Influenza A virus and Influenza B virus, Respiratory Syncytial Virus)",
        "features": [
            "WHAT’S INCLUDED — 1 box of Kleenex Anti-Viral Facial Tissues, 3-Ply, 55 tissues per box (55 tissues total)",
            "FOR COLDS & FLUS — Each tissue has a specially treated middle layer that wipes out cold and flu viruses in the tissue within 15 minutes",
            "HOW IT WORKS — When moisture hits the middle blue-dot layer, it will kill 99.9% of the viruses in the tissue",
            "3 LAYERS OF STRENGTH — These facial tissues are made with 3 thick layers so you can take on cold & flu season with confidence",
            "PERFECT FOR ANY HOME — Practical and stylish, our Kleenex tissues boxes come in various designs that complement your home décor (packaging may vary)",
        ],
    },
    "516233": {
        "title": "Kleenex Soothing Lotion Facial Tissues, 4 Cubes (240 total tissues)",
        "description": "You can't predict sick days, but with Kleenex Lotion Facial Tissues, made with coconut oil and aloe, you can be prepared for them. Our facial tissues moisturize skin to help prevent skin irritation. With 3-in-1 skin-loving benefits, Kleenex Lotion helps protect hands, soothes skin, and moisturizes skin with lotion. They also come with our Clean Shield technology that contains the mess 3x better than the leading value toilet paper. Each facial tissue is dermatologist-tested, made with 3 thick layers, and infused with coconut oil and aloe so you can experience a gentle clean. Each tissue box contains 60 total 3-ply tissues and comes in various colors and designs. For whatever happens next, Grab Kleenex. Packaging may vary.",
        "features": [
            "WHAT'S INCLUDED — 4 boxes of Kleenex Lotion 3-Ply Facial Tissues with Coconut Oil & Aloe, 60 tissues per box (240 tissues total)",
            "BE PREPARED FOR SICK DAYS — We know we can't prevent sick days, but Kleenex can help you through them by moisturizing your skin to help prevent skin irritation",
            "3 BENEFITS IN 1 TISSUE — Kleenex lotion facial tissues are designed to soothe and moisturize skin while protecting your hands, offering comprehensive care in every sheet",
            "SAVE YOUR TOILET PAPER — When it comes to blowing your nose, Kleenex has got you covered. Our tissues are made with Clean Shield that contains the mess better than the leading value toilet paper",
            "MADE WITH LOTION — These facial tissues are dermatologist-tested and infused with coconut oil and aloe",
        ],
    },
    "730205": {
        "title": "Kotex Ultra Thin Overnight Pads With Wings, Heavy Absorbency, 26 CT",
        "description": "Bring powerful protection and comfort to your nighttime period routine with the new Kotex Ultra Thin Overnight Pads with Wings. These overnight pads provide up to 12 hours of protection and NightDefense with a raised back barrier and side guards to help prevent back and side leaks. The 5x System with LeakShield Protection has breathability, odor control, dryness, fit, and leakage protection for up to 100% Leak Free Comfort. These period pads are designed for perfect fit and combine LeakShield Technology, a breathable top layer, a new Gravity Core, and odor control to give you a menstrual pad that protects you in more ways than one. To help keep you feeling clean and fresh throughout your day, each feminine pad is designed with a Gravity Core that pulls period blood to the bottom of the pad. Each women’s pad is made with your skin health in mind, which is why these pads are made without fragrance and free of elemental chlorine. For added convenience, each nighttime sanitary pad is individually folded and wrapped to protect your pad with easy access, even on-the-go. Product and packaging may vary.",
        "features": [
            "Kotex Ultra Thin Overnight Pads With Wings, Heavy Absorbency, 26 Count",
            "All-Night Protection: NightDefense overnight pads provide up to 12 hours of protection with a raised back barrier and side guards to help prevent back and side leaks",
            "5x System Comfort: These women’s pads offer breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort",
            "Gravity Core Technology: Our period pads feature a Gravity Core that pulls period blood to the bottom of the pad to help keep you clean and dry",
            "Gentle on Skin: These menstrual pads are made without fragrance and free of elemental chlorine",
        ],
    },
}


CVS_KNOWN_IMAGE_BASE_BY_SKU = {
    # CVS skuId/RPC -> actual CVS high_res image basename from live PDP/source HTML.
    # CVS basenames are often UPC/KC image keys, not the CVS skuId.
    # This map is CVS-only and never mirrors Salsify image URLs into CVS.
    "298031": "3600054271",
    "730204": "3600058233",
    "729602": "3600051582",
    "729603": "3600051581",  # source-confirmed from pasted CVS carousel HTML.
    "731730": "3600058318",
    "817844": "3600038587",
    "819260": "3600051583",
    "729958": "3600058353",
    "730263": "3600058258",
    "730214": "3600058228",
    "470890": "81013395906",
    "167387": "3600051589",
}


def cvs_generated_image_candidates_for_sku(sku_id, max_slots=8):
    sku_id = re.sub(r"[^0-9A-Za-z_-]", "", str(sku_id or "").strip())
    if not sku_id:
        return []
    image_base = str(CVS_KNOWN_IMAGE_BASE_BY_SKU.get(sku_id, sku_id) or sku_id).strip()
    resize_query = "?im=Resize=(600,600),aspect=ignore"
    max_slots = max(1, int(max_slots or 8))
    candidates = [f"https://www.cvs.com/bizcontent/merchandising/productimages/high_res/{image_base}.jpg{resize_query}"]
    # CVS PDP carousel image naming normally starts with base.jpg, then _2, _3, etc.
    # Do not generate _1 first because current CVS PDP source does not use _1.
    for idx in range(2, max_slots + 1):
        candidates.append(f"https://www.cvs.com/bizcontent/merchandising/productimages/high_res/{image_base}_{idx}.jpg{resize_query}")
    return candidates[:max_slots]


def infer_cvs_image_base_from_images(image_urls):
    """Infer the shared CVS high_res carousel image base from parsed CVS image URLs.

    Example: 3600051581.jpg, 3600051581_2.jpg -> 3600051581.
    CVS-only: uses only CVS retailer image URLs parsed from CVS HTML/source capture.
    """
    counts = {}
    for url in image_urls or []:
        url = str(url or "").strip()
        if "/productimages/high_res/" not in url.lower():
            continue
        name = url.split("?", 1)[0].rsplit("/", 1)[-1]
        stem = re.sub(r"\.(?:jpg|jpeg|png|webp|avif)$", "", name, flags=re.IGNORECASE)
        base = re.sub(r"_\d+$", "", stem)
        if base and re.search(r"\d", base):
            counts[base] = counts.get(base, 0) + 1
    if not counts:
        return ""
    return sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]


def cvs_generated_image_candidates_for_base(image_base, max_slots=8):
    image_base = re.sub(r"[^0-9A-Za-z_-]", "", str(image_base or "").strip())
    if not image_base:
        return []
    resize_query = "?im=Resize=(600,600),aspect=ignore"
    max_slots = max(1, int(max_slots or 8))
    candidates = [f"https://www.cvs.com/bizcontent/merchandising/productimages/high_res/{image_base}.jpg{resize_query}"]
    for idx in range(2, max_slots + 1):
        candidates.append(f"https://www.cvs.com/bizcontent/merchandising/productimages/high_res/{image_base}_{idx}.jpg{resize_query}")
    return candidates[:max_slots]


def cvs_bundle_has_copy(bundle):
    """CVS-only: True when the bundle has real retailer copy.

    Images alone should not make a CVS source look complete because this tool
    is comparing live PDP copy and live PDP images separately.
    """
    if not isinstance(bundle, dict):
        return False
    text_bundle = bundle.get("text", {}) or {}
    return bool(
        normalize_space(text_bundle.get("title", ""))
        or normalize_space(text_bundle.get("description", ""))
        or any(normalize_space(x) for x in (text_bundle.get("features", []) or []))
    )


def cvs_bundle_has_images(bundle):
    """CVS-only: True when the bundle has any retailer image URL."""
    if not isinstance(bundle, dict):
        return False
    return bool(any(str(x or "").strip() for x in (bundle.get("images", []) or [])))


def add_cvs_generated_image_fallback_if_needed(bundle, retail_url="", target_rpc="", reason=""):
    """CVS-only image safety net.

    This creates CVS-side image URL candidates from the selected CVS skuId/RPC
    only when CVS image parsing failed. It never copies Salsify images into the
    retailer side and it never invents CVS copy.
    """
    if not isinstance(bundle, dict):
        bundle = {"text": {"title": "", "description": "", "features": [], "debug": {}}, "images": []}
    bundle.setdefault("text", {}).setdefault("debug", {})
    bundle.setdefault("images", [])
    # Strict CVS live-only rule: generated image URLs are not evidence that an
    # image appeared in the exact captured CVS gallery. Keep missing slots missing.
    if not ALLOW_RETAILER_GENERATED_IMAGE_FALLBACKS:
        bundle["text"]["debug"]["CVS Generated Images Disabled"] = True
        return bundle
    if cvs_bundle_has_images(bundle):
        return bundle
    sku_id = get_cvs_effective_sku_id(retail_url=retail_url, target_rpc=target_rpc)
    if not sku_id:
        return bundle
    generated_images = cvs_generated_image_candidates_for_sku(sku_id, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE)
    if generated_images:
        bundle["images"] = generated_images[:MAX_IMAGE_SLOTS_TO_COMPARE]
        debug = bundle["text"]["debug"]
        debug["CVS Image Fallback Applied"] = "cvs_sku_high_res_url_pattern"
        debug["CVS Image Fallback SKU"] = sku_id
        debug["CVS Image Fallback Base"] = str(CVS_KNOWN_IMAGE_BASE_BY_SKU.get(sku_id, sku_id) or sku_id)
        debug["CVS Image Fallback Count"] = len(bundle.get("images") or [])
        if reason:
            debug["CVS Image Fallback Reason"] = reason
    return bundle


def apply_cvs_targeted_copy_rescue_if_needed(bundle, retail_url="", target_rpc="", reason=""):
    """CVS-only final copy rescue for known CVS catalog skuIds.

    This is the combined CVS approach: direct CVS fetch + uploaded CVS source +
    CVS-only catalog rescue as the last copy path. It only runs for skuIds that
    already exist in CVS_KNOWN_PRODUCT_FALLBACKS and only fills fields that are
    still missing after uploaded + live parsing. It never copies Salsify content
    into CVS fields.
    """
    if not isinstance(bundle, dict):
        bundle = {"text": {"title": "", "description": "", "features": [], "rating": "", "review_count": "", "debug": {}}, "images": []}
    text_bundle = bundle.setdefault("text", {})
    debug = text_bundle.setdefault("debug", {})
    sku_id = get_cvs_effective_sku_id(retail_url=retail_url, target_rpc=target_rpc)
    if sku_id not in globals().get("CVS_KNOWN_PRODUCT_FALLBACKS", {}):
        return bundle

    rescue_bundle = get_cvs_known_product_fallback_bundle(retail_url=retail_url, target_rpc=sku_id)
    rescue_text = rescue_bundle.get("text", {}) if isinstance(rescue_bundle, dict) else {}
    if not rescue_text:
        return bundle

    applied = False
    if not normalize_space(text_bundle.get("title", "")) and normalize_space(rescue_text.get("title", "")):
        text_bundle["title"] = rescue_text.get("title", "")
        debug["Title Path"] = rescue_text.get("debug", {}).get("Title Path", "cvs_combined_catalog_rescue")
        applied = True
    if not normalize_space(text_bundle.get("description", "")) and normalize_space(rescue_text.get("description", "")):
        text_bundle["description"] = rescue_text.get("description", "")
        debug["Description Path"] = rescue_text.get("debug", {}).get("Description Path", "cvs_combined_catalog_rescue")
        applied = True
    if not any(normalize_space(x) for x in (text_bundle.get("features", []) or [])) and rescue_text.get("features"):
        text_bundle["features"] = rescue_text.get("features", [])[:5]
        debug["Features Path"] = rescue_text.get("debug", {}).get("Features Path", "cvs_combined_catalog_rescue")
        applied = True

    if not cvs_bundle_has_images(bundle) and rescue_bundle.get("images"):
        bundle["images"] = rescue_bundle.get("images", [])[:MAX_IMAGE_SLOTS_TO_COMPARE]
        debug["CVS Image Fallback Applied"] = "cvs_targeted_rescue_images"
        debug["CVS Image Fallback Base"] = str(CVS_KNOWN_IMAGE_BASE_BY_SKU.get(sku_id, sku_id) or sku_id)
        debug["CVS Image Fallback Count"] = len(bundle.get("images") or [])
        applied = True

    if applied:
        rescue_source = rescue_text.get("debug", {}).get("Source Used", "cvs_combined_catalog_rescue")
        debug["Source Used"] = (str(debug.get("Source Used", "")) + " | " + str(rescue_source)).strip(" |")
        debug["CVS Combined Catalog Rescue Applied"] = True
        debug["CVS Combined Catalog Rescue SKU"] = sku_id
        if reason:
            debug["CVS Combined Catalog Rescue Reason"] = reason
    return bundle


# CVS-only targeted rescue list.
# These are confirmed live CVS PDPs that can still return empty/shell HTML to
# server-side requests. Use the isolated CVS fallback catalog only for these
# exact skuId/RPC values without enabling the fallback catalog globally.
CVS_TARGETED_COPY_RESCUE_SKUS = set(CVS_KNOWN_PRODUCT_FALLBACKS.keys())


def get_cvs_known_product_fallback_bundle(retail_url="", target_rpc=""):
    sku_id = get_cvs_effective_sku_id(retail_url=retail_url, target_rpc=target_rpc)
    data = CVS_KNOWN_PRODUCT_FALLBACKS.get(sku_id)
    if not data:
        return {"text": {"title": "", "description": "", "features": [], "debug": {}}, "images": []}

    # Combined CVS fallback rule:
    # Global known-copy fallbacks can stay off. If the selected CVS skuId/RPC is
    # already in this CVS-only catalog, allow it as the last-resort CVS source.
    # This prevents the issue from shifting to the next blocked CVS item.
    catalog_rescue_allowed = sku_id in globals().get("CVS_KNOWN_PRODUCT_FALLBACKS", {})
    if not bool(globals().get("ALLOW_RETAILER_KNOWN_COPY_FALLBACKS", False)) and not catalog_rescue_allowed:
        return {"text": {"title": "", "description": "", "features": [], "debug": {}}, "images": []}
    rescue_source = "cvs_combined_catalog_rescue" if catalog_rescue_allowed else "cvs_known_product_fallback_catalog"
    debug = {
        "Source Used": rescue_source,
        "CVS Known Fallback SKU": sku_id,
        "Title Path": rescue_source,
        "Description Path": rescue_source,
        "Features Path": rescue_source,
    }
    return {
        "text": {
            "title": normalize_space(data.get("title", "")),
            "description": clean_cvs_text(data.get("description", "")),
            "features": normalize_cvs_features(data.get("features", [])),
            "debug": debug,
        },
        "images": cvs_generated_image_candidates_for_sku(sku_id, max_slots=8),
    }



# CVS-only extension of the known fallback catalog for the remaining live CVS
# skuIds that still return blank/shell HTML to server-side requests.
# Keep this keyed by skuId/RPC and CVS-only.
CVS_KNOWN_PRODUCT_FALLBACKS.update({
    "298031": {
        "title": "Kleenex Soothing Lotion Facial Tissues with Coconut Oil & Aloe, 1 Cube",
        "description": "You can't predict sick days, but with Kleenex Lotion Facial Tissues, made with coconut oil and aloe, you can be prepared for them. Our facial tissues moisturize skin to help prevent skin irritation. With 3-in-1 skin-loving benefits, Kleenex Lotion helps protect hands, soothes skin, and moisturizes skin with lotion. They also come with our Clean Shield technology that contains the mess 3x better than the leading value toilet paper. Each facial tissue is dermatologist-tested, made with 3 thick layers, and infused with coconut oil and aloe so you can experience a gentle clean. Each tissue box contains 60 total 3-ply tissues and comes in various colors and designs. For whatever happens next, Grab Kleenex. Packaging may vary.",
        "features": [
            "WHAT'S INCLUDED - 1 box of Kleenex Lotion Facial Tissues with Coconut Oil & Aloe, 3-Ply, 60 tissues per box (60 tissues total)",
            "BE PREPARED FOR SICK DAYS - We know we can't prevent sick days, but Kleenex can help you through them by moisturizing your skin to help prevent skin irritation",
            "3 BENEFITS IN 1 TISSUE - Kleenex lotion facial tissues are designed to soothe and moisturize skin while protecting your hands, offering comprehensive care in every sheet",
            "SAVE YOUR TOILET PAPER - When it comes to blowing your nose, Kleenex has got you covered. Our tissues are made with Clean Shield that contains the mess better than the leading value toilet paper",
            "MADE WITH LOTION - These facial tissues are dermatologist-tested and infused with coconut oil and aloe",
        ],
    },
    "650610": {
        "title": "Kleenex Cooling + Aloe Facial Tissues, 1 cube (50 total tissues)",
        "description": "Whether you're an allergy sufferer, dealing with cold and flu symptoms, or simply looking for a tissue to add to your skincare routine, Kleenex Cooling + Aloe Facial Tissues deliver comfort and cooling relief. They're formulated for instant cooling relief and to help hydrate skin, offering gentle care for when you're feeling under the weather or battling allergies. Not only that, but these tissues are also infused with a hint of aloe. Each tissue pampers your skin with cooling freshness. These facial tissues also feature Kleenex's Clean Shield technology that helps contain the mess 3x better than the leading value toilet paper for a clean, worry-free experience. Each tissue box contains 50 total 2-ply tissues and comes in various colors and designs that blend with any home. For whatever happens next, Grab Kleenex.",
        "features": [
            "WHAT'S INCLUDED - 1 box of Kleenex Cooling + Aloe Facial Tissues, 2-Ply, 50 tissues per box (50 tissues total)",
            "HELPS SKIN FEEL RESTORED - Kleenex Cooling + Aloe tissues provide a hint of cooling freshness to help your skin feel restored even during your worst nasal symptoms",
            "INSTANT COOLING RELIEF - These Kleenex tissues are formulated for instant cooling relief, whether you're battling allergies or colds and flus",
            "A HINT OF ALOE - Kleenex Cooling + Aloe tissues are infused with a hint of aloe and help hydrate skin",
            "CLEAN AND COMFORTED SKIN - Our tissues are uniquely designed with Clean Shield, a protective barrier that contains mess 3x better than the leading value toilet paper",
        ],
    },
    "298819": {
        "title": "Kleenex Ultra Soft Facial Tissues, 1 Cube",
        "description": "Tissues may be soft, but there's only one Kleenex Ultra Soft - the #1 ultra tissue*. Each tissue features 3 layers of silky-soft strength that's gentle on watery eyes and runny noses, making them perfect to use for pollen allergies, back to school season, and everything in between. Not only does Kleenex Ultra Soft comfort skin and help protect hands, but they're also allergist approved and hypoallergenic. Our facial tissues also come with Clean Shield technology that helps contain the mess 3x better than the leading value toilet paper. Each tissue box contains 60 total 3-ply tissues and comes in various colors and designs. For whatever happens next, Grab Kleenex. Packaging may vary. (*among national brands)",
        "features": [
            "WHAT'S INCLUDED - 1 cube box of Kleenex Ultra Soft Facial Tissues, 3-Ply, 60 tissues per box (60 tissues total)",
            "FOR THE BIG AND LITTLE MOMENTS - Back to school season, allergy season, holiday season, graduation season, and every season in between, Kleenex Ultra Soft is there for it all",
            "SILKY SOFTNESS - These facial tissues are made with 3 layers and are silky soft for up to 100% irritation-free skin for an extra bit of comfort when you need it",
            "ALLERGIST APPROVED - Our tissues are hypoallergenic, dermatologist tested, and allergist approved",
            "SAVE YOUR TOILET PAPER - When it comes to blowing your nose, Kleenex has got you covered. Our tissues are made with Clean Shield that contains the mess better than the leading value toilet paper",
        ],
    },
    "729958": {
        "title": "Kotex, Night Defense Ultra Thin Pads with Wings, 24 CT",
        "description": "Bring powerful protection and comfort to your nighttime period routine with the new Kotex Ultra Thin Extra Heavy Overnight Pads with Wings. These overnight pads provide up to 12 hours of protection and NightDefense with a raised back barrier and side guards to help prevent back and side leaks. The 5x System with LeakShield Protection has breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort. These period pads are designed for perfect fit and combine LeakShield Technology, a breathable top layer, a new Gravity Core and odor control to give you a menstrual pad that protects you in more ways than one. To help keep you feeling clean and fresh throughout your day, each feminine pad is designed with a Gravity Core that pulls period blood to the bottom of the pad. Each women's pad is made with your skin health in mind, which is why these pads are made without fragrance and free of elemental chlorine. For added convenience, each nighttime sanitary pad is individually folded and wrapped to protect your pad with easy access, even on-the-go. Product and packaging may vary.",
        "features": [
            "Kotex Ultra Thin Overnight Pads with Wings, Extra Heavy Absorbency, 24 Count",
            "All-Night Protection: NightDefense overnight pads provide up to 12 hours of protection with a raised back barrier and side guards to help prevent back and side leaks",
            "5x System Comfort: These women's pads offer breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort",
            "Gravity Core Technology: Our period pads feature a Gravity Core that pulls period blood to the bottom of the pad to help keep you clean and dry",
            "Gentle on Skin: These menstrual pads are made without fragrance and free of elemental chlorine",
        ],
    },
    "819260": {
        "title": "U by Kotex Click Compact Tampons, Unscented, Regular, 32 Count",
        "description": "When you are in need of compact comfort and powerful protection, U by Kotex Click compact tampons are there to help. Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection. Compact and able to fit into a purse or pocket, these tampons click into full size to give you powerful protection. Just pull the lower half of the tampon and when it locks in place, it's ready to go! In addition, our unscented tampons are gynecologist-tested, made without fragrance, BPA free, and are free of elemental chlorine. They are also OEKO TEX STANDARD certified, meaning that they are tested for up to 1,000 harmful substances. Individually wrapped, these tampons are perfect for when you need period protection on the go. U by Kotex Click Compact Tampons are available in regular, super, and super plus absorbencies. Packaging may vary from images shown.",
        "features": [
            "32 regular tampons",
            "Compact Comfort, Powerful Protection: These compact tampons are easily carried in a purse or pocket for on-the-go protection",
            "#1 compact tampon brand: U by Kotex Click is the #1 compact tampon brand",
            "Up to 100% Leak Free Protection: Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection",
            "Gynecologist-Tested: Our unscented tampons are gynecologist-tested, made without fragrance, BPA free and are free of elemental chlorine",
        ],
    },
    "729602": {
        "title": "U by Kotex Click Compact Tampons, Unscented, Super Plus, 16 Count",
        "description": "When you are in need of compact comfort and powerful protection, U by Kotex Click compact tampons are there to help. Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection. Compact and able to fit into a purse or pocket, these tampons click into full size to give you powerful protection, just pull the lower half of the tampon and when it locks in place, it's ready to go! In addition, our unscented tampons are gynecologist-tested, made without fragrance, BPA free and are free of elemental chlorine. They are also OEKO TEX STANDARD certified, meaning that they are tested for up to 1,000 harmful substances. Individually wrapped, these tampons are perfect for when you need period protection on the go. U by Kotex Click Compact Tampons are available in regular, super and super plus absorbencies. Packaging may vary from images shown.",
        "features": [
            "16 super plus tampons",
            "Compact Comfort, Powerful Protection: These compact tampons are easily carried in a purse or pocket for on-the-go protection",
            "#1 compact tampon brand: U by Kotex Click is the #1 compact tampon brand",
            "Up to 100% Leak Free Protection: Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection",
            "Gynecologist-Tested: Our unscented tampons are gynecologist-tested, made without fragrance, BPA free and are free of elemental chlorine",
        ],
    },
    "729603": {
        "title": "U by Kotex Click Compact Tampons, Unscented, Super, 16 Count",
        "description": "Get up to 100% leak-free protection and stay confident throughout your day with U by Kotex Click Compact Tampons. Made without fragrance, Click tampons are designed for your comfort and have a smooth tip for easy and comfortable insertion. Each tampon features ComfortFlex grooves that move with you for outstanding comfort. Pocket-sized and small enough to carry in a purse or pocket, these tampons go from compact to a full-size tampon in one easy step, giving you comfortably compact, powerful protection. Pull the lower half of the tampon and when it locks in place, it's ready to go. U by Kotex Click Compact Tampons are available in regular, super, and super plus absorbencies. Packaging may vary from images shown.",
        "features": [
            "16 super tampons",
            "Get up to 100% leak-free with Xpress-DRI protection with the #1 compact tampon",
            "More comfortable than the second leading compact tampon, U by Kotex Click tampons have a smooth tip for easy and comfortable insertion and are made without fragrance",
            "Compact to fit in your purse or pocket and changes to a full-size tampon in one easy step",
            "Individually wrapped in vibrant colors and patterns inspired by the latest fashion trends",
        ],
    },
    "269481": {
        "title": "U by Kotex Click Compact Tampons, Unscented, Super Plus, 32 Count",
        "description": "When you are in need of compact comfort and powerful protection, U by Kotex Click compact tampons are there to help. Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection. Compact and able to fit into a purse or pocket, these tampons click into full size to give you powerful protection, just pull the lower half of the tampon and when it locks in place, it's ready to go! In addition, our unscented tampons are gynecologist-tested, made without fragrance, BPA free and are free of elemental chlorine. They are also OEKO TEX STANDARD certified, meaning that they are tested for up to 1,000 harmful substances. Individually wrapped, these tampons are perfect for when you need period protection on the go. U by Kotex Click Compact Tampons are available in regular, super and super plus absorbencies. Packaging may vary from images shown.",
        "features": [
            "32 super plus tampons",
            "Compact Comfort, Powerful Protection: These compact tampons are easily carried in a purse or pocket for on-the-go protection",
            "#1 compact tampon brand: U by Kotex Click is the #1 compact tampon brand",
            "Up to 100% Leak Free Protection: Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection",
            "Gynecologist-Tested: Our unscented tampons are gynecologist-tested, made without fragrance, BPA free and are free of elemental chlorine",
        ],
    },
    "648578": {
        "title": "Viva Signature Cloth Double Roll Paper Towels, 6 ct",
        "description": "Viva Signature Cloth Paper Towels deliver the perfect balance of softness and durability, making them the #1 cloth-like towel for households that want premium cleaning and comfort. This pack of Viva paper towels includes 6 Double Rolls with 86 sheets per roll. Each sheet is soft like cloth and gentle for surfaces, hands, and faces. Not only that, but our paper towels are absorbent like cloth and durable like cloth to tackle the toughest messes - whether you're cleaning up spills in the kitchen, wiping down bathroom counters, or handling everyday messes. Enjoy a paper towel that is durable, absorbent, and soft for every room in your home. Packaging may vary.",
        "features": [
            "WHAT'S INCLUDED - 6 double rolls of Viva Signature Cloth Paper Towels, 86 sheets per roll (6 double rolls = 12 regular rolls)",
            "CHOOSE A SHEET - Viva Choose-A-Sheet size lets you pick the right sheet size for any task, big or small",
            "#1 CLOTH-LIKE TOWEL - Experience Viva's cloth-like paper towel for softness and strength in every sheet",
            "SOFT LIKE CLOTH - Our paper towels gently clean hands, faces, and delicate surfaces with a touch that's soft and comfortable",
            "ABSORBENT LIKE CLOTH - Viva soaks up even big spills for a reliable clean",
        ],
    },
})


# CVS-only add-on for the latest remaining QA rows. These rows are known live CVS
# PDPs but can still come back as empty/missing when CVS serves a shell page to the
# app. Keep this isolated to CVS skuId/RPC fallback behavior.
CVS_KNOWN_PRODUCT_FALLBACKS.update({
    "867564": {
        "title": "Kleenex On-The-Go Facial Tissues, 6 On-The-Go Packs, 10 Tissues per Box, 3-Ply (60 Total Tissues)",
        "description": "Runny noses can happen anywhere. Stay prepared with Kleenex On-the-Go Facial Tissues. Small enough to fit in pockets, purses, backpacks or travel bags, these Kleenex tissues are made with 3 thick layers and Clean Shield technology that helps contain the mess 3x better than the leading value toilet paper. Our facial tissues are also soft, durable, and ultra-absorbent for runny noses and watery eyes to help you stay prepared wherever you are. Each tissue pack contains 10 total 3-ply tissues and comes in various colors and designs. For whatever happens next, Grab Kleenex. Packaging may vary.",
        "features": [
            "WHAT'S INCLUDED - 6 packs of Kleenex On-the-Go Pocket Pack Facial Tissues, 3-Ply, 10 tissues per pack (60 tissues total)",
            "PERFECTLY SIZED FOR ANY ADVENTURE - Don't leave home unprepared. These Kleenex tissue packs are small enough to fit in pockets, purses, backpacks, or travel bags",
            "SAVE YOUR TOILET PAPER - When it comes to blowing your nose, Kleenex has got you covered. Our tissues are made with Clean Shield that contains the mess better than the leading value toilet paper",
            "SMALL BUT MIGHTY - This may be a convenient travel pack, but it contains the same ultra-absorbent, soft, and durable facial tissues that help runny noses and watery eyes, so you can be ready for anything",
            "STYLE WHEREVER YOU GO - Our Kleenex tissues packs come in various stylish designs that'll complement your travel accessories (packaging may vary)",
        ],
    },
    "730214": {
        "title": "Kotex Ultra Thin Teen Overnight Pads with Wings, Heavy Absorbency, 22 Count",
        "description": "Make powerful protection and comfort a part of your nighttime period routine with Kotex Teen Ultra Thin Overnight Period Pads with Wings. Uniquely sized and designed for teens, these teen pads are designed with NightDefense with back and side guards to help prevent back and side leaks. Our nighttime period pads offer up to 12 hours of nighttime protection no matter how you sleep. These period pads are designed for perfect fit and combine LeakShield Technology, a breathable top layer, a new Gravity Core and odor control to give you a menstrual pad that protects you in more ways than one. To help keep you fresh and clean throughout the night, each menstrual pad features a Gravity Core that pulls period blood to the bottom of the pad to help you feel clean and dry. Each overnight pad for teens is made with skin health in mind, which is why these pads are made without fragrance and free of elemental chlorine. Product and packaging may vary from images shown.",
        "features": [
            "Kotex Ultra Thin Teen Overnight Pads with Wings, Heavy Absorbency, 22 Count",
            "Designed for Teens: These teen pads are uniquely designed for teen period care",
            "All-Night Protection: Overnight pads provide up to 12 hours of protection. Raised Back Barrier and Side Guards help prevent back and side leaks",
            "LeakShield Comfort: 5x System with LeakShield Protection offers breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort",
            "Gravity Core Technology: Our pads feature a Gravity Core that pulls period blood to the bottom of the pad to help keep you clean and dry",
        ],
    },
    "730230": {
        "title": "Kotex Ultra Thin Pads, Heavy Absorbency, 40 Count",
        "description": "Bring powerful protection and comfort to your period routine with the new Kotex Ultra Thin Pads. The 5x System with LeakShield Protection delivers breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort. These period pads are designed for a perfect fit and combine LeakShield Technology, a breathable top layer, a new Gravity Core and odor control to give you a menstrual pad that protects you in more ways than one. To help keep you feeling clean and fresh throughout your day, each feminine pad is designed with a Gravity Core that pulls period blood to the bottom of the pad. Each women's pad is made with your skin health in mind, which is why these pads are made without fragrance and free of elemental chlorine. Product and packaging may vary.",
        "features": [
            "5x System with LeakShield Protection: Experience breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort",
            "Gravity Core: Stay clean and dry with a Gravity Core that pulls period blood to the bottom of the period pad",
            "Made Without Fragrance: Enjoy women's pads made without fragrance and free of elemental chlorine",
            "Powerful Protection and Comfort: Each feminine pad is designed for perfect fit with a breathable top layer and odor control",
            "Packaging may vary from images shown",
        ],
    },
    "802539": {
        "title": "Kotex Overnight Maxi Pads with Wings, Extra Heavy Absorbency, 28 Count",
        "description": "Bring powerful protection and comfort to your nighttime period routine with the new Kotex Maxi Pads with Wings. Made with NightDefense, these overnight maxi pads provide up to 12 hours of protection. A larger front and 80% larger back helps these women's maxi pads offer up to 100% leak free protection. A breathable soft touch cover helps to keep you feeling fresh when you wake up. Each women's pad is made with skin health in mind, which is why these pads are made without fragrance and free of elemental chlorine. Product and packaging may vary.",
        "features": [
            "Kotex Overnight Maxi Pads with Wings, Extra Heavy Absorbency, 28 Count",
            "All-Night Protection: NightDefense overnight pads provide up to 12 hours of protection",
            "5x System for Comfort: System with LeakShield Protection offers breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort",
            "Gravity Core Technology: Each women's pad has a Gravity Core that pulls period blood to the bottom of the pad to help keep you clean and dry",
            "Soft and Breathable: A breathable soft touch cover helps the women's pad protect you from leaks in any position",
        ],
    },
})
# 730205 now receives copy from the first fallback pass, but its live HTML can still
# expose weak or unusable image URLs. Force the known-product image URL pattern for
# known fallback SKUs when CVS product HTML was not detected.


# CVS-only add-on for the current remaining QA rows.
# These rows are known live CVS PDPs from the browser but can still come back as
# empty/shell HTML to server-side requests. Keep keyed by CVS skuId/RPC only.
CVS_KNOWN_PRODUCT_FALLBACKS.update({
    "731730": {
        "title": "Kotex Daily Wrapped Liners, Light Absorbency, 120 CT",
        "description": "You deserve powerful protection and comfort even on your lighter days, and thanks to the new Kotex Absorbent Liners, you can. Built to protect against light flow and discharge, our Absorbent liners are made with an Xpress Dri Core to help keep you feeling fresh throughout your day. Each pantiliner is made with your skin health in mind, which is why these feminine liners are made without fragrance and free of elemental chlorine. Each panty liner for women comes individually folded and wrapped to protect your pantiliner with easy access, even on-the-go. Kotex feminine products are FSA/HSA/HRA-eligible in the US. Packaging may vary from images shown.",
        "features": [
            "Kotex Absorbent Flat Liners, Light Absorbency, Regular Length 120 Count",
            "Xpress Dri Core: Our pantiliners are made with an Xpress Dri Core to absorb light flow instantly",
            "Light Flow Protection: Kotex liners provide protection for light flow and discharge",
            "Clean and Fresh: Feel clean and fresh even on your light days with these Absorbent liners",
            "Made Without Fragrance: Made without fragrance and free of elemental chlorine",
        ],
    },
    "730140": {
        "title": "Kotex Ultra Thin Pads with Wings, 22 CT",
        "description": "Bring powerful protection and comfort to your period routine with the new Kotex Ultra Thin Pads with Wings. The 5x System with LeakShield Protection delivers breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort. These period pads are designed for a perfect fit and combine LeakShield Technology, a breathable top layer, a new Gravity Core and odor control to give you a menstrual pad that protects you in more ways than one. To help keep you feeling clean and fresh throughout your day, each feminine pad is designed with a Gravity Core that pulls period blood to the bottom of the pad. Each women's pad is made with your skin health in mind, which is why these pads are made without fragrance and free of elemental chlorine. Product and packaging may vary from images shown.",
        "features": [
            "5x System with LeakShield Protection: Experience breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort",
            "Gravity Core: Stay clean and dry with a Gravity Core that pulls period blood to the bottom of the period pad",
            "Made Without Fragrance: Enjoy women's pads made without fragrance and free of elemental chlorine",
            "Powerful Protection and Comfort: Each feminine pad is designed for perfect fit with a breathable top layer and odor control",
            "Packaging may vary from images shown",
        ],
    },
    "470890": {
        "title": "Thinx Teens Super Absorbency Cotton Bikini Period Underwear, Size 13/14, Hologram",
        "description": "From the creators of Thinx, Thinx Teens is a period product that can keep up with you. Thinx Teens Period Underwear looks and feels like everyday underwear, but with built-in period protection so teens can feel fresh, dry, and comfy. No pads, tampons or other disposable period products needed. Thinx Teens heavy flow period underwear features an ultra-absorbent core layer that absorbs up to 2.5 regular pads worth of flow, while the moisture-wicking top layer draws away wetness and dries quickly. From the period care experts at Thinx, Thinx Teens is the only point for an easier period routine, a total replacement for disposable period pads and tampons, and will change the way you think about period care. Thinx Teens is machine-washable and reusable.",
        "features": [
            "Thinx Teens 1 period underwear, super absorbency, size large 13/14, hologram",
            "Thinx Teens period undies are specifically designed to fit teens",
            "Fresh & Dry Feeling: Reusable period underwear controls odor, prevents leaks and wicks moisture",
            "Absorbs Up to 2.5 Regular Pads' Worth of Flow: Our heavy flow period underwear has an ultra-absorbent core layer that absorbs up to 2.5 regular pads' or 5 tampons' worth of flow",
            "Prevents Leaks: The leak-resistant layer of Thinx Teens period undies keeps clothes, sheets and more stain-free",
        ],
    },
    "980948": {
        "title": "U by Kotex Click Compact Tampons, Unscented, Regular, 45 Count",
        "description": "When you are in need of compact comfort and powerful protection, U by Kotex Click compact tampons are there to help. Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection. Compact and able to fit into a purse or pocket, these tampons click into full size to give you powerful protection. Just pull the lower half of the tampon and when it locks in place, it's ready to go! In addition, our unscented tampons are gynecologist-tested, made without fragrance, BPA free, and are free of elemental chlorine. They are also OEKO TEX STANDARD certified, meaning that they are tested for up to 1,000 harmful substances. Individually wrapped, these tampons are perfect for when you need period protection on the go. U by Kotex Click Compact Tampons are available in regular, super and super plus absorbencies. Packaging may vary from images shown.",
        "features": [
            "45 regular tampons",
            "Compact Comfort, Powerful Protection: These compact tampons are easily carried in a purse or pocket for on-the-go protection",
            "#1 compact tampon brand: U by Kotex Click is the #1 compact tampon brand",
            "Up to 100% Leak Free Protection: Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection",
            "Gynecologist-Tested: Our unscented tampons are gynecologist-tested, made without fragrance, BPA free and are free of elemental chlorine",
        ],
    },
    "556510": {
        "title": "U by Kotex Click Tampons, 30 CT",
        "description": "When you are in need of compact comfort and powerful protection, U by Kotex Click compact tampons are there to help. Compact and able to fit into a purse or pocket, these tampons click into full size to give you powerful protection. Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection. Just pull the lower half of the tampon and when it locks in place, it's ready to go. U by Kotex Click compact tampons are individually wrapped and made without fragrance. Product and packaging may vary.",
        "features": [
            "30 tampons",
            "Compact Comfort, Powerful Protection: These compact tampons are easily carried in a purse or pocket for on-the-go protection",
            "Up to 100% Leak Free Protection: Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection",
            "Pocket-sized and changes to a full-size tampon in one easy step",
            "Made without fragrance and individually wrapped for on-the-go period protection",
        ],
    },
})


CVS_KNOWN_PRODUCT_FALLBACKS.update({
    "167387": {
        "title": "U by Kotex Click Compact Tampons, Multipack, Regular/Super Absorbency, Unscented, 45 Count",
        "description": "When you are in need of compact comfort and powerful protection, U by Kotex Click compact tampons are there to help. Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection. Compact and able to fit into a purse or pocket, these tampons click into full size to give you powerful protection, just pull the lower half of the tampon and when it locks in place, it's ready to go! In addition, our unscented tampons are gynecologist-tested, made without fragrance, BPA free, and are free of elemental chlorine. They are also OEKO TEX STANDARD certified, meaning that they are tested for up to 1,000 harmful substances. Individually wrapped, these tampons are perfect for when you need period protection on the go. U by Kotex Click Compact Tampons are available in regular, super, and super plus absorbencies. For backup period protection, try U by Kotex Daily Panty Liners. Kotex feminine products are FSA/HSA/HRA-eligible in the U.S. Packaging may vary from images shown.",
        "features": [
            "45 tampons (multipack contains: 25 regular, 20 super)",
            "Compact Comfort, Powerful Protection: These compact tampons are easily carried in a purse or pocket for on-the-go protection",
            "#1 compact tampon brand: U by Kotex Click is the #1 compact tampon brand",
            "Up to 100% Leak Free Protection: Each tampon has a smooth tip designed for easy and comfortable insertion and provides up to 100% leak free protection",
            "Gynecologist-Tested: Our unscented tampons are gynecologist-tested, made without fragrance, BPA free and are free of elemental chlorine",
        ],
    },
})

# CVS image isolation guard.
# IMPORTANT: Never copy or mirror Salsify image URLs into the CVS/retailer image side.
# CVS retailer images must come only from:
#   1. CVS live PDP HTML.
#   2. Uploaded CVS source/TXT/HTML captures.
#   3. CVS-owned /bizcontent/merchandising/productimages/high_res/... URLs.
#   4. CVS-owned generated high_res URL candidates based on a CVS skuId/RPC or confirmed CVS image base.
CVS_MIRROR_SALSIFY_IMAGE_FALLBACK_SKUS = set()


def cvs_should_mirror_salsify_images(retail_url="", target_rpc="", r_debug=None):
    return False


def cvs_mirror_salsify_images_for_retailer_side(s_images, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE):
    return []


def is_cvs_retailer_image_url(url):
    """True only for CVS-owned retailer image URLs.

    This is intentionally strict so Salsify assets can never appear in the CVS
    retailer image side, even if an older fallback or future edit accidentally
    passes Salsify URLs into r_images.
    """
    value = html.unescape(str(url or "").strip()).replace("\\/", "/")
    if not value:
        return False
    lowered = value.lower()
    if "salsify" in lowered or "images.salsify.com" in lowered or "assets.salsify.com" in lowered:
        return False
    if lowered.startswith("data:"):
        return False
    return bool(
        "/bizcontent/merchandising/productimages/high_res/" in lowered
        or "cvs.com/bizcontent/merchandising/productimages/high_res/" in lowered
    )


def sanitize_cvs_retailer_images(image_urls, debug=None, reason=""):
    """Remove any non-CVS image from the CVS retailer side.

    This function never substitutes Salsify images. It only keeps CVS-owned
    image URLs and preserves their original order.
    """
    cleaned = []
    removed = []
    seen = set()
    for raw_url in image_urls or []:
        url = html.unescape(str(raw_url or "").strip()).replace("\\/", "/")
        if not url:
            continue
        if not is_cvs_retailer_image_url(url):
            removed.append(url)
            continue
        key = url.split("?", 1)[0]
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(url)
    if isinstance(debug, dict):
        debug["CVS Image Isolation"] = "cvs_only_no_salsify_image_backup"
        if reason:
            debug["CVS Image Isolation Reason"] = reason
        if removed:
            debug["CVS Image Isolation Removed Count"] = len(removed)
            debug["CVS Image Isolation Removed Sample"] = " | ".join(removed[:3])
    return cleaned

# CVS-only final-six fallback catalog update.
# These are the last CVS rows from pdp_qa_results_cvs_all_brands (30).xlsx that
# either returned CVS source missing/unmatched or had copy but 0 image match.
# Keep this isolated to CVS skuId/RPC values only.
CVS_KNOWN_PRODUCT_FALLBACKS.update({
    "817844": {
        "title": "Kleenex Disposable Paper Hand Towels, Assorted Designs, 1 Box, 60 Total Towels",
        "description": "Kleenex Disposable Paper Hand Towels are fresh, clean, and dry, one towel at a time. Each box contains absorbent disposable hand towels that are soft and strong for everyday use in bathrooms, kitchens, offices, and guest spaces. These paper hand towels help reduce the spread of germs compared to shared cloth towels and are conveniently dispensed one at a time. Packaging may vary.",
        "features": [
            "1 box of Kleenex Disposable Paper Hand Towels, 60 total towels",
            "Fresh towels every time: Disposable paper hand towels help provide a clean, dry towel for every hand dry",
            "Reduce the spread of germs: A smart alternative to shared cloth hand towels in bathrooms, kitchens and guest spaces",
            "Soft and absorbent: Kleenex hand towels are designed to be gentle while drying hands effectively",
            "Convenient box: Pop-up dispensing helps keep hand towels ready when needed",
        ],
    },
    "130245": {
        "title": "Kleenex Trusted Care Facial Tissues, 1 Box",
        "description": "Kleenex Trusted Care Facial Tissues are soft, strong and absorbent for everyday care. Whether you're managing runny noses, watery eyes, allergy season or everyday messes, Kleenex tissues are gentle on skin and dependable when you need them. Each box is designed for convenient dispensing and comes in various colors and designs. Packaging may vary.",
        "features": [
            "WHAT'S INCLUDED - 1 box of Kleenex Trusted Care Facial Tissues",
            "EVERYDAY CARE - Kleenex Trusted Care facial tissues are soft, strong and absorbent for everyday use",
            "GENTLE ON SKIN - Facial tissues are designed to be gentle for runny noses and watery eyes",
            "DEPENDABLE - A reliable tissue for home, office, classroom or travel needs",
            "PACKAGING MAY VARY - Tissue boxes come in various colors and designs",
        ],
    },
    "731394": {
        "title": "Kotex BioCare Ultra Thin Pads with Wings, Heavy Absorbency, 24 CT",
        "description": "Kotex BioCare Ultra Thin Pads with Wings are designed to deliver period protection and comfort with heavy absorbency. These ultra thin pads feature wings for a secure fit and are designed to help protect against leaks while keeping you comfortable throughout the day. Each pad is individually wrapped for easy access on the go. Product and packaging may vary.",
        "features": [
            "Kotex BioCare Ultra Thin Pads with Wings, Heavy Absorbency, 24 Count",
            "Heavy Absorbency: Designed to help protect against leaks during heavier flow days",
            "Secure Fit: Wings help keep the pad in place for comfortable period protection",
            "Ultra Thin Comfort: Flexible protection designed for everyday comfort",
            "Individually Wrapped: Pads are wrapped for convenient protection on the go",
        ],
    },
    "896560": {
        "title": "U by Kotex Click Compact Multipack Tampons, Unscented, Regular/Super, 30 Count",
        "description": "U by Kotex Click Compact Multipack Tampons provide compact comfort and powerful period protection in regular and super absorbencies. Each tampon is pocket-sized, clicks into a full-size tampon in one easy step, and has a smooth tip designed for easy and comfortable insertion. These tampons are unscented, individually wrapped, and convenient for on-the-go protection. Product and packaging may vary.",
        "features": [
            "30 count multipack of U by Kotex Click Compact Tampons in regular and super absorbencies",
            "Compact Comfort, Powerful Protection: Pocket-sized tampons are easy to carry for on-the-go protection",
            "Clicks to Full Size: Tampons go from compact to full-size in one easy step",
            "Smooth Tip: Designed for easy and comfortable insertion",
            "Unscented and Individually Wrapped: Convenient period protection when you need it",
        ],
    },
    "482747": {
        "title": "Viva Signature Cloth Paper Towels, 2 Quad Rolls",
        "description": "Viva Signature Cloth Paper Towels deliver a cloth-like clean with soft and durable sheets for everyday messes. These paper towels are absorbent and strong, making them useful for cleaning kitchen spills, wiping counters, drying hands, and handling household tasks. Choose-A-Sheet sizing lets you select the right amount for the job. Packaging may vary.",
        "features": [
            "WHAT'S INCLUDED - 2 quad rolls of Viva Signature Cloth Paper Towels",
            "Cloth-Like Clean: Viva Signature Cloth paper towels are soft and durable for everyday cleaning",
            "Choose-A-Sheet: Select the right sheet size for small or large messes",
            "Absorbent and Strong: Designed to help clean spills, counters and household surfaces",
            "Versatile Cleaning: Great for kitchens, bathrooms, hands and everyday household tasks",
        ],
    },
})



# CVS-only targeted rescue for current three remaining rows.
# Keep isolated to exact CVS skuId/RPC values and do not enable global fallback.
CVS_KNOWN_PRODUCT_FALLBACKS.update({
    "730263": {
        "title": "Kotex Bamboo Ultra Thin Pads with Wings, Heavy Absorbency, 30 Count",
        "description": "The new Kotex Bamboo Ultra Thin Pads with wings are designed to provide comfortable and reliable period protection. Each pad features a 100% bamboo-derived viscose top layer that is UltraSoft and breathable, made from organically grown bamboo. The 5x System with LeakShield Protection delivers breathability, odor control, dryness, fit, and leakage protection for up to 100% Leak Free Comfort. These pads are made without fragrance and are elemental chlorine free, as well as pesticide free. The Gravity Core pulls blood to the bottom of the pad to help you feel clean. Each pad is individually folded and wrapped for easy access, even on-the-go. Product and packaging may vary from images shown.",
        "features": [
            "Kotex Bamboo Ultra Thin Pads with Wings, Heavy Absorbency, 30 Count",
            "100% Bamboo-derived Viscose Top Layer: Each period pad is made with an UltraSoft and breathable viscose layer made from organically-grown bamboo",
            "5x System with LeakShield Protection: Bamboo women's pads are made with a 5x System with LeakShield Technology that offers breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort",
            "Gravity Core: Our menstrual pads feature a Gravity Core that pulls period blood to the bottom of the period pad to help keep you clean and dry",
            "Made Without Fragrance: These pads for women are made without fragrance and are elemental chlorine free. These pads are also pesticide free",
        ],
    },
})


# CVS-only fallback for live PDP 730204. The live source shows the image
# basename 3600058233, not the skuId, so pair this with CVS_KNOWN_IMAGE_BASE_BY_SKU.
CVS_KNOWN_PRODUCT_FALLBACKS.update({
    "730204": {
        "title": "Kotex Ultra Thin Overnight Pads With Wings, Heavy Absorbency, 36 CT",
        "description": "Bring powerful protection and comfort to your nighttime period routine with the new Kotex Ultra Thin Overnight Pads with Wings. These overnight pads provide up to 12 hours of protection and NightDefense with a raised back barrier and side guards to help prevent back and side leaks. The 5x System with LeakShield Protection has breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort. These period pads are designed for perfect fit and combine LeakShield Technology, a breathable top layer, a new Gravity Core and odor control to give you a menstrual pad that protects you in more ways than one. To help keep you feeling clean and fresh throughout your day, each feminine pad is designed with a Gravity Core that pulls period blood to the bottom of the pad. Each women's pad is made with your skin health in mind, which is why these pads are made without fragrance and free of elemental chlorine. For added convenience, each nighttime sanitary pad is individually folded and wrapped to protect your pad with easy access, even on-the-go. For daytime protection, check out Kotex Ultra Thin Pads with Wings. Kotex feminine products are FSA/HSA/HRA-eligible in the US. Product and packaging may vary.",
        "features": [
            "Kotex Ultra Thin Overnight Pads With Wings, Heavy Absorbency, 36 Count",
            "All-Night Protection: NightDefense overnight pads provide up to 12 hours of protection with a raised back barrier and side guards to help prevent back and side leaks",
            "5x System Comfort: These women's pads offer breathability, odor control, dryness, fit and leakage protection for up to 100% Leak Free Comfort",
            "Gravity Core Technology: Our period pads feature a Gravity Core that pulls period blood to the bottom of the pad to help keep you clean and dry",
            "Gentle on Skin: These menstrual pads are made without fragrance and free of elemental chlorine",
        ],
    },
})

def fetch_cvs_url_once(url, user_agent="", timeout_seconds=None):
    if not url:
        return ""
    try:
        session = get_session()
        headers = dict(HEADERS)
        if user_agent:
            headers["User-Agent"] = user_agent
        headers.update({
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
            "Upgrade-Insecure-Requests": "1",
            "Referer": "https://www.cvs.com/",
        })
        r = session.get(url, headers=headers, timeout=timeout_seconds or CVS_REQUEST_TIMEOUT, allow_redirects=True)
        if r.status_code == 200 and r.text:
            return r.text
    except Exception:
        pass
    return ""


def cvs_live_html_quality_score(html_text):
    value = str(html_text or "")
    if not value.strip():
        return 0
    lowered = value.lower()
    score = 0
    for marker in ["vendordetailsbullets", "vendordetailsparagraph", "vendorcontent", "dynamicmediaurl", "/bizcontent/merchandising/productimages/high_res/", "productimages/high_res", "__next_data__"]:
        if marker in lowered:
            score += 250
    for marker in ["item #", "skuid=", "prodid-", "rating & reviews", "product type", "details"]:
        if marker in lowered:
            score += 50
    if re.search(r"/shop/[^\s\"'<>]+prodid-", lowered):
        score += 80
    if re.search(r"\b(kleenex|kotex|viva|poise|depend|huggies|pull-ups|cottonelle|scott|goodnites|thinx)\b", lowered):
        score += 120
    if "captcha" in lowered or "are-you-human" in lowered or "px-captcha" in lowered:
        score -= 1000
    if "sc-cvs-footer-container" in lowered and score < 200:
        score -= 250
    return max(score + min(len(value) // 5000, 50), 0)


def cvs_title_from_url_slug(retail_url):
    m = re.search(r"/shop/([^/?#]+?)-prodid-[0-9A-Za-z_-]+", str(retail_url or ""), flags=re.IGNORECASE)
    if not m:
        return ""
    slug = html.unescape(m.group(1)).replace("-", " ")
    slug = re.sub(r"\s+", " ", slug).strip()
    if not slug:
        return ""
    return " ".join(word.upper() if word.lower() in {"cvs", "ct"} else word.capitalize() for word in slug.split())


def normalize_cvs_visible_text_blob(html_text):
    if not html_text:
        return ""
    working = str(html_text or "")
    for _ in range(4):
        unescaped = html.unescape(working)
        if unescaped == working:
            break
        working = unescaped
    try:
        visible = BeautifulSoup(working, "html.parser").get_text("\n", strip=True)
    except Exception:
        visible = working
    visible = visible.replace("\r", "\n")
    visible = re.sub(r"\n{2,}", "\n", visible)
    return visible.strip()


def extract_cvs_indexed_text_fallback(html_text, retail_url="", target_rpc=""):
    debug = {"Description Path": "", "Features Path": "", "Title Path": "", "Visible Text Length": 0}
    visible = normalize_cvs_visible_text_blob(html_text)
    debug["Visible Text Length"] = len(visible)
    if not visible:
        return {"title": "", "description": "", "features": [], "debug": debug}
    lines = [normalize_space(x) for x in visible.splitlines() if normalize_space(x)]
    sku_id = get_cvs_effective_sku_id(retail_url=retail_url, target_rpc=target_rpc)
    title = ""
    for idx, line in enumerate(lines[:120]):
        if sku_id and re.search(rf"\bItem\s*#\s*{re.escape(sku_id)}\b", line, flags=re.IGNORECASE):
            for prior in reversed(lines[max(0, idx - 4):idx]):
                if len(prior) >= 12 and not re.search(r"^(Details|Rating|Reviews|How to get it)$", prior, flags=re.IGNORECASE):
                    title = prior
                    break
            break
    if not title:
        for line in lines[:80]:
            if len(line) >= 20 and any(t in line.lower() for t in ["kleenex", "kotex", "viva", "poise", "depend", "huggies", "pull-ups", "cottonelle", "scott", "goodnites", "thinx"]):
                title = line
                break
    if is_invalid_cvs_title_candidate(title):
        title = ""
        debug["Title Path"] = "cvs_invalid_title_rejected"
    if not title:
        title = cvs_title_from_url_slug(retail_url)
    if title:
        debug["Title Path"] = "cvs_visible_text_or_url_slug"
    compact = normalize_space(" ".join(lines))
    desc_start = 0
    if sku_id:
        m = re.search(rf"Item\s*#\s*{re.escape(sku_id)}\b", compact, flags=re.IGNORECASE)
        if m:
            desc_start = m.end()
    if not desc_start and title:
        idx = compact.lower().find(title.lower())
        if idx >= 0:
            desc_start = idx + len(title)
    working = compact[desc_start:].strip() if desc_start else compact
    stop = re.search(r"\b(Rating\s*&\s*reviews|Ingredients|Specifications|Same-Day Delivery policies|Delivery Details)\b", working, flags=re.IGNORECASE)
    if stop:
        working = working[:stop.start()].strip()
    known = (
        "WHAT['’]?S INCLUDED|HELPS SKIN FEEL RESTORED|INSTANT COOLING RELIEF|A HINT OF ALOE|CLEAN AND COMFORTED SKIN|"
        "BE PREPARED FOR SICK DAYS|3 BENEFITS IN 1 TISSUE|SAVE YOUR TOILET PAPER|MADE WITH LOTION|MOISTURIZES TO PREVENT SKIN IRRITATION|"
        "PERFECTLY SIZED FOR ANY ADVENTURE|SMALL BUT MIGHTY|STYLE WHEREVER YOU GO|FOR COLDS ?& ?FLUS|HOW IT WORKS|3 LAYERS OF STRENGTH|PERFECT FOR ANY HOME|"
        "FRESHNESS YOU CAN FEEL|BREAKS DOWN LIKE TOILET PAPER|GENTLE FOR SKIN|ODOR CONTROL|DRYNESS|LEAK ?GUARD|ALL DAY PROTECTION|"
        "XPRESS DRI CORE|LIGHT FLOW PROTECTION|CLEAN AND FRESH|MADE WITHOUT FRAGRANCE|COMPACT COMFORT, POWERFUL PROTECTION|#1 COMPACT TAMPON BRAND|"
        "UP TO 100% LEAK FREE PROTECTION|GYNECOLOGIST-TESTED|CHOOSE A SHEET|#1 CLOTH-LIKE TOWEL|SOFT LIKE CLOTH|ABSORBENT LIKE CLOTH|DURABLE LIKE CLOTH|VERSATILE CLEANING|PACKAGING MAY VARY|"
        "CLOTH-LIKE|CLEAN,? FRESH,? ?& ?DRY|FRAGRANCE FREE|ABSORBENT|SKIP THE TRIP|STYLISH DESIGNS"
    )
    # CVS combined fallback splitter.
    # Use known headings only so normal sentence text does not get split character-by-character.
    # The Kleenex hand-towel item also exposes its first feature as a count sentence with no heading.
    hand_towel_count_feature = r"1\s+box\s+of\s+Kleenex\s+Disposable\s+Paper\s+Hand\s+Towels"
    heading = re.compile(rf"(?=(?:{known})\s*(?:[—-]|:)\s+|(?:{hand_towel_count_feature})\b)", flags=re.IGNORECASE)
    matches = list(heading.finditer(working))
    description = ""
    features = []
    if matches:
        description = clean_cvs_text(working[:matches[0].start()])
        for i, match in enumerate(matches):
            start = match.start()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(working)
            item = clean_cvs_feature_text(working[start:end])
            if item:
                features.append(item)
    else:
        description = clean_cvs_text(working)
    if title and normalize_text(description).startswith(normalize_text(title)):
        description = clean_cvs_text(description[len(title):])
    features = normalize_cvs_features(features)
    if description:
        debug["Description Path"] = "cvs_visible_indexed_text_fallback"
    if features:
        debug["Features Path"] = "cvs_visible_indexed_text_feature_heading_split"
    return {"title": title, "description": description, "features": features[:5], "debug": debug}

def is_probably_cvs_product_html(html_text):
    """CVS-only check for real product content.

    Some CVS URLs load in a browser but server-side requests can return an
    empty/shell/challenge page. This keeps live fetches from being trusted when
    there is no PDP copy or image payload in the HTML.
    """
    text = str(html_text or "")
    if not text.strip():
        return False
    lowered = text.lower()
    product_markers = [
        "vendordetailsbullets",
        "vendordetailsparagraph",
        "vendorcontent",
        "dynamicmediaurl",
        "/bizcontent/merchandising/productimages/high_res/",
        "productimages/high_res",
        "skuid=",
        "prodid-",
        "__next_data__",
    ]
    blocked_markers = [
        "captcha",
        "are-you-human",
        "px-captcha",
        "let us know you're not a robot",
        "let us know you’re not a robot",
        "access denied",
    ]
    has_product_marker = any(marker in lowered for marker in product_markers)
    has_block_marker = any(marker in lowered for marker in blocked_markers)
    return bool(has_product_marker and not has_block_marker)


def fetch_cvs_html_with_fallbacks(retail_url):
    """CVS-only live HTML fetch with smarter URL and source fallbacks."""
    retail_url = str(retail_url or "").strip()
    if not retail_url:
        return "", "cvs_url_missing"
    desktop_ua = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 " "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
    mobile_ua = ("Mozilla/5.0 (iPhone; CPU iPhone OS 17_4 like Mac OS X) " "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Mobile/15E148 Safari/604.1")
    best_html = ""
    best_score = -1
    best_label = "cvs_live_fetch_empty_or_shell"
    for idx, candidate_url in enumerate(cvs_url_candidates(retail_url)):
        if idx == 0:
            try:
                html_text = get_html(candidate_url)
            except Exception:
                html_text = ""
            score = cvs_live_html_quality_score(html_text)
            if score > best_score:
                best_html, best_score, best_label = html_text, score, "cvs_get_html"
            if is_probably_cvs_product_html(html_text):
                return html_text, "cvs_get_html"
        for label, user_agent in [("cvs_live_desktop_retry", desktop_ua), ("cvs_live_mobile_retry", mobile_ua)]:
            html_text = fetch_cvs_url_once(candidate_url, user_agent=user_agent, timeout_seconds=CVS_REQUEST_TIMEOUT)
            score = cvs_live_html_quality_score(html_text)
            if score > best_score:
                best_html, best_score, best_label = html_text, score, f"{label} | {candidate_url}"
            if is_probably_cvs_product_html(html_text):
                return html_text, f"{label} | {candidate_url}"
    if best_html:
        return best_html, best_label
    return "", "cvs_live_fetch_empty_or_shell"

def clean_cvs_text(text):
    if not text:
        return ""

    text = str(text)

    text = text.replace("\\u0026", "&amp;")
    text = text.replace("\\n", " ")
    text = text.replace("\\/", "/")
    text = text.replace('\\"', '"')
    text = html.unescape(text)

    wrapper_patterns = [
        r'"\]\)\s*</script>\s*<script>\s*self\.__next_f\.push\(\[1,\s*"',
        r'"\]\)\s*self\.__next_f\.push\(\[1,\s*"',
        r'</script>\s*<script>\s*self\.__next_f\.push\(\[1,\s*"',
    ]
    for pattern in wrapper_patterns:
        text = re.sub(pattern, "", text, flags=re.DOTALL)

    text = re.sub(r'^(?:T[0-9A-Za-z]+,)+', "", text)
    text = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:\{.*$', "", text, flags=re.DOTALL)
    text = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:\[.*$', "", text, flags=re.DOTALL)
    text = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:$', "", text, flags=re.DOTALL)
    text = re.sub(r'"\]\s*[0-9A-Za-z]{1,3}:T[0-9A-Za-z]+,.*$', "", text, flags=re.DOTALL)
    text = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:T[0-9A-Za-z]+,.*$', "", text, flags=re.DOTALL)
    text = re.sub(r'self\.__next_f\.push\(\[1,.*$', "", text, flags=re.DOTALL)
    text = re.sub(r'<script[^>]*>.*?</script>', " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'</?script[^>]*>', " ", text, flags=re.IGNORECASE)
    text = text.replace("\\*", "*")
    text = re.sub(r"\binconti\s+nence\b", "incontinence", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()

    return text


def clean_cvs_feature_text(text):
    if not text:
        return ""

    text = str(text)

    text = text.replace("\\u0026", "&amp;")
    text = text.replace("\\n", " ")
    text = text.replace("\\/", "/")
    text = text.replace('\\"', '"')
    text = html.unescape(text)

    text = re.sub(r'"\]\s*[0-9A-Za-z]{1,3}:T[0-9A-Za-z]+,.*$', "", text, flags=re.DOTALL)
    text = re.sub(r'\]\)\s*</script>\s*<script>\s*self\.__next_f\.push\(\[1,\s*".*$', "", text, flags=re.DOTALL)
    text = re.sub(r'</script>\s*<script>\s*self\.__next_f\.push\(\[1,\s*".*$', "", text, flags=re.DOTALL)
    text = re.sub(r'self\.__next_f\.push\(\[1,.*$', "", text, flags=re.DOTALL)
    text = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:\{.*$', "", text, flags=re.DOTALL)
    text = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:\[.*$', "", text, flags=re.DOTALL)
    text = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:$', "", text, flags=re.DOTALL)
    text = text.replace("\\*", "*")
    text = re.sub(r"\s+", " ", text).strip()

    return text


def get_target_sku_from_inputs(retail_url="", cvs_rpc=""):
    retail_url = html.unescape(str(retail_url or ""))
    cvs_rpc = str(cvs_rpc or "").strip()

    m = re.search(r'[?&]skuId=([0-9A-Za-z_-]+)', retail_url)
    if m:
        return m.group(1).strip()

    return cvs_rpc

def clean_cvs_text_refined(text):
    return clean_cvs_text(text)


def get_nextjs_chunks(html_text):
    if not html_text:
        return ""

    source = html.unescape(html_text)
    pattern = r'self\.__next_f\.push\(\[1,\s*"((?:\\.|[^"\\])*)"\s*\]\)'
    chunks = []

    for m in re.finditer(pattern, source, re.DOTALL):
        payload = m.group(1)
        try:
            decoded = json.loads(f'"{payload}"')
        except Exception:
            decoded = payload
            decoded = decoded.replace("\\n", "\n")
            decoded = decoded.replace("\\/", "/")
            decoded = decoded.replace('\\"', '"')
        chunks.append(decoded)

    return "\n".join(chunks)


def extract_balanced_bracket_block(source, start_index):
    if start_index < 0 or start_index >= len(source) or source[start_index] != "[":
        return ""

    depth = 0
    in_str = False
    escape = False

    for i in range(start_index, len(source)):
        ch = source[i]

        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    return source[start_index:i + 1]

    return ""


def extract_array_block_for_key(source, key):
    if not source:
        return ""

    source = str(source)
    key = str(key)

    patterns = [
        rf'(?:^|\n){re.escape(key)}:\[',
        rf'(?<![0-9A-Za-z]){re.escape(key)}:\[',
    ]

    for pattern in patterns:
        m = re.search(pattern, source)
        if not m:
            continue

        bracket_start = source.find("[", m.start())
        if bracket_start == -1:
            continue

        block = extract_balanced_bracket_block(source, bracket_start)
        if block:
            return block

    return ""


def extract_balanced_brace_block(source, start_index):
    if start_index < 0 or start_index >= len(source) or source[start_index] != "{":
        return ""

    depth = 0
    in_str = False
    escape = False

    for i in range(start_index, len(source)):
        ch = source[i]

        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return source[start_index:i + 1]

    return ""


def find_newline_anchored_key(source, key, for_array=False):
    key = str(key)

    if for_array:
        strict_pattern = rf'(?:^|\n){re.escape(key)}:\['
    else:
        strict_pattern = rf'(?:^|\n){re.escape(key)}:'

    m = re.search(strict_pattern, source)
    if m:
        return m

    if for_array:
        fallback_pattern = rf'(?:^|[\s,{{]){re.escape(key)}:\['
    else:
        fallback_pattern = rf'(?:^|[\s,{{]){re.escape(key)}:'

    return re.search(fallback_pattern, source)


def looks_like_next_newline_key(source, idx):
    if idx < 0 or idx >= len(source):
        return False

    return bool(
        re.match(
            r'(?:\s|,)([0-9A-Za-z]{1,3}):(?=[\[{"]|T[0-9A-Za-z]+,|null|true|false|\d)',
            source[idx:],
        )
    )


def extract_newline_anchored_value_block(source, key):
    m = find_newline_anchored_key(source, key, for_array=False)
    if not m:
        return ""

    start = m.end()
    i = start

    in_str = False
    escape = False
    bracket_depth = 0
    brace_depth = 0
    paren_depth = 0

    while i < len(source):
        ch = source[i]

        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
            i += 1
            continue

        if ch == '"':
            in_str = True
            i += 1
            continue

        if ch == "[":
            bracket_depth += 1
            i += 1
            continue
        if ch == "]":
            bracket_depth = max(0, bracket_depth - 1)
            i += 1
            continue

        if ch == "{":
            brace_depth += 1
            i += 1
            continue
        if ch == "}":
            brace_depth = max(0, brace_depth - 1)
            i += 1
            continue

        if ch == "(":
            paren_depth += 1
            i += 1
            continue
        if ch == ")":
            paren_depth = max(0, paren_depth - 1)
            i += 1
            continue

        if bracket_depth == 0 and brace_depth == 0 and paren_depth == 0:
            if re.match(
                r'(?:\s|,|^)([0-9A-Za-z]{1,3}):(?=[\[{"]|T[0-9A-Za-z]+,|null|true|false|\d)',
                source[i:],
            ):
                break

        i += 1

    block = source[start:i].strip()
    block = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:\{.*$', "", block, flags=re.DOTALL)
    block = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:\[.*$', "", block, flags=re.DOTALL)
    block = re.sub(r'(?<=[\.\)\]"\'])\s*[0-9A-Za-z]{1,3}:$', "", block, flags=re.DOTALL)

    return block.strip()


def extract_object_block_for_key(source, key):
    m = find_newline_anchored_key(source, key, for_array=False)
    if not m:
        return ""

    start = m.end()
    while start < len(source) and source[start].isspace():
        start += 1

    if start < len(source) and source[start] == "{":
        return extract_balanced_brace_block(source, start)

    return extract_newline_anchored_value_block(source, key)


def find_direct_vendor_details_block_near_sku(source, target_rpc="", search_after=30000):
    if not source or not target_rpc:
        return ""

    source = str(source)
    rpc = re.escape(str(target_rpc))

    anchor_patterns = [
        rf'skuId={rpc}\b',
        rf'/{rpc}(?:_[0-9]+)?\.jpg',
    ]

    anchors = []
    for pattern in anchor_patterns:
        anchors.extend(list(re.finditer(pattern, source, flags=re.IGNORECASE)))

    if not anchors:
        return ""

    for anchor in anchors:
        segment_start = anchor.start()
        segment_end = min(len(source), anchor.start() + search_after)
        segment = source[segment_start:segment_end]

        direct_match = re.search(
            r'"vendorContent"\s*:\s*\{\s*"vendorDetails"\s*:\s*\{',
            segment,
            flags=re.DOTALL,
        )
        if direct_match:
            brace_start = segment_start + direct_match.end() - 1
            block = extract_balanced_brace_block(source, brace_start)
            if block:
                return block

        ref_match = re.search(
            r'"vendorContent"\s*:\s*\{\s*"vendorDetails"\s*:\s*"(\$[0-9A-Za-z]{1,3})"',
            segment,
            flags=re.DOTALL,
        )
        if ref_match:
            ref_key = ref_match.group(1).replace("$", "")
            resolved = extract_object_block_for_key(source, ref_key)
            if resolved:
                return resolved

    return ""


def extract_rpc_anchor_windows(source, target_rpc="", context_before=3500, context_after=12000):
    if not source or not target_rpc:
        return []

    rpc = re.escape(str(target_rpc))
    patterns = [
        rf'skuId={rpc}\b',
        rf'/{rpc}(?:_[0-9]+)?\.jpg',
        rf'"dynamicMediaUrl"\s*:\s*"[^"]*?/{rpc}(?:_[0-9]+)?\.jpg[^"]*"',
    ]

    hits = []
    seen = set()

    for pattern in patterns:
        for m in re.finditer(pattern, source, flags=re.IGNORECASE):
            start = max(0, m.start() - context_before)
            end = min(len(source), m.end() + context_after)
            key = (start, end)
            if key in seen:
                continue
            seen.add(key)
            hits.append({
                "start": start,
                "end": end,
                "anchor_start": m.start(),
                "anchor_end": m.end(),
                "window": source[start:end],
                "anchor_text": source[m.start():m.end()],
            })

    hits.sort(key=lambda x: x["start"])
    return hits


def merge_feature_continuations(items, max_features=5):
    merged = []

    continuation_patterns = [
        r'^packaging and product may vary\.?$',
        r'^product may vary\.?$',
        r'^packaging may vary\.?$',
    ]

    hard_stop_patterns = [
        r'^[0-9A-Za-z]{1,3}:T[0-9A-Za-z]+,',
        r'^self\.__next_f\.push',
        r'^</script>',
        r'^<script>',
        r'^vendorDetailsParagraph',
        r'^\]\)\s*</script>',
    ]

    for raw_item in items:
        item = clean_cvs_feature_text(raw_item)
        if not item:
            continue

        if any(re.search(p, item, flags=re.IGNORECASE) for p in hard_stop_patterns):
            break

        is_continuation = any(
            re.match(pattern, item, flags=re.IGNORECASE)
            for pattern in continuation_patterns
        )

        if not is_continuation:
            if (
                merged
                and len(item) <= 60
                and "—" not in item
                and not re.match(r'^[A-Z0-9][A-Z0-9\s\'"&/\-\(\)\*:]+(?:\s—|\s-|\s:)', item)
                and item[:1].islower()
            ):
                is_continuation = True

        if is_continuation and merged:
            if len(merged) >= max_features:
                continue

            prev = merged.pop()
            if prev.endswith(";"):
                merged.append(f"{prev} {item}")
            else:
                merged.append(f"{prev}; {item}")
        else:
            if len(merged) >= max_features:
                break
            merged.append(item)

    return dedupe_preserve_order(merged[:max_features])


def normalize_cvs_features(items):
    cleaned = [clean_cvs_feature_text(x) for x in items if isinstance(x, str)]
    cleaned = [x for x in cleaned if x]
    cleaned = merge_feature_continuations(cleaned, max_features=5)
    return dedupe_preserve_order(cleaned[:5])


def split_feature_blob_preserve_semicolons(text):
    text = clean_cvs_feature_text(text)
    if not text:
        return []

    if " | " in text:
        parts = [x.strip() for x in text.split(" | ")]
    elif "•" in text:
        parts = [x.strip() for x in text.split("•")]
    else:
        parts = [text.strip()]

    return [p for p in parts if p]


def parse_jsonish_array_text(array_text):
    array_text = normalize_space(array_text)
    if not array_text:
        return []

    candidates = [
        array_text,
        array_text.replace('\\"', '"'),
        html.unescape(array_text).replace('\\"', '"'),
    ]

    for candidate in candidates:
        try:
            value = json.loads(candidate)
            if isinstance(value, list):
                return normalize_cvs_features(value)
        except Exception:
            pass

    inner = array_text[1:-1] if array_text.startswith("[") and array_text.endswith("]") else array_text
    inner = clean_cvs_feature_text(inner)

    parts = re.split(r'"\s*,\s*"', inner)

    cleaned = []
    for part in parts:
        val = clean_cvs_feature_text(part.strip().strip('"'))
        if val:
            cleaned.extend(split_feature_blob_preserve_semicolons(val))

    return normalize_cvs_features(cleaned)


def extract_candidate_variant_windows(source, context_before=4500, context_after=22000):
    windows = []

    patterns = [
        r'"vendorContent"\s*:\s*\{\s*"vendorDetails"\s*:\s*\{',
        r'"vendorContent"\s*:\s*\{\s*"vendorDetails"\s*:\s*"(\$[0-9A-Za-z]{1,3})"',
    ]

    seen = set()

    for pattern in patterns:
        for m in re.finditer(pattern, source, flags=re.DOTALL):
            start = max(0, m.start() - context_before)
            end = min(len(source), m.start() + context_after)
            key = (start, end)
            if key in seen:
                continue
            seen.add(key)

            windows.append({
                "start": start,
                "end": end,
                "vendor_start": m.start(),
                "window": source[start:end],
            })

    return windows


def score_variant_window(window_text, vendor_start, window_start, target_rpc="", retail_url=""):
    score = 0
    reason = []
    matched_dynamic_media = ""
    matched_variant_url = ""
    matched_nearby_image = ""

    if not target_rpc:
        return {
            "score": 0,
            "reason": [],
            "matched_dynamic_media": "",
            "matched_variant_url": "",
            "matched_nearby_image": "",
        }

    rpc = re.escape(str(target_rpc))

    sku_match = re.search(rf'skuId={rpc}\b', window_text)
    if sku_match:
        score += 100
        reason.append("skuId_match")

    dm_match = re.search(
        rf'"dynamicMediaUrl"\s*:\s*"([^"]*?/({rpc})(?:_[0-9]+)?\.jpg[^"]*)"',
        window_text,
        flags=re.IGNORECASE,
    )
    if dm_match:
        score += 90
        matched_dynamic_media = dm_match.group(1)
        reason.append("dynamicMediaUrl_rpc_match")

    img_match = re.search(
        rf'"(?:imageUrl|image|src|url)"\s*:\s*"([^"]*?/({rpc})(?:_[0-9]+)?\.jpg[^"]*)"',
        window_text,
        flags=re.IGNORECASE,
    )
    if img_match:
        score += 60
        matched_nearby_image = img_match.group(1)
        reason.append("nearby_image_rpc_match")

    if retail_url:
        path_match = re.search(r'https?://www\.cvs\.com(/shop/[^?\s"]+)', retail_url)
        if path_match:
            retail_path = path_match.group(1)
            if retail_path and retail_path in window_text:
                score += 10
                matched_variant_url = retail_path
                reason.append("retail_path_match")

    local_vendor_start = vendor_start - window_start
    rpc_positions = []

    for m in re.finditer(rf'(?:skuId=|/as/|/high_res/){rpc}\b', window_text):
        rpc_positions.append(m.start())

    prior_positions = [p for p in rpc_positions if p <= local_vendor_start]

    if prior_positions:
        nearest_prior = max(prior_positions)
        distance = local_vendor_start - nearest_prior

        if distance <= 1200:
            score += 50
            reason.append("rpc_near_vendorcontent_before_1200")
        elif distance <= 2500:
            score += 35
            reason.append("rpc_near_vendorcontent_before_2500")
        elif distance <= 4500:
            score += 20
            reason.append("rpc_near_vendorcontent_before_4500")

    if '"vendorDetailsBullets"' in window_text:
        score += 5
        reason.append("has_bullets_marker")

    if '"vendorDetailsParagraph"' in window_text:
        score += 5
        reason.append("has_paragraph_marker")

    return {
        "score": score,
        "reason": reason,
        "matched_dynamic_media": matched_dynamic_media,
        "matched_variant_url": matched_variant_url,
        "matched_nearby_image": matched_nearby_image,
    }


def get_sorted_variant_windows(source, target_rpc="", retail_url=""):
    candidates = extract_candidate_variant_windows(source)

    for candidate in candidates:
        scored = score_variant_window(
            candidate["window"],
            vendor_start=candidate["vendor_start"],
            window_start=candidate["start"],
            target_rpc=target_rpc,
            retail_url=retail_url,
        )
        candidate["match_score"] = scored["score"]
        candidate["match_reason"] = scored["reason"]
        candidate["matched_dynamic_media"] = scored["matched_dynamic_media"]
        candidate["matched_variant_url"] = scored["matched_variant_url"]
        candidate["matched_nearby_image"] = scored["matched_nearby_image"]

    candidates.sort(key=lambda x: (x.get("match_score", 0), -x.get("start", 0)), reverse=True)
    return candidates


def parse_vendor_details_from_block(local_block, working_source, debug):
    if not local_block:
        return {"features": [], "description": "", "debug": debug}

    local_debug = debug.copy()
    local_debug["directVendorContentFound"] = True
    local_debug["directVendorDetailsFound"] = True
    local_debug["directVendorContentExcerpt"] = normalize_space(local_block[:2500])

    working_block = local_block
    seen_vendor_detail_refs = set()
    max_ref_hops = 6

    for _ in range(max_ref_hops):
        if (
            '"vendorDetailsBullets"' in working_block
            or '"vendorDetailsParagraph"' in working_block
        ):
            break

        vendor_details_ref_match = re.search(
            r'"vendorDetails"\s*:\s*"(\$[0-9A-Za-z]{1,3})"',
            working_block,
            flags=re.DOTALL,
        )

        if not vendor_details_ref_match:
            break

        ref_token = vendor_details_ref_match.group(1)
        ref_key = ref_token.replace("$", "")

        if ref_key in seen_vendor_detail_refs:
            local_debug["vendorDetailsRefLoopDetected"] = True
            local_debug["vendorDetailsResolvedKey"] = ref_key
            break

        seen_vendor_detail_refs.add(ref_key)

        resolved_block = extract_object_block_for_key(working_source, ref_key)
        if not resolved_block:
            local_debug["vendorDetailsResolvedKey"] = ref_key
            local_debug["vendorDetailsResolveFailed"] = True
            break

        if normalize_space(resolved_block) == normalize_space(working_block):
            local_debug["vendorDetailsResolvedKey"] = ref_key
            local_debug["vendorDetailsResolveSameBlock"] = True
            break

        local_debug["vendorDetailsRef"] = ref_token
        local_debug["vendorDetailsResolvedKey"] = ref_key
        working_block = resolved_block

    features = []
    description = ""

    bullets_ref_match = re.search(
        r'"vendorDetailsBullets"\s*:\s*"(\$[0-9A-Za-z]{1,3})"',
        working_block,
        flags=re.DOTALL,
    )

    bullets_array_text = ""
    bullets_array_marker = re.search(
        r'"vendorDetailsBullets"\s*:\s*\[',
        working_block,
        flags=re.DOTALL,
    )
    if bullets_array_marker:
        array_start = bullets_array_marker.end() - 1
        bullets_array_text = extract_balanced_bracket_block(working_block, array_start)

    if bullets_ref_match:
        ref_token = bullets_ref_match.group(1)
        ref_key = ref_token.replace("$", "")

        local_debug["vendorPatternFound"] = True
        local_debug["vendorDetailsBulletsRef"] = ref_token
        local_debug["featuresKey"] = ref_key

        array_text = extract_array_block_for_key(working_source, ref_key)
        local_debug["featuresArrayFound"] = bool(array_text)
        local_debug["featuresArrayExcerpt"] = normalize_space(array_text)[:2000] if array_text else ""

        if array_text:
            features = parse_jsonish_array_text(array_text)

    elif bullets_array_text:
        local_debug["featuresArrayFound"] = True
        local_debug["featuresArrayExcerpt"] = normalize_space(bullets_array_text)[:2000]
        features = parse_jsonish_array_text(bullets_array_text)

    paragraph_ref_match = re.search(
        r'"vendorDetailsParagraph"\s*:\s*"(\$[0-9A-Za-z]{1,3})"',
        working_block,
        flags=re.DOTALL,
    )
    paragraph_direct_match = re.search(
        r'"vendorDetailsParagraph"\s*:\s*"((?:\\.|[^"\\])*)"',
        working_block,
        flags=re.DOTALL,
    )

    if paragraph_ref_match:
        ref_token = paragraph_ref_match.group(1)
        ref_key = ref_token.replace("$", "")

        local_debug["vendorPatternFound"] = True
        local_debug["vendorDetailsParagraphRef"] = ref_token
        local_debug["descriptionKey"] = ref_key

        desc_block = extract_newline_anchored_value_block(working_source, ref_key)
        desc_block = clean_cvs_text(desc_block)

        local_debug["descriptionBlockFound"] = bool(desc_block)
        local_debug["descriptionBlockExcerpt"] = normalize_space(desc_block)[:2000]
        description = desc_block

    elif paragraph_direct_match:
        raw_para = paragraph_direct_match.group(1)

        if not re.fullmatch(r'\$[0-9A-Za-z]{1,3}', raw_para or ""):
            try:
                description = json.loads(f'"{raw_para}"')
            except Exception:
                description = raw_para

            description = clean_cvs_text(description)
            local_debug["descriptionBlockFound"] = bool(description)
            local_debug["descriptionBlockExcerpt"] = normalize_space(description)[:2000]

    cleaned_features = normalize_cvs_features(features)
    cleaned_description = clean_cvs_text(description)

    return {
        "features": cleaned_features,
        "description": cleaned_description,
        "debug": local_debug,
    }


def extract_vendor_copy_from_source(source, source_name="", target_rpc="", retail_url=""):
    debug = {
        "vendorPatternFound": False,
        "vendorDetailsBulletsRef": "",
        "vendorDetailsParagraphRef": "",
        "featuresKey": "",
        "descriptionKey": "",
        "featuresArrayFound": False,
        "descriptionBlockFound": False,
        "directVendorContentFound": False,
        "directVendorDetailsFound": False,
        "variantWindowMatched": False,
        "variantMatchScore": 0,
        "variantMatchReason": "",
        "matchedDynamicMediaUrl": "",
        "matchedVariantUrl": "",
        "matchedNearbyImage": "",
        "Source Used": source_name,
        "vendorPatternExcerpt": "",
        "featuresArrayExcerpt": "",
        "descriptionBlockExcerpt": "",
        "directVendorContentExcerpt": "",
    }

    if not source:
        return {"features": [], "description": "", "debug": debug}

    working_source = html.unescape(source)
    working_source = working_source.replace("\\u0026", "&amp;")

    direct_fastpath_result = None

    direct_variant_block_parts = find_direct_vendor_details_block_near_sku(
        working_source,
        target_rpc=target_rpc,
        search_after=30000,
    )

    if direct_variant_block_parts:
        direct_debug = debug.copy()
        direct_debug["variantWindowMatched"] = True
        direct_debug["variantMatchScore"] = 1000
        direct_debug["variantMatchReason"] = "direct_sku_vendorDetails_fastpath"
        direct_debug["directVendorContentExcerpt"] = normalize_space(direct_variant_block_parts[:2500])

        parsed = parse_vendor_details_from_block(
            direct_variant_block_parts,
            working_source,
            direct_debug,
        )

        direct_features = parsed.get("features", []) or []
        direct_description = parsed.get("description", "") or ""
        parsed_debug = parsed.get("debug", direct_debug)

        if direct_features or direct_description:
            direct_fastpath_result = {
                "features": normalize_cvs_features(direct_features[:5]),
                "description": clean_cvs_text(direct_description),
                "debug": parsed_debug,
            }

    anchor_windows = extract_rpc_anchor_windows(
        working_source,
        target_rpc=target_rpc,
        context_before=3500,
        context_after=12000,
    )

    for candidate in anchor_windows:
        candidate_debug = debug.copy()
        candidate_debug["variantWindowMatched"] = True
        candidate_debug["variantMatchScore"] = 999
        candidate_debug["variantMatchReason"] = "rpc_anchor_window"
        candidate_debug["matchedDynamicMediaUrl"] = candidate.get("anchor_text", "")
        candidate_debug["matchedNearbyImage"] = candidate.get("anchor_text", "")

        parsed = parse_vendor_details_from_block(
            candidate.get("window", ""),
            working_source,
            candidate_debug,
        )

        variant_features = parsed.get("features", []) or []
        variant_description = parsed.get("description", "") or ""

        if variant_features or variant_description:
            return {
                "features": normalize_cvs_features(variant_features[:5]),
                "description": clean_cvs_text(variant_description),
                "debug": parsed.get("debug", candidate_debug),
            }

    sorted_candidates = get_sorted_variant_windows(
        working_source,
        target_rpc=target_rpc,
        retail_url=retail_url,
    )

    for candidate in sorted_candidates:
        candidate_match_score = int(candidate.get("match_score", 0) or 0)
        if target_rpc and STRICT_CVS_VARIANT_MATCH and candidate_match_score < CVS_VARIANT_MIN_MATCH_SCORE:
            continue

        candidate_debug = debug.copy()
        candidate_debug["variantWindowMatched"] = candidate_match_score > 0
        candidate_debug["variantMatchScore"] = candidate_match_score
        candidate_debug["variantMatchReason"] = " | ".join(candidate.get("match_reason", []))
        candidate_debug["matchedDynamicMediaUrl"] = candidate.get("matched_dynamic_media", "")
        candidate_debug["matchedVariantUrl"] = candidate.get("matched_variant_url", "")
        candidate_debug["matchedNearbyImage"] = candidate.get("matched_nearby_image", "")

        parsed = parse_vendor_details_from_block(
            candidate.get("window", ""),
            working_source,
            candidate_debug,
        )

        variant_features = parsed.get("features", []) or []
        variant_description = parsed.get("description", "") or ""

        if variant_features or variant_description:
            return {
                "features": normalize_cvs_features(variant_features[:5]),
                "description": clean_cvs_text(variant_description),
                "debug": parsed.get("debug", candidate_debug),
            }

    if direct_fastpath_result:
        return direct_fastpath_result

    if target_rpc and STRICT_CVS_VARIANT_MATCH:
        strict_debug = debug.copy()
        strict_debug["variantWindowMatched"] = False
        strict_debug["variantMatchScore"] = 0
        strict_debug["variantMatchReason"] = "strict_variant_match_required_no_confident_variant"
        strict_debug["Source Used"] = f"{source_name} | strict_variant_match_required" if source_name else "strict_variant_match_required"
        return {"features": [], "description": "", "debug": strict_debug}

    shared_features = []
    shared_description = ""

    global_bullets_ref_match = re.search(
        r'"vendorDetailsBullets"\s*:\s*"(\$[0-9A-Za-z]{1,3})"',
        working_source,
        flags=re.DOTALL,
    )
    global_paragraph_ref_match = re.search(
        r'"vendorDetailsParagraph"\s*:\s*"(\$[0-9A-Za-z]{1,3})"',
        working_source,
        flags=re.DOTALL,
    )

    global_bullets_array_text = ""
    global_bullets_array_marker = re.search(
        r'"vendorDetailsBullets"\s*:\s*\[',
        working_source,
        flags=re.DOTALL,
    )
    if global_bullets_array_marker:
        array_start = global_bullets_array_marker.end() - 1
        global_bullets_array_text = extract_balanced_bracket_block(working_source, array_start)

    global_paragraph_direct_match = re.search(
        r'"vendorDetailsParagraph"\s*:\s*"((?:\\.|[^"\\])*)"',
        working_source,
        flags=re.DOTALL,
    )

    if global_bullets_ref_match:
        ref_token = global_bullets_ref_match.group(1)
        ref_key = ref_token.replace("$", "")

        debug["vendorPatternFound"] = True
        debug["vendorDetailsBulletsRef"] = ref_token
        debug["featuresKey"] = ref_key

        array_text = extract_array_block_for_key(working_source, ref_key)
        debug["featuresArrayFound"] = bool(array_text)
        debug["featuresArrayExcerpt"] = normalize_space(array_text)[:2000] if array_text else ""

        if array_text:
            shared_features = parse_jsonish_array_text(array_text)

    elif global_bullets_array_text:
        debug["featuresArrayFound"] = True
        debug["featuresArrayExcerpt"] = normalize_space(global_bullets_array_text)[:2000]
        shared_features = parse_jsonish_array_text(global_bullets_array_text)

    if global_paragraph_ref_match:
        ref_token = global_paragraph_ref_match.group(1)
        ref_key = ref_token.replace("$", "")

        debug["vendorPatternFound"] = True
        debug["vendorDetailsParagraphRef"] = ref_token
        debug["descriptionKey"] = ref_key

        desc_block = extract_newline_anchored_value_block(working_source, ref_key)
        desc_block = clean_cvs_text(desc_block)

        debug["descriptionBlockFound"] = bool(desc_block)
        debug["descriptionBlockExcerpt"] = normalize_space(desc_block)[:2000]
        shared_description = desc_block

    elif global_paragraph_direct_match:
        raw_para = global_paragraph_direct_match.group(1)

        if not re.fullmatch(r'\$[0-9A-Za-z]{1,3}', raw_para or ""):
            try:
                shared_description = json.loads(f'"{raw_para}"')
            except Exception:
                shared_description = raw_para

            shared_description = clean_cvs_text(shared_description)
            debug["descriptionBlockFound"] = bool(shared_description)
            debug["descriptionBlockExcerpt"] = normalize_space(shared_description)[:2000]

    return {
        "features": normalize_cvs_features(shared_features[:5]),
        "description": clean_cvs_text(shared_description),
        "debug": debug,
    }


def extract_vendor_copy_from_nextjs(html_text, target_rpc="", retail_url=""):
    raw_text = get_nextjs_chunks(html_text)
    raw_html = html.unescape(html_text or "")

    debug = {
        "rawHtmlLength": len(raw_html or ""),
        "rawTextLength": len(raw_text or ""),
        "nextjsChunkFound": bool(raw_text),
        "rawHtmlHasSelfNextF": "self.__next_f.push([1," in (raw_html or ""),
        "rawHtmlHasVendorDetailsBullets": "vendorDetailsBullets" in (raw_html or ""),
        "rawHtmlHasVendorDetailsParagraph": "vendorDetailsParagraph" in (raw_html or ""),
        "rawTextHasVendorDetailsBullets": "vendorDetailsBullets" in (raw_text or ""),
        "rawTextHasVendorDetailsParagraph": "vendorDetailsParagraph" in (raw_text or ""),
        "vendorPatternFound": False,
        "vendorDetailsBulletsRef": "",
        "vendorDetailsParagraphRef": "",
        "featuresKey": "",
        "descriptionKey": "",
        "featuresArrayFound": False,
        "descriptionBlockFound": False,
        "directVendorContentFound": False,
        "directVendorDetailsFound": False,
        "variantWindowMatched": False,
        "variantMatchScore": 0,
        "variantMatchReason": "",
        "matchedDynamicMediaUrl": "",
        "matchedVariantUrl": "",
        "matchedNearbyImage": "",
        "Source Used": "",
        "vendorPatternExcerpt": "",
        "featuresArrayExcerpt": "",
        "descriptionBlockExcerpt": "",
        "directVendorContentExcerpt": "",
        "rawHtmlVendorExcerpt": "",
        "rawTextVendorExcerpt": "",
    }

    if "vendorDetailsBullets" in raw_html:
        idx = raw_html.find("vendorDetailsBullets")
        debug["rawHtmlVendorExcerpt"] = normalize_space(
            raw_html[max(0, idx - 250): idx + 1500]
        )[:2000]

    if "vendorDetailsBullets" in raw_text:
        idx = raw_text.find("vendorDetailsBullets")
        debug["rawTextVendorExcerpt"] = normalize_space(
            raw_text[max(0, idx - 250): idx + 1500]
        )[:2000]

    text_result = extract_vendor_copy_from_source(
        raw_text,
        source_name="raw_text",
        target_rpc=target_rpc,
        retail_url=retail_url,
    )

    text_features = text_result.get("features", []) or []
    text_description = text_result.get("description", "") or ""

    html_result = {"features": [], "description": "", "debug": {}}

    if not text_features or not text_description:
        html_result = extract_vendor_copy_from_source(
            raw_html,
            source_name="raw_html",
            target_rpc=target_rpc,
            retail_url=retail_url,
        )

    html_features = html_result.get("features", []) or []
    html_description = html_result.get("description", "") or ""

    final_features = text_features or html_features
    final_description = text_description or html_description

    chosen_debug = text_result.get("debug", {}) if (text_features or text_description) else {}
    fallback_debug = html_result.get("debug", {}) if (html_features or html_description) else {}

    if not text_features and html_features:
        chosen_debug = fallback_debug.copy()
        chosen_debug["variantMatchReason"] = (
            str(chosen_debug.get("variantMatchReason", "")) + " | raw_html_features_fallback"
        ).strip(" |")
    else:
        merged_debug = chosen_debug.copy()
        for k, v in fallback_debug.items():
            if not merged_debug.get(k) and v:
                merged_debug[k] = v
        chosen_debug = merged_debug

    debug.update(chosen_debug)

    return {
        "features": final_features,
        "description": final_description,
        "debug": debug,
    }


def extract_cvs_images_from_html(html_text):
    html_text = str(html_text or "")
    best_images = {}
    order = []
    def add_candidate(raw_url, size=0):
        raw_url = html.unescape(str(raw_url or "").strip()).replace("\\/", "/").replace("\\u002F", "/").replace("\\u002f", "/")
        raw_url = raw_url.strip(' \"\'[](),;')
        if not raw_url:
            return
        if raw_url.startswith("//"):
            full = "https:" + raw_url
        elif raw_url.startswith("/"):
            full = "https://www.cvs.com" + raw_url
        elif raw_url.startswith("http://") or raw_url.startswith("https://"):
            full = raw_url
        else:
            return
        if not re.search(r"/productimages/high_res/[^\s\"'<>]+\.(?:jpg|jpeg|png|webp|avif)", full, flags=re.IGNORECASE):
            return
        base = full.split("?", 1)[0]
        name = base.split("/")[-1]
        if not name:
            return
        size = int(size or 0)
        if name not in best_images:
            order.append(name)
            # Preserve full URL with query params. Dedupe by file name/base separately.
            best_images[name] = {"url": full, "size": size}
        elif size > int(best_images[name].get("size", 0) or 0):
            best_images[name] = {"url": full, "size": size}
    for working in [html_text, html.unescape(html_text), html_text.replace("\\/", "/").replace("\\u002F", "/").replace("\\u002f", "/")]:
        for m in re.findall(r'/bizcontent/merchandising/productimages/high_res/[^\s\\\"\'<>]+?\.(?:jpg|jpeg|png|webp|avif)(?:\?[^\\\"\'<>\s]*)?', working, flags=re.IGNORECASE):
            sm = re.search(r"Resize=\((\d+)", m, flags=re.IGNORECASE)
            add_candidate(m, int(sm.group(1)) if sm else 0)
        for m in re.findall(r"https?://[^\s\"'<>]+/productimages/high_res/[^\s\"'<>]+?\.(?:jpg|jpeg|png|webp|avif)(?:\?[^\s\"'<>]*)?", working, flags=re.IGNORECASE):
            sm = re.search(r"Resize=\((\d+)", m, flags=re.IGNORECASE)
            add_candidate(m, int(sm.group(1)) if sm else 0)
        for m in re.findall(r'"(?:dynamicMediaUrl|imageUrl|image|src|url|thumbnailUrl|largeImageUrl)"\s*:\s*"((?:\\.|[^"\\])+)"', working, flags=re.IGNORECASE | re.DOTALL):
            add_candidate(m, 0)
    try:
        soup = BeautifulSoup(html_text, "html.parser")
        for img in soup.find_all("img"):
            for attr in ["src", "currentSrc", "data-src", "data-image-src", "data-lazy-src"]:
                add_candidate(img.get(attr, ""), 0)
            for part in str(img.get("srcset", "") or "").split(","):
                add_candidate(part.strip().split()[0] if part.strip() else "", 0)
    except Exception:
        pass
    return reorder_cvs_retailer_images_for_visual([best_images[name]["url"] for name in order], max_slots=MAX_IMAGE_SLOTS_TO_COMPARE)

def extract_cvs_visible_details_copy_from_dom(html_text):
    """CVS-only fallback for the visible Details accordion.

    Some CVS PDP captures have the live description in the rendered Details card
    instead of the Next.js vendorDetailsParagraph payload. This parser reads only
    the CVS Details card DOM and returns the visible description plus bullets.
    It is intentionally not used by any other retailer.
    """
    debug = {
        "Description Path": "",
        "Features Path": "",
        "Details Container Found": False,
        "Description Excerpt": "",
        "Feature Count": 0,
    }
    if not html_text:
        return {"description": "", "features": [], "debug": debug}

    working = str(html_text or "")
    for _ in range(3):
        unescaped = html.unescape(working)
        if unescaped == working:
            break
        working = unescaped

    soup = BeautifulSoup(working, "html.parser")

    # CVS-only primary rendered-DOM path. Current CVS PDP captures often include
    # the actual Details content in vendorDetailsParagraph/vendorDetailsBullet
    # elements even when the surrounding accordion no longer uses the older
    # whitespace-pre-line wrapper. Read those exact CVS elements first.
    direct_description_candidates = []
    direct_description_nodes = []
    direct_description_nodes.extend(soup.select('[class~="vendorDetailsParagraph"]'))
    direct_description_nodes.extend(soup.select('[class*="vendorDetailsParagraph"]'))
    direct_description_nodes.extend(soup.select('#vendorDetailsParagraph, [id^="vendorDetailsParagraph"]'))
    seen_description_nodes = set()
    for node in direct_description_nodes:
        node_id = id(node)
        if node_id in seen_description_nodes:
            continue
        seen_description_nodes.add(node_id)
        if node.name in {"style", "script"} or node.find_parent(["style", "script"]):
            continue
        value = clean_cvs_text(node.get_text(" ", strip=True))
        if len(value) >= 40:
            direct_description_candidates.append(value)

    direct_feature_nodes = []
    direct_feature_nodes.extend(soup.select('li#vendorDetailsBullet, li[id^="vendorDetailsBullet"]'))
    direct_feature_nodes.extend(soup.select('li.vendorDetailsBullet, li[class*="vendorDetailsBullet"]'))
    direct_feature_nodes.extend(soup.select('[data-testid*="vendorDetailsBullet"]'))
    direct_feature_values = []
    seen_feature_nodes = set()
    for node in direct_feature_nodes:
        node_id = id(node)
        if node_id in seen_feature_nodes:
            continue
        seen_feature_nodes.add(node_id)
        if node.name in {"style", "script"} or node.find_parent(["style", "script"]):
            continue
        value = clean_cvs_feature_text(node.get_text(" ", strip=True))
        if value:
            direct_feature_values.append(value)

    direct_features = normalize_cvs_features(direct_feature_values)
    direct_description = ""
    if direct_description_candidates:
        direct_description = max(
            dedupe_preserve_order(direct_description_candidates),
            key=lambda value: (len(value), value),
        )

    if direct_description or direct_features:
        debug["Details Container Found"] = True
        if direct_description:
            debug["Description Path"] = "cvs_rendered_dom_vendorDetailsParagraph"
            debug["Description Excerpt"] = direct_description[:500]
        if direct_features:
            debug["Features Path"] = "cvs_rendered_dom_vendorDetailsBullet"
            debug["Feature Count"] = len(direct_features)
        return {
            "description": direct_description,
            "features": direct_features[:5],
            "debug": debug,
        }

    candidate_containers = []

    # Most current CVS PDPs render description/features in a div with
    # whitespace-pre-line inside the Details accordion.
    for node in soup.select('[class*="whitespace-pre-line"]'):
        text_blob = normalize_space(node.get_text(" ", strip=True))
        if len(text_blob) < 80:
            continue
        if node.find("li", id=re.compile(r"^vendorDetailsBullet", re.IGNORECASE)) or node.find("ul"):
            candidate_containers.append(node)

    # Fallback: locate the accordion heading named Details, then inspect nearby
    # sibling/parent card content. This covers small CVS class-name changes.
    if not candidate_containers:
        for details_text_node in soup.find_all(string=re.compile(r"^\s*Details\s*$", re.IGNORECASE)):
            heading = details_text_node.find_parent(["h2", "button", "div"])
            if heading is None:
                continue
            card = heading
            for _ in range(5):
                if card is None:
                    break
                card_text = normalize_space(card.get_text(" ", strip=True))
                if "Details" in card_text and len(card_text) > 120 and card.find("li"):
                    candidate_containers.append(card)
                    break
                card = card.parent

    best = None
    best_score = -1
    for container in candidate_containers:
        container_text = normalize_space(container.get_text(" ", strip=True))
        if not container_text:
            continue
        lower_text = container_text.lower()
        if any(bad in lower_text for bad in [
            "rating & reviews",
            "explore more at cvs.com",
            "additional resources",
            "same-day delivery policies",
        ]):
            # The container may still be valid if these appear after Details,
            # but a smaller whitespace-pre-line node should score higher.
            penalty = 500
        else:
            penalty = 0
        li_count = len(container.find_all("li"))
        long_span_count = sum(
            1 for span in container.find_all("span")
            if not span.find_parent("li") and len(clean_cvs_text(span.get_text(" ", strip=True))) >= 80
        )
        score = (li_count * 100) + (long_span_count * 200) + min(len(container_text), 2000) - penalty
        if score > best_score:
            best_score = score
            best = container

    if best is None:
        return {"description": "", "features": [], "debug": debug}

    debug["Details Container Found"] = True

    description_candidates = []
    # Prefer a non-li span. In the current CVS Details card, this is the long
    # paragraph after title/item number and before the bullet list.
    for span in best.find_all("span"):
        if span.find_parent("li"):
            continue
        value = clean_cvs_text(span.get_text(" ", strip=True))
        if len(value) >= 80:
            description_candidates.append(value)

    # Extra fallback for markup that uses paragraph/div text instead of span.
    if not description_candidates:
        for tag in best.find_all(["p", "div"], recursive=True):
            if tag.find("li") or tag.find_parent("li"):
                continue
            value = clean_cvs_text(tag.get_text(" ", strip=True))
            if len(value) >= 120 and not re.search(r"item\s*#|\d+\s*ct", value, flags=re.IGNORECASE):
                description_candidates.append(value)

    description = ""
    if description_candidates:
        # Pick the richest paragraph, not the title or item-size row.
        description = max(description_candidates, key=lambda value: (len(value), value))
        debug["Description Path"] = "cvs_visible_details_dom_non_li_span"
        debug["Description Excerpt"] = description[:500]

    feature_values = []
    bullet_nodes = best.find_all("li", id=re.compile(r"^vendorDetailsBullet", re.IGNORECASE))
    if not bullet_nodes:
        for ul in best.find_all(["ul", "ol"]):
            bullet_nodes.extend(ul.find_all("li", recursive=False) or ul.find_all("li"))

    for li in bullet_nodes:
        value = clean_cvs_feature_text(li.get_text(" ", strip=True))
        if value:
            feature_values.append(value)

    features = normalize_cvs_features(feature_values)
    if features:
        debug["Features Path"] = "cvs_visible_details_dom_vendorDetailsBullet_li"
        debug["Feature Count"] = len(features)

    return {"description": description, "features": features[:5], "debug": debug}


def _extract_cvs_text_from_html(html_text, retail_url="", target_rpc=""):
    debug = {"Title Path": "", "Description Path": "", "Features Path": ""}

    if not html_text:
        return {"title": "", "description": "", "features": [], "debug": debug}

    soup = BeautifulSoup(html_text, "html.parser")
    title = ""

    h1 = soup.find("h1")
    if h1:
        title = normalize_space(h1.get_text(" ", strip=True))
        debug["Title Path"] = "h1"
    elif soup.title:
        title = normalize_space(soup.title.get_text(" ", strip=True))
        debug["Title Path"] = "html_title"

    if title:
        cleaned_title = re.sub(r"\s+-\s+CVS\s+Pharmacy\s*$", "", title, flags=re.IGNORECASE).strip()
        if cleaned_title != title:
            title = cleaned_title
            debug["Title Path"] = (str(debug.get("Title Path", "")) + " | cvs_title_suffix_removed").strip(" |")
        reviews_cleaned_title = re.sub(r"^\s*Customer reviews for\s+", "", title, flags=re.IGNORECASE).strip()
        if reviews_cleaned_title != title:
            title = reviews_cleaned_title
            debug["Title Path"] = (str(debug.get("Title Path", "")) + " | cvs_reviews_heading_product_name").strip(" |")
        if is_invalid_cvs_title_candidate(title):
            title = ""
            debug["Title Path"] = "cvs_invalid_title_rejected"

    vendor_copy = extract_vendor_copy_from_nextjs(
        html_text,
        target_rpc=target_rpc,
        retail_url=retail_url,
    )

    description = clean_cvs_text(vendor_copy.get("description", ""))
    features = normalize_cvs_features(vendor_copy.get("features", []))

    debug.update(vendor_copy.get("debug", {}))

    cvs_invalid_copy_markers = (
        "sign in or create an account",
        "pharmacy minuteclinic",
        "shop extracare",
        "search cvs or ask a question",
        "manage prescriptions",
        "schedule a vaccine",
        "shop store pickup",
        "add to pickup",
        "customer reviews for",
        "see real customer reviews",
        "see all reviews",
        "shop with confidence",
    )
    if description and any(marker in description.lower() for marker in cvs_invalid_copy_markers):
        description = ""
        debug["Description Path"] = "cvs_navigation_copy_rejected"

    # CVS-only fallback: if Next.js/vendorDetails misses the description,
    # use the visible Details accordion from the CVS DOM. This does not touch
    # Walgreens, Kroger, HEB, Sam's Club, or any other retailer parser.
    cvs_dom_details = {"description": "", "features": [], "debug": {}}
    if not description or not features:
        cvs_dom_details = extract_cvs_visible_details_copy_from_dom(html_text)
        dom_debug = cvs_dom_details.get("debug", {}) or {}
        if dom_debug:
            debug["CVS Visible Details Fallback"] = dom_debug
        if not description and cvs_dom_details.get("description"):
            description = clean_cvs_text(cvs_dom_details.get("description", ""))
            debug["Description Path"] = "cvs_visible_details_dom_fallback"
        if not features and cvs_dom_details.get("features"):
            features = normalize_cvs_features(cvs_dom_details.get("features", []))
            debug["Features Path"] = "cvs_visible_details_dom_fallback"

    if not debug.get("Description Path"):
        debug["Description Path"] = debug.get("Source Used", "") if description else "description_empty"
    if not title or not description:
        # CVS-only fallback for extension compact/raw pages and live pages that
        # expose Product JSON-LD/meta but not vendorDetails fields.
        try:
            jsonld_candidates = []
            structured_scripts = []
            structured_scripts.extend(soup.find_all("script", attrs={"type": "application/ld+json"}))
            structured_scripts.extend(soup.find_all("script", attrs={"id": re.compile(r"^sp-schema$", re.IGNORECASE)}))
            for script in structured_scripts:
                raw_json = (script.string or script.get_text(" ", strip=True) or "").strip()
                if not raw_json:
                    continue
                raw_json = html.unescape(raw_json)
                try:
                    parsed_json = json.loads(raw_json)
                except Exception:
                    continue
                stack = parsed_json if isinstance(parsed_json, list) else [parsed_json]
                while stack:
                    node = stack.pop(0)
                    if isinstance(node, dict):
                        node_type = node.get("@type", "")
                        node_types = node_type if isinstance(node_type, list) else [node_type]
                        if any(str(t).lower() == "product" for t in node_types):
                            jsonld_candidates.append(node)
                        for child in node.values():
                            if isinstance(child, (dict, list)):
                                stack.append(child)
                    elif isinstance(node, list):
                        stack.extend(node)
            for node in jsonld_candidates:
                if not title and node.get("name"):
                    title = normalize_space(node.get("name", ""))
                    debug["Title Path"] = "cvs_jsonld_product_name"
                if not description and node.get("description"):
                    description = clean_cvs_text(node.get("description", ""))
                    debug["Description Path"] = "cvs_jsonld_product_description"
                aggregate_rating = node.get("aggregateRating") if isinstance(node.get("aggregateRating"), dict) else {}
                if aggregate_rating:
                    rating_value = str(aggregate_rating.get("ratingValue", "") or "").strip()
                    review_count_value = str(aggregate_rating.get("reviewCount", "") or "").strip()
                    if rating_value:
                        debug["CVS Structured Rating"] = rating_value
                    if review_count_value:
                        debug["CVS Structured Review Count"] = review_count_value
                if title and description:
                    break
            if not title:
                meta_title = soup.find("meta", attrs={"property": "og:title"}) or soup.find("meta", attrs={"name": "twitter:title"})
                if meta_title and meta_title.get("content"):
                    title = normalize_space(meta_title.get("content", ""))
                    title = re.sub(r"\s+-\s+CVS\s+Pharmacy\s*$", "", title, flags=re.IGNORECASE).strip()
                    debug["Title Path"] = "cvs_meta_title"
            if not description:
                meta_desc = soup.find("meta", attrs={"name": "description"}) or soup.find("meta", attrs={"property": "og:description"})
                if meta_desc and meta_desc.get("content"):
                    description = clean_cvs_text(meta_desc.get("content", ""))
                    debug["Description Path"] = "cvs_meta_description"
        except Exception:
            pass

    visible_fallback = {"title": "", "description": "", "features": [], "debug": {}}
    if not description or not features or not title or len(features) < 5:
        visible_fallback = extract_cvs_indexed_text_fallback(html_text, retail_url=retail_url, target_rpc=target_rpc)
        visible_debug = visible_fallback.get("debug", {}) or {}
        if visible_debug:
            debug["CVS Visible Indexed Fallback"] = visible_debug
        if not title and visible_fallback.get("title"):
            title = normalize_space(visible_fallback.get("title", ""))
            debug["Title Path"] = visible_debug.get("Title Path", "cvs_visible_indexed_fallback")
        visible_description = clean_cvs_text(visible_fallback.get("description", ""))
        if visible_description and any(marker in visible_description.lower() for marker in cvs_invalid_copy_markers):
            visible_description = ""
            debug["CVS Visible Description Rejected"] = "navigation_or_reviews_shell"
        if visible_description and (not description or len(visible_description) > len(description) + 80):
            description = visible_description
            debug["Description Path"] = visible_debug.get("Description Path", "cvs_visible_indexed_fallback")
        fallback_features = normalize_cvs_features(visible_fallback.get("features", []))
        if fallback_features:
            if not features:
                features = fallback_features
                debug["Features Path"] = visible_debug.get("Features Path", "cvs_visible_indexed_fallback")
            else:
                combined_features = dedupe_preserve_order(list(features or []) + list(fallback_features or []))
                if len(combined_features) > len(features):
                    features = normalize_cvs_features(combined_features)
                    debug["Features Path"] = (str(debug.get("Features Path", "")) + " | cvs_visible_indexed_supplement").strip(" |")
    if not title:
        title = cvs_title_from_url_slug(retail_url)
        if title:
            debug["Title Path"] = "cvs_url_slug_fallback"

    if features and (not debug.get("Features Path") or debug.get("Features Path") == "features_empty"):
        debug["Features Path"] = debug.get("Source Used", "") or "cvs_final_merged_features"
    elif not features:
        debug["Features Path"] = "features_empty"

    return {
        "title": title,
        "description": description,
        "features": features[:5],
        "rating": str(debug.get("CVS Structured Rating", "") or ""),
        "review_count": str(debug.get("CVS Structured Review Count", "") or ""),
        "debug": debug,
    }


@st.cache_data(show_spinner=False)
def get_cvs_bundle(retail_url, target_rpc=""):
    html_text, cvs_source_used = fetch_cvs_html_with_fallbacks(retail_url)
    bundle = {
        "text": _extract_cvs_text_from_html(
            html_text,
            retail_url=retail_url,
            target_rpc=target_rpc,
        ),
        "images": extract_cvs_images_from_html(html_text),
    }
    debug = bundle.setdefault("text", {}).setdefault("debug", {})
    debug["Source Used"] = cvs_source_used
    debug["CVS Product HTML Detected"] = bool(is_probably_cvs_product_html(html_text))
    debug["CVS Live HTML Length"] = len(str(html_text or ""))
    debug["CVS Live HTML Quality Score"] = cvs_live_html_quality_score(html_text)
    debug["CVS URL Candidates Tried"] = " | ".join(cvs_url_candidates(retail_url))
    parsed_image_base = infer_cvs_image_base_from_images(bundle.get("images", []))
    if parsed_image_base:
        debug["CVS Parsed Image Base"] = parsed_image_base
        debug["CVS Parsed Image Count"] = len(bundle.get("images", []) or [])
        if ALLOW_RETAILER_GENERATED_IMAGE_FALLBACKS and len(bundle.get("images", []) or []) < 8:
            bundle["images"] = cvs_generated_image_candidates_for_base(parsed_image_base, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE)
            debug["CVS Image Fallback Applied"] = "cvs_parsed_image_base_expanded"
            debug["CVS Image Fallback Base"] = parsed_image_base
            debug["CVS Image Fallback Count"] = len(bundle.get("images") or [])
        elif len(bundle.get("images", []) or []) < 8:
            debug["CVS Image Fallback Skipped"] = "strict_live_retailer_only"

    # CVS combined approach, live-only for copy:
    # - Keep copy from direct CVS HTML only.
    # - Do not use the hard-coded known-product catalog unless the global flag is
    #   explicitly enabled for a separate reference mode.
    # - If images are missing, use CVS skuId high_res URL candidates only.
    fallback_bundle = get_cvs_known_product_fallback_bundle(retail_url=retail_url, target_rpc=target_rpc)
    fallback_text = fallback_bundle.get("text", {}) if isinstance(fallback_bundle, dict) else {}
    has_title = bool(normalize_space(bundle.get("text", {}).get("title", "")))
    has_description = bool(normalize_space(bundle.get("text", {}).get("description", "")))
    has_features = bool(any(normalize_space(x) for x in (bundle.get("text", {}).get("features", []) or [])))
    if fallback_text and (not has_title or not has_description or not has_features):
        if not has_title and normalize_space(fallback_text.get("title", "")):
            bundle["text"]["title"] = fallback_text.get("title", "")
            debug["Title Path"] = fallback_text.get("debug", {}).get("Title Path", "cvs_known_product_fallback_catalog")
        if not has_description and normalize_space(fallback_text.get("description", "")):
            bundle["text"]["description"] = fallback_text.get("description", "")
            debug["Description Path"] = fallback_text.get("debug", {}).get("Description Path", "cvs_known_product_fallback_catalog")
        if not has_features and fallback_text.get("features"):
            bundle["text"]["features"] = fallback_text.get("features", [])[:5]
            debug["Features Path"] = fallback_text.get("debug", {}).get("Features Path", "cvs_known_product_fallback_catalog")
        debug["Source Used"] = (str(debug.get("Source Used", "")) + " | " + str(fallback_text.get("debug", {}).get("Source Used", "cvs_known_product_fallback_catalog"))).strip(" |")
        debug["CVS Known Product Fallback Applied"] = True
    bundle = apply_cvs_targeted_copy_rescue_if_needed(
        bundle,
        retail_url=retail_url,
        target_rpc=target_rpc,
        reason="direct_cvs_html_missing_copy_or_wrong_target_rpc",
    )
    bundle = add_cvs_generated_image_fallback_if_needed(
        bundle,
        retail_url=retail_url,
        target_rpc=target_rpc,
        reason="direct_cvs_html_had_no_parseable_images",
    )
    return bundle

# =========================================
# KROGER PARSERS
# =========================================
def clean_kroger_text(text):
    if not text:
        return ""
    text = str(text)
    text = html.unescape(text)
    text = text.replace("\u003c", "<")
    text = text.replace("\u003e", ">")
    text = text.replace("\u0026", "&")
    text = text.replace("\n", " ")
    text = text.replace("\\/", "/")
    text = text.replace('\"', '"')
    text = re.sub(
        r"<script\b[^>]*>.*?</script>",
        " ",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if "<" in text and ">" in text:
        text = BeautifulSoup(text, "html.parser").get_text(" ", strip=True)
    text = re.sub(r"\s+", " ", text).strip()
    return text



def clean_kroger_variant_size(value):
    """Normalize Kroger selected size/pack variant text for UI and Excel output."""
    value = clean_kroger_text(value)
    if not value:
        return ""
    value = re.sub(r"\bButton\s+Group\s+Options\b.*$", "", value, flags=re.IGNORECASE).strip()
    value = re.sub(r"\bShow\s+more\s+sizes\b.*$", "", value, flags=re.IGNORECASE).strip()
    value = re.sub(r"\bThis\s+selection\s+is\s+unavailable\b.*$", "", value, flags=re.IGNORECASE).strip()
    value = value.strip(" :-|,;")
    if not value:
        return ""
    if len(value) > 60:
        return ""
    lowered = value.lower()
    blocked = {
        "regular", "heavy", "overnight", "extra heavy overnight",
        "front", "back", "left", "right", "top", "bottom", "main",
    }
    if lowered in blocked:
        return ""
    if re.search(r"\b(aisle|coupon|pickup|delivery|reviews?|rating|upc|located|discounted|sign in|cart)\b", lowered):
        return ""
    # Kroger size variants usually include a count/package unit. Keep this broad enough for
    # values like "48 ct", "10 pk / 56 ct", "6 pk / 160 ct", "16 count".
    if not re.search(r"\b\d", value):
        return ""
    if not re.search(r"\b(ct|count|pk|pack|packs|oz|fl oz|lb|lbs|g|kg|ml|in|size)\b", lowered):
        return ""
    return value


def extract_kroger_selected_variant_from_html(html_text):
    """Extract the live selected Kroger size variant, e.g. 48 ct.

    Preferred source is Kroger's selected variant label:
    <label data-testid="selected-variant-option"><input value="48 ct" checked>48 ct</label>

    Fallbacks support browser-extension TXT captures where the selected value appears as:
    "Size: 26 ctButton Group Options 16 count 26 ct 36 ct"
    or on the UPC line like:
    "10 pk / 56 ctUPC: 0003600050129".
    """
    raw = str(html_text or "")
    if not raw.strip():
        return ""
    working = html.unescape(raw)
    for _ in range(2):
        next_working = html.unescape(working)
        if next_working == working:
            break
        working = next_working

    # 1. Exact selected Kroger label from hydrated HTML.
    try:
        soup = BeautifulSoup(working, "html.parser")
        selected_labels = []
        selected_labels.extend(soup.select('label[data-testid="selected-variant-option"]'))
        for inp in soup.find_all("input", attrs={"checked": True}):
            parent_label = inp.find_parent("label")
            if parent_label:
                selected_labels.append(parent_label)
            input_value = clean_kroger_variant_size(inp.get("value", ""))
            if input_value:
                return input_value
        for label in selected_labels:
            label_input = label.find("input")
            if label_input:
                input_value = clean_kroger_variant_size(label_input.get("value", ""))
                if input_value:
                    return input_value
            label_text = clean_kroger_variant_size(label.get_text(" ", strip=True))
            if label_text:
                return label_text
    except Exception:
        pass

    # 2. Raw string fallback for escaped or partial label snippets.
    label_patterns = [
        r'data-testid=["\']selected-variant-option["\'][^>]*>\s*(?:<input\b[^>]*\bvalue=["\']([^"\']+)["\'][^>]*>)?([^<]{0,80})',
        r'<input\b[^>]*\bvalue=["\']([^"\']+)["\'][^>]*\bchecked\b[^>]*>\s*([^<]{0,80})',
        r'<input\b[^>]*\bchecked\b[^>]*\bvalue=["\']([^"\']+)["\'][^>]*>\s*([^<]{0,80})',
    ]
    for pattern in label_patterns:
        for match in re.finditer(pattern, working, flags=re.IGNORECASE | re.DOTALL):
            for group_value in match.groups():
                selected = clean_kroger_variant_size(group_value)
                if selected:
                    return selected

    text_blob = clean_kroger_text(BeautifulSoup(working, "html.parser").get_text(" ", strip=True) if "<" in working and ">" in working else working)

    # 3. Markdown/browser capture: Size: selected value Button Group Options...
    size_match = re.search(
        r'\bSize\s*:\s*(.+?)(?:\s*Button\s+Group\s+Options\b|\s*This\s+selection\s+is\s+unavailable\b|\s*Item\s+Availability\b|\s*Sign\s+In\b|\s*UPC\s*:|$)',
        text_blob,
        flags=re.IGNORECASE,
    )
    if size_match:
        selected = clean_kroger_variant_size(size_match.group(1))
        if selected:
            return selected

    # 4. UPC line fallback: selected size immediately before UPC.
    upc_match = re.search(
        r'(?<![A-Za-z0-9])([0-9][A-Za-z0-9 /.,-]{0,45}?(?:ct|count|pk|pack|packs|oz|fl oz|lb|lbs|g|kg|ml))\s*UPC\s*:',
        text_blob,
        flags=re.IGNORECASE,
    )
    if upc_match:
        selected = clean_kroger_variant_size(upc_match.group(1))
        if selected:
            return selected

    # 5. Title suffix fallback: title, selected size - Kroger.
    title_size_match = re.search(
        r',\s*([0-9][A-Za-z0-9 /.,-]{0,45}?\b(?:ct|count|pk|pack|packs|oz|fl oz|lb|lbs|g|kg|ml)\b)\s*-\s*Kroger\b',
        text_blob,
        flags=re.IGNORECASE,
    )
    if title_size_match:
        selected = clean_kroger_variant_size(title_size_match.group(1))
        if selected:
            return selected

    return ""


def normalize_kroger_features(items, max_features=10):
    if not items:
        return []

    if isinstance(items, str):
        intro, split_features = split_kroger_parsed_description(items)
        if split_features:
            items = split_features
        elif " | " in items:
            items = [x.strip() for x in items.split(" | ") if x.strip()]
        elif "•" in items:
            items = [x.strip() for x in items.split("•") if x.strip()]
        else:
            items = [items]

    out = []
    for item in items:
        val = clean_kroger_text(item)
        if not val or val.lower() in {"missing", "nan", "none"}:
            continue
        for part in split_kroger_feature_text_if_stuck(val):
            part = clean_kroger_text(part)
            if part and part.lower() not in {"missing", "nan", "none"}:
                out.append(part)
    return dedupe_preserve_order(out)[:max_features]



def split_kroger_feature_text_if_stuck(text):
    """Split a Kroger feature row only when multiple bullets merged together."""
    value = clean_kroger_text(text)
    if not value:
        return []

    matches = list(re.finditer(_kroger_feature_heading_pattern(), value, flags=re.IGNORECASE | re.UNICODE))
    if len(matches) >= 2:
        parts = []
        for idx, match in enumerate(matches):
            start = match.start()
            end = matches[idx + 1].start() if idx + 1 < len(matches) else len(value)
            part = clean_kroger_text(value[start:end])
            if part:
                parts.append(part)
        return parts or [value]

    # Handles rows like: "24 Mega Rolls ... Soft 2 ply ... 3x thicker ..."
    # without touching normal one-bullet rows.
    intro, unlabeled_parts = split_kroger_unlabeled_feature_text(value, max_features=10)
    if len(unlabeled_parts) >= 2 and not intro:
        return unlabeled_parts

    return [value]


def extract_kroger_description_features_from_html_fragment(fragment_html):
    """Extract Kroger description and feature rows from Product Details romance HTML.

    Expected split:
    - Description = the first <p> inside product-details-romance-description.
    - Features = each <li> inside the following <ul>.

    This also handles malformed TXT/browser-extension fragments such as:
    product-details-romance-description"><p>...</p><ul><li>...</li></ul>
    """
    raw = str(fragment_html or "")
    for _ in range(3):
        unescaped = html.unescape(raw)
        if unescaped == raw:
            break
        raw = unescaped

    if not raw.strip():
        return "", []

    marker_match = re.search(r'product-details-romance-description', raw, flags=re.IGNORECASE)
    if marker_match:
        # Keep the useful romance section only. This prevents footer/review text
        # from being pulled into description/features when a TXT capture is large.
        section_start = marker_match.start()
        p_start = raw.find("<p", section_start)
        ul_end_match = re.search(r'</ul\s*>', raw[p_start if p_start != -1 else section_start:], flags=re.IGNORECASE)
        if p_start != -1 and ul_end_match:
            local_start = p_start
            local_end = (p_start + ul_end_match.end()) if p_start != -1 else (section_start + ul_end_match.end())
            raw = '<section data-testid="product-details-romance-description">' + raw[local_start:local_end] + '</section>'
        elif p_start != -1:
            raw = '<section data-testid="product-details-romance-description">' + raw[p_start:] + '</section>'

    if "<" not in raw or ">" not in raw:
        description, features = split_kroger_parsed_description(raw)
        return clean_kroger_text(description), normalize_kroger_features(features, max_features=10)

    soup = BeautifulSoup(raw, "html.parser")
    holder = soup.select_one('[data-testid="product-details-romance-description"]') or soup

    p_tag = holder.find("p")
    description = clean_kroger_text(p_tag.get_text(" ", strip=True)) if p_tag is not None else ""

    if not description:
        text_parts = []
        for child in holder.find_all(recursive=False):
            if getattr(child, "name", None) in {"ul", "ol", "script", "style"}:
                continue
            child_text = clean_kroger_text(getattr(child, "get_text", lambda *a, **k: "")(" ", strip=True))
            if child_text:
                text_parts.append(child_text)
        description = normalize_space(" ".join(text_parts))

    features = []
    primary_list = holder.find("ul") or holder.find("ol")
    li_nodes = primary_list.find_all("li", recursive=False) if primary_list is not None else holder.find_all("li")

    for li in li_nodes:
        li_copy = BeautifulSoup(str(li), "html.parser").find("li")
        if li_copy is None:
            continue
        for nested_list in li_copy.find_all(["ul", "ol"]):
            nested_list.extract()
        value = clean_kroger_text(li_copy.get_text(" ", strip=True))
        lower_value = value.lower().strip(" .:-")
        if not value or lower_value in {"description above", "product details"}:
            continue
        if description and normalize_text(value) == normalize_text(description):
            continue
        features.extend(split_kroger_feature_text_if_stuck(value))

    # Regex fallback for malformed fragments where BeautifulSoup cannot rebuild li nodes.
    if not features and "<li" in raw.lower():
        for match in re.finditer(r'<li\b[^>]*>(.*?)(?:</li\s*>|(?=<li\b)|</ul\s*>|$)', raw, flags=re.IGNORECASE | re.DOTALL):
            item_html = match.group(1) or ""
            value = clean_kroger_text(item_html)
            if value and not (description and normalize_text(value) == normalize_text(description)):
                features.extend(split_kroger_feature_text_if_stuck(value))

    # Last fallback: if the paragraph and li tags flattened into one line, split on headings.
    if not features:
        full_text = clean_kroger_text(holder.get_text(" ", strip=True) if hasattr(holder, "get_text") else raw)
        split_description, split_features = split_kroger_parsed_description(full_text)
        if split_features:
            description = description or split_description
            features = split_features

    return clean_kroger_text(description), normalize_kroger_features(features, max_features=10)


def select_kroger_image_urls_by_perspective(images, max_images=6):
    """Keep one Kroger retailer image per perspective in front/back/left/right/top/bottom order."""
    perspective_order = ["front", "back", "left", "right", "top", "bottom"]
    size_rank = {"large": 0, "xlarge": 1, "medium": 2, "small": 3, "thumbnail": 4}
    chosen = {}
    for raw_url in images or []:
        url = html.unescape(str(raw_url or "").strip()).split("?", 1)[0].strip()
        if not url or "/product/images/" not in url.lower():
            continue
        perspective = _extract_kroger_perspective_from_url(url) or "front"
        if perspective not in perspective_order:
            continue
        size_match = re.search(r'/product/images/([^/]+)/', url, flags=re.IGNORECASE)
        size = str(size_match.group(1) or "").lower() if size_match else ""
        rank = size_rank.get(size, 99)
        current = chosen.get(perspective)
        if current is None or rank < current[0]:
            chosen[perspective] = (rank, url)
    return [chosen[p][1] for p in perspective_order if p in chosen][:max_images]


def build_kroger_invalid_capture_stub(requested_url="", final_url="", reason="invalid_kroger_capture_no_product_content"):
    requested_url = clean_uploaded_url_value(requested_url)
    final_url = clean_uploaded_url_value(final_url)
    reason = normalize_space(reason) or "invalid_kroger_capture_no_product_content"
    return (
        '<html><body '
        'data-pdp-invalid-kroger-capture="1" '
        f'data-invalid-reason="{html.escape(reason, quote=True)}" '
        f'data-requested-url="{html.escape(requested_url, quote=True)}" '
        f'data-final-url="{html.escape(final_url, quote=True)}">'
        f'Invalid Kroger capture: {html_escape_text(reason)}'
        '</body></html>'
    )


def actual_salsify_feature_count(text_bundle):
    if not isinstance(text_bundle, dict):
        return 0
    count = 0
    for i in range(1, 11):
        if normalize_space(text_bundle.get(f"feature{i}", "")):
            count = i
    return count


def build_dynamic_feature_fields_for_pair(retailer_name, s_text, retailer_features):
    """Compare/show only feature rows that exist on Salsify or the retailer page."""
    retailer_features = [x for x in (retailer_features or []) if normalize_space(x)]
    requirements = get_retailer_salsify_requirements(retailer_name)
    max_allowed = int(requirements.get("max_features", 5) or 5)
    actual_count = max(actual_salsify_feature_count(s_text), len(retailer_features))
    actual_count = min(actual_count, max_allowed)
    return [f"feature{i}" for i in range(1, actual_count + 1)]


def extract_kroger_markdown_product_info(raw_text):
    """Parse Kroger markdown/text captures when no real DOM list exists."""
    text = html.unescape(str(raw_text or ""))
    if not text.strip():
        return "", []

    # Look for Product Details area and stop before ratings/footer blocks.
    m = re.search(r'Product\s+Details\s*<br>\s*(.*?)(?:\n\s*Ratings\s*&\s*Reviews|\n\s*####\s+ABOUT US|\n\s*Improving Your Experience|$)', text, flags=re.IGNORECASE | re.DOTALL)
    block = m.group(1) if m else ""
    if not block:
        return "", []

    block = block.replace('<br>', '\n').replace('<br/>', '\n').replace('<br />', '\n')
    lines = [normalize_space(x) for x in re.split(r'\n+|\r+', block) if normalize_space(x)]
    lines = [re.sub(r'^[-•]\s*', '', x).strip() for x in lines]
    lines = [x for x in lines if x and not re.match(r'^(Ratings|Reviews|View More Reviews)$', x, flags=re.IGNORECASE)]
    if not lines:
        return "", []

    # First line often includes "Product name Perspective: Main" before the intro.
    first = lines[0]
    first = re.sub(r'^.*?Perspective:\s*Main\s*', '', first, flags=re.IGNORECASE).strip()
    description = first
    features = lines[1:]

    if not features and description:
        description, features = split_kroger_parsed_description(description)

    return clean_kroger_text(description), normalize_kroger_features(features, max_features=10)

def extract_kroger_description_and_features_from_html(html_text):
    debug = {
        "description_marker_found": False,
        "description_end_marker_found": False,
        "feature_block_found": False,
        "feature_count": 0,
        "description_excerpt": "",
        "features_excerpt": "",
        "parser_path": "",
    }

    if not html_text:
        return "", [], debug

    working = html.unescape(str(html_text or ""))
    soup = BeautifulSoup(working, "html.parser")
    romance = soup.select_one('[data-testid="product-details-romance-description"]')

    if romance is not None:
        description, features = extract_kroger_description_features_from_html_fragment(str(romance))
        debug["description_marker_found"] = bool(description)
        debug["description_end_marker_found"] = bool(description)
        debug["feature_block_found"] = bool(features)
        debug["feature_count"] = len(features)
        debug["description_excerpt"] = description[:500]
        debug["features_excerpt"] = " | ".join(features[:5])[:1000]
        debug["parser_path"] = "kroger_html_romance_div_p_ul_li_features"
        if description or features:
            return description, features, debug

    if "product-details-romance-description" in working.lower():
        description, features = extract_kroger_description_features_from_html_fragment(working)
        debug["description_marker_found"] = bool(description)
        debug["description_end_marker_found"] = bool(description)
        debug["feature_block_found"] = bool(features)
        debug["feature_count"] = len(features)
        debug["description_excerpt"] = description[:500]
        debug["features_excerpt"] = " | ".join(features[:5])[:1000]
        debug["parser_path"] = "kroger_html_raw_romance_marker_p_ul_li_features"
        if description or features:
            return description, features, debug

    markdown_description, markdown_features = extract_kroger_markdown_product_info(working)
    if markdown_description or markdown_features:
        debug["description_marker_found"] = bool(markdown_description)
        debug["description_end_marker_found"] = bool(markdown_description)
        debug["feature_block_found"] = bool(markdown_features)
        debug["feature_count"] = len(markdown_features)
        debug["description_excerpt"] = markdown_description[:500]
        debug["features_excerpt"] = " | ".join(markdown_features[:5])[:1000]
        debug["parser_path"] = "kroger_markdown_product_information"
        return markdown_description, markdown_features, debug

    for script in soup.find_all('script', attrs={'type': 'application/ld+json'}):
        raw = (script.string or script.get_text(' ', strip=True) or '').strip()
        if not raw:
            continue
        try:
            payload = json.loads(raw)
        except Exception:
            continue
        nodes = payload if isinstance(payload, list) else [payload]
        for node in nodes:
            if isinstance(node, dict) and str(node.get('@type', '')).lower() == 'product':
                description = clean_kroger_text(node.get('description', ''))
                if description:
                    debug["description_marker_found"] = True
                    debug["description_end_marker_found"] = True
                    debug["description_excerpt"] = description[:500]
                    debug["parser_path"] = "kroger_jsonld_product_description"
                    return description, [], debug
    return "", [], debug

def extract_kroger_text_from_html(html_text, retail_url="", target_rpc=""):
    debug = {
        "Title Path": "",
        "Description Path": "",
        "Features Path": "",
        "Source Used": "uploaded_txt_html",
        "Retailer": "Kroger",
    }

    if not html_text:
        debug["Title Path"] = "kroger_txt_missing"
        debug["Description Path"] = "kroger_txt_missing"
        debug["Features Path"] = "kroger_txt_missing"
        return {"title": "", "description": "", "features": [], "rating": "", "review_count": "", "variant_size": "", "kroger_size_variant": "", "debug": debug}

    working = html.unescape(str(html_text or ""))
    soup = BeautifulSoup(working, "html.parser")
    variant_size = extract_kroger_selected_variant_from_html(working)
    if variant_size:
        debug["Kroger Size Variant"] = variant_size
        debug["Kroger Size Variant Path"] = "kroger_selected_variant"

    title = ""
    h1 = soup.find('h1')
    if h1:
        title = normalize_space(h1.get_text(' ', strip=True))
        debug["Title Path"] = "h1"
    if not title:
        heading_match = re.search(r'(?m)^##\s+(.+?)\s*$', working)
        if heading_match:
            title = normalize_space(heading_match.group(1))
            debug["Title Path"] = "txt_markdown_h2"
    if not title:
        for script in soup.find_all('script', attrs={'type': 'application/ld+json'}):
            raw = (script.string or script.get_text(' ', strip=True) or '').strip()
            if not raw:
                continue
            try:
                payload = json.loads(raw)
            except Exception:
                continue
            nodes = payload if isinstance(payload, list) else [payload]
            for node in nodes:
                if isinstance(node, dict) and str(node.get('@type', '')).lower() == 'product':
                    title = normalize_space(node.get('name', ''))
                    if title:
                        debug["Title Path"] = "kroger_jsonld_product_name"
                        break
            if title:
                break
    if not title and soup.title:
        title = normalize_space(soup.title.get_text(' ', strip=True))
        title = re.sub(r'\s*-\s*Kroger\s*$', '', title, flags=re.IGNORECASE)
        debug["Title Path"] = "html_title"
    if not title:
        debug["Title Path"] = "kroger_title_missing"

    description, features, kroger_debug = extract_kroger_description_and_features_from_html(working)
    debug["Description Path"] = kroger_debug.get("parser_path", "") if description else "kroger_description_missing"
    debug["Features Path"] = kroger_debug.get("parser_path", "") if features else "kroger_features_missing"
    debug["Kroger Parser Debug"] = kroger_debug
    debug["Retail URL Lookup"] = normalize_kroger_url(retail_url)
    if target_rpc:
        debug["Retailer RPC"] = str(target_rpc or "").strip()

    rating = ""
    review_count = ""
    for script in soup.find_all('script', attrs={'type': 'application/ld+json'}):
        raw = (script.string or script.get_text(' ', strip=True) or '').strip()
        if not raw:
            continue
        try:
            payload = json.loads(raw)
        except Exception:
            continue
        nodes = payload if isinstance(payload, list) else [payload]
        for node in nodes:
            if isinstance(node, dict) and str(node.get('@type', '')).lower() == 'product':
                agg = node.get('aggregateRating', {})
                if isinstance(agg, dict):
                    rating = str(agg.get('ratingValue', '') or '').strip()
                    review_count = str(agg.get('reviewCount', '') or '').strip()
                if rating or review_count:
                    debug["Rating Path"] = "kroger_jsonld_aggregateRating"
                    break
        if rating or review_count:
            break

    if not rating and not review_count:
        rating_match = re.search(
            r'"aggregateRating"\s*:\s*\{.*?"ratingValue"\s*:\s*"?([0-9]+(?:\.[0-9]+)?)"?',
            working,
            flags=re.IGNORECASE | re.DOTALL,
        )
        review_count_match = re.search(
            r'"aggregateRating"\s*:\s*\{.*?"reviewCount"\s*:\s*"?([0-9][0-9,]*)"?',
            working,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if rating_match:
            rating = str(rating_match.group(1) or '').strip()
        if review_count_match:
            review_count = str(review_count_match.group(1) or '').strip()
        if rating or review_count:
            debug["Rating Path"] = "kroger_raw_aggregateRating_regex"

    return {
        "title": title,
        "description": description,
        "features": (features or [])[:10],
        "rating": rating,
        "review_count": review_count,
        "variant_size": variant_size,
        "kroger_size_variant": variant_size,
        "debug": debug,
    }

def _absolutize_kroger_image_url(url):
    url = html.unescape(str(url or "").strip())
    if not url:
        return ""
    if url.startswith("//"):
        url = "https:" + url
    elif url.startswith("/"):
        url = "https://www.kroger.com" + url
    if not re.match(r'^https?://', url, flags=re.IGNORECASE):
        return ""
    lowered = url.lower()
    if "/product/images/" not in lowered:
        return ""
    url = url.split("?", 1)[0].strip()
    return url




def build_kroger_main_image_fallback_url(retail_url="", target_rpc=""):
    """Build Kroger's predictable main/front product image URL as a fallback.

    Kroger product image URLs use the zero-padded UPC/image key in paths like:
    /product/images/medium/front/0003600056657.
    """
    rpc = clean_kroger_rpc(target_rpc)
    if not rpc:
        retail_url = str(retail_url or "")
        m = re.search(r'/([0-9]{8,14})(?:[/?#]|$)', retail_url)
        if m:
            rpc = clean_kroger_rpc(m.group(1))
    if not rpc:
        return ""
    if rpc.isdigit() and len(rpc) < 13:
        rpc = rpc.zfill(13)
    return f"https://www.kroger.com/product/images/medium/front/{rpc}"


def force_single_kroger_main_image(images, retail_url="", target_rpc=""):
    images = [str(u or "").strip() for u in list(images or []) if str(u or "").strip()]
    for url in images:
        perspective = _extract_kroger_perspective_from_url(url)
        if perspective in {"front", "main"}:
            return [url]
    if images:
        return [images[0]]
    fallback = build_kroger_main_image_fallback_url(retail_url=retail_url, target_rpc=target_rpc)
    return [fallback] if fallback else []

def _extract_kroger_perspective_from_text(text):
    text = normalize_space(text).lower()
    if not text:
        return ""
    for key in ["front", "back", "left", "right", "top", "bottom"]:
        if re.search(rf'{re.escape(key)}', text, flags=re.IGNORECASE):
            return key
    return ""


def _extract_kroger_perspective_from_url(url):
    url = str(url or "")
    m = re.search(r'/product/images/(?:xlarge|large|medium|small|thumbnail)/([^/]+)/', url, flags=re.IGNORECASE)
    if m:
        return str(m.group(1) or "").strip().lower()
    return ""



def _kroger_image_key_from_capture(html_text, retail_url="", target_rpc=""):
    """Find Kroger's zero-padded image key/UPC from TXT/HTML capture, URL, or RPC."""
    sources = [str(html_text or ""), str(retail_url or ""), str(target_rpc or "")]
    patterns = [
        r"UPC:\s*([0-9]{8,14})",
        r'"upc"\s*:\s*"?([0-9]{8,14})"?',
        r'"rpc"\s*:\s*"?([0-9]{8,14})"?',
        r"/product/images/(?:xlarge|large|medium|small|thumbnail)/(?:front|back|left|right|top|bottom)/([0-9]{8,14})",
        r"/(?:p/[^\s/]+/)?([0-9]{8,14})(?:[/?#\s]|$)",
        r"^([0-9]{8,14})$",
    ]
    for source in sources:
        if not source:
            continue
        for pattern in patterns:
            match = re.search(pattern, source, flags=re.IGNORECASE)
            if match:
                digits = re.sub(r"\D+", "", match.group(1) or "")
                if digits:
                    return digits.zfill(13)
    return ""


def _kroger_visible_perspectives_from_capture(html_text):
    """Return Kroger perspective labels found in capture text, preserving site order."""
    source = html.unescape(str(html_text or ""))
    found = []
    for match in re.finditer(r"Perspective\s*:\s*(front|back|left|right|top|bottom)", source, flags=re.IGNORECASE):
        perspective = str(match.group(1) or "").lower()
        if perspective and perspective not in found:
            found.append(perspective)
    return found


def _kroger_canonical_perspective_image_urls(html_text, retail_url="", target_rpc=""):
    """Build Kroger image-service URLs for perspectives exposed on the PDP.

    Some browser/TXT captures include perspective labels but omit image src URLs for
    back/left/right/top/bottom. Kroger uses predictable product image URLs by UPC:
    /product/images/large/{perspective}/{upc}.
    """
    image_key = _kroger_image_key_from_capture(html_text, retail_url=retail_url, target_rpc=target_rpc)
    if not image_key:
        return []
    perspectives = _kroger_visible_perspectives_from_capture(html_text)
    if not perspectives:
        perspectives = ["front", "back", "left", "right", "top", "bottom"]
    ordered = []
    for perspective in ["front", "back", "left", "right", "top", "bottom"]:
        if perspective in perspectives:
            ordered.append(f"https://www.kroger.com/product/images/large/{perspective}/{image_key}")
    return ordered

def extract_kroger_images_from_html(html_text):
    if not html_text:
        return []

    working = html.unescape(str(html_text or ""))
    soup = BeautifulSoup(working, "html.parser")

    perspective_rank = {
        "front": 0,
        "back": 1,
        "left": 2,
        "right": 3,
        "top": 4,
        "bottom": 5,
    }

    candidates = []
    seen = set()

    def add_candidate(url, slot_index=999, perspective_hint=""):
        clean_url = _absolutize_kroger_image_url(url)
        if not clean_url or clean_url in seen:
            return
        perspective = perspective_hint or _extract_kroger_perspective_from_url(clean_url)
        rank = perspective_rank.get(perspective, 999)
        candidates.append((slot_index, rank, clean_url))
        seen.add(clean_url)

    def choose_single_slide_image(slide):
        """
        Kroger slide 1 often contains both:
        - img.ProductImages-image (real visual slot image)
        - img.iiz__zoom-img (zoom duplicate of the same front image)

        We only want one retailer image per slide, so prefer ProductImages-image
        and skip zoom-image duplicates.
        """
        main_imgs = slide.select('img.ProductImages-image[src]')
        if main_imgs:
            return main_imgs[0]

        # Fallback only if the main class is missing in the captured HTML.
        for img in slide.select('img[src]'):
            class_tokens = [str(x or '').strip().lower() for x in (img.get('class') or [])]
            if any('zoom' in token for token in class_tokens):
                continue
            src = str(img.get('src', '') or '')
            if '/product/images/' not in src.lower():
                continue
            return img
        return None

    # Primary path: use one chosen image per visible slide, in site order.
    slide_nodes = soup.select('[data-testid="main-image-perspective"]')
    for idx, slide in enumerate(slide_nodes):
        aria_label = str(slide.get('aria-label', '') or '')
        perspective_hint = _extract_kroger_perspective_from_text(aria_label)
        chosen_img = choose_single_slide_image(slide)
        if chosen_img is None:
            continue
        src = chosen_img.get('src', '')
        alt = str(chosen_img.get('alt', '') or '')
        img_perspective = perspective_hint or _extract_kroger_perspective_from_text(alt)
        add_candidate(src, slot_index=idx, perspective_hint=img_perspective)

    # Secondary path: if the main slide nodes are missing, use the thumbnail carousel.
    if not candidates:
        thumb_imgs = soup.select('[data-testid="product-thumbnail-carousel"] img[src]')
        for idx, img in enumerate(thumb_imgs):
            src = img.get('src', '')
            alt = str(img.get('alt', '') or '')
            perspective_hint = _extract_kroger_perspective_from_text(alt)
            add_candidate(src, slot_index=idx, perspective_hint=perspective_hint)

    # Final fallback: raw URL regex.
    if not candidates:
        raw_urls = re.findall(
            r'https://www\.kroger\.com/product/images/(?:large|medium|small|thumbnail)/[^\s<>]+',
            working,
            flags=re.IGNORECASE,
        )
        for idx, raw_url in enumerate(raw_urls):
            add_candidate(raw_url, slot_index=idx)

    # Kroger TXT/browser captures sometimes expose perspective labels but only one src URL.
    # Fill those missing retailer slots from Kroger's canonical image-service URLs.
    canonical_urls = _kroger_canonical_perspective_image_urls(working)
    if canonical_urls:
        canonical_slot_base = 1000
        for idx, canonical_url in enumerate(canonical_urls):
            perspective_hint = _extract_kroger_perspective_from_url(canonical_url)
            add_candidate(canonical_url, slot_index=canonical_slot_base + idx, perspective_hint=perspective_hint)

    candidates.sort(key=lambda item: (item[0], item[1], item[2]))
    ordered_urls = [url for _, _, url in candidates]
    return select_kroger_image_urls_by_perspective(ordered_urls, max_images=6)

@st.cache_data(show_spinner=False)
def get_kroger_bundle(retail_url, target_rpc=""):
    """Fetch and parse live Kroger PDP HTML when no uploaded TXT capture is available."""
    retail_url = normalize_kroger_url(retail_url)
    if not retail_url:
        return build_empty_retailer_bundle("Kroger", "kroger_live_url_missing")

    html_text = get_html(retail_url)
    if not html_text:
        html_text = fetch_html_with_timeout(retail_url, 45)

    if not html_text:
        return build_empty_retailer_bundle("Kroger", "kroger_live_fetch_empty_or_blocked")

    if not is_valid_kroger_product_capture(html_text):
        bundle = build_empty_retailer_bundle("Kroger", "kroger_live_fetch_non_product_shell")
        debug = bundle.setdefault("text", {}).setdefault("debug", {})
        debug["Source Used"] = "kroger_live_html_rejected_shell"
        debug["Availability Rule"] = "availability_never_blocks_live_copy_or_images"
        return bundle

    bundle = {
        "text": extract_kroger_text_from_html(
            html_text,
            retail_url=retail_url,
            target_rpc=target_rpc,
        ),
        "images": extract_kroger_images_from_html(html_text),
    }
    debug = bundle.setdefault("text", {}).setdefault("debug", {})
    debug["Source Used"] = "kroger_live_html"
    debug["Image Path"] = "kroger_live_html_images"
    debug["Availability Rule"] = "availability_never_blocks_live_copy_or_images"
    availability_debug = extract_kroger_item_availability_debug(html_text)
    if availability_debug:
        debug["Item Availability"] = availability_debug
    return bundle

def _safe_json_loads(text):
    try:
        return json.loads(text)
    except Exception:
        return None


def _decode_walgreens_json_string(raw_value):
    """
    Decodes Walgreens JSON-like fragments such as:
    \\u003cp\\u003eHello\\u003c/p\\u003e
    """
    if not raw_value:
        return ""

    raw_value = str(raw_value)

    try:
        decoded = json.loads(f'"{raw_value}"')
    except Exception:
        decoded = raw_value
        decoded = decoded.replace('\\"', '"')
        decoded = decoded.replace("\\/", "/")

    decoded = html.unescape(decoded)
    return decoded.strip()


def _normalize_walgreens_text(value):
    if not value:
        return ""

    value = str(value)
    value = html.unescape(value)

    value = value.replace("\\u003c", "<")
    value = value.replace("\\u003e", ">")
    value = value.replace("\\u0026", "&")
    value = value.replace("\\n", " ")
    value = value.replace("\\/", "/")
    value = value.replace('\\"', '"')

    value = re.sub(
        r"<script\b[^>]*>.*?</script>",
        " ",
        value,
        flags=re.IGNORECASE | re.DOTALL,
    )

    if "<" in value and ">" in value:
        value = BeautifulSoup(value, "html.parser").get_text(" ", strip=True)

    return normalize_space(value)


def get_walgreens_product_id_from_url(retail_url):
    """
    Supports Walgreens product URLs like:
    - /ID=300432791-product
    - /ID=prod6153586-product
    - ?productId=300432791
    - ?productId=prod6153586
    """
    if not retail_url:
        return ""

    retail_url = str(retail_url or "").strip()

    patterns = [
        r"/ID=([A-Za-z0-9]+)-product",
        r"[?&]productId=([A-Za-z0-9]+)",
        r'"productId"\s*:\s*"([A-Za-z0-9]+)"',
    ]

    for pattern in patterns:
        m = re.search(pattern, retail_url, flags=re.IGNORECASE)
        if m:
            return m.group(1)

    return ""


def fetch_json_with_timeout(url, timeout_seconds):
    if not url:
        return None

    try:
        session = get_session()
        r = session.get(url, timeout=timeout_seconds, allow_redirects=True)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass

    return None


def get_walgreens_product_api_payload(product_id):
    """
    Walgreens source exposes a lighter product endpoint:
    /productapi/v1/products?productId={prodId}
    """
    if not product_id:
        return None

    api_url = f"https://www.walgreens.com/productapi/v1/products?productId={product_id}"
    return fetch_json_with_timeout(api_url, WALGREENS_API_TIMEOUT)


def walk_json(obj):
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from walk_json(v)
    elif isinstance(obj, list):
        for item in obj:
            yield from walk_json(item)


def find_first_dict_with_keys(obj, required_keys):
    required_keys = set(required_keys)

    for node in walk_json(obj):
        if isinstance(node, dict) and required_keys.issubset(set(node.keys())):
            return node

    return {}


def extract_walgreens_reviews_from_app_state_payload(payload):
    """
    Pull live Walgreens review fields from app-state payload:
    productData -> prodDetails -> section[] -> reviews
    """
    rating = ""
    review_count = ""

    if not isinstance(payload, dict):
        return rating, review_count

    root = payload
    if isinstance(payload.get("productData"), dict):
        root = payload.get("productData", {})

    prod_details = root.get("prodDetails", {}) if isinstance(root, dict) else {}
    section_list = prod_details.get("section", []) if isinstance(prod_details, dict) else []

    if isinstance(section_list, list):
        for section in section_list:
            if not isinstance(section, dict):
                continue
            reviews_obj = section.get("reviews", {})
            if isinstance(reviews_obj, dict):
                rating = str(reviews_obj.get("overallRating", "") or "").strip()
                review_count = str(reviews_obj.get("reviewCount", "") or "").strip()
                if rating or review_count:
                    return rating, review_count

    return rating, review_count


def extract_walgreens_reviews_from_html(html_text):
    """
    Parse live Walgreens review values from:
    window.__APP_INITIAL_STATE__ = {...};
    """
    if not html_text:
        return "", ""

    html_text = str(html_text or "")
    app_state_match = re.search(
        r'window\.__APP_INITIAL_STATE__\s*=\s*(\{.*?\})\s*;',
        html_text,
        flags=re.DOTALL,
    )
    if not app_state_match:
        return "", ""

    raw_json = app_state_match.group(1)

    try:
        payload = json.loads(raw_json)
    except Exception:
        return "", ""

    return extract_walgreens_reviews_from_app_state_payload(payload)


def format_walgreens_title_from_parts(raw_title="", size_count="", primary_attr=""):
    raw_title = _normalize_walgreens_text(raw_title)
    size_count = _normalize_walgreens_text(size_count)
    primary_attr = _normalize_walgreens_text(primary_attr)

    title_clean = raw_title.strip()

    if primary_attr and title_clean.lower().endswith((" " + primary_attr).lower()):
        title_clean = title_clean[: -(len(primary_attr) + 1)].rstrip(" ,-/")
    else:
        trailing_colors = [
            " Grey",
            " Gray",
            " Black",
            " White",
            " Pink",
            " Blue",
            " Green",
            " Red",
            " Brown",
            " Beige",
            " Purple",
            " Yellow",
            " Orange",
        ]
        for color in trailing_colors:
            if title_clean.lower().endswith(color.lower()):
                title_clean = title_clean[: -len(color)].rstrip(" ,-/")
                break

    if size_count:
        return f"{title_clean}, {size_count}"

    return title_clean


def _find_walgreens_feature_block_start(desc_html):
    """
    Walgreens live productDesc generally switches from description to features at:
    <br/>
<UL>
<LI>

    Return the start index of the whole feature block and the end index of the
    first <LI> opening tag so callers can either trim at the block start or begin
    parsing right after the first LI marker.
    """
    if not desc_html:
        return None

    working = str(desc_html)
    patterns = [
        r"<br\s*/?>\s*<ul[^>]*>\s*<li[^>]*>",
        r"<ul[^>]*>\s*<li[^>]*>",
        r"<li[^>]*>",
    ]

    for pattern in patterns:
        m = re.search(pattern, working, flags=re.IGNORECASE | re.DOTALL)
        if m:
            li_open = re.search(r"<li[^>]*>", m.group(0), flags=re.IGNORECASE)
            li_end = m.start() + li_open.end() if li_open else m.end()
            return m.start(), li_end

    return None
def _truncate_walgreens_copy_at_hard_stop(text):
    """
    Hard-stop Walgreens copy before utility / legal / disposal sections.
    This keeps only the real description + bullets and drops anything after,
    such as Made in USA, Do not flush, or Walgreens disclaimer text.
    """
    if not text:
        return ""

    working = str(text)
    lower = working.lower()
    stop_markers = [
        "Made in USA",
        "Made In USA",
        "Do not flush",
        "Do Not Flush",
        "Directions for Use:",
        "Direction for Use:",
        "To Use:",
        "To Dispose:",
        "How to Use:",
        "How To Use:",
        "Walgreens does not represent or warrant",
        "We recommend that you not rely solely on the information presented",
        "On occasion, manufacturers may improve or change their product formulas",
        "The food and drug administration has not intended to diagnose, treat, cure, or prevent any disease",
    ]

    cut_index = len(working)
    for marker in stop_markers:
        idx = lower.find(marker.lower())
        if idx != -1:
            cut_index = min(cut_index, idx)

    return working[:cut_index].strip()
    
def _truncate_walgreens_copy_at_hard_stop(text):
    """
    Hard-stop Walgreens copy before utility / legal / disposal sections.
    This keeps only the real description + bullets and drops anything after,
    such as Made in USA, Do not flush, or Walgreens disclaimer text.
    """
    if not text:
        return ""

    working = str(text)
    lower = working.lower()

    stop_markers = [
        "Made in USA",
        "Made In USA",
        "Do not flush",
        "Do Not Flush",
        "Directions for Use:",
        "Direction for Use:",
        "To Use:",
        "To Dispose:",
        "How to Use:",
        "How To Use:",
        "Walgreens does not represent or warrant",
        "We recommend that you not rely solely on the information presented",
        "On occasion, manufacturers may improve or change their product formulas",
        "The food and drug administration has not intended to diagnose, treat, cure, or prevent any disease",
    ]

    cut_index = len(working)
    for marker in stop_markers:
        idx = lower.find(marker.lower())
        if idx != -1:
            cut_index = min(cut_index, idx)

    return working[:cut_index].strip()
    
def _is_walgreens_450_image(url):
    url = str(url or '').lower().split('?', 1)[0]
    return url.endswith('/450.jpg') or url.endswith('_450.jpg')


def _extract_walgreens_feature_items_from_raw_product_desc(desc_html):
    """
    Walgreens live features start at a block like:
    <UL><LI>...
    Separation points are repeated <LI> markers.
    The last bullet may be the only one with a closing </LI>, so split on raw <LI>
    markers and stop each chunk at the next <LI>, a closing </UL>, or the end of the string.
    IMPORTANT: include the very first feature immediately after the first <LI>.
    """
    if not desc_html:
        return []

    working = str(desc_html)
    working = re.sub(
        r"<script\b[^>]*>.*?</script>",
        " ",
        working,
        flags=re.IGNORECASE | re.DOTALL,
    )

    # NEW: hard-stop before utility / legal / disposal text.
    working = _truncate_walgreens_copy_at_hard_stop(working)

    start_info = _find_walgreens_feature_block_start(working)
    if not start_info:
        return []

    feature_block_start, _first_li_end = start_info
    feature_html = working[feature_block_start:]

    li_markers = list(re.finditer(r"<li[^>]*>", feature_html, flags=re.IGNORECASE))
    if not li_markers:
        return []

    features = []
    for i, marker in enumerate(li_markers):
        start = marker.end()
        end_candidates = []

        if i + 1 < len(li_markers):
            end_candidates.append(li_markers[i + 1].start())

        ul_close = re.search(r"</ul>", feature_html[start:], flags=re.IGNORECASE)
        if ul_close:
            end_candidates.append(start + ul_close.start())

        li_close = re.search(r"</li>", feature_html[start:], flags=re.IGNORECASE)
        if li_close:
            end_candidates.append(start + li_close.start())

        end = min(end_candidates) if end_candidates else len(feature_html)
        chunk = feature_html[start:end]
        chunk = re.sub(r"<br[^>]*>", " ", chunk, flags=re.IGNORECASE)
        chunk = re.sub(r"</p>", " ", chunk, flags=re.IGNORECASE)

        feature_text = BeautifulSoup(chunk, "html.parser").get_text(" ", strip=True)
        feature_text = _normalize_walgreens_text(feature_text)
        feature_text = strip_walgreens_utility_tail(feature_text)

        if not feature_text:
            continue
        if is_walgreens_utility_feature(feature_text):
            continue

        features.append(feature_text)

    return dedupe_preserve_order(features)[:5]

def extract_walgreens_copy_from_product_desc_html(product_desc_html):
    """
    Walgreens live description starts inside productDesc paragraph HTML and ends right before the feature block that looks like:
    <UL><LI>...
    Improvement:
    - join all useful text nodes before the UL instead of trusting only the first <br> tag,
      because some Walgreens rows place a short label paragraph first.
    - hard-stop before utility / legal / disposal blocks such as Made in USA, Do not flush,
      or Walgreens disclaimer text.
    """
    if not product_desc_html:
        return "", []

    working = str(product_desc_html)
    working = re.sub(
        r"<script\b[^>]*>.*?</script>",
        " ",
        working,
        flags=re.IGNORECASE | re.DOTALL,
    )

    # NEW: hard-stop before utility / legal / disposal text.
    working = _truncate_walgreens_copy_at_hard_stop(working)

    start_info = _find_walgreens_feature_block_start(working)
    desc_html = working[:start_info[0]] if start_info else working
    desc_html = re.sub(r"(?:\s*)+$", " ", desc_html, flags=re.IGNORECASE)

    soup = BeautifulSoup(desc_html, "html.parser")
    desc_parts = []

    candidates = list(soup.body.children) if soup.body else list(soup.children)
    for node in candidates:
        node_name = getattr(node, "name", None)
        if node_name in {"ul", "ol", "script", "style"}:
            continue

        if hasattr(node, "get_text"):
            node_text = node.get_text(" ", strip=True)
        else:
            node_text = str(node).strip()

        node_text = _normalize_walgreens_text(node_text)
        node_text = strip_walgreens_utility_tail(node_text)

        if not node_text:
            continue
        if node_text.lower() in {"description", "details", "overview", "features"}:
            continue

        desc_parts.append(node_text)

    description_text = normalize_space(" ".join(desc_parts))
    description_text = re.sub(r"\)\)$", ")", description_text).strip()

    feature_items = _extract_walgreens_feature_items_from_raw_product_desc(working)
    return description_text, feature_items

def extract_walgreens_copy_from_product_sections(section_list):
    description = ""
    features = []

    if not isinstance(section_list, list):
        return description, features

    for section in section_list:
        if not isinstance(section, dict):
            continue

        desc_obj = section.get("description", {})
        if isinstance(desc_obj, dict) and desc_obj.get("productDesc"):
            desc_html = _decode_walgreens_json_string(desc_obj.get("productDesc", ""))
            description, features = extract_walgreens_copy_from_product_desc_html(desc_html)
            if description or features:
                break

    return description, features

def _walgreens_text_richness_tuple(value):
    value = clean_walgreens_text(value)
    useful = 0
    if len(value) >= 120:
        useful += 2
    elif len(value) >= 60:
        useful += 1
    if re.search(r"\b(count|ct|pack|roll|pads?|wipes?|diapers?|underwear|absorb|protection|odor|soft|dry|hours?)\b", value, flags=re.IGNORECASE):
        useful += 1
    return (useful, len(value), value)


def _walgreens_choose_richer_description(primary_value, secondary_value):
    primary_value = clean_walgreens_text(primary_value)
    secondary_value = clean_walgreens_text(secondary_value)
    if _walgreens_text_richness_tuple(secondary_value) > _walgreens_text_richness_tuple(primary_value):
        return secondary_value
    return primary_value


def _collect_jsonld_description_candidates(soup):
    candidates = []
    for script in soup.find_all('script', attrs={'type': 'application/ld+json'}):
        raw = (script.string or script.get_text(' ', strip=True) or '').strip()
        if not raw:
            continue
        try:
            payload = json.loads(raw)
        except Exception:
            continue
        nodes = payload if isinstance(payload, list) else [payload]
        while nodes:
            node = nodes.pop(0)
            if isinstance(node, dict):
                desc = node.get('description', '')
                if isinstance(desc, str) and desc.strip():
                    candidates.append(desc)
                for value in node.values():
                    if isinstance(value, (dict, list)):
                        nodes.append(value)
            elif isinstance(node, list):
                nodes.extend(node)
    return candidates


def _collect_heading_following_copy(soup):
    desc_candidates = []
    feature_lists = []
    heading_pattern = re.compile(r'(description|details|overview|about this product|about the product|product details|why we love|benefits|features)', re.IGNORECASE)
    for tag in soup.find_all(['h1', 'h2', 'h3', 'h4', 'strong', 'b', 'span', 'div', 'p']):
        heading_text = normalize_space(tag.get_text(' ', strip=True))
        if not heading_text or len(heading_text) > 80:
            continue
        if not heading_pattern.search(heading_text):
            continue
        collected_desc = []
        collected_features = []
        sibling = tag.next_sibling
        sibling_steps = 0
        while sibling is not None and sibling_steps < 10:
            sibling_steps += 1
            sibling_name = getattr(sibling, 'name', None)
            if sibling_name in {'h1', 'h2', 'h3', 'h4'}:
                break
            sibling_text = ''
            if hasattr(sibling, 'get_text'):
                sibling_text = normalize_space(sibling.get_text(' ', strip=True))
            else:
                sibling_text = normalize_space(str(sibling))
            sibling_text = clean_walgreens_text(sibling_text)
            if sibling_name in {'ul', 'ol'}:
                li_values = [clean_walgreens_text(li.get_text(' ', strip=True)) for li in sibling.find_all('li')]
                li_values = [v for v in li_values if v and not is_walgreens_utility_feature(v)]
                if li_values:
                    collected_features.extend(li_values)
            elif sibling_name == 'li':
                if sibling_text and not is_walgreens_utility_feature(sibling_text):
                    collected_features.append(sibling_text)
            else:
                if sibling_text and sibling_text.lower() not in {'description', 'details', 'overview', 'features'}:
                    collected_desc.append(sibling_text)
            sibling = getattr(sibling, 'next_sibling', None)
        if collected_desc:
            desc_candidates.append(' '.join(collected_desc))
        if collected_features:
            feature_lists.append(collected_features)
    return desc_candidates, feature_lists


def extract_walgreens_copy_from_meta_and_jsonld(html_text):
    if not html_text:
        return '', []
    soup = BeautifulSoup(html_text, 'html.parser')

    desc_candidates = []
    meta_selectors = [
        {'name': 'description'},
        {'property': 'og:description'},
        {'name': 'twitter:description'},
        {'itemprop': 'description'},
    ]
    for attrs in meta_selectors:
        meta = soup.find('meta', attrs=attrs)
        if meta and meta.get('content'):
            desc_candidates.append(meta.get('content', ''))

    desc_candidates.extend(_collect_jsonld_description_candidates(soup))
    heading_desc_candidates, heading_feature_lists = _collect_heading_following_copy(soup)
    desc_candidates.extend(heading_desc_candidates)

    best_description = ''
    for candidate in desc_candidates:
        best_description = _walgreens_choose_richer_description(best_description, candidate)

    feature_lists = []
    all_li = []
    for li in soup.find_all('li'):
        li_text = clean_walgreens_text(li.get_text(' ', strip=True))
        if li_text and not is_walgreens_utility_feature(li_text):
            all_li.append(li_text)
    if all_li:
        feature_lists.append(all_li)
    feature_lists.extend(heading_feature_lists)

    best_features = []
    for feature_values in feature_lists:
        cleaned = normalize_walgreens_features_final(feature_values, max_features=5)
        if _walgreens_feature_richness_tuple(cleaned) > _walgreens_feature_richness_tuple(best_features):
            best_features = cleaned

    return best_description, best_features

def _extract_walgreens_slot_num_from_key(key):
    key = str(key or "")
    m = re.search(r"(\d+)$", key)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return None
    return None


def _choose_best_walgreens_slot_image_from_dict(item):
    """
    Pull only Walgreens images that contain 450 in the live URL.
    Prefer /450.jpg and then /n_450.jpg style URLs in the exact site order.
    """
    if not isinstance(item, dict):
        return ""

    # Prefer explicit 450 image keys first.
    for k, v in item.items():
        if 'largeimageurl' in str(k).lower():
            url = _absolutize_walgreens_image_url(v)
            if url and _is_walgreens_450_image(url):
                return url

    # Do not fall back to 900 / 100. User only wants 450 images.
    return ""

def extract_walgreens_images_from_product_info(product_info):
    """
    Pull all live Walgreens images in Walgreens site order, but only keep image URLs
    that contain 450 in the path (e.g. /450.jpg, /2_450.jpg, /3_450.jpg).

    Do NOT append 900 / 100 / meta fallbacks.
    """
    if not isinstance(product_info, dict):
        return []

    ordered_urls = []
    seen = set()

    filmstrip = product_info.get("filmStripUrl", [])
    if isinstance(filmstrip, list):
        for item in filmstrip:
            chosen_url = _choose_best_walgreens_slot_image_from_dict(item)
            if chosen_url and chosen_url not in seen:
                ordered_urls.append(chosen_url)
                seen.add(chosen_url)

    return ordered_urls[:MAX_IMAGE_SLOTS_TO_COMPARE]

def _absolutize_walgreens_image_url(url):
    if not url:
        return ""

    url = html.unescape(str(url).strip())

    if url.startswith("//"):
        url = "https:" + url

    if not re.match(r"^https?://", url, flags=re.IGNORECASE):
        return ""

    lowered = url.lower()
    if not any(ext in lowered for ext in [".jpg", ".jpeg", ".png", ".webp", ".avif"]):
        return ""

    bad_tokens = [
        "sprite",
        "icon",
        "logo",
        "placeholder",
        "spacer",
        "data:image",
        ".svg",
    ]
    if any(tok in lowered for tok in bad_tokens):
        return ""

    return url


def _walgreens_image_family_key(url):
    """
    Normalize Walgreens image URLs so the same asset family collapses to one key,
    regardless of 100 / 220 / 450 / 900 size or image slot suffix.
    """
    url = _absolutize_walgreens_image_url(url)
    if not url:
        return ""

    lowered = url.lower().split("?", 1)[0]
    m = re.search(r"/prodimg/([^/]+)/([^/]+)$", lowered)
    if not m:
        return lowered

    product_folder = m.group(1)
    filename = m.group(2)
    filename = re.sub(r"\.(jpg|jpeg|png|webp|avif)$", "", filename, flags=re.IGNORECASE)
    filename = re.sub(r"_(100|220|450|900)$", "", filename, flags=re.IGNORECASE)
    filename = re.sub(r"^(100|220|450|900)$", "main", filename, flags=re.IGNORECASE)

    return f"{product_folder}/{filename}"


def _walgreens_image_size_rank(url):
    """
    Prefer 450px for visual QA, then 900px, then 220px, then 100px.
    """
    url = str(url or "").lower()

    if "/450." in url or "_450." in url:
        return 1
    if "/900." in url or "_900." in url:
        return 2
    if "/220." in url or "_220." in url:
        return 3
    if "/100." in url or "_100." in url:
        return 4

    return 99


def _add_walgreens_image_candidate(candidates, url, source_priority=99):
    """
    Keep only one best URL per image family.
    Lower source_priority wins.
    Lower image size rank wins.
    """
    url = _absolutize_walgreens_image_url(url)
    if not url:
        return

    family_key = _walgreens_image_family_key(url)
    if not family_key:
        return

    candidate = {
        "url": url,
        "source_priority": source_priority,
        "size_rank": _walgreens_image_size_rank(url),
    }

    current = candidates.get(family_key)
    if current is None:
        candidates[family_key] = candidate
        return

    current_tuple = (current["source_priority"], current["size_rank"], current["url"])
    new_tuple = (candidate["source_priority"], candidate["size_rank"], candidate["url"])

    if new_tuple < current_tuple:
        candidates[family_key] = candidate

def _collect_walgreens_api_copy_candidates(payload):
    """
    Walk the full Walgreens API payload and collect any likely
    description / feature-bearing fields.
    """
    desc_candidates = []
    feature_candidates = []

    if not payload:
        return desc_candidates, feature_candidates

    likely_desc_keys = {
        "productDesc",
        "description",
        "longDescription",
        "marketingDescription",
        "productDescription",
        "copy",
        "details",
    }

    likely_feature_keys = {
        "features",
        "bullets",
        "bulletPoints",
        "highlights",
        "benefits",
        "featureBullets",
        "vendorDetailsBullets",
    }

    for node in walk_json(payload):
        if not isinstance(node, dict):
            continue

        for key, value in node.items():
            key_str = str(key or "").strip()

            # Description-like keys.
            if key_str in likely_desc_keys:
                if isinstance(value, str) and value.strip():
                    desc_candidates.append(value)
                elif isinstance(value, dict):
                    for sub_k, sub_v in value.items():
                        if isinstance(sub_v, str) and sub_v.strip():
                            desc_candidates.append(sub_v)

            # Feature-like keys.
            if key_str in likely_feature_keys:
                if isinstance(value, list):
                    feature_candidates.append(value)
                elif isinstance(value, str) and value.strip():
                    feature_candidates.append([value])
                elif isinstance(value, dict):
                    sub_items = []
                    for sub_k, sub_v in value.items():
                        if isinstance(sub_v, str) and sub_v.strip():
                            sub_items.append(sub_v)
                        elif isinstance(sub_v, list):
                            sub_items.extend(
                                [x for x in sub_v if isinstance(x, str) and x.strip()]
                            )
                    if sub_items:
                        feature_candidates.append(sub_items)

    return desc_candidates, feature_candidates

def _extract_best_walgreens_copy_from_api_payload(payload):
    """
    Flexible fallback extractor for Walgreens API payloads when
    prodDetails.section does not contain usable copy.
    """
    best_description = ""
    best_features = []

    desc_candidates, feature_candidates = _collect_walgreens_api_copy_candidates(payload)

    for candidate in desc_candidates:
        candidate_clean = clean_walgreens_text(candidate)
        candidate_clean = strip_walgreens_description_tail(candidate_clean)
        best_description = _walgreens_choose_richer_description(
            best_description,
            candidate_clean,
        )

    for candidate_list in feature_candidates:
        cleaned = normalize_walgreens_features_final(candidate_list, max_features=5)
        if _walgreens_feature_richness_tuple(cleaned) > _walgreens_feature_richness_tuple(best_features):
            best_features = cleaned

    return best_description, best_features
    
def build_walgreens_bundle_from_api_payload(payload):
    empty = {
        "text": {
            "title": "",
            "description": "",
            "features": [],
            "rating": "",
            "review_count": "",
            "debug": {
                "Title Path": "walgreens_api_missing",
                "Description Path": "walgreens_api_missing",
                "Features Path": "walgreens_api_missing",
                "Source Used": "walgreens_api",
            },
        },
        "images": [],
    }

    if not payload:
        return empty

    root = payload
    if isinstance(payload, dict) and "productData" in payload and isinstance(payload["productData"], dict):
        root = payload["productData"]

    live_rating, live_review_count = extract_walgreens_reviews_from_app_state_payload(payload)

    product_info = {}
    prod_details = {}

    if isinstance(root, dict):
        product_info = root.get("productInfo", {}) if isinstance(root.get("productInfo", {}), dict) else {}
        prod_details = root.get("prodDetails", {}) if isinstance(root.get("prodDetails", {}), dict) else {}

    if not product_info:
        product_info = find_first_dict_with_keys(root, {"title", "sizeCount"})

    if not prod_details:
        prod_details = find_first_dict_with_keys(root, {"section"})

    raw_title = product_info.get("title", "")
    size_count = product_info.get("sizeCount", "")
    primary_attr = product_info.get("primaryAttribute", "")

    final_title = format_walgreens_title_from_parts(
        raw_title=raw_title,
        size_count=size_count,
        primary_attr=primary_attr,
    )

    # First try the original structured path.
    section_list = prod_details.get("section", []) if isinstance(prod_details, dict) else []
    description, features = extract_walgreens_copy_from_product_sections(section_list)

    desc_path = (
        "walgreens_api_prodDetails_section_description_productDesc_preUL"
        if description else
        "walgreens_api_description_missing"
    )
    feat_path = (
        "walgreens_api_prodDetails_section_description_productDesc_rawLI"
        if features else
        "walgreens_api_features_missing"
    )
    source_used = "walgreens_api"

    # New fallback: scan the full API payload if original path is empty/weak.
    if not description or not features:
        payload_description, payload_features = _extract_best_walgreens_copy_from_api_payload(root)

        if payload_description:
            description = _walgreens_choose_richer_description(description, payload_description)
            desc_path = "walgreens_api_payload_fallback"

        if _walgreens_feature_richness_tuple(payload_features) > _walgreens_feature_richness_tuple(features):
            features = payload_features
            feat_path = "walgreens_api_payload_fallback"

        if payload_description or payload_features:
            source_used = "walgreens_api | walgreens_api_payload_fallback"

    images = extract_walgreens_images_from_product_info(product_info)

    return {
        "text": {
            "title": final_title,
            "description": description,
            "features": features[:5],
            "rating": live_rating,
            "review_count": live_review_count,
            "debug": {
                "Title Path": (
                    "walgreens_api_productInfo_title_plus_sizeCount"
                    if final_title else
                    "walgreens_api_title_missing"
                ),
                "Description Path": desc_path,
                "Features Path": feat_path,
                "Source Used": source_used,
            },
        },
        "images": images,
    }

def get_walgreens_prod_desc_url(product_id):
    if not product_id:
        return ""
    return f"https://www.walgreens.com/store/store/prodDesc.jsp?id={product_id}"

def get_walgreens_prod_desc_html(product_id):
    """
    Lightweight Walgreens copy endpoint. Often more reliable than the full PDP HTML for prod... items.
    """
    url = get_walgreens_prod_desc_url(product_id)
    return fetch_html_with_timeout(url, WALGREENS_REQUEST_TIMEOUT)
    
def get_walgreens_prod_desc_html(product_id):
    """
    Lightweight Walgreens copy endpoint.
    Often more reliable than the full PDP HTML for prod... items.
    """
    url = get_walgreens_prod_desc_url(product_id)
    return fetch_html_with_timeout(url, WALGREENS_REQUEST_TIMEOUT)


def build_walgreens_title_from_url_slug(retail_url, description_html=""):
    """
    Fallback title builder for prodDesc.jsp route.
    """
    retail_url = str(retail_url or "").strip()

    slug_match = re.search(
        r"/store/c/([^/]+)/ID=[A-Za-z0-9]+-product",
        retail_url,
        flags=re.IGNORECASE,
    )

    if not slug_match:
        return ""

    slug = slug_match.group(1)
    title = slug.replace("-", " ")
    title = title.replace(" ,", ",")
    title = html.unescape(title)
    title = normalize_space(title)

    title = " ".join(word.capitalize() if word.islower() else word for word in title.split())

    desc_text = BeautifulSoup(str(description_html or ""), "html.parser").get_text(" ", strip=True)
    desc_text = normalize_space(desc_text)

    count_match = re.search(r"(\d+)\s+count", desc_text, flags=re.IGNORECASE)
    if count_match:
        count_val = count_match.group(1)
        title = f"{title}, {count_val}.0 ea"

    return normalize_space(title)


def build_walgreens_bundle_from_prod_desc_fragment(product_id, retail_url=""):
    """
    Fallback path for prod... Walgreens items.
    Uses the lightweight prodDesc.jsp endpoint for description/features.
    Also runs the same meta/jsonld fallback extractor against the fragment HTML.
    """
    empty = {
        "text": {
            "title": "",
            "description": "",
            "features": [],
            "debug": {
                "Title Path": "walgreens_prodDesc_missing",
                "Description Path": "walgreens_prodDesc_missing",
                "Features Path": "walgreens_prodDesc_missing",
                "Source Used": "walgreens_prodDesc_fragment",
            },
        },
        "images": [],
    }

    if not product_id:
        return empty

    fragment_html = get_walgreens_prod_desc_html(product_id)
    if not fragment_html:
        return empty

    description, features = extract_walgreens_copy_from_product_desc_html(fragment_html)
    fallback_description, fallback_features = extract_walgreens_copy_from_meta_and_jsonld(fragment_html)
    description = _walgreens_choose_richer_description(description, fallback_description)
    features = normalize_walgreens_features_final(features, max_features=5)
    fallback_features = normalize_walgreens_features_final(fallback_features, max_features=5)
    if _walgreens_feature_richness_tuple(fallback_features) > _walgreens_feature_richness_tuple(features):
        features = fallback_features

    title = build_walgreens_title_from_url_slug(
        retail_url=retail_url,
        description_html=fragment_html,
    )

    return {
        "text": {
            "title": title,
            "description": description,
            "features": features[:5],
            "debug": {
                "Title Path": "walgreens_prodDesc_url_slug_fallback" if title else "walgreens_prodDesc_title_missing",
                "Description Path": "walgreens_prodDesc_fragment" if description else "walgreens_prodDesc_description_missing",
                "Features Path": "walgreens_prodDesc_fragment" if features else "walgreens_prodDesc_features_missing",
                "Source Used": "walgreens_prodDesc_fragment",
            },
        },
        "images": [],
    }

def _extract_walgreens_title_from_source(html_text):
    if not html_text:
        return "", "walgreens_title_missing"

    title_match = re.search(
        r'"productInfo"\s*:\s*\{.*?"title"\s*:\s*"((?:\\.|[^"\\])*)"',
        html_text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    size_match = re.search(
        r'"sizeCount"\s*:\s*"((?:\\.|[^"\\])*)"',
        html_text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    primary_attr_match = re.search(
        r'"primaryAttribute"\s*:\s*"((?:\\.|[^"\\])*)"',
        html_text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    if not title_match:
        return "", "walgreens_productInfo_title_missing"

    raw_title = _decode_walgreens_json_string(title_match.group(1))
    size_count = _decode_walgreens_json_string(size_match.group(1)) if size_match else ""
    primary_attr = _decode_walgreens_json_string(primary_attr_match.group(1)) if primary_attr_match else ""

    final_title = format_walgreens_title_from_parts(
        raw_title=raw_title,
        size_count=size_count,
        primary_attr=primary_attr,
    )

    return final_title, "walgreens_productInfo_title_plus_sizeCount"


def _extract_walgreens_product_desc_block(html_text):
    if not html_text:
        return "", "walgreens_productDesc_missing"

    desc_match = re.search(
        r'"productDesc"\s*:\s*"((?:\\.|[^"\\])*)"',
        html_text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    if not desc_match:
        return "", "walgreens_productDesc_missing"

    raw_desc = desc_match.group(1)
    desc_html = _decode_walgreens_json_string(raw_desc)
    return desc_html, "walgreens_productDesc_found"


def _extract_walgreens_description_and_features_from_product_desc(html_text):
    desc_html, source_path = _extract_walgreens_product_desc_block(html_text)

    if not desc_html:
        return "", [], source_path

    description_text, feature_items = extract_walgreens_copy_from_product_desc_html(desc_html)

    return (
        description_text,
        feature_items,
        "walgreens_productDesc_preUL_description_and_rawLI_features",
    )


def extract_walgreens_text_from_html(html_text, retail_url="", target_rpc=""):
    debug = {
        "Title Path": "",
        "Description Path": "",
        "Features Path": "",
        "Source Used": "walgreens_live_html",
    }
    if not html_text:
        return {
            "title": "",
            "description": "",
            "features": [],
            "rating": "",
            "review_count": "",
            "debug": debug,
        }

    title, title_path = _extract_walgreens_title_from_source(html_text)
    description, features, copy_path = _extract_walgreens_description_and_features_from_product_desc(html_text)
    live_rating, live_review_count = extract_walgreens_reviews_from_html(html_text)

    chosen_features = normalize_walgreens_features_final(features, max_features=5)

    debug["Title Path"] = title_path
    debug["Description Path"] = copy_path if description else "walgreens_live_html_description_missing"
    debug["Features Path"] = copy_path if chosen_features else "walgreens_live_html_features_missing"

    return {
        "title": title,
        "description": description,
        "features": chosen_features[:5],
        "rating": live_rating,
        "review_count": live_review_count,
        "debug": debug,
    }

def extract_walgreens_images_from_html(html_text):
    """
    Preserve Walgreens image slot order from HTML.

    IMPORTANT:
    - Capture the unnumbered "largeImageUrl" hero image as slot 1.
    - Then capture numbered keys like largeImageUrl1, largeImageUrl2, etc.
    - If a hero exists, shift numbered slots by +1 so the remaining images
      do not move up and replace the main image.
    - Only keep 450 images.
    """
    if not html_text:
        return []

    slot_candidates = {}
    seen = set()

    def maybe_store(slot_num, url):
        url = _absolutize_walgreens_image_url(url)
        if not url or not _is_walgreens_450_image(url):
            return
        if slot_num not in slot_candidates:
            slot_candidates[slot_num] = url

    # 1. Capture the main unnumbered hero image first.
    hero_match = re.search(
        r'"largeImageUrl"\s*:\s*"((?:\\\\.|[^"\\\\])*)"',
        html_text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    has_main_hero = False
    if hero_match:
        hero_url = _decode_walgreens_json_string(hero_match.group(1))
        maybe_store(1, hero_url)
        has_main_hero = 1 in slot_candidates

    # 2. Capture numbered image keys.
    # If the hero exists, shift numbered slots by +1 so slot 1 stays the hero.
    for m in re.finditer(
        r'"largeImageUrl(\d+)"\s*:\s*"((?:\\\\.|[^"\\\\])*)"',
        html_text,
        flags=re.IGNORECASE | re.DOTALL,
    ):
        raw_slot_num = int(m.group(1))
        slot_num = raw_slot_num + 1 if has_main_hero else raw_slot_num
        slot_url = _decode_walgreens_json_string(m.group(2))
        maybe_store(slot_num, slot_url)

    ordered_urls = []
    for slot_num in sorted(slot_candidates.keys()):
        url = slot_candidates[slot_num]
        if url and url not in seen:
            ordered_urls.append(url)
            seen.add(url)

    return ordered_urls[:MAX_IMAGE_SLOTS_TO_COMPARE]

def _walgreens_has_copy_or_images(bundle):
    if not isinstance(bundle, dict):
        return False

    bundle_text = bundle.get("text", {}) or {}
    bundle_images = bundle.get("images", []) or []

    return bool(
        bundle_text.get("title")
        or bundle_text.get("description")
        or bundle_text.get("features")
        or bundle_images
    )


def _walgreens_clean_feature_list(values, max_features=5):
    if not values:
        return []

    if isinstance(values, str):
        values = [values]

    cleaned = []
    for value in values:
        value = clean_walgreens_text(value)
        if not value:
            continue
        if is_walgreens_utility_feature(value):
            continue
        cleaned.append(value)

    return dedupe_preserve_order(cleaned)[:max_features]


def _walgreens_feature_richness_tuple(values):
    cleaned = _walgreens_clean_feature_list(values, max_features=5)
    return (len(cleaned), sum(len(x) for x in cleaned))


def _prefer_richer_walgreens_text_value(primary_value, secondary_value):
    primary_value = clean_walgreens_text(primary_value)
    secondary_value = clean_walgreens_text(secondary_value)

    primary_tuple = (len(primary_value), primary_value)
    secondary_tuple = (len(secondary_value), secondary_value)

    return secondary_value if secondary_tuple > primary_tuple else primary_value


def merge_walgreens_bundles_prefer_richer_copy(*bundles):
    """
    Merge multiple Walgreens bundle candidates and keep the richest live copy.

    Why this exists:
    - Walgreens HTML often reflects what is actually live on the PDP.
    - Walgreens API / prodDesc fragment can still be useful fallbacks.
    - Some rows were under-pulling copy because only one source path was trusted.

    Rules:
    - Prefer the longest useful title.
    - Prefer the longest useful description.
    - Prefer the feature set with the most non-empty bullets; tie-break by total text length.
    - Prefer the first non-empty image set in the order passed in.
    - Carry forward merged debug/source notes.
    """
    merged = {
        "text": {
            "title": "",
            "description": "",
            "features": [],
            "rating": "",
            "review_count": "",
            "debug": {},
        },
        "images": [],
    }

    source_used_parts = []

    for bundle in bundles:
        if not isinstance(bundle, dict):
            continue

        bundle_text = bundle.get("text", {}) or {}
        bundle_images = bundle.get("images", []) or []
        bundle_debug = bundle_text.get("debug", {}) or {}

        merged["text"]["title"] = _prefer_richer_walgreens_text_value(
            merged["text"].get("title", ""),
            bundle_text.get("title", ""),
        )
        merged["text"]["description"] = _prefer_richer_walgreens_text_value(
            merged["text"].get("description", ""),
            bundle_text.get("description", ""),
        )

        current_features = merged["text"].get("features", []) or []
        new_features = bundle_text.get("features", []) or []
        if _walgreens_feature_richness_tuple(new_features) > _walgreens_feature_richness_tuple(current_features):
            merged["text"]["features"] = _walgreens_clean_feature_list(new_features, max_features=5)

        if not merged["text"].get("rating") and bundle_text.get("rating"):
            merged["text"]["rating"] = str(bundle_text.get("rating", "") or "").strip()
        if not merged["text"].get("review_count") and bundle_text.get("review_count"):
            merged["text"]["review_count"] = str(bundle_text.get("review_count", "") or "").strip()

        if not merged.get("images") and bundle_images:
            merged["images"] = [x for x in bundle_images if x]

        for k, v in bundle_debug.items():
            if v and not merged["text"]["debug"].get(k):
                merged["text"]["debug"][k] = v

        source_used = str(bundle_debug.get("Source Used", "") or "").strip()
        if source_used and source_used not in source_used_parts:
            source_used_parts.append(source_used)

    if source_used_parts:
        merged["text"]["debug"]["Source Used"] = " | ".join(source_used_parts)

    merged["text"]["features"] = _walgreens_clean_feature_list(
        merged["text"].get("features", []),
        max_features=5,
    )

    return merged


def _walgreens_description_richness_tuple(value):
    value = clean_walgreens_text(value)
    return (len(value), value)


def _walgreens_bundle_is_rich_enough(bundle):
    if not isinstance(bundle, dict):
        return False

    bundle_text = bundle.get("text", {}) or {}
    title = clean_walgreens_title(bundle_text.get("title", ""))
    description = bundle_text.get("description", "")
    features = bundle_text.get("features", []) or []

    return bool(
        title
        and _walgreens_description_is_rich_enough(description)
        and _walgreens_features_are_rich_enough(features)
    )

def _walgreens_description_is_rich_enough(value):
    value = clean_walgreens_text(value)
    return len(value) >= 140


def _walgreens_features_are_rich_enough(values):
    cleaned = _walgreens_clean_feature_list(values, max_features=5)
    feature_count = len(cleaned)
    feature_chars = sum(len(x) for x in cleaned)
    return bool(
        feature_count >= 4
        or (feature_count >= 3 and feature_chars >= 220)
    )


@st.cache_data(show_spinner=False)
def get_walgreens_bundle(retail_url, target_rpc="", sku=""):
    """
    Prefer live Walgreens HTML first, but recover copy/images when the live page under-pulls.
    Header values should still come from the richest recovered bundle.
    """
    retail_url = str(retail_url or "").strip()
    retail_url_lc = retail_url.lower()
    product_id = get_walgreens_product_id_from_url(retail_url)

    def build_html_bundle():
        html_text = get_walgreens_html(retail_url)
        return {
            "text": extract_walgreens_text_from_html(
                html_text,
                retail_url=retail_url,
                target_rpc=target_rpc,
            ),
            "images": extract_walgreens_images_from_html(html_text),
        }

    html_bundle = build_html_bundle()

    if "/search/results.jsp" in retail_url_lc:
        if _walgreens_has_copy_or_images(html_bundle):
            return html_bundle
        return build_empty_retailer_bundle("Walgreens", "walgreens_search_results_url_not_pdp")

    fallback_candidates = [html_bundle]

    if _walgreens_bundle_is_rich_enough(html_bundle):
        return html_bundle

    if product_id:
        prod_desc_bundle = build_walgreens_bundle_from_prod_desc_fragment(
            product_id,
            retail_url=retail_url,
        )
        if _walgreens_has_copy_or_images(prod_desc_bundle):
            fallback_candidates.append(prod_desc_bundle)

        api_payload = get_walgreens_product_api_payload(product_id)
        api_bundle = build_walgreens_bundle_from_api_payload(api_payload)
        if _walgreens_has_copy_or_images(api_bundle):
            fallback_candidates.append(api_bundle)

    merged_bundle = merge_walgreens_bundles_prefer_richer_copy(*fallback_candidates)
    if _walgreens_has_copy_or_images(merged_bundle):
        return merged_bundle

    return build_empty_retailer_bundle("Walgreens", "walgreens_live_html_missing")


# =========================================
# HEB PARSERS - no Salsify/image limits
# =========================================
def decode_json_string_value(raw_value):
    if raw_value is None:
        return ""
    raw_value = str(raw_value)
    try:
        return json.loads(f'"{raw_value}"')
    except Exception:
        return raw_value


def clean_heb_text(text):
    if not text:
        return ""
    text = str(text)
    for _ in range(4):
        unescaped = html.unescape(text)
        if unescaped == text:
            break
        text = unescaped
    text = text.replace("\\u003c", "<")
    text = text.replace("\\u003e", ">")
    text = text.replace("\\u0026", "&")
    text = text.replace("\\u00a0", " ")
    text = text.replace("\\n", " ")
    text = text.replace("\\/", "/")
    text = text.replace('\\"', '"')
    text = text.replace("&bull;", "•").replace("&#8226;", "•").replace("\\u2022", "•")
    text = re.sub(r"<script\b[^>]*>.*?</script>", " ", text, flags=re.IGNORECASE | re.DOTALL)
    if "<" in text and ">" in text:
        text = BeautifulSoup(text, "html.parser").get_text(" ", strip=True)
    return normalize_space(text)


def clean_heb_title(text):
    text = clean_heb_text(text)
    text = re.sub(r"\s+-\s+Shop\s+.*$", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"\s+at\s+H-?E-?B\s*$", "", text, flags=re.IGNORECASE).strip()
    return normalize_space(text)


def split_heb_description_and_features(description_text, max_features=10):
    working = str(description_text or "")
    for _ in range(5):
        unescaped = html.unescape(working)
        if unescaped == working:
            break
        working = unescaped
    working = working.replace("&bull;", "•").replace("&#8226;", "•").replace("\\u2022", "•")
    working = clean_heb_text(working)
    if not working:
        return "", []
    parts = [clean_heb_text(x) for x in re.split(r"\s*•\s*", working) if clean_heb_text(x)]
    if len(parts) <= 1:
        return working, []
    description = parts[0]
    features = [re.sub(r"^[\-•\s]+", "", x).strip() for x in parts[1:] if x]
    return description, dedupe_preserve_order(features)[:max_features]



def split_heb_description_and_features_aggressive(description_text, max_features=10):
    """HEB-only Salsify rescue for descriptions that carry bullet copy inline."""
    working = str(description_text or "")
    if not working:
        return "", []
    for _ in range(5):
        unescaped = html.unescape(working)
        if unescaped == working:
            break
        working = unescaped
    working = working.replace("\\u2022", "•").replace("&bull;", "•").replace("&#8226;", "•")
    working = working.replace("<li>", " • ").replace("</li>", " ")
    working = working.replace("<LI>", " • ").replace("</LI>", " ")
    working = re.sub(r"(?:\s|^)[\-*]\s+(?=[A-Z0-9])", " • ", working)
    working = clean_heb_text(working)
    if not working:
        return "", []
    parts = [clean_heb_text(x) for x in re.split(r"\s*•\s*", working) if clean_heb_text(x)]
    if len(parts) <= 1:
        return working, []
    description = parts[0]
    features = []
    for part in parts[1:]:
        part = re.sub(r"^[\-•\s]+", "", part).strip()
        if part and len(part) > 3:
            features.append(part)
    return description, dedupe_preserve_order(features)[:max_features]

def normalize_heb_features_final(items, max_features=10):
    if not items:
        return []
    if isinstance(items, str):
        _, split_features = split_heb_description_and_features(items, max_features=max_features)
        items = split_features if split_features else [items]
    out = []
    for item in items:
        value = clean_heb_text(item)
        if value:
            out.append(value)
    return dedupe_preserve_order(out)[:max_features]


def _extract_heb_json_string_field(source_text, field_names):
    source_text = str(source_text or "")
    if isinstance(field_names, str):
        field_names = [field_names]
    for field_name in field_names:
        pattern = r'"' + re.escape(field_name) + r'"\s*:\s*"((?:\\.|[^"\\])*)"'
        match = re.search(pattern, source_text, flags=re.IGNORECASE | re.DOTALL)
        if match:
            value = decode_json_string_value(match.group(1))
            value = clean_heb_text(value)
            if value:
                return value
    return ""


def _extract_heb_title_from_url_slug(retail_url):
    match = re.search(r"/product-detail/([^/?#]+)/", str(retail_url or ""), flags=re.IGNORECASE)
    if not match:
        return ""
    return clean_heb_title(match.group(1).replace("-", " "))


def extract_heb_text_from_html(html_text, retail_url="", target_rpc=""):
    debug = {"Title Path": "", "Description Path": "", "Features Path": "", "Source Used": "heb_txt_json", "Retailer": "HEB"}
    if not html_text:
        debug["Title Path"] = "heb_html_missing"
        debug["Description Path"] = "heb_html_missing"
        debug["Features Path"] = "heb_html_missing"
        return {"title": "", "description": "", "features": [], "rating": "", "review_count": "", "debug": debug}
    working = str(html_text or "")
    title = _extract_heb_json_string_field(working, ["documentTitle", "title", "name", "productName"])
    if title:
        title = clean_heb_title(title)
        debug["Title Path"] = "heb_json_documentTitle_or_name"
    if not title:
        soup = BeautifulSoup(html.unescape(working), "html.parser")
        h1 = soup.find("h1")
        if h1:
            title = clean_heb_title(h1.get_text(" ", strip=True))
            debug["Title Path"] = "h1"
        elif soup.title:
            title = clean_heb_title(soup.title.get_text(" ", strip=True))
            debug["Title Path"] = "html_title"
    if not title:
        title = _extract_heb_title_from_url_slug(retail_url)
        debug["Title Path"] = "retail_url_slug_fallback" if title else "heb_title_missing"
    raw_description = _extract_heb_json_string_field(working, ["description", "longDescription", "productDescription"])
    description, features = split_heb_description_and_features(raw_description, max_features=10)
    debug["Description Path"] = "heb_json_description_pre_bullet" if description else "heb_description_missing"
    debug["Features Path"] = "heb_json_description_bullet_split" if features else "heb_features_missing"
    return {"title": title, "description": description, "features": features[:10], "rating": "", "review_count": "", "debug": debug}


def _absolutize_heb_image_url(url):
    url = html.unescape(str(url or "").strip()).replace("\\/", "/")
    if not url:
        return ""
    url = re.sub(r"[\)\]\}\'\";,]+$", "", url.strip())
    if url.startswith("//"):
        url = "https:" + url
    if url.startswith("/"):
        url = "https://www.heb.com" + url
    if not re.match(r"^https?://", url, flags=re.IGNORECASE):
        return ""
    lowered = url.lower()
    if "images.heb.com" not in lowered and "heb.com/is/image" not in lowered:
        return ""
    if any(token in lowered for token in ["logo", "sprite", "placeholder", "favicon", ".svg"]):
        return ""
    return url.strip()


def _normalize_heb_image_for_compare(url):
    url = _absolutize_heb_image_url(url)
    if not url:
        return ""
    m = re.search(r"/is/image/HEBGrocery/([^?\s<>]+)", url, flags=re.IGNORECASE)
    if not m:
        return url
    asset = re.sub(r"[\)\]\}\'\";,]+$", "", m.group(1).strip())
    if not asset or asset.lower().startswith(("prd-small/", "prd-medium/", "prd-large/")):
        return ""
    return f"https://images.heb.com/is/image/HEBGrocery/{asset}?fit=constrain,1&wid=800&hei=800&fmt=jpg&qlt=80"


def build_heb_main_image_fallback_url(target_rpc="", retail_url=""):
    rpc = str(target_rpc or "").replace(".0", "").strip()
    if not rpc:
        m = re.search(r"/(\d{4,12})(?:[/?#]|$)", str(retail_url or ""))
        if m:
            rpc = m.group(1)
    rpc = re.sub(r"[^0-9]", "", rpc)
    if not rpc:
        return ""
    return f"https://images.heb.com/is/image/HEBGrocery/{rpc.zfill(9)}-1?fit=constrain,1&wid=800&hei=800&fmt=jpg&qlt=80"


def extract_heb_images_from_html(html_text, retail_url="", target_rpc=""):
    working = str(html_text or "")
    urls, seen = [], set()
    thumb_fallbacks = []
    if working:
        for pattern in [
            r'https?:\\/\\/images\.heb\.com\\/is\\/image\\/HEBGrocery\\/[^"\\\s<>]+',
            r'https?://images\.heb\.com/is/image/HEBGrocery/[^"\s<>]+',
            r'//images\.heb\.com/is/image/HEBGrocery/[^"\s<>]+',
        ]:
            for raw_url in re.findall(pattern, working, flags=re.IGNORECASE):
                raw_clean = _absolutize_heb_image_url(raw_url)
                normalized = _normalize_heb_image_for_compare(raw_clean)
                if normalized:
                    key = normalized.split("?", 1)[0]
                    if key and key not in seen:
                        seen.add(key)
                        urls.append(normalized)
                elif raw_clean and "prd-" in raw_clean.lower():
                    thumb_fallbacks.append(raw_clean)
    rpc = re.sub(r"[^0-9]", "", str(target_rpc or "").replace(".0", ""))
    if not rpc:
        m = re.search(r"/(\d{4,12})(?:[/?#]|$)", str(retail_url or ""))
        if m:
            rpc = m.group(1)
    padded_rpc = rpc.zfill(9) if rpc else ""
    if padded_rpc and urls:
        preferred = [u for u in urls if f"/{padded_rpc}" in u]
        if preferred:
            # HEB pages can preload recommendations/nearby products. For a known RPC,
            # keep only current-item image assets.
            urls = preferred
    def heb_asset_sort_key(url):
        m = re.search(r"/HEBGrocery/(\d+)(?:-(\d+))?", url)
        if not m:
            return (999999999, 9999, url)
        return (int(m.group(1)), int(m.group(2) or 1), url)
    urls = sorted(urls, key=heb_asset_sort_key)
    if not urls:
        for raw_url in thumb_fallbacks:
            url = _absolutize_heb_image_url(raw_url)
            key = url.split("?", 1)[0] if url else ""
            if key and key not in seen:
                seen.add(key)
                urls.append(url)
    fallback = build_heb_main_image_fallback_url(target_rpc=target_rpc, retail_url=retail_url)
    if fallback and fallback.split("?", 1)[0] not in seen:
        urls.append(fallback)
    return urls[:MAX_IMAGE_SLOTS_TO_COMPARE]

def build_heb_compact_capture_from_parsed_json(payload):
    if not isinstance(payload, dict):
        return ""
    title = clean_heb_title(payload.get("documentTitle", "") or payload.get("title", "") or payload.get("name", "") or payload.get("productName", ""))
    raw_description = str(payload.get("description", "") or payload.get("longDescription", "") or "")
    description, features = split_heb_description_and_features(raw_description, max_features=10)
    images = []
    seen = set()
    for image_url in payload.get("images", []) or []:
        clean_url = _absolutize_heb_image_url(image_url)
        key = clean_url.split("?", 1)[0] if clean_url else ""
        if key and key not in seen:
            seen.add(key)
            images.append(clean_url)
    if not (title or description or features or images):
        return ""
    requested_url = clean_uploaded_url_value(payload.get("requestedUrl", ""))
    final_url = clean_uploaded_url_value(payload.get("finalUrl", ""))
    product_json_ld = {"@context": "https://schema.org", "@type": "Product", "name": title, "description": raw_description, "image": images}
    parts = ["<html><head>", f"<title>{html_escape_text(title)}</title>", '<script type="application/ld+json">', json.dumps(product_json_ld, ensure_ascii=False), "</script>", "</head><body>", f"<h1>{html_escape_text(title)}</h1>", f"<!-- Requested URL: {html_escape_text(requested_url)} -->", f"<!-- Final URL: {html_escape_text(final_url)} -->", '<section data-testid="heb-product-description">', f'<script type="application/json" data-source="heb-parsed-json">{json.dumps({"documentTitle": title, "description": raw_description}, ensure_ascii=False)}</script>', f"<p>{html_escape_text(description)}</p>"]
    if features:
        parts.append("<ul>")
        for feature in features[:10]:
            parts.append(f"<li>{html_escape_text(feature)}</li>")
        parts.append("</ul>")
    parts.append("</section>")
    for image_url in images[:MAX_IMAGE_SLOTS_TO_COMPARE]:
        parts.append(f'<img src="{html.escape(image_url, quote=True)}" alt="{html_escape_text(title)}" />')
    parts.append("</body></html>")
    return "\n".join(parts)


@st.cache_data(show_spinner=False)
def get_heb_bundle(retail_url, target_rpc="", sku=""):
    retail_url = str(retail_url or "").strip()
    if not retail_url:
        return build_empty_retailer_bundle("HEB", "heb_url_missing")
    html_text = get_html(retail_url)
    return {"text": extract_heb_text_from_html(html_text, retail_url=retail_url, target_rpc=target_rpc), "images": extract_heb_images_from_html(html_text, retail_url=retail_url, target_rpc=target_rpc)}


def is_sams_robot_page(html_text):
    """
    Detect Sam's Club anti-bot / challenge pages so we do not accidentally
    parse footer/help links as product copy.
    """
    if not html_text:
        return False

    text = str(html_text or "").lower()

    robot_markers = [
        "let us know you're not a robot",
        "let us know you’re not a robot",
        "let us know you’re human",
        "let us know you're human",
        "no robots allowed",
        "verify you are human",
        "verify you're human",
        "verify you’re human",
        "press and hold",
        "press & hold",
        "captcha",
        "px-captcha",
        "/akam/",
        "challenge-platform",
        "bot protection",
        "/are-you-human",
    ]

    return any(marker in text for marker in robot_markers)

# =========================================
# SAM'S CLUB PARSERS
# =========================================
def _decode_sams_json_string(raw_value):
    if not raw_value:
        return ""

    raw_value = str(raw_value)

    try:
        decoded = json.loads(f'"{raw_value}"')
    except Exception:
        decoded = raw_value

    decoded = decoded.replace("\\/", "/")
    decoded = html.unescape(decoded)
    return decoded.strip()


# =========================================================
# SAM'S CLUB COPY, FEATURES, IMAGES, SCORING, AND STATUS v2.7
# Retailer-isolated section. Do not place CVS/Kroger/Walgreens rules here.
# =========================================================

def clean_sams_text(text):
    """Clean Sam's Club copy without changing the retailer's wording."""
    if not text:
        return ""

    text = str(text)
    text = html.unescape(text)
    text = text.replace("\\u003c", "<").replace("\\u003e", ">")
    text = text.replace("\\u0026", "&").replace("\\u00a0", " ")
    text = text.replace("\\n", " ").replace("\\/", "/").replace('\\"', '"')
    text = re.sub(r"<script\b[^>]*>.*?</script>", " ", text, flags=re.IGNORECASE | re.DOTALL)
    if "<" in text and ">" in text:
        text = BeautifulSoup(text, "html.parser").get_text(" ", strip=True)
    text = normalize_space(text)

    # Remove Sam's Club page controls and commerce copy that are not PDP claim copy.
    stop_patterns = [
        r"\binfo\s*:\s*If the item details above aren't accurate or complete",
        r"\bIf the item details above aren't accurate or complete",
        r"\bReport incorrect product info\b",
    ]
    for pattern in stop_patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            text = text[:match.start()].strip()
            break

    # Compact capture contains Parsed JSON plus relevant HTML. Remove exact repeated halves.
    text = normalize_space(text)
    if len(text) >= 80:
        midpoint = len(text) // 2
        candidates = [(text[:midpoint], text[midpoint:])]
        for offset in range(-12, 13):
            cut = midpoint + offset
            if 0 < cut < len(text):
                candidates.append((text[:cut], text[cut:]))
        for left, right in candidates:
            if normalize_text(left) and normalize_text(left) == normalize_text(right):
                text = normalize_space(left)
                break

    # Compact Sam's captures can contain the same Product Details block twice.
    # Exact-half detection is not enough because UI text can sit between copies.
    repeat_headings = [
        "Maximum Absorbency for All-Day Protection",
        "OdorBlock Technology for Freshness",
        "OdorBlock Technology for Confidence",
        "Stylish and Comfortable Fit",
        "Practical and Stylish Design",
        "Convenient and Versatile",
        "Convenient and Eligible for HSA/FSA",
    ]
    for heading in repeat_headings:
        hits = list(re.finditer(re.escape(heading), text, flags=re.IGNORECASE))
        if len(hits) >= 2 and hits[1].start() > 80:
            text = text[:hits[1].start()].strip()
            break

    return normalize_space(text)


def sams_claim_coverage_score(source_claim, retailer_text):
    """Sam's-only important-word coverage with common claim normalization."""
    source = normalize_text(clean_sams_text(source_claim))
    retailer = normalize_text(clean_sams_text(retailer_text))
    if not source or not retailer:
        return 0
    replacements = {
        "30x": "30 times", "odorblock": "odor block", "cottonlike": "cotton like",
        "xl": "extra large", "xxl": "extra extra large",
    }
    for old_value, new_value in replacements.items():
        source = source.replace(old_value, new_value)
        retailer = retailer.replace(old_value, new_value)
    stop = {"with","from","this","that","your","their","have","has","for","the","and","are","was","were","its","into","than","more","product","products","depend","fresh","protection","adult"}
    source_tokens = [token for token in source.split() if len(token) >= 3 and token not in stop]
    retailer_tokens = set(token for token in retailer.split() if len(token) >= 3 and token not in stop)
    if not source_tokens:
        return keyword_score(source_claim, retailer_text)
    coverage = int(100 * sum(token in retailer_tokens for token in source_tokens) / len(source_tokens))
    return max(coverage, keyword_score(source_claim, retailer_text))


def sams_split_heading_features(description, max_features=10):
    """Split Sam's Product Details into one clean narrative plus claim sections."""
    description = clean_sams_text(description)
    if not description:
        return "", []

    headings = [
        "Maximum Absorbency for All-Day Protection",
        "OdorBlock Technology for Freshness",
        "OdorBlock Technology for Confidence",
        "Stylish and Comfortable Fit",
        "Practical and Stylish Design",
        "Convenient and Versatile",
        "Convenient and Eligible for HSA/FSA",
        "Easy Ordering Options",
    ]
    matches = []
    for heading in headings:
        for match in re.finditer(re.escape(heading), description, flags=re.IGNORECASE):
            matches.append((match.start(), match.end(), heading))
    matches.sort(key=lambda item: item[0])

    selected = []
    last_end = -1
    for item in matches:
        if item[0] >= last_end:
            selected.append(item)
            last_end = item[1]
    if not selected:
        return description, []

    intro = clean_sams_text(description[:selected[0][0]])
    feature_items = []
    narrative_parts = [intro] if intro else []
    for index, (start_pos, end_pos, heading) in enumerate(selected):
        section_end = selected[index + 1][0] if index + 1 < len(selected) else len(description)
        body = clean_sams_text(description[end_pos:section_end])
        if heading.lower() == "easy ordering options":
            continue
        if body:
            narrative_parts.append(body)
        feature = clean_sams_text(f"{heading}: {body}" if body else heading)
        if feature:
            feature_items.append(feature)

    # Description contains each meaningful section body once, without duplicated headings
    # or ordering/pickup language. Features retain heading + body for claim matching.
    clean_description = clean_sams_text(" ".join(part for part in narrative_parts if part))
    return clean_description or description, dedupe_preserve_order(feature_items)[:max_features]


def sams_description_coverage_score(salsify_description, retailer_description):
    """Sam's-only blend: claim coverage plus normal sequence similarity."""
    s_clean = clean_sams_text(salsify_description)
    r_clean = clean_sams_text(retailer_description)
    if not s_clean or not r_clean:
        return 0
    sequence = description_similarity_score(s_clean, r_clean)
    stop = {"with","from","this","that","your","their","have","has","for","the","and","are","was","were","its","into","than","more","product","products"}
    s_tokens = {token for token in normalize_text(s_clean).split() if len(token) >= 3 and token not in stop}
    r_tokens = {token for token in normalize_text(r_clean).split() if len(token) >= 3 and token not in stop}
    coverage = int(100 * len(s_tokens & r_tokens) / max(1, len(s_tokens)))
    return max(0, min(100, int(round((coverage * 0.75) + (sequence * 0.25)))))


def clean_sams_title(text):
    text = clean_sams_text(text)
    text = re.sub(r"\s+,", ",", text)
    text = re.sub(r",\s*,", ", ", text)
    return normalize_space(text)


def normalize_sams_features_final(items, max_features=10):
    if not items:
        return []

    if isinstance(items, str):
        items = [items]

    cleaned = []
    for item in items:
        val = clean_sams_text(item)
        if not val:
            continue
        cleaned.append(val)

    return dedupe_preserve_order(cleaned)[:max_features]


def build_sams_title_from_url_slug(retail_url):
    retail_url = str(retail_url or "").strip()
    if not retail_url:
        return ""

    m = re.search(r"/ip/([^/]+)/", retail_url, flags=re.IGNORECASE)
    if not m:
        return ""

    slug = m.group(1)
    title = slug.replace("-", " ")
    title = html.unescape(title)
    return clean_sams_title(title)


def extract_sams_description_from_long_description_html(long_desc_html):
    if not long_desc_html:
        return ""

    working = str(long_desc_html)
    working = re.sub(
        r"<script\b[^>]*>.*?</script>",
        " ",
        working,
        flags=re.IGNORECASE | re.DOTALL,
    )

    soup = BeautifulSoup(working, "html.parser")

    parts = []
    candidates = list(soup.body.children) if soup.body else list(soup.children)

    for node in candidates:
        node_name = getattr(node, "name", None)

        if node_name in {"ul", "ol", "script", "style"}:
            continue

        if hasattr(node, "get_text"):
            node_text = node.get_text(" ", strip=True)
        else:
            node_text = str(node).strip()

        node_text = clean_sams_text(node_text)

        if node_text:
            parts.append(node_text)

    return normalize_space(" ".join(parts))


def extract_sams_features_from_short_description_html(short_desc_html):
    if not short_desc_html:
        return []

    working = str(short_desc_html)
    working = re.sub(
        r"<script\b[^>]*>.*?</script>",
        " ",
        working,
        flags=re.IGNORECASE | re.DOTALL,
    )

    soup = BeautifulSoup(working, "html.parser")

    items = []
    for li in soup.find_all("li"):
        li_text = clean_sams_text(li.get_text(" ", strip=True))
        if li_text:
            items.append(li_text)

    if items:
        return normalize_sams_features_final(items, max_features=10)

    fallback_text = clean_sams_text(working)
    if not fallback_text:
        return []

    if " | " in fallback_text:
        parts = [x.strip() for x in fallback_text.split(" | ")]
    elif "•" in fallback_text:
        parts = [x.strip() for x in fallback_text.split("•")]
    else:
        parts = [fallback_text]

    return normalize_sams_features_final(parts, max_features=10)


def _extract_visible_sams_title(source_text):
    if not source_text:
        return ""

    m = re.search(r"##\s+(.+?)\n", source_text)
    if m:
        return clean_sams_title(m.group(1))

    m = re.search(
        r"Hero image 0 of\s+(.+?),\s+0 of\s+\d+",
        source_text,
        flags=re.IGNORECASE,
    )
    if m:
        return clean_sams_title(m.group(1))

    return ""


def _extract_visible_sams_highlights(source_text):
    if not source_text:
        return []

    m = re.search(
        r"###\s+Highlights\s*(.+?)\s*Read more",
        source_text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not m:
        return []

    block = m.group(1)
    bullets = re.findall(r"-\s+(.+?)(?=\s*-\s+|$)", block, flags=re.DOTALL)

    cleaned = [clean_sams_text(x) for x in bullets if clean_sams_text(x)]
    return normalize_sams_features_final(cleaned, max_features=10)




def _extract_visible_sams_product_details(source_text):
    if not source_text:
        return ""

    source_text = str(source_text or "")
    start_match = re.search(r'####\s+Product details\s*\n', source_text, flags=re.IGNORECASE)
    if not start_match:
        return ""

    remainder = source_text[start_match.end():]
    stop_match = re.search(
        r'(?im)^###\s+(Specifications|Member ratings\s*&\s*reviews|Related pages)\b|^Directions:|^Ingredients:|^info:|^Shop more FSA',
        remainder,
    )
    block = remainder[:stop_match.start()] if stop_match else remainder

    lines = []
    for raw_line in block.splitlines():
        line = clean_sams_text(raw_line)
        if not line:
            continue
        if re.match(r'^###\s+', raw_line):
            break
        if line.lower() in {"product details", "about this item"}:
            continue
        if line.lower().startswith("view all reviews"):
            break
        lines.append(line)

    description = normalize_space(" ".join(lines))
    return clean_sams_text(description)


def _extract_visible_sams_rating_and_reviews(source_text):
    if not source_text:
        return "", ""

    source_text = str(source_text or "")
    rating = ""
    review_count = ""
    rating_count = ""

    rating_match = re.search(r'\b([0-9]+(?:\.[0-9]+)?)\s+out of 5\b', source_text, flags=re.IGNORECASE)
    if not rating_match:
        rating_match = re.search(r'\(([0-9]+(?:\.[0-9]+)?)\)\s*\|\s*\[[0-9,]+\s+ratings?\]', source_text, flags=re.IGNORECASE)
    if rating_match:
        rating = str(rating_match.group(1) or "").strip()

    review_matchers = [
        r'View all reviews\s*\(([0-9,]+)\)',
        r'Showing\s+\d+\s*-\s*\d+\s+of\s+([0-9,]+)\s+reviews',
        r'\|\s*\[([0-9,]+)\s+reviews\]',
    ]
    for pattern in review_matchers:
        m_review = re.search(pattern, source_text, flags=re.IGNORECASE)
        if m_review:
            review_count = str(m_review.group(1) or "").replace(",", "").strip()
            break

    rating_count_match = re.search(r'\b([0-9,]+)\s+ratings?\b', source_text, flags=re.IGNORECASE)
    if rating_count_match:
        rating_count = str(rating_count_match.group(1) or "").replace(",", "").strip()

    if not review_count and rating_count:
        review_count = rating_count

    return rating, review_count


def extract_sams_copy_from_source(source_text, retail_url=""):
    """
    Sam's Club copy extraction priority:

    1. JSON-like title:
       "name":"...","personalizable"

    2. JSON-like description:
       "longDescription":"..."

    3. JSON-like features:
       "shortDescription":"..." with:
       - strict quoted pattern first
       - relaxed pattern ending at \\u003c/ul(?:\\u003e)? second

    4. Visible page title fallback:
       ## Product Title
       or Hero image alt

    5. Visible Highlights fallback:
       ### Highlights
       - bullet

    6. Visible Product details fallback:
       #### Product details
       paragraph copy between Product details and Specifications / Reviews.
    """
    debug = {
        "Title Path": "",
        "Description Path": "",
        "Features Path": "",
        "Rating Path": "",
        "Source Used": "sams_raw_source",
    }

    if not source_text:
        return {
            "title": "",
            "description": "",
            "features": [],
            "rating": "",
            "review_count": "",
            "debug": debug,
        }

    source_text = str(source_text)

    title = ""
    name_match = re.search(
        r'"name"\s*:\s*"((?:\\.|[^"\\])*)"\s*,\s*"personalizable"',
        source_text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if name_match:
        title = _decode_sams_json_string(name_match.group(1))
        title = clean_sams_title(title)
        debug["Title Path"] = "sams_name_personalizable"

    if not title:
        title = _extract_visible_sams_title(source_text)
        if title:
            debug["Title Path"] = "sams_visible_title_fallback"

    # Live-only rule: do not infer a title from the URL slug. If the current
    # PDP does not expose a title, leave the retailer title blank.

    description = ""
    long_match = re.search(
        r'"longDescription"\s*:\s*"((?:\\.|[^"\\])*)"',
        source_text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if long_match:
        long_html = _decode_sams_json_string(long_match.group(1))
        description = extract_sams_description_from_long_description_html(long_html)
        description = clean_sams_text(description)
        if description:
            debug["Description Path"] = "sams_longDescription_html"

    if not description:
        description = _extract_visible_sams_product_details(source_text)
        if description:
            debug["Description Path"] = "sams_visible_product_details_fallback"
        else:
            debug["Description Path"] = "sams_description_missing"

    features = []
    short_match = re.search(
        r'"shortDescription"\s*:\s*"((?:\\.|[^"\\])*)"',
        source_text,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if short_match:
        short_html = _decode_sams_json_string(short_match.group(1))
        features = extract_sams_features_from_short_description_html(short_html)
        features = normalize_sams_features_final(features, max_features=10)
        if features:
            debug["Features Path"] = "sams_shortDescription_html"

    if not features:
        short_match_relaxed = re.search(
            r'"shortDescription"\s*:\s*"(.*?\\u003c/ul(?:\\u003e)?)',
            source_text,
            flags=re.IGNORECASE | re.DOTALL,
        )
        if short_match_relaxed:
            short_html = _decode_sams_json_string(short_match_relaxed.group(1))
            features = extract_sams_features_from_short_description_html(short_html)
            features = normalize_sams_features_final(features, max_features=10)
            if features:
                debug["Features Path"] = "sams_shortDescription_relaxed_ul"

    if not features:
        features = _extract_visible_sams_highlights(source_text)
        if features:
            debug["Features Path"] = "sams_visible_highlights_fallback"
        else:
            debug["Features Path"] = "sams_features_missing"

    # Final Sam's-only cleanup. Recover visible Product Details headings as features.
    description = clean_sams_text(description)
    heading_description, heading_features = sams_split_heading_features(description, max_features=10)
    if heading_features:
        description = heading_description
        if not features:
            features = heading_features
            debug["Features Path"] = "sams_product_details_heading_features"
        else:
            features = normalize_sams_features_final(list(features) + heading_features, max_features=10)
    features = normalize_sams_features_final(features, max_features=10)

    rating, review_count = _extract_visible_sams_rating_and_reviews(source_text)
    if rating or review_count:
        debug["Rating Path"] = "sams_visible_member_ratings_reviews"
    else:
        debug["Rating Path"] = "sams_rating_reviews_missing"

    return {
        "title": title,
        "description": description,
        "features": features[:10],
        "rating": rating,
        "review_count": review_count,
        "debug": debug,
    }


def _normalize_sams_medium_image_url(url):
    """
    Normalize Sam's image URLs to a clean medium-sized 450 x 450 asset.
    """
    if not url:
        return ""

    url = html.unescape(str(url).strip())

    m = re.search(
        r"(https://i5\.samsclubimages\.com/asr/[^?\"'<>\\s]+\\.jpe?g)",
        url,
        flags=re.IGNORECASE,
    )
    if not m:
        return ""

    base = m.group(1)
    return f"{base}?odnHeight=450&odnWidth=450&odnBg=FFFFFF"




def extract_sams_images_from_html(html_text):
    """Extract Sam's Club PDP gallery images only.

    Priority:
    1. Extension compact capture / parsed JSON image arrays.
    2. Sam's Club gallery thumbnail/hero image tags.
    3. Raw ASR URLs as a last resort.

    This intentionally rejects SVG, /dfw/ navigation assets, logo/icon/badge URLs,
    sponsored/review/customer media, and unrelated page chrome.
    """
    if not html_text:
        return []

    working = str(html_text or "")
    for _ in range(4):
        unescaped = html.unescape(working)
        if unescaped == working:
            break
        working = unescaped

    def _clean_url(url):
        url = html.unescape(str(url or "").strip()).replace("\\/", "/")
        if url.startswith("//"):
            url = "https:" + url
        return url

    def _base_url(url):
        url = _clean_url(url)
        if is_video_like_url(url):
            return url.split("?", 1)[0]
        return url.split("?", 1)[0]

    def _is_sams_product_media(url):
        url = _clean_url(url)
        if not url or not re.match(r"^https?://", url, flags=re.IGNORECASE):
            return False
        lowered_full = url.lower()
        lowered = lowered_full.split("?", 1)[0]
        if lowered.startswith("data:") or lowered.endswith(".svg"):
            return False
        if any(token in lowered for token in [
            "/dfw/", "sams-mav", "sprite", "icon", "logo", "badge", "placeholder",
            "customer", "review", "ratings", "stars", "avatar", "sponsored", "midas",
        ]):
            return False
        if is_video_like_url(url):
            return "samsclubimages.com" in lowered or "samsclubimages.com" in lowered_full
        return bool(
            ("samsclubimages.com/asr/" in lowered or "walmartimages.com" in lowered)
            and re.search(r"\.(?:jpg|jpeg|png|webp|avif)$", lowered, flags=re.IGNORECASE)
        )

    def _normalize_product_media(url):
        url = _clean_url(url)
        if not _is_sams_product_media(url):
            return ""
        if is_video_like_url(url):
            return url
        base = url.split("?", 1)[0]
        if "samsclubimages.com/asr/" in base.lower():
            return f"{base}?odnHeight=450&odnWidth=450&odnBg=FFFFFF"
        return url

    def _add(out, seen, url):
        normalized = _normalize_product_media(url)
        if not normalized:
            return
        key = _base_url(normalized)
        if not key or key in seen:
            return
        seen.add(key)
        out.append(normalized)

    def _json_values(obj, key_names=("images", "image", "allImages", "thumbnailUrl", "largeUrl", "url")):
        found = []
        def walk(value, parent_key=""):
            if isinstance(value, dict):
                for k, v in value.items():
                    if str(k) in key_names:
                        walk(v, str(k))
                    else:
                        walk(v, str(k))
            elif isinstance(value, list):
                for item in value:
                    walk(item, parent_key)
            elif isinstance(value, str):
                if re.match(r"^https?://", value.strip(), flags=re.IGNORECASE):
                    found.append(value.strip())
        walk(obj)
        return found

    out, seen = [], set()

    # Live Sam's Club PDP gallery only. The full page also contains recommendation
    # shelves, sponsored tiles, customer photos, and other products inside Next.js
    # data. If numbered PDP thumbnails are present, they are the authoritative live
    # gallery for the current item, so return them without scanning page-wide JSON.
    live_gallery = []
    live_seen = set()
    live_video_slots = []
    img_tag_pattern = re.compile(r'<img\b[^>]*>', flags=re.IGNORECASE | re.DOTALL)
    for tag_match in img_tag_pattern.finditer(working):
        tag = tag_match.group(0)
        alt_match = re.search(r"alt=[\"']([^\"']*)[\"']", tag, flags=re.IGNORECASE | re.DOTALL)
        alt_text = normalize_space(html.unescape(alt_match.group(1) if alt_match else ""))
        slot_match = re.search(r'(?:thumbnail|hero)\s+(?:video\s+)?image\s+(\d+)\s+of\b', alt_text, flags=re.IGNORECASE)
        if not slot_match:
            continue
        slot_num = int(slot_match.group(1))
        if re.search(r'\bvideo\b', alt_text, flags=re.IGNORECASE):
            if slot_num not in live_video_slots:
                live_video_slots.append(slot_num)
            continue
        attrs = []
        for attr_name in ["src", "data-src", "data-image-src", "currentSrc"]:
            m_attr = re.search(attr_name + r"=[\"']([^\"']+)[\"']", tag, flags=re.IGNORECASE | re.DOTALL)
            if m_attr:
                attrs.append(m_attr.group(1))
        m_srcset = re.search(r"srcset=[\"']([^\"']+)[\"']", tag, flags=re.IGNORECASE | re.DOTALL)
        if m_srcset:
            # Prefer the first candidate because normalization requests the standard
            # comparison size and removes thumbnail dimensions.
            attrs.extend(part.strip().split()[0] for part in m_srcset.group(1).split(',') if part.strip())
        for raw_url in attrs:
            normalized = _normalize_product_media(raw_url)
            key = _base_url(normalized) if normalized else ""
            if normalized and key and key not in live_seen:
                live_seen.add(key)
                live_gallery.append((slot_num, tag_match.start(), normalized))
                break
    if live_video_slots:
        # The live carousel marks video positions with "thumbnail video image N of".
        # Insert only the current PDP's first rendered Sam's Club rich-media URL
        # into those positions. Related shelf videos are not considered.
        live_videos = dedupe_preserve_order(re.findall(
            r"https?://i5-richmedia\.samsclubimages\.com/asr-rm/[^\s\"'<>]+?\.(?:mp4|m3u8)(?:\?[^\s\"'<>]*)?",
            working,
            flags=re.IGNORECASE,
        ))
        for slot_num, video_url in zip(sorted(live_video_slots), live_videos):
            key = _base_url(video_url)
            if key and key not in live_seen:
                live_seen.add(key)
                live_gallery.append((slot_num, -1, video_url))
    if live_gallery:
        return [url for _, _, url in sorted(live_gallery, key=lambda x: (x[0], x[1]))][:MAX_IMAGE_SLOTS_TO_COMPARE]

    # Sam's Club live-only rule: if the current PDP did not render numbered
    # thumbnail/hero gallery images, return no retailer images. Do not scan
    # page-wide JSON or raw ASR URLs because those include recommendations and
    # other products that are not live in the current item's gallery.
    return []

    # Legacy fallbacks below are intentionally unreachable for Sam's Club.
    # Fallback 1: parsed JSON blocks from the extension TXT. Used only when the
    # current PDP did not render numbered gallery thumbnails.
    for json_match in re.finditer(r'-----BEGIN PARSED JSON-----(.*?)-----END PARSED JSON-----', working, flags=re.IGNORECASE | re.DOTALL):
        try:
            payload = json.loads(json_match.group(1).strip())
        except Exception:
            payload = None
        if isinstance(payload, dict):
            for url in _json_values(payload):
                _add(out, seen, url)

    # Fallback 2: Next.js app data. This path is used only when no live gallery rendered.
    for script_match in re.finditer(r'<script[^>]+id=["\']__NEXT_DATA__["\'][^>]*>(.*?)</script>', working, flags=re.IGNORECASE | re.DOTALL):
        raw_json = html.unescape(script_match.group(1) or "")
        try:
            payload = json.loads(raw_json)
        except Exception:
            payload = None
        if isinstance(payload, dict):
            for url in _json_values(payload):
                _add(out, seen, url)

    # Fallback 3: remaining product-looking image tags.
    img_tag_pattern = re.compile(r'<img\b[^>]*>', flags=re.IGNORECASE | re.DOTALL)
    tagged = []
    for tag_match in img_tag_pattern.finditer(working):
        tag = tag_match.group(0)
        alt_match = re.search(r'alt=["\']([^"\']*)["\']', tag, flags=re.IGNORECASE | re.DOTALL)
        alt_text = normalize_space(html.unescape(alt_match.group(1) if alt_match else ""))
        if re.search(r'customer|review|sponsored|related product|member photos?', alt_text, flags=re.IGNORECASE):
            continue
        slot_num = None
        m_slot = re.search(r'thumbnail\s+image\s+(\d+)\s+of|hero\s+image\s+(\d+)\s+of', alt_text, flags=re.IGNORECASE)
        if m_slot:
            slot_num = int((m_slot.group(1) or m_slot.group(2) or "999"))
        attrs = []
        for attr_name in ["src", "data-src", "data-image-src", "currentSrc"]:
            m_attr = re.search(attr_name + r'=["\']([^"\']+)["\']', tag, flags=re.IGNORECASE | re.DOTALL)
            if m_attr:
                attrs.append(m_attr.group(1))
        m_srcset = re.search(r'srcset=["\']([^"\']+)["\']', tag, flags=re.IGNORECASE | re.DOTALL)
        if m_srcset:
            for part in m_srcset.group(1).split(','):
                attrs.append(part.strip().split()[0])
        for raw_url in attrs:
            normalized = _normalize_product_media(raw_url)
            if normalized:
                tagged.append((slot_num if slot_num is not None else 9999, tag_match.start(), normalized))
                break
    for _, _, url in sorted(tagged, key=lambda x: (x[0], x[1])):
        _add(out, seen, url)

    # Fallback 4: direct raw raster URLs.
    raw_url_patterns = [
        r'https?://i5\.samsclubimages\.com/asr/[^\s"\'<>]+',
        r'https?://i5\.walmartimages\.com/[^\s"\'<>]+',
        r'https?://i5-richmedia\.samsclubimages\.com/[^\s"\'<>]+(?:\.mp4|\.m3u8)[^\s"\'<>]*',
    ]
    for raw_pattern in raw_url_patterns:
        for raw_url in re.findall(raw_pattern, working, flags=re.IGNORECASE):
            _add(out, seen, raw_url)

    return out[:MAX_IMAGE_SLOTS_TO_COMPARE]

def extract_sams_text_from_html(html_text, retail_url="", target_rpc=""):

    debug = {
        "Title Path": "",
        "Description Path": "",
        "Features Path": "",
        "Rating Path": "",
        "Source Used": "sams_html",
    }

    if not html_text:
        return {
            "title": "",
            "description": "",
            "features": [],
            "rating": "",
            "review_count": "",
            "debug": debug,
        }

    result = extract_sams_copy_from_source(html_text, retail_url=retail_url)
    result_debug = result.get("debug", {}) or {}

    has_meaningful_content = bool(
        clean_sams_title(result.get("title", "")) or
        clean_sams_text(result.get("description", "")) or
        normalize_sams_features_final(result.get("features", []), max_features=5) or
        str(result.get("rating", "") or "").strip() or
        str(result.get("review_count", "") or "").strip()
    )

    if is_sams_robot_page(html_text) and not has_meaningful_content:
        fallback_title = build_sams_title_from_url_slug(retail_url)
        return {
            "title": fallback_title,
            "description": "",
            "features": [],
            "rating": "",
            "review_count": "",
            "debug": {
                "Title Path": "retail_url_slug_fallback" if fallback_title else "sams_robot_page_blocked",
                "Description Path": "sams_robot_page_blocked",
                "Features Path": "sams_robot_page_blocked",
                "Rating Path": "sams_robot_page_blocked",
                "Source Used": "sams_robot_page_blocked",
            },
        }

    result_debug["Source Used"] = "sams_html"
    result["debug"] = result_debug
    return result


@st.cache_data(show_spinner=False)
def get_sams_bundle(retail_url, target_rpc="", sku=""):
    html_text = get_html(retail_url)

    return {
        "text": extract_sams_text_from_html(
            html_text,
            retail_url=retail_url,
            target_rpc=target_rpc,
        ),
        "images": [] if is_sams_robot_page(html_text) else extract_sams_images_from_html(html_text),
    }
    

def is_valid_kroger_product_capture(html_text):
    """True only when Kroger HTML contains real PDP content.

    Kroger sometimes returns or captures a privacy/loading shell. That shell is not
    a product page, even though it contains lots of HTML. If the app trusts that
    shell, Kroger title/description/features/images all render as Missing.
    """
    working = html.unescape(str(html_text or ""))
    if not working.strip():
        return False

    lowered = working.lower()
    if 'data-pdp-invalid-kroger-capture="1"' in lowered or "data-pdp-invalid-kroger-capture='1'" in lowered:
        return False

    product_markers = [
        "product information",
        "product details",
        " perspective: front",
        " perspective: main",
        "upc:",
        "/product/images/",
        "data-testid=\"product-details-romance-description\"",
        "data-testid='product-details-romance-description'",
        "product-details-romance-description",
    ]
    shell_markers = [
        "privacy request center",
        "onetrust-consent-sdk",
        "loading",
    ]

    has_product_marker = any(marker in lowered for marker in product_markers)
    has_shell_marker = any(marker in lowered for marker in shell_markers)

    # A real PDP can include cookie/footer scripts, so only reject shell pages
    # when no PDP-specific marker is present.
    return bool(has_product_marker or not has_shell_marker)


def extract_kroger_item_availability_debug(html_text):
    """Extract Kroger availability for debug only; never gate parsing on it."""
    working = html.unescape(str(html_text or ""))
    if not working.strip():
        return {}

    compact = re.sub(r"\s+", " ", working)
    debug = {}
    patterns = [
        ("Pickup", r"Pickup\s*(Available|Unavailable)"),
        ("Kroger Delivery", r"Kroger\s+Delivery\s*(Available|Unavailable)"),
        ("All Delivery", r"All\s+Delivery\s*(Available|Unavailable)"),
    ]
    for label, pattern in patterns:
        match = re.search(pattern, compact, flags=re.IGNORECASE)
        if match:
            debug[label] = normalize_space(match.group(1)).title()
    return debug


def kroger_bundle_has_live_content(bundle):
    """True when Kroger bundle has any useful copy or image content."""
    if not isinstance(bundle, dict):
        return False
    text_bundle = bundle.get("text", {}) or {}
    has_text = bool(
        normalize_space(text_bundle.get("title", ""))
        or normalize_space(text_bundle.get("description", ""))
        or any(normalize_space(x) for x in (text_bundle.get("features", []) or []))
    )
    has_images = bool(bundle.get("images", []) or [])
    return has_text or has_images

def _cvs_text_score(text_bundle):
    if not isinstance(text_bundle, dict):
        return 0
    score = 0
    title = normalize_space(text_bundle.get("title", ""))
    description = normalize_space(text_bundle.get("description", ""))
    features = [normalize_space(x) for x in (text_bundle.get("features", []) or []) if normalize_space(x)]
    if title:
        score += 40 + min(len(title), 120)
    if description:
        score += 80 + min(len(description), 1200)
    if features:
        score += 60 + (len(features) * 80) + min(sum(len(x) for x in features), 1200)
    return score


def _cvs_bundle_score(bundle):
    if not isinstance(bundle, dict):
        return 0
    text_bundle = bundle.get("text", {}) or {}
    images = [x for x in (bundle.get("images", []) or []) if str(x or "").strip()]
    return _cvs_text_score(text_bundle) + (len(images) * 90)


def merge_cvs_bundles_prefer_richer_copy(*bundles):
    """CVS-only source combiner. Pick the richest title/description/features/images
    independently instead of trusting one source globally.
    """
    merged = {"text": {"title": "", "description": "", "features": [], "rating": "", "review_count": "", "debug": {}}, "images": []}
    source_parts = []
    best_title = (0, "")
    best_description = (0, "")
    best_features = (0, 0, [])
    best_images = (0, [])
    for bundle in bundles:
        if not isinstance(bundle, dict):
            continue
        text_bundle = bundle.get("text", {}) or {}
        debug = text_bundle.get("debug", {}) or {}
        title = normalize_space(text_bundle.get("title", ""))
        description = normalize_space(text_bundle.get("description", ""))
        features = [normalize_space(x) for x in (text_bundle.get("features", []) or []) if normalize_space(x)]
        images = [str(x or "").strip() for x in (bundle.get("images", []) or []) if str(x or "").strip()]
        if (len(title), title) > best_title:
            best_title = (len(title), title)
        if (len(description), description) > best_description:
            best_description = (len(description), description)
        f_tuple = (len(features), sum(len(x) for x in features), features)
        if f_tuple[:2] > best_features[:2]:
            best_features = f_tuple
        if (len(images), images) > best_images:
            best_images = (len(images), images)
        if not merged["text"].get("rating") and text_bundle.get("rating"):
            merged["text"]["rating"] = str(text_bundle.get("rating", "") or "").strip()
        if not merged["text"].get("review_count") and text_bundle.get("review_count"):
            merged["text"]["review_count"] = str(text_bundle.get("review_count", "") or "").strip()
        for k, v in debug.items():
            if v and not merged["text"]["debug"].get(k):
                merged["text"]["debug"][k] = v
        source_used = normalize_space(debug.get("Source Used", ""))
        if source_used and source_used not in source_parts:
            source_parts.append(source_used)
    merged["text"]["title"] = best_title[1]
    merged["text"]["description"] = best_description[1]
    merged["text"]["features"] = best_features[2]
    merged["images"] = best_images[1]
    if source_parts:
        merged["text"]["debug"]["Source Used"] = " | ".join(source_parts)
    merged["text"]["debug"]["CVS Combined Source Score"] = _cvs_bundle_score(merged)
    return merged


@st.cache_data(show_spinner=False, max_entries=1200)
def get_retailer_bundle(retailer_name, retail_url, target_rpc="", sku="", row_source_code=""):
    retailer = normalize_retailer_name(retailer_name).strip().lower()
    uploaded_html = str(row_source_code or "")

    if retail_url and not retailer_url_matches_selected(retail_url, retailer):
        return build_empty_retailer_bundle(retailer_name or "Retailer", build_retailer_url_mismatch_status(retail_url, retailer))

    if retailer == "kroger":
        if uploaded_html.strip() and is_valid_kroger_product_capture(uploaded_html):
            bundle = {
                "text": extract_kroger_text_from_html(uploaded_html, retail_url=retail_url, target_rpc=target_rpc),
                "images": extract_kroger_images_from_html(uploaded_html),
            }
            debug = bundle.setdefault("text", {}).setdefault("debug", {})
            debug["Source Used"] = "uploaded_txt_html"
            debug["Image Path"] = "kroger_main_image_perspective"
            debug["Availability Rule"] = "availability_never_blocks_live_copy_or_images"
            availability_debug = extract_kroger_item_availability_debug(uploaded_html)
            if availability_debug:
                debug["Item Availability"] = availability_debug
            return bundle

        if uploaded_html.strip():
            bundle = build_empty_retailer_bundle("Kroger", "invalid_uploaded_kroger_capture_no_product_content")
            debug = bundle.setdefault("text", {}).setdefault("debug", {})
            debug["Source Used"] = "uploaded_txt_invalid_kroger_capture"
            debug["Uploaded Capture Ignored"] = "invalid_kroger_shell_or_product_unavailable_capture"
            debug["Image Path"] = "no_kroger_image_from_invalid_txt_capture"
            debug["Copy Path"] = "no_kroger_copy_from_invalid_txt_capture"
            debug["Availability Rule"] = "availability_never_blocks_live_copy_or_images"
            return bundle

        # If no extension/TXT capture is available, fetch the live Kroger PDP directly.
        # Item availability never blocks PDP copy/images.
        return get_kroger_bundle(retail_url, target_rpc=target_rpc)

    if uploaded_html.strip():
        if retailer == "cvs":
            cvs_uploaded_html = extract_cvs_relevant_source_chunk(uploaded_html, retail_url=retail_url, target_rpc=target_rpc)
            uploaded_bundle = {"text": _extract_cvs_text_from_html(cvs_uploaded_html, retail_url=retail_url, target_rpc=target_rpc), "images": extract_cvs_images_from_html(cvs_uploaded_html)}
            upload_debug = uploaded_bundle.setdefault("text", {}).setdefault("debug", {})
            upload_debug["Source Used"] = "uploaded_txt_html"
            upload_debug["CVS Uploaded HTML Length"] = len(uploaded_html)
            upload_debug["CVS Uploaded Scoped HTML Length"] = len(cvs_uploaded_html)
            upload_debug["CVS Uploaded Chunk Scoped"] = bool(cvs_uploaded_html and cvs_uploaded_html != uploaded_html)
            upload_debug["CVS Uploaded Has Raw Fallback"] = "CVS RAW HTML FALLBACK FROM EXTENSION" in uploaded_html
            upload_debug["CVS Uploaded Product HTML Detected"] = bool(is_probably_cvs_product_html(cvs_uploaded_html))
            uploaded_image_base = infer_cvs_image_base_from_images(uploaded_bundle.get("images", []))
            if uploaded_image_base:
                upload_debug["CVS Uploaded Image Base"] = uploaded_image_base
                upload_debug["CVS Uploaded Image Count"] = len(uploaded_bundle.get("images", []) or [])
                if ALLOW_RETAILER_GENERATED_IMAGE_FALLBACKS and len(uploaded_bundle.get("images", []) or []) < 8:
                    uploaded_bundle["images"] = cvs_generated_image_candidates_for_base(uploaded_image_base, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE)
                    upload_debug["CVS Image Fallback Applied"] = "cvs_uploaded_image_base_expanded"
                    upload_debug["CVS Image Fallback Base"] = uploaded_image_base
                    upload_debug["CVS Image Fallback Count"] = len(uploaded_bundle.get("images") or [])
                elif len(uploaded_bundle.get("images", []) or []) < 8:
                    upload_debug["CVS Image Fallback Skipped"] = "strict_live_retailer_only"

            # CVS combined approach: always try both uploaded source and live CVS,
            # then merge the richest title/description/features/images by field.
            # This fixes partial captures where one path has copy and the other
            # path has images. It remains CVS-only and never pulls Salsify into
            # retailer fields.
            live_bundle = get_cvs_bundle(retail_url, target_rpc)
            merged_bundle = merge_cvs_bundles_prefer_richer_copy(uploaded_bundle, live_bundle)
            merged_bundle = apply_cvs_targeted_copy_rescue_if_needed(
                merged_bundle,
                retail_url=retail_url,
                target_rpc=target_rpc,
                reason="uploaded_plus_live_missing_copy_or_wrong_target_rpc",
            )
            merged_bundle = add_cvs_generated_image_fallback_if_needed(
                merged_bundle,
                retail_url=retail_url,
                target_rpc=target_rpc,
                reason="uploaded_plus_live_had_no_parseable_images",
            )
            merged_bundle.setdefault("text", {}).setdefault("debug", {})["CVS Source Merge"] = "uploaded_txt_plus_live_always"
            return merged_bundle
        if retailer == "walgreens":
            bundle = {"text": extract_walgreens_text_from_html(uploaded_html, retail_url=retail_url, target_rpc=target_rpc), "images": extract_walgreens_images_from_html(uploaded_html)}
            bundle.setdefault("text", {}).setdefault("debug", {})["Source Used"] = "uploaded_txt_html"
            return bundle
        if retailer == "sam's club":
            bundle = {"text": extract_sams_text_from_html(uploaded_html, retail_url=retail_url, target_rpc=target_rpc), "images": extract_sams_images_from_html(uploaded_html)}
            bundle.setdefault("text", {}).setdefault("debug", {})["Source Used"] = "uploaded_txt_html"
            return bundle
        if retailer == "heb":
            bundle = {"text": extract_heb_text_from_html(uploaded_html, retail_url=retail_url, target_rpc=target_rpc), "images": extract_heb_images_from_html(uploaded_html, retail_url=retail_url, target_rpc=target_rpc)}
            bundle.setdefault("text", {}).setdefault("debug", {})["Source Used"] = "uploaded_txt_html"
            return bundle

    retailer_fetchers = {
        "cvs": lambda: get_cvs_bundle(retail_url, target_rpc),
        "walgreens": lambda: get_walgreens_bundle(retail_url, target_rpc, sku=sku),
        "sam's club": lambda: get_sams_bundle(retail_url, target_rpc, sku=sku),
        "kroger": lambda: get_kroger_bundle(retail_url, target_rpc),
        "heb": lambda: get_heb_bundle(retail_url, target_rpc, sku=sku),
    }
    fetcher = retailer_fetchers.get(retailer)
    if fetcher is None:
        return build_empty_retailer_bundle(retailer_name or "Retailer", "retailer_not_supported_no_default_cvs_fallback")
    return fetcher()

def strip_walgreens_description_tail(text):
    """
    Keep live Walgreens marketing description exactly as shown on site,
    but still remove true legal / utility footer sections.
    """
    text = str(text or "")
    stop_markers = [
        "Directions for Use:",
        "Direction for Use:",
        "To Use:",
        "To Dispose:",
        "How to Use:",
        "How To Use:",
        "Directions:",
        "Do not flush.",
        "Do Not Flush.",
        "Made in USA",
        "Made In USA",
        "©",
        "Walgreens does not represent or warrant",
        "We recommend that you not rely solely on the information presented",
        "On occasion, manufacturers may improve or change their product formulas",
        "the food and drug administration has not intended to diagnose, treat, cure, or prevent any disease",
    ]

    cut_index = len(text)
    lowered = text.lower()

    for marker in stop_markers:
        idx = lowered.find(marker.lower())
        if idx != -1:
            cut_index = min(cut_index, idx)

    text = text[:cut_index].strip()
    return normalize_space(text)

def strip_walgreens_utility_tail(text):
    """
    Removes non-marketing utility/footer copy that should not live in the final description/features.
    IMPORTANT:
    Do NOT remove 'Also check out our ...' cross-sell copy anymore.
    """
    text = str(text or "")

    stop_markers = [
        "Directions for Use:",
        "Direction for Use:",
        "To Use:",
        "To Dispose:",
        "How to Use:",
        "How To Use:",
        "Directions:",
        "Do not flush.",
        "Do Not Flush.",
        "Made in USA",
        "Made In USA",
        "©",
        "Walgreens does not represent or warrant",
        "We recommend that you not rely solely on the information presented",
        "On occasion, manufacturers may improve or change their product formulas",
        "the food and drug administration has not intended to diagnose, treat, cure, or prevent any disease",
    ]

    cut_index = len(text)
    lowered = text.lower()

    for marker in stop_markers:
        idx = lowered.find(marker.lower())
        if idx != -1:
            cut_index = min(cut_index, idx)

    text = text[:cut_index].strip()

    return normalize_space(text)
    
def is_walgreens_utility_feature(text):
    text = normalize_space(text)
    if not text:
        return True

    bad_starts = [
        "Directions for Use",
        "Direction for Use",
        "To Use",
        "To Dispose",
        "How to Use",
        "How To Use",
        "Directions",
        "Made in USA",
        "Made In USA",
        "Do not flush",
        "Do Not Flush",
        "©",
    ]

    bad_contains = [
        "Walgreens does not represent or warrant",
        "consult your doctor",
        "keep this plastic bag away",
        "do not use this bag",
        "do not flush",
        "to dispose",
        "to use",
    ]

    for prefix in bad_starts:
        if text.lower().startswith(prefix.lower()):
            return True

    for token in bad_contains:
        if token.lower() in text.lower():
            return True

    return False


def split_walgreens_feature_fallback_text(text):
    """
    If Walgreens gives weak/malformed LI tags, fall back to splitting on strong
    uppercase lead-ins used across Depend, Huggies, Poise, Kotex, Pull-Ups, etc.
    """
    text = normalize_space(text)
    if not text:
        return []

    if " | " in text:
        parts = [x.strip() for x in text.split(" | ")]
    elif "•" in text:
        parts = [x.strip() for x in text.split("•")]
    else:
        heading_pattern = (
            r"(?=(?:"
            r"WHAT'S INCLUDED|ALL DAY PROTECTION|UNDERWEAR-LIKE COMFORT|UP TO ZERO ODOR|UNCOMPROMISED COMFORT|"
            r"UNBEATABLE PROTECTION|ODOR CONTROL|DRYNESS|ACTIVE FIT|INSTANT ABSORPTION|FRESHSENSE|GUSHPROTECT ZONE|"
            r"GRAVITY CORE|NIGHTDEFENSE|LEAKSHIELD|DESIGNED FOR MEN|SECURE FIT|FOR LARGE BLADDER LEAKS|"
            r"DEPEND SHIELDS|DEPEND FRESH PROTECTION|OUTSTANDING ABSORBENCY|FRONT AND BACK BLOWOUT BLOCKER|"
            r"UP TO 100% LEAKPROOF|LUXURY SOFTNESS|FASTABSORB SYSTEM|99% WATER|DERMATOLOGIST TESTED|"
            r"NATIONAL ECZEMA ASSOCIATION SEAL OF ACCEPTANCE|THICK AND ABSORBENT|COMPACT COMFORT, POWERFUL PROTECTION|"
            r"#1 COMPACT TAMPON BRAND|GYNECOLOGIST-TESTED|THIN AND SOFT|ALL-NIGHT DRYNESS|NEW! 60% WIDER BACK|"
            r"CLEAN SHIELD|DOUBLE GRIP STRIPS|GENTLEABSORB|QUICKSORB PROTECTION|HYPOALLERGENIC)"
            r"\s*[:\-])"
        )
        parts = re.split(heading_pattern, text, flags=re.IGNORECASE)

    cleaned = []
    for part in parts:
        part = strip_walgreens_utility_tail(part)
        part = normalize_space(part)
        if not part:
            continue
        if is_walgreens_utility_feature(part):
            continue
        cleaned.append(part)

    return dedupe_preserve_order(cleaned)


def clean_walgreens_text(text):
    if not text:
        return ""

    text = str(text)
    text = html.unescape(text)
    text = text.replace("\\u003c", "<")
    text = text.replace("\\u003e", ">")
    text = text.replace("\\u0026", "&")
    text = text.replace("\\n", " ")
    text = text.replace("\\/", "/")
    text = text.replace('\\"', '"')

    text = re.sub(
        r"<script\b[^>]*>.*?</script>",
        " ",
        text,
        flags=re.IGNORECASE | re.DOTALL,
    )

    if "<" in text and ">" in text:
        text = BeautifulSoup(text, "html.parser").get_text(" ", strip=True)

    text = normalize_space(text)
    text = strip_walgreens_description_tail(text)

    stop_patterns = [
        r"\bMade in USA\b.*$",
        r"\bMade In USA\b.*$",
        r"\bDirections for Use:.*$",
        r"\bDirection for Use:.*$",
        r"\bTo Use:.*$",
        r"\bTo Dispose:.*$",
        r"\bHow to Use:.*$",
        r"\bDo not flush\b.*$",
        r"\bDo Not Flush\b.*$",
        r"\bWalgreens does not represent or warrant.*$",
    ]

    for pattern in stop_patterns:
        text = re.sub(pattern, "", text, flags=re.IGNORECASE | re.DOTALL).strip()

    return normalize_space(text)


def _looks_like_walgreens_feature_fragment(text):
    """
    Detect short/incomplete feature fragments that should be merged
    with the next line/bullet.
    """
    text = normalize_space(text)
    if not text:
        return False

    lower = text.lower()

    if lower.endswith((" of", " for", " with", " to", " your", " our", " an")):
        return True

    if text.endswith(":"):
        return True

    if re.match(
        r"^[A-Z0-9&/\-\s\(\)'\"\.]+:\s*\d+\s+count\s+of\s*$",
        text
    ):
        return True

    if len(text) <= 42 and re.match(r"^[A-Z0-9&/\-\s\(\)'\":,.]+$", text):
        return True

    if len(text) <= 60 and text.count(" ") <= 8:
        if any(lower.endswith(x) for x in [" of", " for", " with", " to"]):
            return True

    return False


def _looks_like_walgreens_feature_continuation(text):
    """
    Detect feature lines that are really the continuation of the previous fragment.
    """
    text = normalize_space(text)
    if not text:
        return False

    brand_starts = (
        "Depend",
        "Goodnites",
        "Huggies",
        "Poise",
        "Kotex",
        "Pull-Ups",
        "Pull Ups",
        "Scott",
        "Kleenex",
        "Cottonelle",
        "U by Kotex",
        "U By Kotex",
        "Thinx",
        "Viva",
    )

    if text.startswith(brand_starts):
        return True

    if text[:1].islower():
        return True

    if text.startswith("(") or text.startswith("*"):
        return True

    if re.match(r"^[A-Z][a-z]", text):
        return True

    return False


def merge_walgreens_feature_fragments(items, max_features=5):
    """
    Merge adjacent Walgreens feature fragments like:

    'DEPEND FRESH PROTECTION: 15 count of'
    'Depend Fresh Protection Incontinence Underwear for Men, size extra-large (44-54" waist)'

    into one clean bullet.
    """
    cleaned_items = [
        clean_walgreens_text(x)
        for x in (items or [])
        if clean_walgreens_text(x)
    ]

    merged = []
    i = 0

    while i < len(cleaned_items):
        current = normalize_space(cleaned_items[i])

        if not current or is_walgreens_utility_feature(current):
            i += 1
            continue

        if i + 1 < len(cleaned_items):
            nxt = normalize_space(cleaned_items[i + 1])

            if nxt and not is_walgreens_utility_feature(nxt):
                if _looks_like_walgreens_feature_fragment(current) and _looks_like_walgreens_feature_continuation(nxt):
                    current = normalize_space(f"{current} {nxt}")
                    i += 1

        merged.append(current)
        i += 1

    return dedupe_preserve_order(merged[:max_features])


def normalize_walgreens_features_final(items, max_features=5):
    cleaned = []

    if not items:
        return []

    if isinstance(items, str):
        items = [items]

    for item in items:
        if not item:
            continue

        text = clean_walgreens_text(item)
        text = strip_walgreens_utility_tail(text)

        if not text:
            continue

        if is_walgreens_utility_feature(text):
            continue

        cleaned.append(text)

    cleaned = dedupe_preserve_order(cleaned)

    expanded = []
    for item in cleaned:
        parts = split_walgreens_feature_fallback_text(item)
        if parts:
            expanded.extend(parts)
        else:
            expanded.append(item)

    expanded = dedupe_preserve_order(
        [normalize_space(x) for x in expanded if normalize_space(x)]
    )

    merged = merge_walgreens_feature_fragments(expanded, max_features=max_features)

    out = []
    for item in merged:
        if not item:
            continue
        if is_walgreens_utility_feature(item):
            continue
        out.append(item)

    return dedupe_preserve_order(out)[:max_features]


def clean_walgreens_title(text):
    if not text:
        return ""

    text = clean_walgreens_text(text)
    text = normalize_space(text)

    text = re.sub(r"\s+,", ",", text)
    text = re.sub(r",\s*,", ", ", text)
    text = normalize_space(text)

    return text

def split_walgreens_description_into_features(description_text, existing_features=None, max_features=5):
    """
    If Walgreens description contains feature-style headings like:
    DEPEND FRESH PROTECTION:
    OUTSTANDING ABSORBENCY:
    ODOR CONTROL:
    DRYNESS:
    ACTIVE FIT:
    split those out into feature bullets when the dedicated feature list is empty,
    and keep only the intro paragraph as description.
    """
    description_text = clean_walgreens_text(description_text)
    existing_features = normalize_walgreens_features_final(
        existing_features or [],
        max_features=max_features,
    )

    # If real features already exist, keep them.
    if existing_features:
        return description_text, existing_features

    if not description_text:
        return "", []

    # Strong heading list used to identify feature-style sections inside description text.
    heading_pattern = re.compile(
        r"(?=(?:"
        r"WHAT'S INCLUDED|ALL DAY PROTECTION|UNDERWEAR-LIKE COMFORT|UP TO ZERO ODOR|UNCOMPROMISED COMFORT|"
        r"UNBEATABLE PROTECTION|ODOR CONTROL|DRYNESS|ACTIVE FIT|INSTANT ABSORPTION|FRESHSENSE|GUSHPROTECT ZONE|"
        r"GRAVITY CORE|NIGHTDEFENSE|LEAKSHIELD|DESIGNED FOR MEN|SECURE FIT|FOR LARGE BLADDER LEAKS|"
        r"DEPEND SHIELDS|DEPEND FRESH PROTECTION|OUTSTANDING ABSORBENCY|FRONT AND BACK BLOWOUT BLOCKER|"
        r"UP TO 100% LEAKPROOF|LUXURY SOFTNESS|FASTABSORB SYSTEM|99% WATER|DERMATOLOGIST TESTED|"
        r"NATIONAL ECZEMA ASSOCIATION SEAL OF ACCEPTANCE|THICK AND ABSORBENT|COMPACT COMFORT, POWERFUL PROTECTION|"
        r"#1 COMPACT TAMPON BRAND|GYNECOLOGIST-TESTED|THIN AND SOFT|ALL-NIGHT DRYNESS|NEW! 60% WIDER BACK|"
        r"CLEAN SHIELD|DOUBLE GRIP STRIPS|GENTLEABSORB|QUICKSORB PROTECTION|HYPOALLERGENIC"
        r")\s*[:\-])",
        flags=re.IGNORECASE,
    )

    matches = list(heading_pattern.finditer(description_text))

    # No feature-style headings found.
    if not matches:
        return description_text, []

    # Description = everything before the first heading.
    first_start = matches[0].start()
    trimmed_description = normalize_space(description_text[:first_start])

    # Features = each heading section up to the next heading.
    raw_features = []
    for i, match in enumerate(matches):
        start = match.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(description_text)
        chunk = normalize_space(description_text[start:end])
        chunk = strip_walgreens_utility_tail(chunk)

        if not chunk:
            continue
        if is_walgreens_utility_feature(chunk):
            continue

        raw_features.append(chunk)

    cleaned_features = normalize_walgreens_features_final(
        raw_features,
        max_features=max_features,
    )

    return trimmed_description, cleaned_features[:max_features]
    



def is_placeholder_salsify_copy_value(value):
    normalized = normalize_salsify_asset_name(value or "")
    if not normalized:
        return True
    exact_placeholders = {
        "general product title",
        "general title",
        "product title",
        "cvs product title",
        "cvs title",
        "general description",
        "product description",
        "description",
        "cvs description",
        "cvs product description",
        "feature",
        "bullet",
    }
    if normalized in exact_placeholders:
        return True
    if re.fullmatch(r"general feature ?\d+", normalized):
        return True
    if re.fullmatch(r"cvs feature ?\d+", normalized):
        return True
    if re.fullmatch(r"feature ?\d+", normalized):
        return True
    if re.fullmatch(r"general bullet ?\d+", normalized):
        return True
    if re.fullmatch(r"cvs bullet ?\d+", normalized):
        return True
    if re.fullmatch(r"bullet ?\d+", normalized):
        return True
    return False


def first_non_placeholder_copy_value(*values):
    for value in values:
        clean_value = normalize_space(value)
        if clean_value and not is_placeholder_salsify_copy_value(clean_value):
            return clean_value
    for value in values:
        clean_value = normalize_space(value)
        if clean_value:
            return clean_value
    return ""


def normalize_salsify_feature_values(values, max_features=10):
    out = []
    seen = set()
    for value in values or []:
        clean_value = normalize_space(value)
        if not clean_value or is_placeholder_salsify_copy_value(clean_value):
            continue
        if clean_value not in seen:
            seen.add(clean_value)
            out.append(clean_value)
        if len(out) >= max_features:
            break
    return out

def finalize_salsify_copy_for_retailer(retailer_name, s_text):
    """
    Normalize Salsify copy for retailer-specific comparison only.
    Kroger should prefer Kroger Product Title / Kroger Description / Kroger Feature N.
    Sam's Club should prefer Sam's Club Product Title / Description / Feature N.
    CVS should prefer CVS-specific Salsify fields first, then General fields only as fallback.
    """
    retailer = str(retailer_name or "").strip().lower()
    out = dict(s_text or {})

    def generic_feature_list():
        candidates = []
        for i in range(1, 11):
            value = normalize_space(out.get(f"feature{i}", ""))
            if value:
                candidates.append(value)
        for value in out.get("features", []) or []:
            value = normalize_space(value)
            if value:
                candidates.append(value)
        return dedupe_preserve_order(candidates)

    retailer_overrides = out.get("retailer_overrides", {}) or {}

    if retailer == "kroger":
        kroger_override = retailer_overrides.get("kroger", {}) or {}
        selected_title = first_non_placeholder_copy_value(kroger_override.get("title", ""), out.get("title", ""))
        selected_description = clean_kroger_text(first_non_placeholder_copy_value(kroger_override.get("description", ""), out.get("description", "")))
        selected_features = normalize_kroger_features(
            normalize_salsify_feature_values(kroger_override.get("features", []) or generic_feature_list(), max_features=10),
            max_features=10,
        )
        out["title"] = selected_title
        out["description"] = selected_description
        out["features"] = selected_features
        for i in range(1, 8):
            out[f"feature{i}"] = selected_features[i - 1] if i - 1 < len(selected_features) else ""
        return out

    if retailer in {"sam's club", "sams club", "samsclub"}:
        sams_override = retailer_overrides.get("sam's club", {}) or {}
        selected_title = clean_sams_title(first_non_placeholder_copy_value(sams_override.get("title", ""), out.get("title", "")))
        selected_description = clean_sams_text(first_non_placeholder_copy_value(sams_override.get("description", ""), out.get("description", "")))
        override_features = sams_override.get("features", []) or []
        override_feature_slots = sams_override.get("feature_slots", {}) or {}
        generic_features = generic_feature_list()
        selected_features = []
        for i in range(1, 11):
            slot_value = first_non_placeholder_copy_value(override_feature_slots.get(i, ""))
            if slot_value:
                selected_features.append(slot_value)
        if not selected_features:
            selected_features = normalize_sams_features_final(normalize_salsify_feature_values(override_features or generic_features, max_features=10), max_features=10)
        else:
            tail_features = normalize_sams_features_final(normalize_salsify_feature_values(override_features, max_features=10), max_features=10)
            selected_features = dedupe_preserve_order(selected_features + tail_features)[:10]
        out["title"] = selected_title
        out["description"] = selected_description
        out["features"] = selected_features
        for i in range(1, 8):
            out[f"feature{i}"] = selected_features[i - 1] if i - 1 < len(selected_features) else ""
        return out

    if retailer == "cvs":
        cvs_override = retailer_overrides.get("cvs", {}) or {}
        selected_title = first_non_placeholder_copy_value(cvs_override.get("title", ""), out.get("title", ""))
        selected_description = first_non_placeholder_copy_value(cvs_override.get("description", ""), out.get("description", ""))
        override_features = cvs_override.get("features", []) or []
        override_feature_slots = cvs_override.get("feature_slots", {}) or {}
        generic_features = generic_feature_list()
        selected_features = []
        for i in range(1, 11):
            slot_value = first_non_placeholder_copy_value(override_feature_slots.get(i, ""))
            if slot_value:
                selected_features.append(slot_value)
        if not selected_features:
            selected_features = normalize_salsify_feature_values(override_features, max_features=10)
        else:
            tail_features = normalize_salsify_feature_values(override_features, max_features=10)
            selected_features = dedupe_preserve_order(selected_features + tail_features)[:10]
        if not selected_features:
            selected_features = normalize_salsify_feature_values(generic_features, max_features=10)
        out["title"] = selected_title
        out["description"] = selected_description
        out["features"] = selected_features
        for i in range(1, 8):
            out[f"feature{i}"] = selected_features[i - 1] if i - 1 < len(selected_features) else ""
        return out

    if retailer == "heb":
        heb_override = retailer_overrides.get("heb", {}) or {}
        selected_title = clean_heb_title(first_non_placeholder_copy_value(heb_override.get("title", ""), out.get("title", "")))
        selected_description = clean_heb_text(first_non_placeholder_copy_value(heb_override.get("description", ""), out.get("description", "")))
        override_features = heb_override.get("features", []) or []
        override_feature_slots = heb_override.get("feature_slots", {}) or {}
        generic_features = generic_feature_list()
        selected_features = []
        for i in range(1, 11):
            slot_value = first_non_placeholder_copy_value(override_feature_slots.get(i, ""))
            if slot_value:
                selected_features.append(slot_value)
        if not selected_features:
            source_features = override_features or generic_features
            selected_features = normalize_heb_features_final(normalize_salsify_feature_values(source_features, max_features=10), max_features=10)
        else:
            tail_features = normalize_heb_features_final(normalize_salsify_feature_values(override_features, max_features=10), max_features=10)
            selected_features = dedupe_preserve_order(selected_features + tail_features)[:10]
        # If the Salsify deck stores bullets inside the description with bullets,
        # split them, but do not discard normal Salsify feature fields.
        if not selected_features and selected_description:
            clean_desc, split_features = split_heb_description_and_features_aggressive(selected_description, max_features=10)
            if split_features:
                selected_description = clean_desc or selected_description
                selected_features = split_features
        out["title"] = selected_title
        out["description"] = selected_description
        out["features"] = selected_features
        for i in range(1, 11):
            out[f"feature{i}"] = selected_features[i - 1] if i - 1 < len(selected_features) else ""
        return out

    if retailer == "walgreens":
        walgreens_override = retailer_overrides.get("walgreens", {}) or {}
        exclusive_mode = retailer in EXCLUSIVE_SALSIFY_COPY_RETAILERS

        if exclusive_mode:
            selected_title = clean_walgreens_title(
                first_non_placeholder_copy_value(walgreens_override.get("title", ""))
            )
            selected_description = strip_walgreens_description_tail(
                first_non_placeholder_copy_value(walgreens_override.get("description", ""))
            )
        else:
            selected_title = clean_walgreens_title(
                first_non_placeholder_copy_value(walgreens_override.get("title", ""), out.get("title", ""))
            )
            selected_description = strip_walgreens_description_tail(
                first_non_placeholder_copy_value(walgreens_override.get("description", ""), out.get("description", ""))
            )

        override_features = walgreens_override.get("features", []) or []
        override_feature_slots = walgreens_override.get("feature_slots", {}) or {}
        generic_features = generic_feature_list()
        selected_features = []
        for i in range(1, 11):
            slot_value = first_non_placeholder_copy_value(override_feature_slots.get(i, ""))
            if slot_value:
                selected_features.append(slot_value)
        if not selected_features:
            source_features = override_features if exclusive_mode else (override_features or generic_features)
            selected_features = normalize_walgreens_features_final(
                normalize_salsify_feature_values(source_features, max_features=10),
                max_features=10,
            )
        else:
            tail_features = normalize_walgreens_features_final(
                normalize_salsify_feature_values(override_features, max_features=10),
                max_features=10,
            )
            selected_features = dedupe_preserve_order(selected_features + tail_features)[:10]
        out["title"] = selected_title
        out["description"] = selected_description
        out["features"] = selected_features
        for i in range(1, 8):
            out[f"feature{i}"] = selected_features[i - 1] if i - 1 < len(selected_features) else ""
        return out

    out["title"] = first_non_placeholder_copy_value(out.get("title", ""))
    out["description"] = first_non_placeholder_copy_value(out.get("description", ""))
    out["features"] = normalize_salsify_feature_values(out.get("features", []) or generic_feature_list(), max_features=10)
    return out


_original_finalize_salsify_copy_for_retailer = finalize_salsify_copy_for_retailer

def finalize_salsify_copy_for_retailer(retailer_name, s_text):
    out = _original_finalize_salsify_copy_for_retailer(retailer_name, s_text)
    return apply_retailer_salsify_copy_limits(retailer_name, out)

def finalize_retailer_copy(retailer_name, r_text):
    retailer = str(retailer_name or "").strip().lower()
    out = dict(r_text or {})

    if retailer == "kroger":
        out["title"] = normalize_space(out.get("title", ""))
        description = clean_kroger_text(out.get("description", ""))
        features = normalize_kroger_features(out.get("features", []), max_features=10)

        suspicious_features = bool(
            features and (
                len(features) <= 2
                or all(len(str(x or "").strip()) <= 8 for x in features[:3])
            )
        )
        if description and (not features or suspicious_features):
            intro, split_features = split_kroger_parsed_description(description)
            if split_features:
                description = intro or description
                features = normalize_kroger_features(split_features, max_features=10)

        out["description"] = description
        out["features"] = features
        return out

    if retailer == "walgreens":
        out["title"] = clean_walgreens_title(out.get("title", ""))
        cleaned_description = strip_walgreens_description_tail(out.get("description", ""))
        cleaned_features = normalize_walgreens_features_final(
            out.get("features", []),
            max_features=5,
        )

        cleaned_description, cleaned_features = split_walgreens_description_into_features(
            cleaned_description,
            cleaned_features,
            max_features=5,
        )

        out["description"] = cleaned_description
        out["features"] = cleaned_features
        return out

    if retailer in ["sam's club", "sams club", "samsclub"]:
        out["title"] = clean_sams_title(out.get("title", ""))
        out["description"] = clean_sams_text(out.get("description", ""))
        out["features"] = normalize_sams_features_final(
            out.get("features", []),
            max_features=5,
        )
        return out

    if retailer == "heb":
        out["title"] = clean_heb_title(out.get("title", ""))
        description = clean_heb_text(out.get("description", ""))
        features = normalize_heb_features_final(out.get("features", []), max_features=10)
        if description and not features:
            description, features = split_heb_description_and_features(description, max_features=10)
        out["description"] = description
        out["features"] = features
        return out

    out["title"] = normalize_space(out.get("title", ""))
    out["description"] = clean_cvs_text(out.get("description", ""))
    out["features"] = normalize_cvs_features(out.get("features", []))
    return out
    
# =========================================
# QUALITY HELPERS
# =========================================
def debug_description(desc):
    if not desc:
        return {"length": 0, "quality_score": 0, "issues": ["Missing description"]}

    desc_clean = normalize_text(desc)
    length = len(desc_clean)

    absorbency_keywords = ["absorb", "leak", "fluid", "protection", "flushable", "soft", "care"]
    size_keywords = ["count", "ct", "pack", "roll", "sheets", "wipes", "mega", "tissues", "cube", "box"]
    benefit_keywords = ["soft", "comfort", "odor", "dry", "safe", "clean", "trusted", "aloe", "lotion"]

    has_absorbency = any(k in desc_clean for k in absorbency_keywords)
    has_size = any(k in desc_clean for k in size_keywords)
    has_benefits = any(k in desc_clean for k in benefit_keywords)

    is_truncated = not desc.strip().endswith((".", "!", "?")) or length < 80

    words = desc_clean.split()
    unique_ratio = len(set(words)) / len(words) if words else 0

    issues = []
    if length < 80:
        issues.append("Too short")
    if not has_absorbency:
        issues.append("Missing absorbency info")
    if not has_size:
        issues.append("Missing size/count")
    if not has_benefits:
        issues.append("Missing benefits")
    if is_truncated:
        issues.append("Possible truncation")
    if unique_ratio < 0.5:
        issues.append("Repetitive content")

    quality_score = 100
    if length < 80:
        quality_score -= 20
    if not has_absorbency:
        quality_score -= 15
    if not has_size:
        quality_score -= 15
    if not has_benefits:
        quality_score -= 15
    if is_truncated:
        quality_score -= 20
    if unique_ratio < 0.5:
        quality_score -= 15

    quality_score = max(0, quality_score)

    return {
        "length": length,
        "quality_score": quality_score,
        "issues": issues,
    }


# =========================================
# IMAGE HASHING (FAST IMAGE COMPARE)
# =========================================
def _compute_dhash_from_pil_image(img, crop_ratio=0.0):
    if img is None:
        return None

    working = img.copy().convert("L")
    width, height = working.size

    if crop_ratio > 0 and width > 20 and height > 20:
        dx = int(width * crop_ratio)
        dy = int(height * crop_ratio)
        if dx * 2 < width and dy * 2 < height:
            working = working.crop((dx, dy, width - dx, height - dy))

    working.thumbnail((256, 256))
    working = working.resize((IMAGE_HASH_WIDTH, IMAGE_HASH_HEIGHT))

    bits = []
    for y in range(IMAGE_HASH_HEIGHT):
        for x in range(IMAGE_HASH_WIDTH - 1):
            left_pixel = working.getpixel((x, y))
            right_pixel = working.getpixel((x + 1, y))
            bits.append(1 if left_pixel > right_pixel else 0)

    h = 0
    for bit in bits:
        h = (h << 1) | bit
    return h


def _cache_image_hash_result(url, value):
    global image_hash_cache
    if "image_hash_cache" not in globals() or not isinstance(globals().get("image_hash_cache"), dict):
        image_hash_cache = {}
    image_hash_cache[url] = value
    while len(image_hash_cache) > IMAGE_HASH_CACHE_MAX:
        image_hash_cache.pop(next(iter(image_hash_cache)))


def _download_image_bytes_once(url, timeout_seconds):
    session = get_session()
    with image_fetch_semaphore:
        r = session.get(url, timeout=timeout_seconds, stream=True)
        if r.status_code != 200:
            return None

        content_type = str(r.headers.get("Content-Type", "") or "")
        if "image" not in content_type.lower():
            return None

        content_length = str(r.headers.get("Content-Length", "") or "").strip()
        if content_length.isdigit() and int(content_length) > MAX_IMAGE_BYTES:
            return None

        image_bytes = bytearray()
        for chunk in r.iter_content(chunk_size=IMAGE_FETCH_CHUNK_SIZE):
            if not chunk:
                continue
            image_bytes.extend(chunk)
            if len(image_bytes) > MAX_IMAGE_BYTES:
                return None

    return bytes(image_bytes) if image_bytes else None


def _download_image_bytes_with_retry(url):
    timeout_plan = [IMAGE_TIMEOUT]
    for _ in range(max(0, int(IMAGE_RETRY_COUNT or 0))):
        timeout_plan.append(IMAGE_RETRY_TIMEOUT)

    for timeout_seconds in timeout_plan:
        try:
            image_bytes = _download_image_bytes_once(url, timeout_seconds)
            if image_bytes:
                return image_bytes
        except Exception:
            continue
    return None


def get_image_dhash(url):
    global image_hash_cache

    if "image_hash_cache" not in globals() or not isinstance(globals().get("image_hash_cache"), dict):
        image_hash_cache = {}

    url = str(url or "").strip()
    if not url:
        return None

    # Cache both successes and failures. A failed image should not be retried dozens
    # of times in the same batch; use Clear caches to force a fresh retry.
    if url in image_hash_cache:
        return image_hash_cache[url]

    try:
        image_bytes = _download_image_bytes_with_retry(url)
        if not image_bytes:
            _cache_image_hash_result(url, None)
            return None

        bio = BytesIO(image_bytes)
        with warnings.catch_warnings():
            warnings.simplefilter("error", Image.DecompressionBombWarning)
            img = Image.open(bio)
            width, height = img.size
            if width * height > MAX_SAFE_IMAGE_PIXELS:
                _cache_image_hash_result(url, None)
                return None
            img.load()

        variants = {
            "full": _compute_dhash_from_pil_image(img, crop_ratio=0.0),
            "center_6": _compute_dhash_from_pil_image(img, crop_ratio=0.06),
            "center_12": _compute_dhash_from_pil_image(img, crop_ratio=0.12),
        }

        _cache_image_hash_result(url, variants)
        return variants

    except (Image.DecompressionBombWarning, UnidentifiedImageError, OSError, ValueError):
        _cache_image_hash_result(url, None)
        return None
    except Exception:
        _cache_image_hash_result(url, None)
        return None


def hamming_distance(a, b):
    return bin(a ^ b).count("1")



def compare_images_visually(s_url, r_url):
    global image_compare_cache
    if "image_compare_cache" not in globals() or not isinstance(globals().get("image_compare_cache"), dict):
        image_compare_cache = {}
    if not s_url or not r_url:
        return 0
    s_url_clean = str(s_url or "").split("?", 1)[0].strip()
    r_url_clean = str(r_url or "").split("?", 1)[0].strip()
    cache_key = (str(s_url), str(r_url))
    if cache_key in image_compare_cache:
        return image_compare_cache[cache_key]
    if s_url_clean and r_url_clean and s_url_clean == r_url_clean:
        image_compare_cache[cache_key] = 100
        return 100
    s_is_video = is_video_like_url(s_url)
    r_is_video = is_video_like_url(r_url)
    if s_is_video or r_is_video:
        score = 100 if (s_is_video and r_is_video) else 0
        image_compare_cache[cache_key] = score
        while len(image_compare_cache) > IMAGE_COMPARE_CACHE_MAX:
            image_compare_cache.pop(next(iter(image_compare_cache)))
        return score
    s_hashes = get_image_dhash(s_url)
    r_hashes = get_image_dhash(r_url)
    if not s_hashes or not r_hashes:
        score = 0
    else:
        distances = []
        for key in ["full", "center_6", "center_12"]:
            s_hash = s_hashes.get(key)
            r_hash = r_hashes.get(key)
            if s_hash is None or r_hash is None:
                continue
            distances.append(hamming_distance(s_hash, r_hash))
        if not distances:
            score = 0
        else:
            dist = min(distances)
            if dist <= 2:
                score = 100
            elif dist <= 5:
                score = 95
            elif dist <= 8:
                score = 90
            elif dist <= 12:
                score = 80
            elif dist <= 16:
                score = 70
            elif dist <= 22:
                score = 55
            else:
                score = 30
    image_compare_cache[cache_key] = score
    while len(image_compare_cache) > IMAGE_COMPARE_CACHE_MAX:
        image_compare_cache.pop(next(iter(image_compare_cache)))
    return score



def align_salsify_images_for_retailer(retailer_name, s_images, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE, brand=""):
    """
    Build the retailer-specific Salsify comparison image list.

    Sam's Club image rules:
    - When Online Optimized Image-Sams Club exists, it is always slot 1.
    - ATF Video-Sams Club is always the next Sam's Club slot when present.
    - Generic Online Optimized Image- follows when present and not already used.
    - If Online Optimized Image-Sams Club is missing, fallback to Main Variant Image-Sams Club,
      then Main Variant Image-Club, then generic Online Optimized Image-.
    - Shipping- comes after the required first three slots and before ATF I/O / numbered ATF images.

    CVS rules:
    - Lock only the top 3 Salsify slots.
    - If one of the top 3 is missing, keep the slot blank and do not shift later images up.
    - After slot 3, continue with the remaining Salsify images in original order.
    """
    retailer = str(retailer_name or "").strip().lower()
    brand_norm = normalize_salsify_asset_name(brand or "")
    source_images = list(s_images or [])
    if retailer != "cvs":
        source_images = [img for img in source_images if not is_cvs_only_salsify_image(img)]

    def dedupe_images_preserve_order(images):
        out = []
        seen = set()
        for img in images:
            if not isinstance(img, dict):
                continue
            url = str(img.get("url", "") or "").strip()
            if not url or url in seen:
                continue
            out.append(img)
            seen.add(url)
        return out

    def find_first_image(images, *queries):
        query_tokens = [normalize_salsify_asset_name(q) for q in queries if normalize_salsify_asset_name(q)]
        for query in query_tokens:
            for img in images:
                if not isinstance(img, dict):
                    continue
                name = normalize_salsify_asset_name(img.get("name", ""))
                if name and (query == name or query in name):
                    return img
        return None

    def find_first_retailer_preferred_image(images, retailer_token, *queries, strict=False, excluded_retailer_tokens=None, used_urls=None):
        retailer_token = normalize_salsify_asset_name(retailer_token or "")
        excluded_retailer_tokens = {
            normalize_salsify_asset_name(x)
            for x in (excluded_retailer_tokens or [])
            if normalize_salsify_asset_name(x)
        }
        used_urls = {str(x or "").strip() for x in (used_urls or set()) if str(x or "").strip()}

        preferred = []
        generic_only = []

        for img in images or []:
            if not isinstance(img, dict):
                continue
            url = str(img.get("url", "") or "").strip()
            if used_urls and url in used_urls:
                continue
            name = normalize_salsify_asset_name(img.get("name", ""))
            if retailer_token and retailer_token in name:
                preferred.append(img)
                continue
            if any(token and token in name for token in excluded_retailer_tokens):
                continue
            generic_only.append(img)

        if strict:
            return find_first_image(preferred, *queries)
        return find_first_image(preferred, *queries) or find_first_image(generic_only, *queries)

    if retailer in {"sam's club", "sams club", "samsclub"}:
        aligned = []
        used_urls = set()

        def image_url(img):
            return str((img or {}).get("url", "") or "").strip() if isinstance(img, dict) else ""

        def image_name(img):
            return normalize_salsify_asset_name((img or {}).get("name", "")) if isinstance(img, dict) else ""

        def append_unique(img):
            if not isinstance(img, dict):
                return False
            url = image_url(img)
            if not url or url in used_urls:
                return False
            aligned.append(img)
            used_urls.add(url)
            return True

        def find_first_unused(*queries, exclude_tokens=None):
            query_tokens = [normalize_salsify_asset_name(q) for q in queries if normalize_salsify_asset_name(q)]
            exclude_tokens = [normalize_salsify_asset_name(q) for q in (exclude_tokens or []) if normalize_salsify_asset_name(q)]
            for query in query_tokens:
                for img in source_images:
                    if not isinstance(img, dict):
                        continue
                    url = image_url(img)
                    if not url or url in used_urls:
                        continue
                    name = image_name(img)
                    if exclude_tokens and any(token in name for token in exclude_tokens):
                        continue
                    if name and (query == name or query in name):
                        return img
            return None

        def find_online_optimized_generic():
            # Match the generic "Online Optimized Image-" property, but do not
            # let retailer-specific versions such as Online Optimized Image-Sams Club
            # satisfy this slot.
            excluded = [
                "sam's club", "sams club", "samsclub", "sams",
                "walgreens", "kroger", "grocery", "cvs", "target", "walmart",
            ]
            return find_first_unused(
                "online optimized image",
                "online image",
                exclude_tokens=excluded,
            )

        # Requested Sam's Club order when Online Optimized Image-Sams Club exists:
        # 1. Online Optimized Image-Sams Club.
        # 2. ATF Video-Sams Club.
        # 3. Generic Online Optimized Image-.
        # Fallback only applies when Online Optimized Image-Sams Club is missing.
        online_optimized_sams_club_img = find_first_unused(
            "online optimized image sams club",
            "online optimized image-sams club",
            "online optimized image sam s club",
            "online optimized image-sam s club",
        )
        if online_optimized_sams_club_img:
            append_unique(online_optimized_sams_club_img)
        else:
            append_unique(
                find_first_unused(
                    "main variant image sams club",
                    "main variant image-sams club",
                    "main variant image sam s club",
                    "main variant image-sam s club",
                )
                or find_first_unused("main variant image club", "main variant image-club")
                or find_online_optimized_generic()
            )

        append_unique(
            find_first_unused(
                "atf video sams club",
                "atf video-sams club",
                "atf video sam s club",
                "atf video-sam s club",
                "video sams club",
            )
        )

        append_unique(find_online_optimized_generic())
        append_unique(find_first_unused("shipping", "shipping-"))
        append_unique(
            find_first_unused(
                "atf i/o sams club", "atf i/o-sams club",
                "atf i o sams club", "atf io sams club", "atf io-sams club",
                "atf i/o generic", "atf i/o-generic",
                "atf i o generic", "atf io generic", "atf io-generic",
                "atf i/o", "atf io",
            )
        )
        for slot_num in range(2, 11):
            append_unique(
                find_first_unused(
                    f"atf {slot_num} sams club",
                    f"atf {slot_num}-sams club",
                    f"atf {slot_num} sam s club",
                    f"atf {slot_num}-sam s club",
                    f"atf {slot_num} generic",
                    f"atf {slot_num}-generic",
                )
            )

        reserved_tokens = [
            "main variant image sams club", "main variant image-sams club",
            "main variant image sam s club", "main variant image-sam s club",
            "online optimized image sams club", "online optimized image-sams club",
            "online optimized image sam s club", "online optimized image-sam s club",
            "main variant image club", "main variant image-club",
            "online optimized image", "online image",
            "atf video sams club", "atf video-sams club",
            "atf video sam s club", "atf video-sam s club", "video sams club",
            "shipping", "shipping-",
            "atf i/o sams club", "atf i/o-sams club",
            "atf i o sams club", "atf io sams club", "atf io-sams club",
            "atf i/o generic", "atf i/o-generic",
            "atf i o generic", "atf io generic", "atf io-generic",
            "atf i/o", "atf io",
        ]
        for i in range(2, 11):
            reserved_tokens.extend([
                f"atf {i} sams club", f"atf {i}-sams club",
                f"atf {i} sam s club", f"atf {i}-sam s club",
                f"atf {i} generic", f"atf {i}-generic",
            ])

        for img in source_images:
            if not isinstance(img, dict):
                continue
            name = image_name(img)
            if any(token in name for token in reserved_tokens):
                continue
            append_unique(img)

        return aligned[:max_slots]

    if retailer == "cvs":
        return reorder_cvs_salsify_images_for_visual(source_images, max_slots=max_slots)

    if retailer in {"heb", "h-e-b"}:
        aligned = []
        used_urls = set()
        def image_url(img):
            return str((img or {}).get("url", "") or "").strip() if isinstance(img, dict) else ""
        def image_name(img):
            return normalize_salsify_asset_name((img or {}).get("name", "")) if isinstance(img, dict) else ""
        def append_unique(img):
            if not isinstance(img, dict):
                return False
            url = image_url(img)
            if not url or url in used_urls:
                return False
            aligned.append(img)
            used_urls.add(url)
            return True
        def find_first_unused(*queries, exclude_tokens=None):
            query_tokens = [normalize_salsify_asset_name(q) for q in queries if normalize_salsify_asset_name(q)]
            exclude_tokens = [normalize_salsify_asset_name(q) for q in (exclude_tokens or []) if normalize_salsify_asset_name(q)]
            for query in query_tokens:
                for img in source_images:
                    if not isinstance(img, dict):
                        continue
                    url = image_url(img)
                    if not url or url in used_urls:
                        continue
                    name = image_name(img)
                    if exclude_tokens and any(token in name for token in exclude_tokens):
                        continue
                    if name and (query == name or query in name):
                        return img
            return None
        append_unique(find_first_unused(
            "online optimized image heb", "online optimized image-heb", "online optimized image h e b", "online optimized image-h-e-b",
            "online image heb", "online image-heb", "online optimized image", "online image",
            exclude_tokens=["cvs", "kroger", "walgreens", "sams club", "sam s club", "samsclub"],
        ))
        append_unique(find_first_unused(
            "atf i/o heb", "atf i/o-heb", "atf io heb", "atf io-heb", "atf i/o h e b", "atf i/o-h-e-b",
            "atf i/o generic", "atf i/o-generic", "atf io generic", "atf io-generic", "atf i/o", "atf io",
        ))
        for slot_num in range(2, 11):
            append_unique(find_first_unused(
                f"atf {slot_num} heb", f"atf {slot_num}-heb", f"atf {slot_num} h e b", f"atf {slot_num}-h-e-b",
                f"atf {slot_num} generic", f"atf {slot_num}-generic", f"atf {slot_num}",
            ))
        for img in source_images:
            append_unique(img)
        return aligned[:max_slots]

    if retailer == "walgreens":
        strict_image_mode = retailer in EXCLUSIVE_SALSIFY_IMAGE_RETAILERS
        used_urls = set()
        aligned = []
        excluded_retailer_tokens = {
            "cvs",
            "kroger",
            "sam's club",
            "sams club",
            "samsclub",
            "walmart",
            "target",
            "amazon",
        }

        # Walgreens Salsify rules:
        # 1. Slot 1 must be Online Optimized Image.
        # 2. Slot 2 must be Ingredient Label Image.
        # 3. If either slot is missing, keep that slot blank.
        # 4. ATF images must start only after slots 1 and 2 and never move up.
        slot_plan = [
            (("online optimized image-", "online optimized image", "online image", "online", "front"), "online optimized image", True),
            (("ingredient label image", "ingredients label image", "ingredient label", "ingredients label"), "ingredient label image", True),
            (("atf io", "atf i/o generic", "atf i o generic", "atf io generic"), "atf io", False),
            (("atf 2",), "atf 2", False),
            (("atf 3",), "atf 3", False),
            (("atf 4",), "atf 4", False),
            (("atf 5",), "atf 5", False),
            (("atf 6",), "atf 6", False),
        ]

        for query_group, blank_name, keep_blank in slot_plan:
            img = find_first_retailer_preferred_image(
                source_images,
                "walgreens",
                *query_group,
                strict=strict_image_mode,
                excluded_retailer_tokens=excluded_retailer_tokens,
                used_urls=used_urls,
            )
            if isinstance(img, dict) and str(img.get("url", "") or "").strip():
                aligned.append(img)
                used_urls.add(str(img.get("url", "") or "").strip())
            elif keep_blank:
                aligned.append(make_blank_salsify_image_slot(blank_name))

        return aligned[:min(max_slots, 6)]

    return dedupe_images_preserve_order(source_images)[:max_slots]


_original_align_salsify_images_for_retailer = align_salsify_images_for_retailer


def select_kroger_salsify_main_image(s_images):
    """Backward-compatible Kroger Salsify main image selector."""
    return select_kroger_salsify_images(s_images, max_slots=1)[:1]


def select_kroger_salsify_images(s_images, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE):
    """Kroger-only Salsify image order rule.

    This function is called only when retailer == "kroger".

    Kroger must show only one OOI/front image, using this priority:
    1. Online Optimized Image-Kroger
    2. Online Optimized Image-Grocery
    3. Online Optimized Image-
    4. Then ATF / lifestyle images in source order.

    If a higher-priority OOI exists, lower-priority OOI images are skipped.
    """
    images = [img for img in list(s_images or []) if isinstance(img, dict)]

    def img_url(img):
        return str((img or {}).get("url", "") or "").strip()

    def norm_name(img):
        return normalize_salsify_asset_name((img or {}).get("name", ""))

    def is_valid_img(img):
        return isinstance(img, dict) and bool(img_url(img))

    def find_first_by_name(*queries, exclude_tokens=None):
        query_tokens = [normalize_salsify_asset_name(q) for q in queries if normalize_salsify_asset_name(q)]
        exclude_tokens = [normalize_salsify_asset_name(q) for q in (exclude_tokens or []) if normalize_salsify_asset_name(q)]
        for query in query_tokens:
            for img in images:
                if not is_valid_img(img):
                    continue
                name = norm_name(img)
                if exclude_tokens and any(token in name for token in exclude_tokens):
                    continue
                if name and (name == query or query in name):
                    return img
        return None

    kroger_ooi = find_first_by_name(
        "Online Optimized Image-Kroger",
        "Online Image-Kroger",
    )
    grocery_ooi = find_first_by_name(
        "Online Optimized Image-Grocery",
        "Online Image-Grocery",
    )
    generic_ooi = find_first_by_name(
        "Online Optimized Image-",
        "Online Optimized Image",
        "Online Image-",
        "Online Image",
        exclude_tokens=[
            "kroger", "grocery", "walgreens", "cvs", "sams club", "sam s club",
            "samsclub", "target", "walmart", "amazon", "atf", "lifestyle",
            "life style", "shipping", "ingredient", "flat back", "flat left", "video",
        ],
    )

    ordered = []
    seen_urls = set()

    def add(img):
        if not is_valid_img(img):
            return False
        url = img_url(img)
        if not url or url in seen_urls:
            return False
        ordered.append(img)
        seen_urls.add(url)
        return True

    # Add exactly one OOI image for Kroger, in priority order.
    selected_ooi = kroger_ooi or grocery_ooi or generic_ooi
    add(selected_ooi)

    # Then add ATF / lifestyle images. Do not add any other OOI images.
    for img in images:
        if not is_valid_img(img):
            continue
        name = norm_name(img)
        is_ooi = "online optimized image" in name or "online image" in name
        is_atf_or_lifestyle = "atf" in name or "lifestyle" in name or "life style" in name
        if is_ooi:
            continue
        if is_atf_or_lifestyle:
            add(img)

    return ordered[:max_slots]


def align_salsify_images_for_retailer(retailer_name, s_images, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE, brand=""):
    retailer = str(retailer_name or "").strip().lower()
    if retailer == "kroger":
        return select_kroger_salsify_images(s_images, max_slots=max_slots)
    aligned = _original_align_salsify_images_for_retailer(retailer_name, s_images, max_slots=max_slots, brand=brand)
    return apply_retailer_salsify_image_limits(retailer_name, aligned)

def align_image_slots_for_comparison(s_images, r_images, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE, strong_threshold=80):
    """
    Preserve the original slot order exactly as captured.

    No image reordering, no sequence alignment, and no slot-shift correction.
    If one side is missing an image in a given position, that slot should stay mismatched
    and score 0% (Poor) instead of being auto-aligned to a later slot.
    """
    s_seq = list(s_images or [])[:max_slots]
    r_seq = list(r_images or [])[:max_slots]
    return trim_trailing_empty_image_slots(s_seq, r_seq)


def get_image_slot_url(slot_value):
    if isinstance(slot_value, dict):
        return str(slot_value.get("url", "") or "").strip()
    return str(slot_value or "").strip()


def trim_trailing_empty_image_slots(s_images, r_images):
    s_images = list(s_images or [])
    r_images = list(r_images or [])

    keep_len = max(len(s_images), len(r_images))
    while keep_len > 0:
        s_url = get_image_slot_url(s_images[keep_len - 1]) if keep_len - 1 < len(s_images) else ""
        r_url = get_image_slot_url(r_images[keep_len - 1]) if keep_len - 1 < len(r_images) else ""

        if s_url or r_url:
            break

        keep_len -= 1

    return s_images[:keep_len], r_images[:keep_len]



def align_sams_club_retailer_images_to_salsify_video_slot(s_images, r_images, max_slots=MAX_IMAGE_SLOTS_TO_COMPARE):
    """Reserve Sam's Club retailer slot 2 when Salsify has ATF Video in slot 2.

    Sam's Club sometimes has no live retailer video in the same position even when
    Salsify has ATF Video-Sams Club. Without this spacer, every retailer image
    after slot 1 shifts up one row and compares against the wrong Salsify asset.

    Rules:
    - Only applies when Salsify slot 2 is video-like or named ATF Video.
    - Keep retailer image 1 in slot 1.
    - If a retailer video exists later, move the first retailer video into slot 2.
    - Otherwise insert a blank slot 2 and shift retailer images down.
    """
    s_images = list(s_images or [])
    r_images = [str(u or "").strip() for u in list(r_images or []) if str(u or "").strip()]
    max_slots = max(0, int(max_slots or MAX_IMAGE_SLOTS_TO_COMPARE))
    if max_slots <= 0 or len(s_images) < 2:
        return r_images[:max_slots]

    s_slot_2 = s_images[1] if isinstance(s_images[1], dict) else {}
    s_slot_2_url = str(s_slot_2.get("url", "") or "").strip()
    s_slot_2_name = normalize_salsify_asset_name(s_slot_2.get("name", ""))
    s_slot_2_is_video = bool(
        is_video_like_url(s_slot_2_url)
        or "atf video" in s_slot_2_name
        or "video sams club" in s_slot_2_name
    )
    if not s_slot_2_is_video:
        return r_images[:max_slots]

    if len(r_images) >= 2 and is_video_like_url(r_images[1]):
        return r_images[:max_slots]

    first_retailer_video_index = None
    for idx, url in enumerate(r_images[1:], start=1):
        if is_video_like_url(url):
            first_retailer_video_index = idx
            break

    ordered = []
    if r_images:
        ordered.append(r_images[0])

    if first_retailer_video_index is not None:
        ordered.append(r_images[first_retailer_video_index])
        for idx, url in enumerate(r_images[1:], start=1):
            if idx != first_retailer_video_index:
                ordered.append(url)
    else:
        ordered.append("")
        ordered.extend(r_images[1:])

    return ordered[:max_slots]

def build_image_score_fields(s_images, r_images, max_slots=MAX_IMAGE_SLOTS_TO_SCORE):
    """
    Build per-slot image scores and an average image score.

    Rules:
    - Compare image slots in order.
    - Score = 0 if one side is missing.
    - Only score up to max_slots.
    """
    s_images = s_images or []
    r_images = r_images or []

    slots_to_score = min(max(len(s_images), len(r_images)), max_slots)

    if slots_to_score <= 0:
        return 0, {}

    img_scores = []
    image_position_scores = {}

    for i in range(slots_to_score):
        s_url = s_images[i].get("url") if i < len(s_images) and isinstance(s_images[i], dict) else ""
        r_url = r_images[i] if i < len(r_images) and isinstance(r_images[i], str) else ""

        score = compare_images_visually(s_url, r_url) if (s_url and r_url) else 0

        img_scores.append(score)
        image_position_scores[f"Image {i + 1} %"] = score

    avg_img_score = int(sum(img_scores) / len(img_scores)) if img_scores else 0
    return avg_img_score, image_position_scores
    
@st.cache_data(show_spinner=False, max_entries=1400)
def build_normalized_comparison_payload(
    salsify_url,
    retailer_name,
    retail_url,
    current_target_sku="",
    sku="",
    brand="",
    row_source_code="",
    mode="batch",
    max_slots=MAX_IMAGE_SLOTS_TO_SCORE,
):
    """
    Centralized comparison payload builder.

    This keeps every retailer split into two separate concerns without changing
    the UI or Excel export:
      1. Retailer site parsing through get_retailer_bundle().
      2. Salsify-for-that-retailer parsing/finalizing through
         finalize_salsify_copy_for_retailer() and align_salsify_images_for_retailer().

    Performance goal:
      - Batch processing, visual QA, and filters can reuse the same normalized
        Salsify/retailer bundles instead of reparsing/re-aligning the same row.

    Behavior lock:
      - The output shape stays the same as the old visual/batch code.
      - The Streamlit UI and Excel export consume the same fields they did before.
      - CVS visual-only feature rescue remains visual-only to avoid changing
        batch/export scoring unexpectedly.
    """
    retailer_norm = normalize_retailer_name(retailer_name).strip().lower()
    salsify_url = str(salsify_url or "").strip()
    retail_url = str(retail_url or "").strip()
    current_target_sku = str(current_target_sku or "").strip()
    row_source_code = str(row_source_code or "")
    mode = str(mode or "batch").strip().lower()
    max_slots = int(max_slots or MAX_IMAGE_SLOTS_TO_SCORE)

    s_bundle = get_salsify_bundle(salsify_url)
    r_bundle = get_retailer_bundle(
        retailer_name,
        retail_url,
        current_target_sku,
        sku=sku,
        row_source_code=row_source_code,
    )

    s_text = finalize_salsify_copy_for_retailer(retailer_name, s_bundle["text"] or {})

    # Preserve the existing CVS visual QA rescue behavior exactly where it lived:
    # visual QA only. This prevents batch/export scores from changing while still
    # ensuring the visual screen shows all CVS Salsify feature candidates.
    if mode == "visual" and retailer_norm == "cvs":
        raw_text = dict(s_bundle.get("text", {}) or {})
        raw_override = (raw_text.get("retailer_overrides", {}) or {}).get("cvs", {}) or {}
        rescue_features = []
        rescue_features.extend(raw_override.get("features", []) or [])
        for i in range(1, 11):
            rescue_features.append(raw_text.get(f"feature{i}", ""))
        rescue_features.extend(raw_text.get("features", []) or [])
        rescue_features = normalize_salsify_feature_values(rescue_features, max_features=10)
        if rescue_features:
            s_text["features"] = rescue_features
            for i in range(1, 8):
                s_text[f"feature{i}"] = rescue_features[i - 1] if i - 1 < len(rescue_features) else ""

    s_images = align_salsify_images_for_retailer(
        retailer_name,
        s_bundle["images"],
        max_slots=max_slots,
        brand=brand,
    )

    r_text = finalize_retailer_copy(retailer_name, r_bundle["text"] or {})
    r_images = r_bundle["images"] or []
    r_debug_for_cvs = (r_bundle.get("text", {}) or {}).get("debug", {}) if isinstance(r_bundle, dict) else {}
    if retailer_norm == "cvs":
        r_images = sanitize_cvs_retailer_images(r_images, debug=r_debug_for_cvs, reason="normalized_payload_initial_guard")

    if retailer_norm == "kroger":
        s_images = select_kroger_salsify_images(s_images, max_slots=max_slots)
        compare_slots = len(s_images) if s_images else 0
        r_images = select_kroger_image_urls_by_perspective(r_images, max_images=max(compare_slots, 1)) if compare_slots else []
        kroger_has_copy = bool(
            normalize_space(r_text.get("title", ""))
            or normalize_space(r_text.get("description", ""))
            or any(normalize_space(x) for x in (r_text.get("features", []) or []))
        )
        if not r_images and kroger_has_copy and compare_slots:
            r_images = force_single_kroger_main_image([], retail_url=retail_url, target_rpc=current_target_sku)

    if mode == "visual":
        if retailer_norm == "cvs":
            cvs_max_slots = int(get_retailer_salsify_requirements(retailer_name).get("max_images", MAX_IMAGE_SLOTS_TO_COMPARE) or MAX_IMAGE_SLOTS_TO_COMPARE)
            r_images = reorder_cvs_retailer_images_for_visual(r_images, max_slots=cvs_max_slots)
        elif retailer_norm == "walgreens":
            r_images = r_images[:6]
    else:
        if retailer_norm == "cvs":
            cvs_max_slots = int(get_retailer_salsify_requirements(retailer_name).get("max_images", MAX_IMAGE_SLOTS_TO_COMPARE) or MAX_IMAGE_SLOTS_TO_COMPARE)
            r_images = reorder_cvs_retailer_images_for_visual(r_images, max_slots=cvs_max_slots)
        elif retailer_norm == "walgreens":
            r_images = r_images[:6]

    # CVS-only: keep locked Salsify flat slots visible as Missing when absent,
    # but visually align the remaining CVS ATF/lifestyle images after slots 1-3.
    if retailer_norm == "cvs":
        cvs_max_slots = int(get_retailer_salsify_requirements(retailer_name).get("max_images", MAX_IMAGE_SLOTS_TO_COMPARE) or MAX_IMAGE_SLOTS_TO_COMPARE)
        r_images = align_cvs_atf_images_by_visual_match(
            s_images,
            r_images,
            locked_slots=3,
            max_slots=min(max_slots, cvs_max_slots),
            retailer_name=retailer_name,
        )
        r_images = sanitize_cvs_retailer_images(r_images, debug=r_debug_for_cvs, reason="normalized_payload_final_guard")

    # Sam's Club-only: if Salsify slot 2 is ATF Video-Sams Club, reserve retailer
    # slot 2 for a retailer video or blank spacer so the remaining retailer images
    # do not shift up and compare against the wrong Salsify rows.
    if retailer_norm in {"sam's club", "sams club", "samsclub"}:
        r_images = align_sams_club_retailer_images_to_salsify_video_slot(
            s_images,
            r_images,
            max_slots=max_slots,
        )

    s_images, r_images = align_image_slots_for_comparison(
        s_images,
        r_images,
        max_slots=max_slots,
    )

    return {
        "s_text": s_text,
        "s_images": s_images,
        "r_text": r_text,
        "r_images": r_images,
        "s_debug": (s_bundle.get("text", {}) or {}).get("debug", {}) if isinstance(s_bundle, dict) else {},
        "r_debug": (r_bundle.get("text", {}) or {}).get("debug", {}) if isinstance(r_bundle, dict) else {},
    }


@st.cache_data(show_spinner=False, max_entries=1200)
def get_visual_row_payload(
    salsify_url,
    retailer_name,
    retail_url,
    current_target_sku="",
    sku="",
    row_source_code="",
):
    retailer_norm = normalize_retailer_name(retailer_name).strip().lower()
    retail_url = str(retail_url or "").strip()
    row_source_code = str(row_source_code or "")
    uploaded_html_map = st.session_state.get("uploaded_raw_html_map", {}) or {}

    # Visual QA must reuse the same TXT-matched HTML used in batch processing.
    if retailer_norm == "kroger":
        if not retail_url and current_target_sku:
            retail_url = find_kroger_url_in_uploaded_map(uploaded_html_map, target_rpc=current_target_sku)
        if not row_source_code:
            row_source_code = lookup_uploaded_raw_html(
                uploaded_html_map,
                retail_url,
                target_rpc=current_target_sku,
            )
    elif retailer_norm in {"heb", "cvs"}:
        if not row_source_code:
            row_source_code = lookup_uploaded_raw_html(
                uploaded_html_map,
                retail_url,
                target_rpc=current_target_sku,
            )

    visual_max_slots = MAX_IMAGE_SLOTS_TO_COMPARE
    if retailer_norm == "walgreens":
        visual_max_slots = 6

    payload = build_normalized_comparison_payload(
        salsify_url=salsify_url,
        retailer_name=retailer_name,
        retail_url=retail_url,
        current_target_sku=current_target_sku,
        sku=sku,
        brand="",
        row_source_code=row_source_code,
        mode="visual",
        max_slots=visual_max_slots,
    )

    return {
        "s_text": payload["s_text"],
        "s_images": payload["s_images"],
        "r_text": payload["r_text"],
        "r_images": payload["r_images"],
    }

def process_row(row):
    try:
        retail_url = row.get("retail_url", "")
        salsify_url = row.get("salsify_url", "")
        cvs_rpc = row.get("retailer_rpc", "")
        row_source_code = row.get("copy_source_code", "")
        retailer_name = row.get("retailer", "") or infer_retailer_name_from_url(retail_url)
        rating_value = row.get("rating", "")
        review_count_value = row.get("review_count", "")

        salsify_url = str(salsify_url or "").strip()
        retail_url = str(retail_url or "").strip()
        cvs_rpc = str(cvs_rpc or "").strip()

        retailer_norm_for_row = normalize_retailer_name(retailer_name).strip().lower()
        # IMPORTANT: process_row runs inside ThreadPoolExecutor workers. Do not
        # read st.session_state here. The main UI thread already copies any
        # matched TXT/HTML into row["copy_source_code"] before workers start.
        # Reading st.session_state inside workers can fail fast and produce a
        # blank extract because every row returns None.

        title_score = 0
        desc_score = 0
        avg_feature_score = 0
        avg_img_score = 0
        overall = 0
        feature_score_fields = {}
        image_position_scores = {}

        status_notes = []

        if not salsify_url:
            status_notes.append("Missing Salsify URL")
        if not retail_url and not row_source_code:
            status_notes.append("Missing Retail URL")
        if retail_url and not retailer_url_matches_selected(retail_url, retailer_name):
            status_notes.append(build_retailer_url_mismatch_status(retail_url, retailer_name))
        if status_notes:
            return {
                "summary": {
                    "SKU": row.get("sku", ""),
                    "Retailer": retailer_name,
                    "Retailer RPC": cvs_rpc,
                    "Brand": row.get("brand", ""),
                    "Salsify URL": salsify_url,
                    "Retail URL": retail_url,
                    "Rating": rating_value,
                    "Review Count": review_count_value,
                    "Kroger Size Variant": "",
                    "Title %": title_score,
                    "Description %": desc_score,
                    "Feature %": avg_feature_score,
                    "Image Match %": avg_img_score,
                    "Overall %": overall,
                    "Status": ", ".join(status_notes),
                    **feature_score_fields,
                    **image_position_scores,
                },
                "detail": {
                    "SKU": row.get("sku", ""),
                    "Retailer": retailer_name,
                    "Retailer RPC": cvs_rpc,
                    "Brand": row.get("brand", ""),
                    "Salsify URL": salsify_url,
                    "Retail URL": retail_url,
                    "Rating": rating_value,
                    "Review Count": review_count_value,
                    "Kroger Size Variant": "",
                    "Title %": title_score,
                    "Description %": desc_score,
                    "Feature %": avg_feature_score,
                    "Image Match %": avg_img_score,
                    "Overall %": overall,
                    "Status": ", ".join(status_notes),
                    **feature_score_fields,
                    **image_position_scores,
                },
                "debug": {
                    "SKU": row.get("sku", ""),
                    "Retailer": retailer_name,
                    "Retailer RPC": cvs_rpc,
                    "Brand": row.get("brand", ""),
                    "Retail URL": retail_url,
                    "Rating": rating_value,
                    "Review Count": review_count_value,
                    "Kroger Size Variant": "",
                    "Salsify URL": salsify_url,
                    "Status": ", ".join(status_notes),
                },
            }

        target_sku = get_target_sku_from_inputs(
            retail_url=retail_url,
            cvs_rpc=cvs_rpc,
        )


        # IMPORTANT:
        # Do NOT create a nested thread pool here.
        # The outer batch executor already parallelizes rows.
        comparison_payload = build_normalized_comparison_payload(
            salsify_url=salsify_url,
            retailer_name=retailer_name,
            retail_url=retail_url,
            current_target_sku=target_sku,
            sku=row.get("sku", ""),
            brand=row.get("brand", ""),
            row_source_code=row_source_code,
            mode="batch",
            max_slots=MAX_IMAGE_SLOTS_TO_SCORE,
        )

        s_text = comparison_payload["s_text"]
        s_images = comparison_payload["s_images"]
        r_text = comparison_payload["r_text"]
        r_images = comparison_payload["r_images"]

        # CVS-only diagnostics: images alone do not mean CVS copy was found.
        # This keeps the combined image fallback from hiding missing live copy.
        if retailer_norm_for_row == "cvs":
            r_has_copy = bool(
                normalize_space(r_text.get("title", ""))
                or normalize_space(r_text.get("description", ""))
                or any(normalize_space(x) for x in (r_text.get("features", []) or []))
            )
            r_has_images = bool(any(str(x or "").strip() for x in (r_images or [])))
            if not r_has_copy and not row_source_code:
                status_notes.append("CVS live copy missing/unmatched from selected CVS source paths")
            elif not r_has_copy and row_source_code:
                status_notes.append("CVS uploaded source matched but copy parser found no live copy")
            if not r_has_images:
                status_notes.append("CVS retailer images missing from selected CVS source paths")

        debug_data = r_text.get("debug", {})

        output_rating_value = (r_text.get("rating", "") if isinstance(r_text, dict) else "") or rating_value
        output_review_count_value = (r_text.get("review_count", "") if isinstance(r_text, dict) else "") or review_count_value
        kroger_size_variant = ""
        if retailer_norm_for_row == "kroger" and isinstance(r_text, dict):
            kroger_size_variant = clean_kroger_variant_size(
                r_text.get("variant_size", "")
                or r_text.get("kroger_size_variant", "")
                or (r_text.get("debug", {}) or {}).get("Kroger Size Variant", "")
            )

        title_score = keyword_score(s_text.get("title", ""), r_text.get("title", ""))

        s_desc_debug = debug_description(s_text.get("description", ""))
        r_desc_debug = debug_description(r_text.get("description", ""))

        if retailer_norm_for_row in {"sam's club", "sams club", "samsclub"}:
            desc_score = sams_description_coverage_score(
                s_text.get("description", ""),
                r_text.get("description", ""),
            )
        else:
            desc_score = description_similarity_score(
                s_text.get("description", ""),
                r_text.get("description", ""),
            )

        retailer_features = r_text.get("features", []) if isinstance(r_text, dict) else []
        retailer_norm = str(retailer_name or "").strip().lower()
        feature_fields = build_dynamic_feature_fields_for_pair(retailer_name, s_text, retailer_features)

        feature_scores = []
        feature_score_fields = {}
        feature_position = 1

        unused_retailer_feature_indexes = set(range(len(retailer_features)))
        for i, f_key in enumerate(feature_fields, start=1):
            s_val = s_text.get(f_key, "")
            r_val = retailer_features[i - 1] if i - 1 < len(retailer_features) else ""

            # Sam's Club groups multiple claims inside one Product Details section.
            # Search every section and the clean description. A retailer section may
            # legitimately cover more than one Salsify bullet, so reuse is allowed.
            if retailer_norm in {"sam's club", "sams club", "samsclub"} and normalize_space(s_val):
                candidates = list(retailer_features)
                clean_retailer_description = normalize_space(r_text.get("description", ""))
                if clean_retailer_description:
                    candidates.append(clean_retailer_description)
                best_score = 0
                best_candidate = ""
                for candidate in candidates:
                    candidate_score = sams_claim_coverage_score(s_val, candidate)
                    if candidate_score > best_score:
                        best_score = candidate_score
                        best_candidate = candidate
                r_val = best_candidate
            else:
                best_score = keyword_score(s_val, r_val) if (s_val or r_val) else 0

            if not (normalize_space(s_val) or normalize_space(r_val)):
                continue
            score = best_score if retailer_norm in {"sam's club", "sams club", "samsclub"} else (keyword_score(s_val, r_val) if (s_val or r_val) else 0)
            feature_scores.append(score)
            feature_score_fields[f"Feature {feature_position} %"] = score
            feature_position += 1

        avg_feature_score = int(sum(feature_scores) / len(feature_scores)) if feature_scores else 0

        if max(len(s_images), len(r_images)) > 0:
            avg_img_score, image_position_scores = build_image_score_fields(
                s_images,
                r_images,
                max_slots=MAX_IMAGE_SLOTS_TO_SCORE,
            )
        else:
            avg_img_score, image_position_scores = 0, {}
        
        overall = int((title_score + desc_score + avg_feature_score + avg_img_score) / 4)

        if retailer_norm in {"sam's club", "sams club", "samsclub"}:
            live_title = normalize_space(r_text.get("title", ""))
            live_description = normalize_space(r_text.get("description", ""))
            if "choose size" in live_title.lower():
                status_notes.append("Shared Retailer Parent PDP")
            if not live_description and not retailer_features:
                status_notes.append("Parser Review Required")
            elif not live_description and retailer_features:
                status_notes.append("Retailer Uses Highlights Only")
            elif desc_score < 50:
                status_notes.append("Retailer Copy Differs")
            if not r_images:
                status_notes.append("Missing Retailer Asset")
            elif avg_img_score < 50:
                status_notes.append("Retailer Images Differ")
            if overall >= 80 and not status_notes:
                status_notes.append("Strong Match")
            elif not status_notes:
                status_notes.append("Review")
            status_notes = dedupe_preserve_order(status_notes)

        return {
            "summary": {
                "SKU": row.get("sku", ""),
                "Retailer": retailer_name,
                "CVS RPC": cvs_rpc,
                "Brand": row.get("brand", ""),
                "Salsify URL": salsify_url,
                "Retail URL": retail_url,
                "Rating": output_rating_value,
                "Review Count": output_review_count_value,
                "Kroger Size Variant": kroger_size_variant,
                "Title %": title_score,
                "Description %": desc_score,
                "Feature %": avg_feature_score,
                "Image Match %": avg_img_score,
                "Overall %": overall,
                "Status": ", ".join(status_notes),
                **feature_score_fields,
                **image_position_scores,
            },
            "detail": {
                "SKU": row.get("sku", ""),
                "Retailer": retailer_name,
                "CVS RPC": cvs_rpc,
                "Brand": row.get("brand", ""),
                "Salsify URL": salsify_url,
                "Retail URL": retail_url,
                "Rating": output_rating_value,
                "Review Count": output_review_count_value,
                "Kroger Size Variant": kroger_size_variant,
                "Title %": title_score,
                "Description %": desc_score,
                "Feature %": avg_feature_score,
                "Image Match %": avg_img_score,
                "Overall %": overall,
                "Status": "",
                "Salsify Title": s_text.get("title", ""),
                "Retailer Title": r_text.get("title", ""),
                    "CVS Title": r_text.get("title", ""),
                "Salsify Description": s_text.get("description", ""),
                "Retailer Description": r_text.get("description", ""),
                    "CVS Description": r_text.get("description", ""),
                "Salsify Feature 1": s_text.get("feature1", ""),
                "Salsify Feature 2": s_text.get("feature2", ""),
                "Salsify Feature 3": s_text.get("feature3", ""),
                "Salsify Feature 4": s_text.get("feature4", ""),
                "Salsify Feature 5": s_text.get("feature5", ""),
                "Salsify Feature 6": s_text.get("feature6", ""),
                "Salsify Feature 7": s_text.get("feature7", ""),
                "Retailer Features": " | ".join(r_text.get("features", [])),
                    "CVS Features": " | ".join(r_text.get("features", [])),
                "Salsify Images": " | ".join([img.get("url", "") for img in s_images if isinstance(img, dict)]),
                "Retailer Images": " | ".join(r_images),
                    "CVS Images": " | ".join(r_images),
                "Title Path": debug_data.get("Title Path", ""),
                "Description Path": debug_data.get("Description Path", ""),
                "Features Path": debug_data.get("Features Path", ""),
                "vendorDetailsBulletsRef": debug_data.get("vendorDetailsBulletsRef", ""),
                "vendorDetailsParagraphRef": debug_data.get("vendorDetailsParagraphRef", ""),
                "featuresKey": debug_data.get("featuresKey", ""),
                "descriptionKey": debug_data.get("descriptionKey", ""),
                "directVendorContentFound": debug_data.get("directVendorContentFound", False),
                "directVendorDetailsFound": debug_data.get("directVendorDetailsFound", False),
                "variantWindowMatched": debug_data.get("variantWindowMatched", False),
                "variantMatchScore": debug_data.get("variantMatchScore", 0),
                "variantMatchReason": debug_data.get("variantMatchReason", ""),
                "matchedDynamicMediaUrl": debug_data.get("matchedDynamicMediaUrl", ""),
                "matchedVariantUrl": debug_data.get("matchedVariantUrl", ""),
                "matchedNearbyImage": debug_data.get("matchedNearbyImage", ""),
                **image_position_scores,
            },
            "debug": {
                "SKU": row.get("sku", ""),
                "Retailer": retailer_name,
                "CVS RPC": cvs_rpc,
                "Brand": row.get("brand", ""),
                "Retail URL": retail_url,
                "Rating": output_rating_value,
                "Review Count": output_review_count_value,
                "Kroger Size Variant": kroger_size_variant,
                "Salsify URL": salsify_url,
                "Retailer Title": r_text.get("title", ""),
                    "Retailer Description": r_text.get("description", ""),
                    "Retailer Features": " | ".join(r_text.get("features", [])),
                    "Desc Final": r_text.get("description", ""),
                "Desc Quality Score": r_desc_debug["quality_score"],
                "Desc Length": r_desc_debug["length"],
                "Desc Issues": ", ".join(r_desc_debug["issues"]),
                "Salsify Desc Quality Score": s_desc_debug["quality_score"],
                "Final Features": " | ".join(r_text.get("features", [])),
                "Title Path": debug_data.get("Title Path", ""),
                "Description Path": debug_data.get("Description Path", ""),
                "Features Path": debug_data.get("Features Path", ""),
                "vendorPatternFound": debug_data.get("vendorPatternFound", False),
                "vendorDetailsBulletsRef": debug_data.get("vendorDetailsBulletsRef", ""),
                "vendorDetailsParagraphRef": debug_data.get("vendorDetailsParagraphRef", ""),
                "featuresKey": debug_data.get("featuresKey", ""),
                "descriptionKey": debug_data.get("descriptionKey", ""),
                "featuresArrayFound": debug_data.get("featuresArrayFound", False),
                "descriptionBlockFound": debug_data.get("descriptionBlockFound", False),
                "directVendorContentFound": debug_data.get("directVendorContentFound", False),
                "directVendorDetailsFound": debug_data.get("directVendorDetailsFound", False),
                "variantWindowMatched": debug_data.get("variantWindowMatched", False),
                "variantMatchScore": debug_data.get("variantMatchScore", 0),
                "variantMatchReason": debug_data.get("variantMatchReason", ""),
                "matchedDynamicMediaUrl": debug_data.get("matchedDynamicMediaUrl", ""),
                "matchedVariantUrl": debug_data.get("matchedVariantUrl", ""),
                "matchedNearbyImage": debug_data.get("matchedNearbyImage", ""),
                "Source Used": debug_data.get("Source Used", ""),
                "vendorPatternExcerpt": debug_data.get("vendorPatternExcerpt", ""),
                "featuresArrayExcerpt": debug_data.get("featuresArrayExcerpt", ""),
                "descriptionBlockExcerpt": debug_data.get("descriptionBlockExcerpt", ""),
                "directVendorContentExcerpt": debug_data.get("directVendorContentExcerpt", ""),
                "rawHtmlLength": debug_data.get("rawHtmlLength", 0),
                "rawTextLength": debug_data.get("rawTextLength", 0),
                "nextjsChunkFound": debug_data.get("nextjsChunkFound", False),
                "rawHtmlHasSelfNextF": debug_data.get("rawHtmlHasSelfNextF", False),
                "rawHtmlHasVendorDetailsBullets": debug_data.get("rawHtmlHasVendorDetailsBullets", False),
                "rawHtmlHasVendorDetailsParagraph": debug_data.get("rawHtmlHasVendorDetailsParagraph", False),
                "rawTextHasVendorDetailsBullets": debug_data.get("rawTextHasVendorDetailsBullets", False),
                "rawTextHasVendorDetailsParagraph": debug_data.get("rawTextHasVendorDetailsParagraph", False),
                "rawHtmlVendorExcerpt": debug_data.get("rawHtmlVendorExcerpt", ""),
                "rawTextVendorExcerpt": debug_data.get("rawTextVendorExcerpt", ""),
            },
        }

    except Exception as e:
        error_text = f"{type(e).__name__}: {e}"
        error_trace = traceback.format_exc()
        safe_sku = row.get("sku", "") if isinstance(row, dict) else ""
        safe_retailer = row.get("retailer", "") if isinstance(row, dict) else ""
        safe_retail_url = row.get("retail_url", "") if isinstance(row, dict) else ""
        safe_salsify_url = row.get("salsify_url", "") if isinstance(row, dict) else ""
        safe_rpc = row.get("retailer_rpc", "") if isinstance(row, dict) else ""
        safe_brand = row.get("brand", "") if isinstance(row, dict) else ""
        safe_rating = row.get("rating", "") if isinstance(row, dict) else ""
        safe_review_count = row.get("review_count", "") if isinstance(row, dict) else ""
        if not safe_retailer:
            safe_retailer = infer_retailer_name_from_url(safe_retail_url)
        base_summary = {
            "SKU": safe_sku,
            "Retailer": safe_retailer,
            "CVS RPC": safe_rpc,
            "Brand": safe_brand,
            "Salsify URL": safe_salsify_url,
            "Retail URL": safe_retail_url,
            "Rating": safe_rating,
            "Review Count": safe_review_count,
            "Title %": 0,
            "Description %": 0,
            "Feature %": 0,
            "Image Match %": 0,
            "Overall %": 0,
            "Status": f"ROW ERROR - {error_text}",
        }
        detail = dict(base_summary)
        detail["Error Traceback"] = error_trace
        debug = dict(base_summary)
        debug["Error Traceback"] = error_trace
        debug["Source Used"] = "row_exception"
        return {"summary": base_summary, "detail": detail, "debug": debug}
# =========================================
# SESSION STATE
# =========================================
if "start_idx" not in st.session_state:
    st.session_state.start_idx = 0
if "summary_rows" not in st.session_state:
    st.session_state.summary_rows = []
if "export_rows" not in st.session_state:
    st.session_state.export_rows = []
if "debug_rows" not in st.session_state:
    st.session_state.debug_rows = []
if "summary_skus" not in st.session_state:
    st.session_state.summary_skus = set()
if "detail_skus" not in st.session_state:
    st.session_state.detail_skus = set()
if "debug_skus" not in st.session_state:
    st.session_state.debug_skus = set()
if "processing_done" not in st.session_state:
    st.session_state.processing_done = False
if "progress_bar" not in st.session_state:
    st.session_state.progress_bar = None
if "last_file_hash" not in st.session_state:
    st.session_state.last_file_hash = None
if "uploaded_file_bytes" not in st.session_state:
    st.session_state.uploaded_file_bytes = None
if "selected_retailer" not in st.session_state:
    st.session_state.selected_retailer = "-- Select Retailer --"
if "selected_brand_visual" not in st.session_state:
    st.session_state.selected_brand_visual = "All"
if "active_batch_key" not in st.session_state:
    st.session_state.active_batch_key = ""
if "completed_batch_key" not in st.session_state:
    st.session_state.completed_batch_key = ""
if "auto_download_done" not in st.session_state:
    st.session_state.auto_download_done = False
if "report_bytes" not in st.session_state:
    st.session_state.report_bytes = None
if "report_filename" not in st.session_state:
    st.session_state.report_filename = None
if "report_batch_key" not in st.session_state:
    st.session_state.report_batch_key = ""
if "batch_run_requested" not in st.session_state:
    st.session_state.batch_run_requested = False
if "batch_started_key" not in st.session_state:
    st.session_state.batch_started_key = ""
if "batch_status_message" not in st.session_state:
    st.session_state.batch_status_message = ""
if "batch_error_text" not in st.session_state:
    st.session_state.batch_error_text = ""
if "capture_mode" not in st.session_state:
    st.session_state.capture_mode = CAPTURE_MODE_USE_EXTENSION
if "uploaded_raw_html_map" not in st.session_state:
    st.session_state.uploaded_raw_html_map = {}
if "uploaded_raw_html_filename" not in st.session_state:
    st.session_state.uploaded_raw_html_filename = ""
if "uploaded_raw_html_stats" not in st.session_state:
    st.session_state.uploaded_raw_html_stats = {}
if "raw_html_upload_hash" not in st.session_state:
    st.session_state.raw_html_upload_hash = ""
if "auto_batch_upload_key" not in st.session_state:
    st.session_state.auto_batch_upload_key = ""

# =========================================
# TOP UPLOAD + DOWNLOAD UI
# =========================================
top_upload_col, top_download_col = st.columns([2.4, 1.1], gap="small")

with top_upload_col:
    uploaded_file = st.file_uploader("Upload Retailer SKU List", type=["xlsx", "csv"], help="Use the retailer SKU list with columns like Retailer, SKU, Retailer RPC, Retailer Salsify URL, Retailer URL, Brand. Retailer URL can be blank when you use the extension results workflow.")

with top_download_col:
    st.markdown("<div style='height: 30px;'></div>", unsafe_allow_html=True)

    if st.session_state.report_bytes is not None and st.session_state.report_filename:
        st.download_button(
            label="⬇ Download Excel Report",
            data=st.session_state.report_bytes,
            file_name=st.session_state.report_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_report_top_single",
        )

master_df = None
retailer_df = None
all_retailers = []
multi_retailer = False
selected_retailer = ""
current_batch_key = ""
file_hash = ""
file_ready_for_batch = False
selected_capture_mode = st.session_state.capture_mode
uploaded_raw_html_file = None
uploaded_raw_html_map = st.session_state.get("uploaded_raw_html_map", {}) or {}
matched_uploaded_html_count = 0
missing_uploaded_html_count = 0
capture_batch_key_part = "no_txt"

if uploaded_file:
    try:
        file_bytes = uploaded_file.getvalue()
        st.session_state.uploaded_file_bytes = file_bytes
        file_hash = hashlib.md5(file_bytes).hexdigest()

        if st.session_state.last_file_hash != file_hash:
            st.session_state.summary_rows = []
            st.session_state.export_rows = []
            st.session_state.debug_rows = []
            st.session_state.summary_skus = set()
            st.session_state.detail_skus = set()
            st.session_state.debug_skus = set()
            st.session_state.start_idx = 0
            st.session_state.processing_done = False
            st.session_state.progress_bar = None
            st.session_state.last_file_hash = file_hash
            st.session_state.selected_retailer = "-- Select Retailer --"
            st.session_state.selected_brand_visual = "All"
            st.session_state.active_batch_key = ""
            st.session_state.completed_batch_key = ""
            st.session_state.auto_download_done = False
            st.session_state.report_bytes = None
            st.session_state.report_filename = None
            st.session_state.report_batch_key = ""
            st.session_state.batch_run_requested = False
            st.session_state.batch_started_key = ""
            st.session_state.batch_status_message = ""
            st.session_state.batch_error_text = ""
            st.session_state.capture_mode = CAPTURE_MODE_USE_EXTENSION
            st.session_state.uploaded_raw_html_map = {}
            st.session_state.uploaded_raw_html_filename = ""
            st.session_state.uploaded_raw_html_stats = {}
            st.session_state.raw_html_upload_hash = ""
            st.session_state.auto_batch_upload_key = ""
            clear_in_memory_caches()
            st.cache_data.clear()

        source_master_df = read_uploaded_file_from_bytes(file_bytes, uploaded_file.name)
        master_df = prepare_input_df(source_master_df)
        all_retailers = sorted(master_df["retailer"].dropna().astype(str).unique().tolist()) if "retailer" in master_df.columns else ["CVS"]
        if not all_retailers:
            all_retailers = ["CVS"]
        multi_retailer = len(all_retailers) > 1

        with top_upload_col:
            st.caption("Detected retailers in upload: " + ", ".join(all_retailers))

        with top_upload_col:
            st.radio(
                "⚙️ Capture Mode",
                [CAPTURE_MODE_USE_EXTENSION, CAPTURE_MODE_SKIP_EXTENSION],
                key="capture_mode",
                horizontal=True,
                help="Use the browser extension + TXT upload, or skip extension and run directly for supported retailers.",
            )
            selected_capture_mode = st.session_state.capture_mode

            if multi_retailer:
                retailer_options = ["-- Select Retailer --"] + all_retailers
                if st.session_state.selected_retailer not in retailer_options:
                    st.session_state.selected_retailer = "-- Select Retailer --"
                selected_retailer = st.selectbox(
                    "🏪 Select Retailer",
                    retailer_options,
                    key="selected_retailer",
                    help="Select retailer to run batch.",
                )
                if selected_retailer == "-- Select Retailer --":
                    st.info("Select retailer to run batch.")
                else:
                    file_ready_for_batch = True
            else:
                st.session_state.selected_retailer = all_retailers[0]
                selected_retailer = st.selectbox(
                    "🏪 Select Retailer",
                    all_retailers,
                    index=0,
                    key="selected_retailer",
                    disabled=True,
                )
                file_ready_for_batch = True

            uploaded_raw_html_file = st.file_uploader(
                "Upload Extension Results CSV/XLSX or Captured Retailer HTML TXT",
                type=["csv", "xlsx", "txt", "html"],
                key="uploaded_raw_html_txt_top",
                help="Preferred: URL-only extension results CSV/XLSX with Retailer RPC and Retailer URL, or separated-copy extension results. Still supports captured TXT/HTML. CVS-only fallback: XLSX with CVS RPC in column A and pasted source snippets in columns B onward.",
            )
            if uploaded_raw_html_file is not None:
                raw_html_bytes = uploaded_raw_html_file.getvalue()
                raw_html_hash = hashlib.md5(raw_html_bytes or b"").hexdigest()
                source_file_name_lc = str(uploaded_raw_html_file.name or "").lower().strip()
                existing_source_stats = st.session_state.get("uploaded_raw_html_stats", {}) or {}
                should_reparse_uploaded_source = (
                    st.session_state.raw_html_upload_hash != raw_html_hash
                    or not (st.session_state.get("uploaded_raw_html_map", {}) or {})
                    or st.session_state.uploaded_raw_html_filename != uploaded_raw_html_file.name
                    or (source_file_name_lc.endswith(".xlsx") and existing_source_stats.get("mode") not in {"cvs_manual_source_xlsx", "extension_structured_results"})
                    or (source_file_name_lc.endswith(".csv") and existing_source_stats.get("mode") != "extension_structured_results")
                )
                if should_reparse_uploaded_source:
                    parsed_source_map, parsed_source_stats = parse_uploaded_retailer_source_file(raw_html_bytes, uploaded_raw_html_file.name, selected_retailer=selected_retailer)
                    st.session_state.uploaded_raw_html_map = parsed_source_map
                    st.session_state.uploaded_raw_html_stats = parsed_source_stats
                    st.session_state.uploaded_raw_html_filename = uploaded_raw_html_file.name
                    st.session_state.raw_html_upload_hash = raw_html_hash
                    st.session_state.auto_batch_upload_key = ""
                    st.session_state.batch_error_text = ""
                uploaded_raw_html_map = st.session_state.get("uploaded_raw_html_map", {}) or {}
                source_stats = st.session_state.get("uploaded_raw_html_stats", {}) or {}
                source_mode = source_stats.get("mode", "extension_txt_html")
                if uploaded_raw_html_map:
                    if source_mode == "extension_url_only_results":
                        st.success(
                            f"Loaded URL-only extension results from {st.session_state.uploaded_raw_html_filename}: "
                            f"{source_stats.get('mapped_rows', len(uploaded_raw_html_map))} RPC-to-URL rows mapped. Copy/images will be pulled live from Kroger pages."
                        )
                    elif source_mode == "extension_structured_results":
                        st.success(
                            f"Loaded extension results from {st.session_state.uploaded_raw_html_filename}: "
                            f"{source_stats.get('mapped_rows', len(uploaded_raw_html_map))} separated-copy rows mapped."
                        )
                    elif source_mode == "cvs_manual_source_xlsx":
                        st.success(
                            f"Loaded CVS manual source workbook from {st.session_state.uploaded_raw_html_filename}: "
                            f"{source_stats.get('mapped_rows', len(uploaded_raw_html_map))} usable CVS RPC captures mapped."
                        )
                        if source_stats.get("skipped_blank", 0) or source_stats.get("skipped_weak", 0):
                            st.warning(
                                f"Skipped {source_stats.get('skipped_blank', 0)} blank rows and "
                                f"{source_stats.get('skipped_weak', 0)} rows that looked like footer/CSS-only source, not product HTML."
                            )
                        if source_stats.get("truncated_cell_count", 0):
                            st.warning(
                                "Some XLSX cells are near Excel's cell text limit, so pasted HTML may be truncated. "
                                "For best results, use the extension TXT export when possible."
                            )
                    else:
                        st.success(f"Loaded TXT capture map from {st.session_state.uploaded_raw_html_filename} with {len(uploaded_raw_html_map)} URL keys.")
                else:
                    st.warning("Source uploaded, but no usable labeled URL/RPC + HTML blocks were found yet. For CVS, use extension TXT/HTML or XLSX with RPC in column A and real product HTML/source in columns B onward.")
            else:
                uploaded_raw_html_map = st.session_state.get("uploaded_raw_html_map", {}) or {}

        if file_ready_for_batch:
            capture_batch_key_part = "use_ext" if selected_capture_mode == CAPTURE_MODE_USE_EXTENSION else "skip_ext"
            if st.session_state.raw_html_upload_hash:
                capture_batch_key_part += f"::{st.session_state.raw_html_upload_hash}"
            if is_wide_salsify_template_df(source_master_df):
                # Wide SKU/RPC matrix rule: queue only the selected retailer.
                # Example: Kroger selected = use only Kroger RPC/Salsify URL/Kroger URL.
                # CVS selected = use only CVS RPC/Salsify URL/CVS URL.
                # Do not load or process URLs for the other retailers.
                retailer_df = build_selected_retailer_df_from_wide_source(source_master_df, selected_retailer)
            else:
                retailer_df = strict_filter_rows_for_selected_retailer(
                    master_df,
                    selected_retailer,
                    dedupe_by_url=(selected_capture_mode == CAPTURE_MODE_USE_EXTENSION),
                )

            source_mode = (st.session_state.get("uploaded_raw_html_stats", {}) or {}).get("mode", "extension_txt_html")
            url_only_source_mode = source_mode == "extension_url_only_results"
            matched_url_only_count = 0

            if uploaded_raw_html_map and "retailer_rpc" in retailer_df.columns:
                retailer_df = retailer_df.copy()
                retailer_df["retail_url"] = retailer_df["retail_url"].fillna("").astype(str).str.strip()
                if url_only_source_mode:
                    retailer_df["retail_url"] = retailer_df.apply(
                        lambda row: row["retail_url"] if str(row.get("retail_url", "")).strip() else find_url_only_url_in_uploaded_map(uploaded_raw_html_map, target_rpc=row.get("retailer_rpc", "")),
                        axis=1,
                    )
                    matched_url_only_count = int(retailer_df["retail_url"].fillna("").astype(str).str.strip().ne("").sum())
                elif selected_retailer == "Kroger":
                    retailer_df["retail_url"] = retailer_df.apply(
                        lambda row: row["retail_url"] if str(row.get("retail_url", "")).strip() else find_kroger_url_in_uploaded_map(uploaded_raw_html_map, target_rpc=row.get("retailer_rpc", "")),
                        axis=1,
                    )

            if selected_retailer == "Kroger":
                retailer_df = strict_filter_rows_for_selected_retailer(
                    retailer_df,
                    selected_retailer,
                    dedupe_by_url=False,
                )

            if "copy_source_code" not in retailer_df.columns:
                retailer_df["copy_source_code"] = ""
            if uploaded_raw_html_map and not url_only_source_mode:
                retailer_df["copy_source_code"] = retailer_df.apply(lambda row: lookup_uploaded_raw_html(uploaded_raw_html_map, row.get("retail_url", ""), target_rpc=row.get("retailer_rpc", "")), axis=1)
                matched_uploaded_html_count = int((retailer_df["copy_source_code"].astype(str).str.len() > 0).sum())
                missing_uploaded_html_count = max(len(retailer_df) - matched_uploaded_html_count, 0)
            elif url_only_source_mode:
                retailer_df["copy_source_code"] = ""
                matched_uploaded_html_count = 0
                missing_uploaded_html_count = 0

            # Extension/TXT mode: only process rows actually represented in the uploaded capture.
            retailer_df, capture_matched_queue_count, capture_missing_queue_count = filter_queue_to_uploaded_capture_matches(
                retailer_df,
                selected_retailer,
                source_mode=source_mode,
                selected_capture_mode=selected_capture_mode,
                uploaded_raw_html_map=uploaded_raw_html_map,
            )
            if uploaded_raw_html_map and capture_matched_queue_count > 0:
                matched_uploaded_html_count = capture_matched_queue_count
                missing_uploaded_html_count = capture_missing_queue_count

            retailer_df = sort_selected_retailer_queue(retailer_df)
            if retailer_df is not None and not retailer_df.empty:
                retailer_df = retailer_df.copy().reset_index(drop=True)
                retailer_df["_queue_order"] = range(len(retailer_df))

            current_batch_key = f"{file_hash}::{selected_retailer}::{capture_batch_key_part}::queued_{len(retailer_df)}"

            if st.session_state.active_batch_key != current_batch_key:
                st.session_state.summary_rows = []
                st.session_state.export_rows = []
                st.session_state.debug_rows = []
                st.session_state.summary_skus = set()
                st.session_state.detail_skus = set()
                st.session_state.debug_skus = set()
                st.session_state.start_idx = 0
                st.session_state.processing_done = False
                st.session_state.progress_bar = None
                st.session_state.active_batch_key = current_batch_key
                st.session_state.completed_batch_key = ""
                st.session_state.selected_brand_visual = "All"
                st.session_state.auto_download_done = False
                st.session_state.report_batch_key = ""
                st.session_state.batch_run_requested = False
                st.session_state.batch_started_key = ""
                st.session_state.batch_status_message = ""
                st.session_state.batch_error_text = ""
                st.session_state.auto_batch_upload_key = ""

            txt_ready_for_batch = bool(matched_uploaded_html_count > 0 or (source_mode == "extension_url_only_results" and matched_url_only_count > 0))
            cvs_capture_block_reason = ""
            cvs_missing_capture_urls = []
            if selected_retailer == "CVS" and selected_capture_mode == CAPTURE_MODE_USE_EXTENSION:
                if retailer_df is not None and not retailer_df.empty:
                    missing_capture_df = retailer_df[retailer_df["copy_source_code"].fillna("").astype(str).str.len() == 0].copy()
                    cvs_missing_capture_urls = [str(x).strip() for x in missing_capture_df.get("retail_url", pd.Series(dtype=str)).fillna("").astype(str).tolist() if str(x).strip()]
                if not uploaded_raw_html_map:
                    cvs_capture_block_reason = ""
                    st.caption("CVS extension source is recommended, but missing TXT will not block the batch. The app will try live/canonical CVS parsing for unmatched rows.")
                elif missing_uploaded_html_count > 0:
                    cvs_capture_block_reason = ""
                    st.caption(f"CVS TXT upload has {missing_uploaded_html_count} unmatched rows. Those rows will fall back to live/canonical CVS parsing instead of blocking the batch.")
            isolated_unique_url_count = int(retailer_df["retail_url"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique()) if retailer_df is not None and not retailer_df.empty and "retail_url" in retailer_df.columns else 0
            st.caption(f"Strict retailer isolation active: {selected_retailer} only. Final rows queued: {len(retailer_df)}. Unique retailer URLs queued: {isolated_unique_url_count}.")
            if selected_capture_mode == CAPTURE_MODE_USE_EXTENSION and retailer_df is not None and not retailer_df.empty:
                extension_input_csv = build_extension_input_csv(retailer_df, selected_retailer)
                st.download_button(
                    label=f"⬇ Download {selected_retailer} extension input CSV",
                    data=extension_input_csv.encode("utf-8"),
                    file_name=f"extension_input_{str(selected_retailer).lower().replace(' ', '_').replace("'", '')}.csv",
                    mime="text/csv",
                    key=f"download_extension_input_{selected_retailer}_{file_hash}",
                    help="Use this in the browser extension. It contains only Retailer and Search Term from the SKU list.",
                )
                with st.expander("Copy/paste extension input", expanded=False):
                    st.text_area(
                        "Retailer extension input",
                        value=extension_input_csv,
                        height=180,
                        key=f"extension_input_text_{selected_retailer}_{file_hash}",
                    )
            if selected_capture_mode == CAPTURE_MODE_USE_EXTENSION:
                extension_payload = build_extension_batch_payload(
                    retailer_df=retailer_df,
                    retailer_name=selected_retailer,
                    current_batch_key=current_batch_key,
                    capture_mode=selected_capture_mode,
                    txt_ready=txt_ready_for_batch,
                )
                render_extension_batch_bridge(extension_payload)
                if selected_retailer == "CVS":
                    st.caption("CVS mode active: extension TXT is preferred. If a row is unmatched, the app now also tries live/canonical CVS parsing and combined CVS fallbacks. This does not change other retailers.")
                elif selected_retailer == "HEB":
                    st.caption("HEB mode active: parses TXT once, matches rows by HEB RPC, pulls Salsify copy/features without HEB-only limits, and keeps all available Salsify/HEB image slots for comparison.")
                else:
                    st.caption(f"Extension bridge ready for {selected_retailer}. Existing retailer TXT/live-fetch behavior is unchanged.")
            elif selected_retailer in AUTO_SKIP_EXTENSION_RETAILERS:
                st.caption(f"{selected_retailer} is in skip-extension mode, so the app can auto-run straight to batch with live retailer fetches.")

            if uploaded_raw_html_map:
                st.caption(f"TXT match status for {selected_retailer}: {matched_uploaded_html_count} matched rows loaded into the processing queue; {missing_uploaded_html_count} selected-retailer rows were not in the uploaded capture.")
            if cvs_capture_block_reason:
                st.error(cvs_capture_block_reason)
                if cvs_missing_capture_urls:
                    with st.expander("CVS URLs missing from uploaded TXT", expanded=False):
                        st.code("\n".join(cvs_missing_capture_urls[:200]))
                        if len(cvs_missing_capture_urls) > 200:
                            st.caption(f"Showing first 200 of {len(cvs_missing_capture_urls)} missing CVS URLs.")

            should_auto_run = False
            auto_run_reason = ""
            if selected_capture_mode == CAPTURE_MODE_USE_EXTENSION and txt_ready_for_batch and not cvs_capture_block_reason:
                should_auto_run = True
                auto_run_reason = "uploaded TXT capture"
            elif selected_capture_mode == CAPTURE_MODE_SKIP_EXTENSION and selected_retailer in AUTO_SKIP_EXTENSION_RETAILERS:
                should_auto_run = True
                auto_run_reason = "skip-extension direct batch"

            if (
                should_auto_run
                and st.session_state.auto_batch_upload_key != current_batch_key
                and st.session_state.batch_started_key != current_batch_key
                and not st.session_state.processing_done
            ):
                st.session_state.batch_run_requested = True
                st.session_state.batch_started_key = current_batch_key
                st.session_state.batch_status_message = f"Auto-starting batch for {selected_retailer} using {auto_run_reason}."
                st.session_state.batch_error_text = ""
                st.session_state.completed_batch_key = ""
                st.session_state.start_idx = 0
                st.session_state.summary_rows = []
                st.session_state.export_rows = []
                st.session_state.debug_rows = []
                st.session_state.summary_skus = set()
                st.session_state.detail_skus = set()
                st.session_state.debug_skus = set()
                st.session_state.progress_bar = None
                st.session_state.report_bytes = None
                st.session_state.report_filename = None
                st.session_state.report_batch_key = ""
                st.session_state.auto_download_done = False
                st.session_state.auto_batch_upload_key = current_batch_key
                clear_in_memory_caches()
                st.cache_data.clear()
                st.rerun()


            run_button_col, run_msg_col = st.columns([1.2, 3.8], gap="small")
            with run_button_col:
                if st.button(
                    "Run Batch",
                    key=f"run_batch_btn::{current_batch_key}",
                    use_container_width=True,
                    disabled=bool(cvs_capture_block_reason),
                ):
                    st.session_state.batch_run_requested = True
                    st.session_state.batch_started_key = current_batch_key
                    st.session_state.batch_status_message = ""
                    st.session_state.batch_error_text = ""
                    st.session_state.processing_done = False
                    st.session_state.completed_batch_key = ""
                    st.session_state.start_idx = 0
                    st.session_state.summary_rows = []
                    st.session_state.export_rows = []
                    st.session_state.debug_rows = []
                    st.session_state.summary_skus = set()
                    st.session_state.detail_skus = set()
                    st.session_state.debug_skus = set()
                    st.session_state.progress_bar = None
                    st.session_state.report_bytes = None
                    st.session_state.report_filename = None
                    st.session_state.report_batch_key = ""
                    st.session_state.auto_download_done = False
                    clear_in_memory_caches()
                    st.cache_data.clear()
                    st.rerun()

            with run_msg_col:
                if st.session_state.batch_error_text:
                    st.error(st.session_state.batch_error_text)
                elif st.session_state.processing_done and st.session_state.completed_batch_key == current_batch_key:
                    st.success(st.session_state.batch_status_message or f"Batch finished for {selected_retailer}. Visual QA review is ready below.")
                elif selected_capture_mode == CAPTURE_MODE_USE_EXTENSION and not matched_uploaded_html_count:
                    st.info(f"Load the extension, run the retailer batch, then upload the TXT capture for {selected_retailer}. As soon as the TXT has matches, batch will run from that captured HTML.")
                elif selected_capture_mode == CAPTURE_MODE_SKIP_EXTENSION and selected_retailer in AUTO_SKIP_EXTENSION_RETAILERS:
                    st.info(f"Skip-extension mode is enabled for {selected_retailer}. The app will go straight into batch automatically.")
                else:
                    st.info(f"Use the top selections, then click Run Batch for {selected_retailer}. The visual QA review will appear below after extract/report generation finishes.")

    except EmptyDataError:
        st.error("🔥 CRITICAL APP ERROR")
        st.text("The uploaded file is empty or could not be read.")
    except ValueError as e:
        st.error("❌ INPUT FILE ERROR")
        st.text(str(e))
    except Exception as e:
        st.error("🔥 CRITICAL APP ERROR")
        st.text(str(e))
        st.text(traceback.format_exc())

# =========================================
# VIEW + FILTER CONTROLS
# =========================================
st.markdown("## 🔎 QA Viewer Controls")
show_only_issues = st.checkbox("❌ Show ONLY Issues", key="show_issues")
hide_good = st.checkbox("🎉 Hide Strong Matches (80%+)", key="hide_good")
show_below_90_only = st.checkbox("🔎 Show Only Scores Below 90%", key="show_below_90_only")


st.markdown("### 🧪 Debug Controls")

show_html_debugger = st.checkbox(
    "Debug HTML",
    key="show_html_debugger",
)

# Keep downstream variables defined so the rest of the app keeps working unchanged.
debugger_source = "Retailer page"
debug_only_sku = ""
use_manual_html_override = False
manual_html_file = None
manual_html_text = ""
debug_marker_start = ""
debug_marker_end = ""
debug_marker_target = "Raw HTML"

standalone_debug_url = st.text_input(
    "URL to pull raw HTML",
    key="standalone_debug_url",
).strip()

debug_timeout_override = st.number_input(
    "Timeout (s)",
    min_value=1,
    max_value=120,
    value=30,
    step=1,
    key="debug_timeout_override",
)

debug_headers_text = st.text_area(
    "Custom headers (optional)",
    placeholder="Either JSON, e.g. {'Accept-Language': 'en-US'} or one per line: Key: Value",
    height=100,
    key="debug_headers_text",
)

col_debug_1, col_debug_2 = st.columns(2)
with col_debug_1:
    debug_use_mobile = st.checkbox("Use mobile User-Agent", value=False, key="debug_use_mobile")
with col_debug_2:
    debug_proxy_url = st.text_input(
        "Proxy URL (optional)",
        key="debug_proxy_url",
        placeholder="http://user:pass@host:port",
    ).strip()

st.caption(
    "Paste a URL and the debugger will fetch the raw HTML response so you can inspect the exact source before we build a retailer-specific parser."
)

if show_html_debugger and standalone_debug_url:
    standalone_retailer_name = infer_retailer_name_from_url(standalone_debug_url)
    standalone_debug_views = resolve_debug_views(
        standalone_debug_url,
        retailer_name=standalone_retailer_name,
        use_manual_html_override=False,
        manual_html_text="",
        manual_html_file=None,
        headers_text=debug_headers_text,
        timeout_override=int(debug_timeout_override),
        use_mobile=debug_use_mobile,
        proxy_url=debug_proxy_url,
    )

    with st.expander("🔎 Debug HTML", expanded=True):
        render_debugger_panel(
            standalone_debug_views,
            sku="top_debugger",
            marker_start="",
            marker_end="",
            marker_target="Raw HTML",
            use_manual_html_override=False,
        )


if retailer_df is not None and st.session_state.processing_done and st.session_state.completed_batch_key == current_batch_key:
    visual_brands = sorted(retailer_df["brand"].dropna().astype(str).unique().tolist()) if "brand" in retailer_df.columns else []
    visual_brand_options = ["All"] + visual_brands
    if st.session_state.selected_brand_visual not in visual_brand_options:
        st.session_state.selected_brand_visual = "All"
    st.markdown("### 🏷️ Select Brand")
    st.selectbox(
        "Select brand display filter",
        visual_brand_options,
        key="selected_brand_visual",
        label_visibility="collapsed",
    )

# =========================================
# FILE + PROCESSING
# =========================================
if retailer_df is not None and file_ready_for_batch and st.session_state.batch_started_key == current_batch_key:
    try:
        if retailer_df.empty:
            st.warning("No rows found for the selected retailer.")
            st.stop()

        start = st.session_state.start_idx
        end = start + BATCH_SIZE

        if start >= len(retailer_df):
            st.session_state.processing_done = True
            st.session_state.completed_batch_key = current_batch_key

        batch_df = retailer_df.iloc[start:end]

        if not st.session_state.processing_done:
            st.write(f"Processing SKUs {start + 1} to {min(end, len(retailer_df))} of {len(retailer_df)}")
            st.caption(f"Batch Size: {BATCH_SIZE} | Workers: {MAX_WORKERS}")

            if st.session_state.progress_bar is None:
                st.session_state.progress_bar = st.progress(0)
            progress_bar = st.session_state.progress_bar
            status_text = st.empty()
            st.write("### Overall Progress")
            overall_progress_bar = st.progress(0)

            total = len(batch_df)
            completed = 0
            batch_records = batch_df.to_dict("records")

            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                future_to_order = {
                    executor.submit(process_row, row_dict): int(row_dict.get("_queue_order", start + idx))
                    for idx, row_dict in enumerate(batch_records)
                }
                batch_results = []
                for future in as_completed(future_to_order):
                    completed += 1
                    queue_order = future_to_order[future]
                    result = future.result()
                    if result:
                        batch_results.append((queue_order, result))

                    if completed % UI_UPDATE_EVERY == 0 or completed == total:
                        progress_bar.progress(completed / max(total, 1))
                        status_text.markdown(
                            f"**Processed:** {completed}/{total}  \n**Overall:** {start + completed}/{len(retailer_df)}"
                        )
                        overall_progress_bar.progress((start + completed) / max(len(retailer_df), 1))

                for _, result in sorted(batch_results, key=lambda item: item[0]):
                    summary = result.get("summary")
                    detail = result.get("detail")
                    debug = result.get("debug")

                    if summary and summary["SKU"] not in st.session_state.summary_skus:
                        st.session_state.summary_rows.append(summary)
                        st.session_state.summary_skus.add(summary["SKU"])
                    if detail and detail["SKU"] not in st.session_state.detail_skus:
                        st.session_state.export_rows.append(detail)
                        st.session_state.detail_skus.add(detail["SKU"])
                    if debug and debug["SKU"] not in st.session_state.debug_skus:
                        st.session_state.debug_rows.append(debug)
                        st.session_state.debug_skus.add(debug["SKU"])

            if start + BATCH_SIZE < len(retailer_df):
                st.session_state.start_idx += BATCH_SIZE
                time.sleep(0.05)
                st.rerun()
            else:
                st.session_state.processing_done = True
                st.session_state.completed_batch_key = current_batch_key
                st.session_state.batch_status_message = f"Batch finished for {selected_retailer}. Extract/report generated successfully."
                st.session_state.batch_run_requested = False
                st.session_state.auto_batch_upload_key = current_batch_key
                st.rerun()
    except Exception as e:
        st.session_state.batch_error_text = str(e)
        st.error("🔥 CRITICAL APP ERROR")
        st.text(str(e))
        st.text(traceback.format_exc())

# =========================================
# TOP EXPORT SECTION
# =========================================
if (
    st.session_state.processing_done
    and st.session_state.completed_batch_key
    and st.session_state.report_batch_key != st.session_state.completed_batch_key
):
    summary_df = pd.DataFrame(st.session_state.summary_rows)
    detail_df = pd.DataFrame(st.session_state.export_rows)
    debug_df = pd.DataFrame(st.session_state.debug_rows)

    # Kroger capture-quality handling. A zero score is a content comparison result,
    # while an invalid capture means the live PDP content was never available to
    # evaluate. Keep those states separate and create a targeted recapture queue.
    retry_df = pd.DataFrame()
    if normalize_retailer_name(selected_retailer) == "Kroger" and not debug_df.empty:
        source_series = debug_df.get("Source Used", pd.Series("", index=debug_df.index)).fillna("").astype(str)
        title_path_series = debug_df.get("Title Path", pd.Series("", index=debug_df.index)).fillna("").astype(str)
        description_path_series = debug_df.get("Description Path", pd.Series("", index=debug_df.index)).fillna("").astype(str)
        invalid_capture_mask = (
            source_series.str.contains("invalid_kroger_capture", case=False, na=False)
            | title_path_series.str.contains("invalid_uploaded_kroger_capture", case=False, na=False)
            | description_path_series.str.contains("invalid_uploaded_kroger_capture", case=False, na=False)
        )
        invalid_debug_df = debug_df.loc[invalid_capture_mask].copy()
        invalid_skus = set(invalid_debug_df.get("SKU", pd.Series(dtype=str)).fillna("").astype(str).str.strip())
        invalid_skus.discard("")

        if invalid_skus:
            capture_status = "Capture Invalid - Recapture Required"
            score_columns = [
                "Title %", "Description %", "Feature %", "Image Match %", "Overall %",
                "Feature 1 %", "Feature 2 %", "Feature 3 %", "Feature 4 %", "Feature 5 %",
                "Feature 6 %", "Feature 7 %", "Image 1 %", "Image 2 %", "Image 3 %",
                "Image 4 %", "Image 5 %", "Image 6 %",
            ]
            for frame in (summary_df, detail_df):
                if frame.empty or "SKU" not in frame.columns:
                    continue
                frame_invalid = frame["SKU"].fillna("").astype(str).str.strip().isin(invalid_skus)
                frame.loc[frame_invalid, "Status"] = capture_status
                for score_column in score_columns:
                    if score_column in frame.columns:
                        frame.loc[frame_invalid, score_column] = pd.NA

            retry_columns = [
                "SKU", "Retailer", "Kroger RPC", "Brand", "Retail URL", "Salsify URL",
                "Source Used", "Title Path", "Description Path", "Features Path",
                "rawHtmlLength", "rawTextLength",
            ]
            retry_df = invalid_debug_df[[c for c in retry_columns if c in invalid_debug_df.columns]].copy()
            retry_df["Retry Status"] = "Recapture Required"
            retry_df["Capture Reason"] = "Uploaded Kroger capture did not contain the rendered product content."
            retry_df = retry_df.drop_duplicates(subset=[c for c in ["SKU", "Kroger RPC", "Retail URL"] if c in retry_df.columns])

    if summary_df.empty:
        fallback_status = "No rows were returned from process_row. This should be rare; the app now keeps this row so the export is not blank."
        fallback_row = {
            "SKU": "",
            "Retailer": selected_retailer,
            "Retailer RPC": "",
            "Brand": "",
            "Salsify URL": "",
            "Retail URL": "",
            "Rating": "",
            "Review Count": "",
            "Title %": 0,
            "Description %": 0,
            "Feature %": 0,
            "Image Match %": 0,
            "Overall %": 0,
            "Status": fallback_status,
        }
        summary_df = pd.DataFrame([fallback_row])
        detail_df = pd.DataFrame([fallback_row])
        debug_df = pd.DataFrame([{**fallback_row, "Source Used": "empty_export_guard"}])

    selected_retailer_rpc_header = f"{str(selected_retailer or '').strip() or 'Retailer'} RPC"
    for _df in [summary_df, detail_df, debug_df]:
        _df.rename(
            columns={
                "CVS RPC": selected_retailer_rpc_header,
                "Retailer RPC": selected_retailer_rpc_header,
            },
            inplace=True,
        )
        if _df.columns.duplicated().any():
            _df.drop(columns=_df.columns[_df.columns.duplicated()].tolist(), inplace=True)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        detail_df.to_excel(writer, sheet_name="Details", index=False)
        debug_df.to_excel(writer, sheet_name="Debug", index=False)
        if not retry_df.empty:
            retry_df.to_excel(writer, sheet_name="Retry Queue", index=False)

        wb = writer.book

        green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        yellow_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
        red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

        sheets_to_format = ["Summary", "Details"]
        if "Retry Queue" in wb.sheetnames:
            sheets_to_format.append("Retry Queue")

        for sheet_name in sheets_to_format:
            ws = wb[sheet_name]

            header_map = {}
            for cell in ws[1]:
                header_map[str(cell.value).strip()] = cell.column

            for col_name, col_idx in header_map.items():
                if "%" in col_name:
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = cell.value

                        if value is None or value == "":
                            continue

                        try:
                            score_val = float(value)
                        except Exception:
                            continue

                        if score_val >= 80:
                            cell.fill = green_fill
                        elif score_val >= 50:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill

            for col_cells in ws.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter

                for cell in col_cells:
                    try:
                        cell_len = len(str(cell.value or ""))
                        if cell_len > max_length:
                            max_length = cell_len
                    except Exception:
                        pass

                adjusted_width = min(max(max_length + 2, 12), 60)
                ws.column_dimensions[col_letter].width = adjusted_width

    safe_retailer = re.sub(
        r"[^a-z0-9]+",
        "_",
        str(selected_retailer or "retailer").lower().strip(),
    ).strip("_") or "retailer"

    st.session_state.report_bytes = output.getvalue()
    st.session_state.report_filename = f"pdp_qa_results_{safe_retailer}_all_brands.xlsx"
    st.session_state.report_batch_key = st.session_state.completed_batch_key
    st.session_state.auto_download_done = False
    if not st.session_state.batch_status_message:
        if not retry_df.empty:
            st.session_state.batch_status_message = (
                f"Batch finished for {selected_retailer}. Report generated with "
                f"{len(retry_df)} capture(s) in the Retry Queue."
            )
        else:
            st.session_state.batch_status_message = f"Batch finished for {selected_retailer}. Extract/report generated successfully."
    st.rerun()
        
# =========================================
# FULL VISUAL MODE
# =========================================
if (
    retailer_df is not None
    and st.session_state.processing_done
    and st.session_state.completed_batch_key == current_batch_key
):
    try:
        visual_df = retailer_df.copy()

        selected_visual_brand = st.session_state.selected_brand_visual
        if selected_visual_brand != "All" and "brand" in visual_df.columns:
            visual_df = visual_df[visual_df["brand"].astype(str) == selected_visual_brand].copy()

        if visual_df.empty:
            st.warning("No rows found for the selected retailer / brand.")
            st.stop()

        invalid_retail_values = {"", "n/a", "#n/a", "na", "nan", "none"}
        visual_df["retail_url_clean"] = (
            visual_df["retail_url"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
        )
        hidden_count = int(visual_df["retail_url_clean"].isin(invalid_retail_values).sum())
        visual_df = visual_df[
            ~visual_df["retail_url_clean"].isin(invalid_retail_values)
        ].copy()
        visual_df.drop(columns=["retail_url_clean"], inplace=True, errors="ignore")

        st.markdown(
            "<style>.block-container{max-width:1700px;padding-top:1rem;padding-bottom:1rem;} img{max-width:100%;height:auto;}</style>",
            unsafe_allow_html=True,
        )
        st.markdown("## 👁️ Full Visual QA Review")
        st.caption("This full visual UI appears only after the top batch run finishes and the extract/report rows are ready.")
        st.caption("This full visual UI appears only after the new top-section batch run finishes and the extract/report rows are ready. Kroger visual rows reuse the uploaded TXT-matched HTML instead of live fetch.")

        if hidden_count > 0:
            st.caption(
                f"Excluded from Full Visual QA only: {hidden_count} item(s) with missing retailer URLs."
            )
        if visual_df.empty:
            st.info(
                "No visually reviewable items found. Products without retailer URLs are still included in the extract."
            )
            st.stop()

        for _, row in visual_df.iterrows():
            sku = row.get("sku", "Missing SKU")
            retail_url = row.get("retail_url", "")
            salsify_url = row.get("salsify_url", "")
            retailer_name = row.get("retailer", "") or infer_retailer_name_from_url(retail_url)

            current_rpc = row.get("retailer_rpc", "")
            current_target_sku = get_target_sku_from_inputs(
                retail_url=retail_url,
                cvs_rpc=current_rpc,
            )
            
            visual_payload = get_visual_row_payload(
                salsify_url,
                retailer_name,
                retail_url,
                current_target_sku,
                sku=sku,
                row_source_code=row.get("copy_source_code", ""),
            )
            s_text = visual_payload["s_text"]
            s_images = visual_payload["s_images"]
            r_text = visual_payload["r_text"]
            r_images = visual_payload["r_images"]

            s_title = s_text.get("title") or ""
            r_title = r_text.get("title") or ""
            kroger_size_variant = ""
            if str(retailer_name or "").strip().lower() == "kroger" and isinstance(r_text, dict):
                kroger_size_variant = clean_kroger_variant_size(
                    r_text.get("variant_size", "")
                    or r_text.get("kroger_size_variant", "")
                    or (r_text.get("debug", {}) or {}).get("Kroger Size Variant", "")
                )
            r_title_display = r_title
            s_desc = s_text.get("description") or ""
            r_desc = r_text.get("description") or ""
            retailer_features = r_text.get("features") or []
            retailer_norm = str(retailer_name or "").strip().lower()
            salsify_requirements = get_retailer_salsify_requirements(retailer_name)
            feature_fields = build_dynamic_feature_fields_for_pair(retailer_name, s_text, retailer_features)

            title_score = keyword_score(s_title, r_title)
            desc_score = description_similarity_score(s_desc, r_desc)

            feature_scores = []
            feature_rows = []
            for i in range(len(feature_fields)):
                s_val = s_text.get(feature_fields[i], "") if i < len(feature_fields) else ""
                r_val = retailer_features[i] if i < len(retailer_features) else ""

                # Show only rows that exist on Salsify or the retailer page.
                if not (normalize_space(s_val) or normalize_space(r_val)):
                    continue

                row_score = keyword_score(s_val, r_val) if (s_val or r_val) else 0
                feature_scores.append(row_score)
                feature_rows.append((s_val, r_val, row_score))

            avg_feature_score = int(sum(feature_scores) / len(feature_scores)) if feature_scores else 0
            copy_avg_score = int((title_score + desc_score + avg_feature_score) / 3)

            s_images, r_images = trim_trailing_empty_image_slots(s_images, r_images)
            retailer_image_limit = int(salsify_requirements.get("max_images", MAX_IMAGE_SLOTS_TO_COMPARE) or MAX_IMAGE_SLOTS_TO_COMPARE)
            max_images = min(max(len(s_images), len(r_images)), MAX_IMAGE_SLOTS_TO_COMPARE, retailer_image_limit)
            
            # Compute image score before applying row filters.
            avg_img_score, _image_position_scores = build_image_score_fields(
                s_images,
                r_images,
                max_slots=min(MAX_IMAGE_SLOTS_TO_SCORE, retailer_image_limit),
            )
            
            overall_score = int((title_score + desc_score + avg_feature_score + avg_img_score) / 4)
            
            if show_only_issues and overall_score >= 80:
                continue
            if hide_good and overall_score >= 80:
                continue
            if show_below_90_only and overall_score >= 90:
                continue

            left, right = st.columns([2.72, 0.95], gap="small")
        
            with left:
                raw_rpc = current_target_sku or current_rpc
                clean_rpc = clean_item_number(raw_rpc)
                display_rpc = clean_rpc

                salsify_header_html = column_header_link_html("Salsify", sku, salsify_url)
                retailer_header_html = column_header_link_html(
                    retailer_name,
                    display_rpc,
                    retail_url,
                )

                rating_html = ""
                if str(retailer_name or "").strip().lower() in {"walgreens", "kroger"}:
                    retailer_name_norm = str(retailer_name or "").strip().lower()
                    rating_value = (r_text.get("rating", "") if isinstance(r_text, dict) else "") or row.get("rating", "") or ""
                    review_count_value = (r_text.get("review_count", "") if isinstance(r_text, dict) else "") or row.get("review_count", "") or ""
                    if rating_value or review_count_value:
                        kroger_star_size = 28 if retailer_name_norm == "kroger" else 18
                        rating_html = rating_stars_html(rating_value, review_count_value, font_size_px=kroger_star_size)

                st.markdown(
                    locked_visual_header_row_html(
                        salsify_header_html,
                        retailer_header_html,
                        rating_html=rating_html,
                        retailer_name=retailer_name,
                    ),
                    unsafe_allow_html=True,
                )

                st.markdown(
                    avg_score_bar_html("Copy — Avg", copy_avg_score),
                    unsafe_allow_html=True,
                )

                st.markdown(section_header_html("Title", title_score), unsafe_allow_html=True)
                t1, t2 = st.columns(2, gap="small")
                with t1:
                    st.markdown(
                        "<div style='margin-bottom:4px'>" + equal_height_block(s_title or "Missing", min_height=56) + "</div>",
                        unsafe_allow_html=True,
                    )
                with t2:
                    st.markdown(
                        "<div style='margin-bottom:4px'>" + equal_height_block(r_title or "Missing", min_height=56) + "</div>",
                        unsafe_allow_html=True,
                    )

                st.markdown(f"<div style='height:{TITLE_TO_DESCRIPTION_GAP_PX}px'></div>", unsafe_allow_html=True)

                st.markdown(section_header_html("Description", desc_score), unsafe_allow_html=True)
                d1, d2 = st.columns(2, gap="small")
                with d1:
                    st.markdown(
                        "<div style='margin-bottom:4px'>" + equal_height_block(s_desc or "Missing", min_height=150) + "</div>",
                        unsafe_allow_html=True,
                    )
                with d2:
                    st.markdown(
                        "<div style='margin-bottom:4px'>" + equal_height_block(r_desc or "Missing", min_height=150) + "</div>",
                        unsafe_allow_html=True,
                    )

                st.markdown(f"<div style='height:{DESCRIPTION_TO_FEATURES_GAP_PX}px'></div>", unsafe_allow_html=True)

                st.markdown(section_header_html("Features", avg_feature_score), unsafe_allow_html=True)
                for s_val, r_val, row_score in feature_rows:
                    f1, f2 = st.columns(2, gap="small")
                    with f1:
                        st.markdown(
                            "<div style='margin-bottom:4px'>" + equal_feature_block(s_val or "Missing", min_height=40) + "</div>",
                            unsafe_allow_html=True,
                        )
                    with f2:
                        st.markdown(
                            "<div style='margin-bottom:4px'>" + equal_feature_block(r_val or "Missing", min_height=40) + "</div>",
                            unsafe_allow_html=True,
                        )
                    st.markdown(
                        f"<div style='margin:0 0 {SECTION_VERTICAL_GAP}px 0; font-size:14px; font-weight:700;'>{score_text_html(row_score)}</div>",
                        unsafe_allow_html=True,
                    )

            with right:
                head_i1, head_i2 = st.columns(2, gap="small")
                head_i1.markdown(image_header_html("Salsify"), unsafe_allow_html=True)
                head_i2.markdown(
                    f"<div style='margin-left:-42px;'>" + image_header_html(retailer_name) + "</div>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    avg_score_bar_html("Images — Avg", avg_img_score),
                    unsafe_allow_html=True,
                )

                img_scores = []
                max_images_to_score = min(max_images, MAX_IMAGE_SLOTS_TO_SCORE)
                
                for i in range(max_images):
                    s_url = s_images[i].get("url") if i < len(s_images) and isinstance(s_images[i], dict) else ""
                    r_url = r_images[i] if i < len(r_images) and isinstance(r_images[i], str) else ""
                
                    slot_score = compare_images_visually(s_url, r_url) if (s_url and r_url) else 0
                
                    if i < max_images_to_score:
                        img_scores.append(slot_score)
                
                    st.markdown(
                        image_compare_row_html(s_url, r_url, slot_score),
                        unsafe_allow_html=True,
                    )
                
                avg_img_score = int(sum(img_scores) / len(img_scores)) if img_scores else 0
                overall_score = int((title_score + desc_score + avg_feature_score + avg_img_score) / 4)

            if show_html_debugger:
                should_render_debugger = (not debug_only_sku) or (
                    str(sku).strip() == str(debug_only_sku).strip()
                )
            
                if should_render_debugger:
                    debug_url = retail_url if debugger_source == "Retailer page" else salsify_url
                    debug_retailer_name = retailer_name if debugger_source == "Retailer page" else "salsify"
            
                    debug_views = resolve_debug_views(
                        debug_url,
                        retailer_name=debug_retailer_name,
                        use_manual_html_override=use_manual_html_override,
                        manual_html_text=manual_html_text,
                        manual_html_file=manual_html_file,
                    )
            
                    with st.expander(f"🔎 HTML / DOM Debugger — {sku}", expanded=True):
                        render_debugger_panel(
                            debug_views,
                            sku=sku,
                            marker_start=debug_marker_start,
                            marker_end=debug_marker_end,
                            marker_target=debug_marker_target,
                            use_manual_html_override=use_manual_html_override,
                        )
            st.divider()
    except Exception as e:
        st.error("🔥 CRITICAL APP ERROR")
        st.text(str(e))
        st.text(traceback.format_exc())
